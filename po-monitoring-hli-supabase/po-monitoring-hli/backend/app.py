from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
import pandas as pd
import numpy as np
import re
import os
import json
import html
import hashlib
from datetime import datetime, date, timedelta
import io
import time
import threading
from sqlalchemy import func, text, event, case, desc
from sqlalchemy.engine import Engine
from sqlalchemy.orm import load_only
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from apscheduler.schedulers.background import BackgroundScheduler
    from apscheduler.triggers.cron import CronTrigger
    _APSCHEDULER_AVAILABLE = True
except ImportError:
    _APSCHEDULER_AVAILABLE = False
    print('[scheduler] APScheduler not installed – auto copy-sheet disabled.')

app = Flask(__name__)

_HOLIDAY_CACHE = None
_HOLIDAY_CACHE_KEY = None
_HOLIDAY_LOCK = threading.Lock()
_READ_CACHE_LOCK = threading.Lock()
_COMPLETED_CACHE_LOCK = threading.Lock()
_RATE_CACHE_LOCK = threading.Lock()
_SIMILARITY_LOCK = threading.Lock()
_MASTER_PIC_LOCK = threading.Lock()

_COMPLETED_SUMMARY_CACHE = {}
_COMPLETED_SUMMARY_CACHE_TTL_SECONDS = 900
_RUNTIME_CACHE_VERSION = 0

_SIMILARITY_CACHE = {}
_SIMILARITY_CACHE_FILE = os.path.join(os.path.dirname(__file__), 'instance', 'similarity_cache.json')
_MASTER_PIC_CACHE = {'signature': None, 'by_id': {}, 'by_name': {}}

_PID_CATEGORY_CACHE = {}
_PID_CATEGORY_CACHE_LOADED = False

def _pid_category_cache_load():
    global _PID_CATEGORY_CACHE, _PID_CATEGORY_CACHE_LOADED
    mapping = {}
    try:
        for pid, cat in db.session.query(ProductIDDB.product_id, ProductIDDB.category_name).all():
            if not pid: continue
            raw = (cat or '').strip()
            mapping[str(pid).strip()] = raw.split('>')[0].strip() if '>' in raw else raw
    except Exception:
        mapping = {}
    with _MASTER_PIC_LOCK:
        _PID_CATEGORY_CACHE = mapping
        _PID_CATEGORY_CACHE_LOADED = True

def _pid_category_lookup(product_id):
    global _PID_CATEGORY_CACHE_LOADED
    with _MASTER_PIC_LOCK:
        loaded = _PID_CATEGORY_CACHE_LOADED
    if not loaded:
        _pid_category_cache_load()
    pid = str(product_id or '').strip()
    with _MASTER_PIC_LOCK:
        return _PID_CATEGORY_CACHE.get(pid, '')

def _pid_category_cache_invalidate():
    global _PID_CATEGORY_CACHE_LOADED
    with _MASTER_PIC_LOCK:
        _PID_CATEGORY_CACHE_LOADED = False

_READ_RESPONSE_CACHE = {}

def runtime_cache_get(key):
    with _READ_CACHE_LOCK:
        item = _READ_RESPONSE_CACHE.get(key)
        if not item: return None
        expires_at, payload = item
        if expires_at <= datetime.utcnow():
            _READ_RESPONSE_CACHE.pop(key, None)
            return None
        return payload

def runtime_cache_set(key, payload, ttl_seconds=20):
    with _READ_CACHE_LOCK:
        _READ_RESPONSE_CACHE[key] = (datetime.utcnow() + timedelta(seconds=ttl_seconds), payload)

def runtime_cache_key(namespace):
    return (namespace, request.query_string.decode('utf-8', errors='ignore'))

def clear_runtime_caches():
    global _RUNTIME_CACHE_VERSION
    with _READ_CACHE_LOCK:
        _RUNTIME_CACHE_VERSION += 1
        _READ_RESPONSE_CACHE.clear()
    with _COMPLETED_CACHE_LOCK:
        _COMPLETED_SUMMARY_CACHE.clear()
    try: RFQ_CACHE['expires_at'] = None
    except NameError: pass
    try: VENDOR_CONTROL_CACHE['expires_at'] = None
    except NameError: pass

def _holiday_set():
    global _HOLIDAY_CACHE, _HOLIDAY_CACHE_KEY
    today_year = date.today().year
    cache_key = today_year
    if _HOLIDAY_CACHE is not None and _HOLIDAY_CACHE_KEY == cache_key:
        return _HOLIDAY_CACHE
    years = list(range(today_year - 2, today_year + 2))
    try:
        import holidays as _holidays_pkg
        s = set(_holidays_pkg.country_holidays('ID', years=years).keys())
    except Exception:
        s = set()
    extras_path = os.path.join(os.path.dirname(__file__), 'holiday_extras.json')
    try:
        with open(extras_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        items = data.get('dates', []) if isinstance(data, dict) else data
        for ds in items or []:
            try: s.add(date.fromisoformat(str(ds).strip()))
            except (ValueError, TypeError): pass
    except Exception: pass
    _HOLIDAY_CACHE = s
    _HOLIDAY_CACHE_KEY = cache_key
    return s

_HOLIDAY_ARRAY_CACHE = None
_HOLIDAY_ARRAY_CACHE_KEY = None

def _holiday_array():
    global _HOLIDAY_ARRAY_CACHE, _HOLIDAY_ARRAY_CACHE_KEY
    holiday_set = _holiday_set()
    cache_key = (_HOLIDAY_CACHE_KEY, len(holiday_set))
    if _HOLIDAY_ARRAY_CACHE is not None and _HOLIDAY_ARRAY_CACHE_KEY == cache_key:
        return _HOLIDAY_ARRAY_CACHE
    arr = (np.array(sorted(holiday_set), dtype='datetime64[D]') if holiday_set else np.array([], dtype='datetime64[D]'))
    _HOLIDAY_ARRAY_CACHE = arr
    _HOLIDAY_ARRAY_CACHE_KEY = cache_key
    return arr

def is_workday(d):
    return d.weekday() < 5 and d not in _holiday_set()

def count_workdays(start, end):
    if start is None or end is None: return None
    if start == end: return 0
    holidays = _holiday_array()
    if end > start: return int(np.busday_count(start, end, holidays=holidays))
    return -int(np.busday_count(end, start, holidays=holidays))

def workdays_since(past_date, today=None):
    if past_date is None: return None
    if today is None: today = date.today()
    return count_workdays(past_date, today)

def workdays_until(future_date, today=None):
    if future_date is None: return None
    if today is None: today = date.today()
    return count_workdays(today, future_date)

CORS(app, resources={r"/api/*": {
    "origins": "*",
    "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    "allow_headers": ["Content-Type", "Authorization", "Accept"]
}})

_db_url = os.environ.get('DATABASE_URL', '')
if _db_url:
    if _db_url.startswith('postgres://'):
        _db_url = _db_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = _db_url
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'pool_pre_ping': True, 'pool_recycle': 300, 'pool_size': 10, 'max_overflow': 20,
        'connect_args': {'connect_timeout': 10},
    }
else:
    _inst = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance')
    os.makedirs(_inst, exist_ok=True)
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{_inst}/po_database.db'
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'pool_pre_ping': True,
        'connect_args': {'timeout': 30},
    }

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024
db = SQLAlchemy(app)

@event.listens_for(Engine, 'connect')
def _set_sqlite_pragmas(dbapi_connection, connection_record):
    if 'sqlite' in app.config.get('SQLALCHEMY_DATABASE_URI', ''):
        cursor = dbapi_connection.cursor()
        cursor.execute('PRAGMA journal_mode=WAL')
        cursor.execute('PRAGMA synchronous=NORMAL')
        cursor.execute('PRAGMA busy_timeout=30000')
        cursor.execute('PRAGMA temp_store=MEMORY')
        cursor.execute('PRAGMA cache_size=-65536')
        cursor.execute('PRAGMA wal_autocheckpoint=1000')
        cursor.close()
        try:
            dbapi_connection.create_function('REGEXP', 2, lambda pattern, value: 1 if (value is not None and re.search(pattern, str(value))) else 0)
        except Exception: pass

class SOData(db.Model):
    __tablename__ = 'so_data'
    id = db.Column(db.Integer, primary_key=True)
    so_number = db.Column(db.String(50), index=True)
    so_item = db.Column(db.String(100))
    so_status = db.Column(db.String(50))
    operation_unit_name = db.Column(db.String(200))
    client_id = db.Column(db.String(100), index=True)
    vendor_id = db.Column(db.String(100))
    vendor_name = db.Column(db.String(200))
    customer_po_number = db.Column(db.String(200))
    delivery_memo = db.Column(db.Text)
    product_name = db.Column(db.Text)
    specification = db.Column(db.Text)
    manufacturer_name = db.Column(db.String(300))
    product_id = db.Column(db.String(100))
    so_qty = db.Column(db.Float)
    sales_unit = db.Column(db.String(20))
    sales_price = db.Column(db.Float)
    sales_amount = db.Column(db.Float)
    currency = db.Column(db.String(10))
    purchasing_price = db.Column(db.Float)
    purchasing_amount = db.Column(db.Float)
    purchasing_currency = db.Column(db.String(10))
    purchasing_amount_idr = db.Column(db.Float)
    purchasing_amount_idr_cached_at = db.Column(db.DateTime)
    so_create_date = db.Column(db.Date)
    delivery_possible_date = db.Column(db.Date)
    matched_po_number = db.Column(db.String(50))
    delivery_plan_date = db.Column(db.Date)
    remarks = db.Column(db.Text)
    pic_name = db.Column(db.String(100))
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)

class UploadLog(db.Model):
    __tablename__ = 'upload_log'
    id = db.Column(db.Integer, primary_key=True)
    file_type = db.Column(db.String(50))
    filename = db.Column(db.String(255))
    records_count = db.Column(db.Integer)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)

class ExchangeRate(db.Model):
    __tablename__ = 'exchange_rate'
    id = db.Column(db.Integer, primary_key=True)
    rate_date = db.Column(db.Date, nullable=False, unique=True, index=True)
    usd_to_idr = db.Column(db.Float, nullable=False)
    source = db.Column(db.String(50), default='manual')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class ProductIDDB(db.Model):
    __tablename__ = 'product_id_db'
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.String(100), unique=True, nullable=False, index=True)
    category_id = db.Column(db.String(100))
    category_name = db.Column(db.String(255))
    product_name = db.Column(db.Text)
    product_status = db.Column(db.String(100))
    specification = db.Column(db.Text)
    manufacturer_name = db.Column(db.String(255))
    vendor_name = db.Column(db.String(300))
    order_unit = db.Column(db.String(50))
    hub_handling_check = db.Column(db.String(100))
    tax_type = db.Column(db.String(100))
    registration_date = db.Column(db.Date, index=True)
    product_registry_pic = db.Column(db.String(200))
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)

class MasterPIC(db.Model):
    __tablename__ = 'master_pic'
    id = db.Column(db.Integer, primary_key=True)
    category_id = db.Column(db.String(100), unique=True, nullable=False, index=True)
    category_name = db.Column(db.String(255))
    pic_name = db.Column(db.String(100))
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)

class MasterClientPIC(db.Model):
    __tablename__ = 'master_client_pic'
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.String(100), unique=True, nullable=False, index=True)
    client_name = db.Column(db.String(300))
    pic_name = db.Column(db.String(100))
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)

class MasterVendorPIC(db.Model):
    __tablename__ = 'master_vendor_pic'
    id = db.Column(db.Integer, primary_key=True)
    vendor_id = db.Column(db.String(100), unique=True, nullable=False, index=True)
    vendor_name = db.Column(db.String(300))
    pic_name = db.Column(db.String(100))
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)

class MasterBidTypePIC(db.Model):
    __tablename__ = 'master_bid_type_pic'
    id = db.Column(db.Integer, primary_key=True)
    bid_type = db.Column(db.String(255), unique=True, nullable=False, index=True)
    pic_name = db.Column(db.String(100))
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)

class ItemRegistration(db.Model):
    __tablename__ = 'item_registration'
    id = db.Column(db.Integer, primary_key=True)
    proc_status = db.Column(db.String(100))
    req_date = db.Column(db.Date, index=True)
    existing_owner = db.Column(db.String(100))
    client_id = db.Column(db.String(100), index=True)
    client_name = db.Column(db.String(300), index=True)
    operation_unit_name = db.Column(db.String(300), index=True)
    category = db.Column(db.String(255))
    category_id = db.Column(db.String(100))
    pic = db.Column(db.String(200))
    pic_name = db.Column(db.String(200))
    req_no = db.Column(db.String(100), index=True)
    prod_id = db.Column(db.String(100), index=True)
    product_status = db.Column(db.String(100))
    batch_grp_no = db.Column(db.String(100))
    prod_name = db.Column(db.Text)
    spec = db.Column(db.Text)
    mfr_name = db.Column(db.String(300))
    odr_unit = db.Column(db.String(50))
    bid_except_type = db.Column(db.String(255))
    vendor_name = db.Column(db.String(300))
    vendor_id = db.Column(db.String(100), index=True)
    prod_price = db.Column(db.Float)
    curr = db.Column(db.String(20))
    hub_handling_check = db.Column(db.String(100))
    tax_type = db.Column(db.String(50))
    registration_date = db.Column(db.Date, index=True)
    product_registry_pic = db.Column(db.String(200))
    remarks = db.Column(db.Text)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)

class RFQCellEdit(db.Model):
    __tablename__ = 'rfq_cell_edit'
    id = db.Column(db.Integer, primary_key=True)
    row_key = db.Column(db.String(200), nullable=False, index=True)
    field = db.Column(db.String(100), nullable=False)
    value = db.Column(db.Text)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)
    __table_args__ = (db.UniqueConstraint('row_key', 'field', name='uq_rfq_cell_edit_row_field'),)

class RFQDashboardRow(db.Model):
    __tablename__ = 'rfq_dashboard_row'
    id = db.Column(db.Integer, primary_key=True)
    row_key = db.Column(db.String(200), unique=True, nullable=False, index=True)
    sheet_row = db.Column(db.Integer, index=True)
    data_json = db.Column(db.Text, nullable=False, default='{}')
    dirty_fields_json = db.Column(db.Text, nullable=False, default='[]')
    first_seen_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    last_seen_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)

class ImportVendor(db.Model):
    __tablename__ = 'import_vendor'
    id = db.Column(db.Integer, primary_key=True)
    vendor_name = db.Column(db.String(300), unique=True, nullable=False, index=True)
    origin = db.Column(db.String(100))
    top = db.Column(db.String(50))
    non_ski = db.Column(db.String(50))
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)

class ImportDashboardRow(db.Model):
    __tablename__ = 'import_dashboard_row'
    id = db.Column(db.Integer, primary_key=True)
    row_key = db.Column(db.String(120), unique=True, nullable=False, index=True)
    source_key = db.Column(db.String(50), nullable=False, index=True)
    source_label = db.Column(db.String(100))
    source_uid = db.Column(db.String(50), nullable=False, index=True)
    sheet_row = db.Column(db.Integer)
    vendor_name = db.Column(db.String(300), index=True)
    data_json = db.Column(db.Text, nullable=False, default='{}')
    first_seen_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    last_seen_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)

class ImportDashboardMeta(db.Model):
    __tablename__ = 'import_dashboard_meta'
    id = db.Column(db.Integer, primary_key=True)
    meta_key = db.Column(db.String(100), unique=True, nullable=False, index=True)
    value_json = db.Column(db.Text, nullable=False, default='null')
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)

_RATE_CACHE = {}
_FX_RATE_CACHE = {}

def _fetch_rate_from_api(d, currency='USD'):
    try:
        import urllib.request, json as _json
        cur = (currency or 'USD').strip().upper()
        url = f"https://api.frankfurter.dev/v2/rate/{cur}/IDR?date={d.isoformat()}"
        with urllib.request.urlopen(url, timeout=6) as resp:
            data = _json.loads(resp.read())
        return float(data['rate'])
    except Exception: return None

def _get_fallback_rate():
    last = ExchangeRate.query.order_by(ExchangeRate.rate_date.desc()).first()
    return last.usd_to_idr if last else 16000.0

def get_usd_to_idr(d, cache_only=False):
    if d is None: return _get_fallback_rate()
    if d in _RATE_CACHE: return _RATE_CACHE[d]
    rec = ExchangeRate.query.filter_by(rate_date=d).first()
    if rec:
        _RATE_CACHE[d] = rec.usd_to_idr
        return rec.usd_to_idr
    if not cache_only and d <= date.today():
        rate = _fetch_rate_from_api(d)
        if rate:
            try:
                db.session.add(ExchangeRate(rate_date=d, usd_to_idr=rate, source='frankfurter'))
                db.session.commit()
            except Exception: db.session.rollback()
            _RATE_CACHE[d] = rate
            return rate
    nearest = ExchangeRate.query.order_by(func.abs(func.julianday(ExchangeRate.rate_date) - func.julianday(str(d)))).first()
    if nearest:
        _RATE_CACHE[d] = nearest.usd_to_idr
        return nearest.usd_to_idr
    return _get_fallback_rate()

def get_currency_to_idr(currency, d, cache_only=False):
    cur = (currency or 'IDR').strip().upper()
    if cur in ('IDR', ''): return 1.0
    if cur == 'USD': return get_usd_to_idr(d, cache_only=cache_only)
    if d is None: d = date.today()
    key = (cur, d)
    if key in _FX_RATE_CACHE: return _FX_RATE_CACHE[key]
    if not cache_only and d <= date.today():
        rate = _fetch_rate_from_api(d, cur)
        if rate:
            _FX_RATE_CACHE[key] = rate
            return rate
    same_currency_rates = [(rate_date, rate) for (fx_cur, rate_date), rate in _FX_RATE_CACHE.items() if fx_cur == cur]
    if same_currency_rates:
        _nearest_date, nearest_rate = min(same_currency_rates, key=lambda r: abs((r[0] - d).days))
        _FX_RATE_CACHE[key] = nearest_rate
        return nearest_rate
    fallback_rate = _fetch_rate_from_api(date.today(), cur) if not cache_only else None
    if fallback_rate:
        _FX_RATE_CACHE[key] = fallback_rate
        return fallback_rate
    return _get_fallback_rate()

# Cap on parallel FX API calls per prefetch — prevents hammering the upstream
# frankfurter API when a large upload introduces many new unique dates.
_FX_PARALLEL_MAX = 8
_FX_FETCH_CAP_PER_CALL = 40  # max unique dates to fetch in a single prefetch call


def _fetch_rates_parallel(dates_sorted, currency='USD'):
    """Fetch multiple FX rates concurrently. Returns dict {date: rate}."""
    if not dates_sorted:
        return {}
    from concurrent.futures import ThreadPoolExecutor, as_completed
    results = {}
    # Cap to avoid runaway API calls
    capped = list(dates_sorted)[:_FX_FETCH_CAP_PER_CALL]
    try:
        with ThreadPoolExecutor(max_workers=_FX_PARALLEL_MAX) as ex:
            futures = {ex.submit(_fetch_rate_from_api, d, currency): d for d in capped}
            for fut in as_completed(futures):
                d = futures[fut]
                try:
                    rate = fut.result()
                    if rate is not None:
                        results[d] = rate
                except Exception:
                    pass
    except Exception:
        # Fallback to sequential if thread pool fails
        for d in capped:
            rate = _fetch_rate_from_api(d, currency)
            if rate is not None:
                results[d] = rate
    return results


def prefetch_exchange_rates(dates, fetch_missing=True, currency='USD'):
    cur = (currency or 'USD').strip().upper()
    if not dates or cur in ('IDR', ''): return
    if cur != 'USD':
        needed = {d for d in dates if d is not None and (cur, d) not in _FX_RATE_CACHE}
        if fetch_missing:
            today = date.today()
            to_fetch = sorted(x for x in needed if x <= today)
            if to_fetch:
                rates = _fetch_rates_parallel(to_fetch, cur)
                for d, rate in rates.items():
                    _FX_RATE_CACHE[(cur, d)] = rate
                    needed.discard(d)
        if needed:
            same_currency_rates = [(rate_date, rate) for (fx_cur, rate_date), rate in _FX_RATE_CACHE.items() if fx_cur == cur]
            fallback = get_currency_to_idr(cur, date.today(), cache_only=not fetch_missing)
            for d in needed:
                if same_currency_rates:
                    _nearest_date, nearest_rate = min(same_currency_rates, key=lambda r: abs((r[0] - d).days))
                    _FX_RATE_CACHE[(cur, d)] = nearest_rate
                else:
                    _FX_RATE_CACHE[(cur, d)] = fallback
        return
    needed = {d for d in dates if d is not None and d not in _RATE_CACHE}
    if not needed: return
    db_rows = ExchangeRate.query.filter(ExchangeRate.rate_date.in_(list(needed))).all()
    for row in db_rows:
        _RATE_CACHE[row.rate_date] = row.usd_to_idr
    needed -= {row.rate_date for row in db_rows}
    if not needed: return
    if fetch_missing:
        today = date.today()
        to_api = sorted(d for d in needed if d <= today)
        if to_api:
            rates = _fetch_rates_parallel(to_api, 'USD')
            fetched_rows = []
            for d, rate in rates.items():
                _RATE_CACHE[d] = rate
                fetched_rows.append(ExchangeRate(rate_date=d, usd_to_idr=rate, source='frankfurter'))
                needed.discard(d)
            if fetched_rows:
                try:
                    db.session.bulk_save_objects(fetched_rows)
                    db.session.commit()
                except Exception: db.session.rollback()
    if needed:
        fallback = _get_fallback_rate()
        all_rates = ExchangeRate.query.order_by(ExchangeRate.rate_date).all()
        for d in needed:
            if all_rates:
                nearest = min(all_rates, key=lambda r: abs((r.rate_date - d).days))
                _RATE_CACHE[d] = nearest.usd_to_idr
            else:
                _RATE_CACHE[d] = fallback

def convert_to_idr(amount, currency, rate_date=None, cache_only=False):
    if not amount: return 0.0
    cur = (currency or 'IDR').strip().upper()
    if cur in ('IDR', ''): return float(amount)
    if cur in ('USD', 'EUR'): return float(amount) * get_currency_to_idr(cur, rate_date, cache_only=cache_only)
    return float(amount)

def prefetch_convertible_exchange_rates(rows, fetch_missing=False):
    for currency in ('USD', 'EUR'):
        dates = {s.so_create_date for s in rows if s.so_create_date and (s.purchasing_currency or '').strip().upper() == currency}
        prefetch_exchange_rates(dates, fetch_missing=fetch_missing, currency=currency)

def raw_purchase_amount(s):
    raw = float(s.purchasing_amount or 0)
    if raw == 0 and s.purchasing_price: raw = float(s.purchasing_price) * float(s.so_qty or 0)
    return raw

def is_purchase_valid(s):
    """Margin is only valid when purchase price/amount is present AND positive.
    Empty, zero, or negative purchase → margin = None (invalid for KPIs).
    Used across the entire dashboard to keep margin KPIs accurate."""
    try:
        return raw_purchase_amount(s) > 0
    except Exception:
        return False

def purchase_price_idr(s):
    """Convert purchasing_price to IDR using cached exchange rates.
    If currency is already IDR (or empty), return as-is.
    If USD/EUR, convert using the rate for the SO create date.
    Returns 0 if no exchange rate is available (cache_only)."""
    price = float(s.purchasing_price or 0)
    if price <= 0: return 0.0
    cur = (s.purchasing_currency or 'IDR').strip().upper()
    if cur in ('IDR', ''): return price
    if cur in ('USD', 'EUR'):
        rate = get_currency_to_idr(cur, s.so_create_date, cache_only=True)
        return price * rate
    return price  # unknown currency — return as-is

def purchase_amount_idr_for_margin(s):
    """Purchase amount in IDR, used for ALL margin calculations.
    Priority:
    1. purchasing_amount_idr (cached column) if present
    2. Convert raw_purchase_amount to IDR using cached rates
    If currency is IDR → use raw amount directly.
    Returns 0.0 if conversion not possible (no rate cached)."""
    cached = getattr(s, 'purchasing_amount_idr', None)
    if cached is not None and cached > 0: return float(cached)
    raw = raw_purchase_amount(s)
    if raw <= 0: return 0.0
    cur = (s.purchasing_currency or 'IDR').strip().upper()
    if cur in ('IDR', ''): return raw
    if cur in ('USD', 'EUR'):
        rate = get_currency_to_idr(cur, s.so_create_date, cache_only=True)
        return raw * rate
    return raw  # unknown currency — return as-is

def purchase_amount_idr(s, allow_persist=False):
    cached = getattr(s, 'purchasing_amount_idr', None)
    cur = (s.purchasing_currency or 'IDR').strip().upper()
    if cached is not None: return float(cached)
    raw = raw_purchase_amount(s)
    if cur in ('IDR', ''): return raw
    if not allow_persist: return 0.0
    converted = convert_to_idr(raw, s.purchasing_currency, s.so_create_date, cache_only=True)
    s.purchasing_amount_idr = converted
    s.purchasing_amount_idr_cached_at = datetime.utcnow()
    return converted

def dashboard_purchase_sql_expr():
    currency_expr = func.upper(func.trim(func.coalesce(SOData.purchasing_currency, '')))
    raw_purchase_expr = case(
        (func.coalesce(SOData.purchasing_amount, 0) != 0, func.coalesce(SOData.purchasing_amount, 0)),
        else_=func.coalesce(SOData.purchasing_price, 0) * func.coalesce(SOData.so_qty, 0),
    )
    return case(
        (SOData.purchasing_amount_idr.isnot(None), SOData.purchasing_amount_idr),
        (currency_expr.in_(['', 'IDR']), raw_purchase_expr),
        else_=0.0,
    )

def ensure_purchase_amount_idr_cache(rows, fetch_missing=False):
    if not fetch_missing: return 0
    missing = [s for s in rows if getattr(s, 'purchasing_amount_idr', None) is None and (s.purchasing_currency or '').strip().upper() in ('USD', 'EUR') and raw_purchase_amount(s) > 0]
    if not missing: return 0
    prefetch_convertible_exchange_rates(missing, fetch_missing=True)
    for s in missing: purchase_amount_idr(s, allow_persist=True)
    try: db.session.commit()
    except Exception: db.session.rollback(); raise
    return len(missing)

def _ensure_extra_columns():
    is_sqlite = 'sqlite' in app.config['SQLALCHEMY_DATABASE_URI']
    def existing_columns(table_name):
        try:
            if is_sqlite:
                result = db.session.execute(text(f"PRAGMA table_info({table_name})"))
                return {row[1].lower() for row in result}
            result = db.session.execute(text("SELECT column_name FROM information_schema.columns WHERE table_name = '{}'".format(table_name)))
            return {row[0].lower() for row in result}
        except Exception: return set()
    migration_plan = {
        'so_data': [('specification', 'TEXT'), ('product_id', 'VARCHAR(100)'), ('vendor_id', 'VARCHAR(100)'), ('client_id', 'VARCHAR(100)'), ('manufacturer_name', 'VARCHAR(300)'), ('purchasing_currency', 'VARCHAR(10)'), ('purchasing_amount_idr', 'DOUBLE PRECISION'), ('purchasing_amount_idr_cached_at', 'TIMESTAMP'), ('pic_name', 'VARCHAR(100)')],
        'item_registration': [('req_date', 'DATE'), ('existing_owner', 'VARCHAR(100)'), ('client_id', 'VARCHAR(100)'), ('operation_unit_name', 'VARCHAR(300)'), ('bid_except_type', 'VARCHAR(255)'), ('category_id', 'VARCHAR(100)'), ('pic_name', 'VARCHAR(200)'), ('product_status', 'VARCHAR(100)'), ('hub_handling_check', 'VARCHAR(100)'), ('tax_type', 'VARCHAR(50)'), ('registration_date', 'DATE'), ('product_registry_pic', 'VARCHAR(200)'), ('remarks', 'TEXT'), ('vendor_id', 'VARCHAR(100)')],
        'product_id_db': [('specification', 'TEXT'), ('manufacturer_name', 'VARCHAR(255)'), ('vendor_name', 'VARCHAR(300)'), ('order_unit', 'VARCHAR(50)'), ('product_status', 'VARCHAR(100)'), ('hub_handling_check', 'VARCHAR(100)'), ('tax_type', 'VARCHAR(100)'), ('registration_date', 'DATE'), ('product_registry_pic', 'VARCHAR(200)')],
        'import_vendor': [('origin', 'VARCHAR(100)'), ('top', 'VARCHAR(50)'), ('non_ski', 'VARCHAR(50)')],
    }
    for table_name, columns in migration_plan.items():
        cols = existing_columns(table_name)
        for col_name, col_type in columns:
            if col_name.lower() not in cols:
                try:
                    db.session.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {col_name} {col_type}"))
                    db.session.commit()
                except Exception: db.session.rollback()

def _ensure_performance_indexes():
    statements = [
        "CREATE INDEX IF NOT EXISTS idx_so_status_date ON so_data (so_status, so_create_date)",
        "CREATE INDEX IF NOT EXISTS idx_so_op_unit ON so_data (operation_unit_name)",
        "CREATE INDEX IF NOT EXISTS idx_so_pic_name ON so_data (pic_name)",
        "CREATE INDEX IF NOT EXISTS idx_so_vendor_name ON so_data (vendor_name)",
        "CREATE INDEX IF NOT EXISTS idx_so_item ON so_data (so_item)",
        "CREATE INDEX IF NOT EXISTS idx_so_number ON so_data (so_number)",
        "CREATE INDEX IF NOT EXISTS idx_so_product_id ON so_data (product_id)",
        "CREATE INDEX IF NOT EXISTS idx_so_customer_po ON so_data (customer_po_number)",
        "CREATE INDEX IF NOT EXISTS idx_upload_log_type_date ON upload_log (file_type, uploaded_at)",
        "CREATE INDEX IF NOT EXISTS idx_item_reg_proc_client ON item_registration (proc_status, client_name)",
        "CREATE INDEX IF NOT EXISTS idx_item_reg_pic ON item_registration (pic)",
        "CREATE INDEX IF NOT EXISTS idx_item_reg_req_no ON item_registration (req_no)",
        "CREATE INDEX IF NOT EXISTS idx_item_reg_mfr ON item_registration (mfr_name)",
        "CREATE INDEX IF NOT EXISTS idx_item_reg_owner ON item_registration (existing_owner)",
        "CREATE INDEX IF NOT EXISTS idx_product_status_unit ON product_id_db (product_status, order_unit)",
    ]
    for stmt in statements:
        try: db.session.execute(text(stmt))
        except Exception: db.session.rollback()
    try: db.session.commit()
    except Exception: db.session.rollback()

try:
    with app.app_context():
        db.create_all()
        _ensure_extra_columns()
        _ensure_performance_indexes()
        print('DB schema ready.')
except Exception as exc:
    print(f'DB schema setup skipped at startup: {exc}')

CLOSED_STATUSES = {'Delivery Completed', 'SO Cancel', 'Approval Apply', 'Approval Complete', 'Approval Complete Step', 'Approval Reject', 'Approval Hold', 'Return Complete(Vendor)', 'Return Complete(HUB)', 'Customer PO Reject'}
DISCARDABLE_STATUSES = {'SO Cancel', 'Approval Apply', 'Approval Complete Step', 'Approval Reject', 'Approval Hold', 'Return Complete(Vendor)', 'Return Complete(HUB)', 'Customer PO Reject', 'Ship. Order Reject', 'PO Received Reject'}
EXCLUDED_OP_UNITS = {'HLI GREEN POWER (CONSUMABLE)'}

PO_HLI_RE = re.compile(r'(\d{7,})(?:-(\d{1,4}))?(?!\d)')
PO_SHORT_REF_RE = re.compile(r'\bP\s*\.?\s*O\s*\.?\s*[#:.\-]?\s*(\d{2,6})\b', re.IGNORECASE)

def _normalize_item_no(item_no):
    if item_no is None: return set()
    s = str(item_no).strip()
    variants = {s}
    if s.endswith('.0'):
        s = s[:-2]; variants.add(s)
    try:
        n = int(float(s))
        variants.add(str(n)); variants.add(f"{n:02d}"); variants.add(f"{n:03d}")
    except: pass
    return variants

def extract_po_hli(val):
    if not val: return []
    result = set()
    for m in PO_HLI_RE.finditer(str(val).strip()):
        po_num, item_no = m.group(1), m.group(2)
        if po_num.startswith('2'): continue
        result.add(po_num)
        if item_no:
            for item_var in _normalize_item_no(item_no): result.add(f"{po_num}-{item_var}")
    return list(result)

def extract_po_short_refs(val):
    if not val: return []
    refs = set()
    for m in PO_SHORT_REF_RE.finditer(str(val).strip()):
        n = m.group(1)
        if len(n) >= 7: continue
        refs.add(n)
    return list(refs)

def open_so_filter():
    return db.or_(SOData.so_status.is_(None), SOData.so_status.notin_(list(CLOSED_STATUSES)))

def parse_so_date_args(args=None):
    args = args if args is not None else request.args
    date_year = args.get('date_year', '')
    date_from = args.get('date_from', '')
    date_to = args.get('date_to', '')
    if not date_year:
        legacy = args.get('year', '')
        if legacy and legacy != 'all': date_year = legacy
    return date_year, date_from, date_to

def apply_so_create_date_filter(query, date_year='', date_from='', date_to='', is_sqlite=None):
    if date_year:
        try:
            yr = int(date_year)
            start_date = date(yr, 1, 1)
            end_date = date(yr, 12, 31)
            return query.filter(SOData.so_create_date >= start_date, SOData.so_create_date <= end_date)
        except (ValueError, TypeError): pass
    if date_from: query = query.filter(SOData.so_create_date >= date_from)
    if date_to: query = query.filter(SOData.so_create_date <= date_to)
    return query

def apply_item_registration_date_filter(query, date_year='', date_from='', date_to=''):
    if date_year:
        try:
            yr = int(date_year)
            start_date = date(yr, 1, 1)
            end_date = date(yr, 12, 31)
            return query.filter(ItemRegistration.req_date >= start_date, ItemRegistration.req_date <= end_date)
        except (ValueError, TypeError): pass
    df = parse_date(date_from) if date_from else None
    dt = parse_date(date_to) if date_to else None
    if df: query = query.filter(ItemRegistration.req_date >= df)
    if dt: query = query.filter(ItemRegistration.req_date <= dt)
    return query

def utc_isoformat(dt):
    """Format datetime ke ISO 8601 dengan timezone UTC (+00:00 suffix).

    FIX V7: Sebelumnya hanya menambah 'Z' tanpa cek apakah datetime sudah punya
    tzinfo. Backend pakai datetime.utcnow() yang TIDAK punya tzinfo, jadi
    isoformat() menghasilkan "2026-06-25T01:04:00" tanpa suffix.
    Frontend lalu meng-interpret sebagai local time browser → jam tampil salah.

    Sekarang: selalu tambahkan '+00:00' (UTC offset) kalau datetime naive
    (tanpa tzinfo), supaya frontend tahu ini adalah UTC time.
    """
    if dt is None: return None
    # Kalau datetime sudah punya tzinfo, pakai apa adanya
    if dt.tzinfo is not None:
        return dt.isoformat()
    # Kalau naive (tidak punya tzinfo), asumsikan UTC (karena backend pakai utcnow())
    # Tambahkan +00:00 suffix supaya frontend bisa konversi ke local time dengan benar
    return dt.isoformat() + '+00:00'

def has_internal_po_ref(customer_po_number, delivery_memo):
    for field in [customer_po_number, delivery_memo]:
        if not field: continue
        text = str(field).strip()
        for token in re.split(r'[\s,;]+', text):
            token = token.strip()
            if token and token[0] == '2' and re.match(r'^2\d{6,}', token): return True
    return False

def so_is_countable(so_item, so_number=None, customer_po_number=None, delivery_memo=None):
    if has_internal_po_ref(customer_po_number, delivery_memo): return False
    return True

def so_countable_sql_filter():
    pattern = r'(^|[\s,;]+)2\d{6,}'
    uri = app.config.get('SQLALCHEMY_DATABASE_URI', '')
    customer_po_expr = func.coalesce(SOData.customer_po_number, '')
    delivery_memo_expr = func.coalesce(SOData.delivery_memo, '')
    if 'sqlite' in uri:
        internal_ref = db.or_(customer_po_expr.op('REGEXP')(pattern), delivery_memo_expr.op('REGEXP')(pattern))
    else:
        internal_ref = db.or_(customer_po_expr.op('~')(pattern), delivery_memo_expr.op('~')(pattern))
    return db.not_(internal_ref)

def clean(val):
    if val is None: return None
    try:
        if pd.isna(val): return None
    except: pass
    s = str(val).strip()
    return None if s.lower() in ('nan', 'none', '') else s

def clean_product_id(val):
    s = clean(val)
    if not s: return ''
    try:
        f = float(s)
        if f.is_integer(): return str(int(f))
    except: pass
    return re.sub(r'\.0+$', '', s)

def clean_request_number(val):
    s = clean(val)
    if not s: return ''
    s = str(s).strip()
    try:
        from decimal import Decimal, InvalidOperation
        number = Decimal(s)
        if number == number.to_integral_value(): return format(number.quantize(Decimal('1')), 'f')
    except: pass
    return re.sub(r'\.0+$', '', s)

RFQ_SHEET_ID = '1JrdsYWhv1mzeXB-jbukDxDYxBgaeISzpiVKEKdgfQvw'
RFQ_SHEET_NAME = 'Sales Submit-RFQ'
RFQ_CACHE = {'expires_at': None, 'rows': [], 'fetched_at': None}
RFQ_CACHE_TTL_SECONDS = 3600
VENDOR_CONTROL_SHEET_ID = '1N0Jr_h5InHH1X2TyLxRf2SMXgDzAXIJnhswzMv5Wf4E'
VENDOR_CONTROL_SHEET_GID = 723367207
VENDOR_CONTROL_CACHE = {'expires_at': None, 'rows': [], 'fetched_at': None, 'sheet_name': None, 'columns': {}}
VENDOR_CONTROL_CACHE_TTL_SECONDS = 300

RFQ_TEMPLATE_COLUMNS = [
    ('check', 'Check'), ('sheet_status', 'Status'), ('days_left', 'Days Left'), ('no', 'No'),
    ('client_name', 'Nama Client'), ('rfq_date', 'RFQ Date'), ('closing_date', 'Closing Date'),
    ('sales_pic', 'Sales PIC'), ('category_name', 'Category Name'), ('purchase_pic', 'Purchase PIC'),
    ('rfq_code', 'No. RFQ / KODE'), ('item_name', 'Item Name'), ('detail_spec', 'Detail Spec'),
    ('brand_manufacturer', 'Brand/Manufaktur'), ('qty', 'Qty'), ('unit', 'Unit'), ('remark', 'Remark'),
    ('product_id', 'Product ID'), ('request_number', 'Request Number'), ('same_replacement', 'Same/Replacement'),
    ('vendor_name', 'Vendor Name'), ('unit_price_idr', 'Unit Price (IDR)'), ('amt_idr', 'Amt (IDR)'),
    ('quoted_item_name', 'Item Name'), ('quoted_spec', 'Spec'), ('quoted_brand', 'Brand'), ('quoted_unit', 'Unit'),
    ('moq', 'MOQ'), ('lead_time_days', 'Lead Time (Days)'), ('valid_period', 'Valid period'),
    ('photo_url', 'Photo URL (optional)'), ('remarks', 'Remarks'),
    ('private_remarks_1', 'Private Remarks 1'), ('private_remarks_2', 'Private Remarks 2'),
]
RFQ_SIMILARITY_COLUMNS = [
    ('similar_prod_ids', 'Similar Product ID'), ('similar_prod_name', 'Similar Product Name'),
    ('similar_spec', 'Similar Specification'), ('similar_mfr_name', 'Similar Manufacturer'),
    ('similar_odr_unit', 'Similar Unit'), ('similar_score', '%Similarity'),
]
RFQ_EDITABLE_FIELDS = {
    'sheet_status', 'no', 'client_name', 'rfq_date', 'closing_date', 'sales_pic',
    'category_name', 'purchase_pic', 'item_name', 'detail_spec', 'brand_manufacturer',
    'qty', 'unit', 'remark', 'product_id', 'request_number',
    'same_replacement', 'vendor_name', 'unit_price_idr', 'quoted_item_name',
    'quoted_spec', 'quoted_brand', 'quoted_unit', 'moq', 'lead_time_days', 'valid_period',
    'photo_url', 'remarks', 'private_remarks_1', 'private_remarks_2'
}
RFQ_DIRECT_UPDATE_FIELDS = {'product_id'}
RFQ_BATCH_FIELDS = [
    'same_replacement', 'vendor_name', 'unit_price_idr', 'quoted_item_name',
    'quoted_spec', 'quoted_brand', 'quoted_unit', 'moq', 'lead_time_days', 'valid_period',
    'photo_url', 'remarks', 'private_remarks_1', 'private_remarks_2'
]
RFQ_SHEET_COLUMN_BY_FIELD = {
    'sheet_status': 'A', 'no': 'B', 'client_name': 'C', 'rfq_date': 'E', 'closing_date': 'F',
    'sales_pic': 'G', 'request_number': 'R', 'item_name': 'I', 'detail_spec': 'J',
    'brand_manufacturer': 'K', 'qty': 'L', 'unit': 'M', 'remark': 'N', 'category_name': 'P',
    'product_id': 'Q', 'purchase_pic': 'S', 'same_replacement': 'V', 'vendor_name': 'W',
    'unit_price_idr': 'X', 'quoted_item_name': 'Z', 'quoted_spec': 'AA', 'quoted_brand': 'AB',
    'quoted_unit': 'AC', 'moq': 'AD', 'lead_time_days': 'AE', 'valid_period': 'AF',
    'photo_url': 'AG', 'remarks': 'AH',
}

IMPORT_LAYOUT_SHEET_ID = '1i0N4VdF_vMHjr_0gjrUdS7nCKUpxPYvDWW-HOWSanEM'
IMPORT_LAYOUT_GID = '73188127'
IMPORT_LAYOUT_SOURCE_KEY = 'import_layout'
_LEGACY_IMPORT_SOURCE_KEYS = {'source_1', 'source_2', 'import_tracker'}
_IMPORT_VISIBLE_SOURCE_KEYS = (IMPORT_LAYOUT_SOURCE_KEY, 'import_tracker')

IMPORT_SOURCE_SHEETS = [
    {'key': 'source_1', 'spreadsheet_id': '1OSISIb3-D_-oxj2LXH4Q3jcG2IZWnjFGWAmTmdcPBJg', 'gid': '0', 'label': 'Source 1'},
    {'key': 'source_2', 'spreadsheet_id': '17P7_JsUGF5mqlz-j2fdvFZ9-gX8l-WGPqZABjng5Hnc', 'gid': '0', 'label': 'Source 2'},
]
IMPORT_LAYOUT_VENDOR_COLUMNS = (5, 28)
# FIX V9: Hapus IMPORT_FALLBACK_SOURCE_VENDOR_COLUMNS lama (off-by-one bug).
# Sebelumnya (16,) dipakai sebagai 0-indexed di import_row_vendor_candidates,
# padahal 16 (0-indexed) = kolom Q (Vendor Address di sheet 1, kosong di sheet 2).
# Vendor Name ada di P (sheet 1) atau Q (sheet 2) — tidak bisa hardcoded 1 kolom.
# Sekarang pakai IMPORT_FALLBACK_SOURCE_VENDOR_LETTERS = ('P', 'Q') di atas.

IMPORT_STATUS_OPTIONS = ['NEW', 'ON PROCESS', 'ON DELIVERY', 'DELIVERED', 'CANCELED']
IMPORT_CHECKBOX_FIELDS = {'sap_input', 'bl_awb', 'invoice', 'pl', 'hc', 'msds', 'coa', 'coo'}
IMPORT_FORMULA_FIELDS = {'days_left', 'site', 'vendor', 'arrival_check', 'purchase_amount', 'lt_days'}
IMPORT_HYPERLINK_FIELDS = {'soft_copy_doc'}
IMPORT_PAYMENT_DROPDOWN_FIELDS = {'payment'}
IMPORT_PAYMENT_DATE_FIELDS = {'payment_date'}
IMPORT_VENDOR_ATTR_FIELDS = {'origin', 'top'}  # populated from ImportVendor by vendor_name match
IMPORT_DASHBOARD_LOCAL_FIELDS = {'po_send_date'}
IMPORT_SOURCE_MANAGED_FIELDS = {
    'po_date_by_email', 'site', 'yupi_po', 'po_yupi', 'vendor',
    'req_dlv_date', 'source_req_dlv_date', 'so', 'group', 'po_sementara',
    'item_yupi', 'item_name', 'spec', 'remark_yupi', 'reschedule',
    'ord_qty', 'unit', 'unit_price', 'amount', 'vendor_name',
    'purchase_price', 'currency',
    # FIX V9: field-field di bawah DIPINDAH dari IMPORT_SOURCE_MANAGED_FIELDS ke
    # IMPORT_USER_LOCAL_ONLY_FIELDS (lihat bawah). Sebelumnya, field-field ini
    # di-overwrite saat sync dari source sheet — padahal user edit manual field-field
    # ini di dashboard (incoterm, forwarder, bl_number, inv_no, non_ski).
    # Sekarang field-field tersebut TIDAK akan di-overwrite oleh source sheet sync.
    # 'incoterm', 'forwarder', 'bl_number', 'inv_no', 'non_ski',  ← DIPINDAH
}

IMPORT_LOCAL_EDIT_FIELDS = {
    'status', 'days_left', 'po_send_date', '_po_send_date_manual', 'po_date_by_email', 'site', 'yupi_po', 'po_yupi', 'vendor',
    'req_dlv_date', 'source_req_dlv_date', 'etd', 'eta', 'arrival_check', 'import_remarks',
    'so', 'group', 'po_sementara', 'item_yupi', 'item_name', 'spec', 'remark_yupi',
    'reschedule', 'ord_qty', 'unit', 'unit_price', 'amount', 'vendor_name',
    'purchase_price', 'currency', 'purchase_amount', 'lt_days', 'incoterm',
    'forwarder', 'bl_number', 'inv_no', 'non_ski',
    'sap_input', 'bl_awb', 'invoice', 'pl', 'hc', 'msds', 'coa', 'coo',
    'soft_copy_doc',
    # NEW: payment + payment_date — local-only editable fields, never overwritten by source sync
    'payment', 'payment_date',
}

IMPORT_USER_LOCAL_ONLY_FIELDS = IMPORT_LOCAL_EDIT_FIELDS - IMPORT_SOURCE_MANAGED_FIELDS
# FIX V9: Pastikan field-field user-edit manual SELALU di USER_LOCAL_ONLY.
# Field-field ini TIDAK BOLEH di-overwrite oleh source sheet sync:
#   - incoterm, forwarder, bl_number, inv_no, non_ski (diisi user via dashboard)
#   - status (di-set user via dashboard)
#   - sap_input, bl_awb, invoice, pl, hc, msds, coa, coo, soft_copy_doc (checklist user)
#   - etd, eta, import_remarks (diisi user via dashboard)
#   - payment, payment_date (diisi user via dashboard)
#   - po_send_date, _po_send_date_manual (diisi user via dashboard)
assert {'incoterm', 'forwarder', 'bl_number', 'inv_no', 'non_ski',
        'status', 'sap_input', 'bl_awb', 'invoice', 'pl', 'hc', 'msds', 'coa', 'coo',
        'soft_copy_doc', 'etd', 'eta', 'import_remarks', 'payment', 'payment_date',
        'po_send_date', '_po_send_date_manual'}.issubset(IMPORT_USER_LOCAL_ONLY_FIELDS), \
    "Field user-edit harus di IMPORT_USER_LOCAL_ONLY_FIELDS"

IMPORT_SOURCE_ONLY_COLUMNS = [
    {'source_only': True, 'source_sheet_col': 'F', 'sheet_col': 'R', 'field': 'po_yupi', 'label': 'PO YUPI'},
    {'source_only': True, 'source_sheet_col': 'K', 'sheet_col': 'W', 'field': 'source_req_dlv_date', 'label': 'Req. Dlv Date'},
]

IMPORT_SYNC_FIELD_ALIASES = {'yupi_po': 'po_yupi', 'req_dlv_date': 'source_req_dlv_date'}

IMPORT_REFERENCE_VISIBLE_COLUMNS = [
    {'sheet_col': 'A',  'field': 'status',              'label': 'STATUS',                 'width': 130, 'type': 'status'},
    {'sheet_col': 'B',  'field': 'days_left',           'label': 'Days',                   'width': 64,  'formula': True},
    {'sheet_col': 'C',  'field': 'po_send_date',         'label': 'PO Send Date',          'width': 100, 'local': True, 'group_per_item': False, 'blue_text': True},
    {'sheet_col': 'D',  'source_sheet_col': 'B',  'field': 'site',                'label': 'Site',                   'width': 60,  'formula': True, 'blue_text': True},
    {'sheet_col': 'E',  'source_sheet_col': 'F',  'field': 'yupi_po',             'label': 'YUPI PO',                'width': 100, 'blue_text': True},
    {'sheet_col': 'F',  'source_sheet_col': 'Q',  'field': 'vendor',              'label': 'Vendor',                 'width': 140, 'formula': True, 'blue_text': True},
    # NEW: Origin — populated from ImportVendor table by vendor_name match.
    {'sheet_col': '',   'field': 'origin',              'label': 'Origin',                 'width': 80,  'vendor_attr': True, 'blue_text': True},
    {'sheet_col': 'G',  'source_sheet_col': 'K',  'field': 'req_dlv_date',        'label': 'Req Dlv Date',           'width': 100, 'blue_text': True},
    {'sheet_col': 'H',  'field': 'etd',                 'label': 'ETD',                    'width': 110, 'blue_text': True},
    {'sheet_col': 'I',  'field': 'eta',                 'label': 'ETA',                    'width': 110, 'blue_text': True},
    {'sheet_col': 'J',  'field': 'arrival_check',       'label': 'Arrival Check',          'width': 120, 'formula': True, 'blue_text': True},
    {'sheet_col': 'K',  'field': 'import_remarks',      'label': 'Import Remarks',         'width': 180, 'blue_text': True, 'bold_text': True},
    # ── Shipment-level group columns (moved here per user request) ───────────
    # These are GROUP columns (no group_per_item flag) so they merge across
    # rows with the same YUPI PO + Req Dlv Date. Blue text applies from
    # PO Send Date through Inv No.
    {'sheet_col': 'CU', 'field': 'lt_days',             'label': 'LT (Days)',              'width': 70,  'formula': True, 'number': True, 'blue_text': True},
    {'sheet_col': 'CV', 'field': 'incoterm',            'label': 'Incoterm',               'width': 76,  'blue_text': True},
    {'sheet_col': 'CW', 'field': 'forwarder',           'label': 'Forwarder',              'width': 150, 'blue_text': True},
    {'sheet_col': 'CX', 'field': 'bl_number',           'label': 'BL Number',              'width': 150, 'blue_text': True},
    {'sheet_col': 'CY', 'field': 'inv_no',              'label': 'Invoice No',             'width': 130, 'blue_text': True},
    # SAP INPUT moved here (right after Invoice No). NOT part of the checklist
    # group — always visible as a checkmark toggle.
    {'sheet_col': 'DA', 'field': 'sap_input',           'label': 'SAP INPUT',              'width': 76,  'checkbox': True},
    # NEW: TOP — populated from ImportVendor table by vendor_name match.
    {'sheet_col': '',   'field': 'top',                 'label': 'TOP',                    'width': 60,  'vendor_attr': True, 'blue_text': True},
    # NEW: Payment — dropdown with empty / "DONE".
    {'sheet_col': '',   'field': 'payment',             'label': 'Payment',                'width': 86,  'payment_dropdown': True},
    # NEW: Payment Date — date picker. Auto-shows "Overdue" (red) when
    # today > ETA + TOP days AND payment is not DONE.
    {'sheet_col': '',   'field': 'payment_date',        'label': 'Payment Date',           'width': 110, 'payment_date': True},
    # SOFT COPY DOC moved here (right after Payment Date) per user request.
    {'sheet_col': 'DI', 'field': 'soft_copy_doc',       'label': 'SOFT COPY DOC',          'width': 150, 'hyperlink': True},
    # ── Per-item columns (group_per_item=True → never merged across rows) ──────
    {'sheet_col': 'L',                            'field': 'so',                  'label': 'SO',                     'width': 110, 'group_per_item': True},
    {'sheet_col': 'M',  'source_sheet_col': 'A',  'field': 'group',               'label': 'GROUP',                  'width': 90,  'group_per_item': True},
    {'sheet_col': 'O',  'source_sheet_col': 'C', 'field': 'po_date_by_email',    'label': 'PO DATE\n(By Email)',    'width': 100, 'group_per_item': True},
    {'sheet_col': 'Q',  'source_sheet_col': 'E',  'field': 'po_sementara',        'label': 'PO SEMENTARA',           'width': 120, 'group_per_item': True},
    {'sheet_col': 'S',  'source_sheet_col': 'G',  'field': 'item_yupi',           'label': 'Item Yupi',              'width': 100, 'group_per_item': True},
    {'sheet_col': 'T',  'source_sheet_col': 'H',  'field': 'item_name',           'label': 'Item name',              'width': 200, 'group_per_item': True},
    {'sheet_col': 'U',  'source_sheet_col': 'I',  'field': 'spec',                'label': 'Spec',                   'width': 340, 'group_per_item': True},
    {'sheet_col': 'V',  'source_sheet_col': 'J',  'field': 'remark_yupi',         'label': 'REMARK YUPI',            'width': 260},
    {'sheet_col': 'X',  'source_sheet_col': 'L',  'field': 'reschedule',          'label': 'RESCHEDULE',             'width': 96,  'group_per_item': True},
    {'sheet_col': 'Y',  'source_sheet_col': 'M',  'field': 'ord_qty',             'label': "Ord. Q'ty",             'width': 76,  'number': True, 'group_per_item': True},
    {'sheet_col': 'Z',  'source_sheet_col': 'N',  'field': 'unit',                'label': 'Unit',                   'width': 56,  'group_per_item': True},
    {'sheet_col': 'AA', 'source_sheet_col': 'O',  'field': 'unit_price',          'label': 'Unit Price',             'width': 90,  'number': True, 'group_per_item': True},
    {'sheet_col': 'AB', 'source_sheet_col': 'P',  'field': 'amount',              'label': 'AMOUNT',                 'width': 100, 'number': True, 'group_per_item': True},
    # Vendor Name column removed — Vendor column on the left already covers it.
    {'sheet_col': 'AG', 'source_sheet_col': 'U',  'field': 'purchase_price',      'label': 'PURCHASE PRICE',         'width': 96,  'number': True, 'group_per_item': True},
    {'sheet_col': 'AH', 'source_sheet_col': 'V',  'field': 'currency',            'label': 'CURRENCY',               'width': 72,  'group_per_item': True},
    {'sheet_col': 'AJ', 'source_sheet_col': 'X',  'field': 'purchase_amount',     'label': 'PURCHASE\nAMOUNT',       'width': 100, 'formula': True, 'number': True, 'group_per_item': True},
    {'sheet_col': 'DB', 'field': 'bl_awb',              'label': 'BL / AWB',               'width': 76,  'checkbox': True},
    {'sheet_col': 'DC', 'field': 'invoice',             'label': 'INVOICE',                'width': 76,  'checkbox': True},
    {'sheet_col': 'DD', 'field': 'pl',                  'label': 'PL',                     'width': 56,  'checkbox': True},
    {'sheet_col': 'DE', 'field': 'hc',                  'label': 'HC',                     'width': 56,  'checkbox': True},
    {'sheet_col': 'DF', 'field': 'msds',                'label': 'MSDS',                   'width': 64,  'checkbox': True},
    {'sheet_col': 'DG', 'field': 'coa',                 'label': 'COA',                    'width': 56,  'checkbox': True},
    {'sheet_col': 'DH', 'field': 'coo',                 'label': 'COO',                    'width': 56,  'checkbox': True},
    # NON-SKI moved to rightmost column. Regular editable cell (free text
    # input) — NOT a checkbox or dropdown.
    {'sheet_col': 'CZ', 'field': 'non_ski',             'label': 'NON-SKI',                'width': 70},
]

IMPORT_COLUMN_ALIASES = {
    'status': ['status'], 'po_send_date': [], 'po_date_by_email': ['podatebyemail', 'podateemail', 'poemaildate', 'pokirimdate', 'posenddate'],
    'site': ['siteidnkrg', 'site'], 'yupi_po': ['poyupi'], 'po_yupi': ['poyupi'], 'vendor': ['vendor', 'vendorname'],
    'req_dlv_date': ['reqdlvdate'], 'source_req_dlv_date': ['reqdlvdate'], 'etd': ['etd'], 'eta': ['eta'],
    'import_remarks': ['importremarks', 'remarksvo', 'remark'], 'so': ['so', 'noso'], 'group': ['group'],
    'po_sementara': ['posementara'], 'item_yupi': ['itemyupi'], 'item_name': ['itemname'], 'spec': ['spec', 'specification'],
    'remark_yupi': ['remarkyupi'], 'reschedule': ['reschedule'], 'ord_qty': ['ordqty', 'orderqty', 'orderedqty'],
    'unit': ['unit'], 'unit_price': ['unitprice'], 'amount': ['amount'], 'vendor_name': ['vendorname', 'vendor'],
    'purchase_price': ['purchaseprice'], 'currency': ['currency'], 'purchase_amount': ['purchaseamount'],
    'incoterm': ['incoterm'], 'forwarder': ['forwarder'], 'bl_number': ['blnumber', 'blawb', 'awbnumber'],
    'inv_no': ['invno', 'invoiceno', 'invoicenumber'], 'non_ski': ['nonski'], 'sap_input': ['sapinput'],
    'bl_awb': ['blawb'], 'invoice': ['invoice'], 'pl': ['pl'], 'hc': ['hc'], 'msds': ['msds'], 'coa': ['coa'],
    'coo': ['coo'], 'soft_copy_doc': ['softcopydoc', 'softcopy', 'gdrive', 'googledrive', 'documentlink'],
}

IMPORT_COMMON_SOURCE_FALLBACK_COLUMNS = {
    'group': 'A', 'site': 'B', 'po_date_by_email': 'C', 'po_sementara': 'E', 'po_yupi': 'F', 'yupi_po': 'F',
    'item_yupi': 'G', 'item_name': 'H', 'spec': 'I', 'remark_yupi': 'J', 'req_dlv_date': 'K', 'source_req_dlv_date': 'K',
    'reschedule': 'L', 'ord_qty': 'M', 'unit': 'N', 'unit_price': 'O', 'amount': 'P', 'vendor': 'Q', 'vendor_name': 'Q',
    'purchase_price': 'U', 'currency': 'V',
}
# FIX V9: Sheet 1 (source_1) dan Sheet 2 (source_2) punya layout kolom yang BERBEDA.
# Sheet 1: header di row 4, Vendor Name di kolom P, PO YUPI di kolom E
# Sheet 2: header di row 3, Vendor Name di kolom Q, PO YUPI di kolom F
#
# Header detection sudah DINAMIS (import_detect_header_row scan 60 baris),
# jadi kalau user edit sheet dan pindah header ke row lain, masih bisa detect.
# Fallback letters di atas hanya dipakai kalau header detection gagal total.
IMPORT_RM_SOURCE_FALLBACK_COLUMNS = {**IMPORT_COMMON_SOURCE_FALLBACK_COLUMNS, 'purchase_amount': 'X', 'so': 'AK'}
IMPORT_SP_SOURCE_FALLBACK_COLUMNS = {**IMPORT_COMMON_SOURCE_FALLBACK_COLUMNS, 'purchase_amount': 'Y', 'so': 'AM'}
# FIX V9: Vendor column fallback — HANYA Q (Vendor Name).
# Sebelumnya ('P', 'Q') — P didahulukan. Tapi P = AMOUNT (kolom angka)!
# Saat vendor name (Q) kosong di suatu row (merged cell/subtotal),
# fallback ambil P = "168,000,000" → tidak match vendor list → row di-skip.
# Kasus nyata: PO 410100301 dari CURT GEORGI tidak tercopy karena row-nya
# ke-detect vendor = amount, bukan vendor name.
#
# Header detection (import_source_column_map) sudah dapat vendor_name = Q
# dengan benar untuk kedua sheet. Fallback letter hanya dipakai kalau header
# detection gagal total — dan kalau itu terjadi, Q lebih aman daripada P.
IMPORT_FALLBACK_SOURCE_VENDOR_LETTERS = ('Q',)  # hanya Vendor Name, jangan AMOUNT
# Hapus IMPORT_FALLBACK_SOURCE_VENDOR_COLUMNS lama (off-by-one bug: 16 dipakai sebagai
# 0-indexed padahal harusnya 1-indexed P=16, 0-indexed P=15)

def import_source_kind_from_header(df, header_idx):
    try: headers = [import_header_key(v) for v in df.iloc[header_idx].tolist()]
    except: headers = []
    def key_at(letter):
        try:
            idx = column_index_from_letter(letter) - 1
            return headers[idx] if idx < len(headers) else ''
        except: return ''
    if key_at('Y') == 'purchaseamount' or key_at('AM') == 'noso': return 'sp'
    if key_at('X') == 'purchaseamount' or key_at('AK') == 'noso': return 'rm'
    if 'pokiriminput' in headers or 'pokirimdate' in headers: return 'sp'
    if header_idx == 2: return 'sp'
    if header_idx == 3: return 'rm'
    return ''

def import_source_fallback_columns(df, header_idx):
    kind = import_source_kind_from_header(df, header_idx)
    if kind == 'sp': return IMPORT_SP_SOURCE_FALLBACK_COLUMNS
    if kind == 'rm': return IMPORT_RM_SOURCE_FALLBACK_COLUMNS
    return IMPORT_COMMON_SOURCE_FALLBACK_COLUMNS

def import_meta_get(key):
    row = ImportDashboardMeta.query.filter_by(meta_key=key).first()
    if not row: return None
    try: return json.loads(row.value_json or 'null')
    except: return None

def import_meta_set(key, value):
    row = ImportDashboardMeta.query.filter_by(meta_key=key).first()
    if not row:
        row = ImportDashboardMeta(meta_key=key)
        db.session.add(row)
    row.value_json = json.dumps(value, ensure_ascii=False)
    row.updated_at = datetime.utcnow()
    db.session.commit()

def google_csv_url(spreadsheet_id, gid='0'):
    return f'https://docs.google.com/spreadsheets/d/{spreadsheet_id}/gviz/tq?tqx=out:csv&gid={gid}'

def read_public_sheet_csv(spreadsheet_id, gid='0', nrows=None):
    return pd.read_csv(google_csv_url(spreadsheet_id, gid), header=None, dtype=str, keep_default_na=False, nrows=nrows)

def import_clean_header(value, fallback):
    label = (clean(value) or '').replace('\r', '').replace('\n', ' / ')
    return label or fallback

def import_header_key(value):
    return re.sub(r'[^a-z0-9]+', '', (clean(value) or '').lower())

def import_blankish(value):
    raw = clean(value)
    if raw is None: return True
    s = str(raw).strip()
    return not s or s.lower() in ('nan', 'none', 'null', 'n/a', '#n/a') or s in ('-', '–', '—')

def import_nonblank(value):
    return '' if import_blankish(value) else str(clean(value)).strip()

def import_layout_columns(force=False):
    columns = []
    for idx, col in enumerate(IMPORT_REFERENCE_VISIBLE_COLUMNS):
        item = dict(col)
        item['col_idx'] = idx
        item['checkbox'] = bool(item.get('checkbox') or item.get('field') in IMPORT_CHECKBOX_FIELDS)
        item['formula'] = bool(item.get('formula') or item.get('field') in IMPORT_FORMULA_FIELDS)
        item['hyperlink'] = bool(item.get('hyperlink') or item.get('field') in IMPORT_HYPERLINK_FIELDS)
        item['local'] = bool(item.get('local') or item.get('field') in IMPORT_DASHBOARD_LOCAL_FIELDS)
        item['group_per_item'] = bool(item.get('group_per_item', False))
        item['blue_text'] = bool(item.get('blue_text', False))
        item['bold_text'] = bool(item.get('bold_text', False))
        item['yes_no'] = bool(item.get('yes_no', False))
        if item.get('field') == 'status': item['options'] = IMPORT_STATUS_OPTIONS
        columns.append(item)
    return columns

def import_all_mapping_columns(columns=None):
    base = list(columns or import_layout_columns())
    seen = {c.get('field') for c in base}
    for col in IMPORT_SOURCE_ONLY_COLUMNS:
        if col.get('field') not in seen:
            base.append(dict(col)); seen.add(col.get('field'))
    return base

def import_default_vendors_from_layout(force=False):
    cache_key = ('import_default_vendors_from_layout',)
    cached = None if force else runtime_cache_get(cache_key)
    if cached is not None: return cached
    cached = None if force else import_meta_get('default_vendors')
    if cached is not None:
        runtime_cache_set(cache_key, cached, ttl_seconds=900)
        return cached
    try: df = read_public_sheet_csv(IMPORT_LAYOUT_SHEET_ID, IMPORT_LAYOUT_GID)
    except: return []
    vendors = set()
    for row_idx in range(2, len(df)):
        for col_idx in IMPORT_LAYOUT_VENDOR_COLUMNS:
            if col_idx >= df.shape[1]: continue
            name = clean(df.iloc[row_idx, col_idx])
            if not name or name.lower() in ('vendor', 'vendor name'): continue
            vendors.add(name)
    vendors = sorted(vendors, key=lambda s: s.lower())
    import_meta_set('default_vendors', vendors)
    runtime_cache_set(cache_key, vendors, ttl_seconds=900)
    return vendors

def import_uploaded_vendor_names():
    rows = ImportVendor.query.order_by(ImportVendor.vendor_name.asc()).all()
    return [r.vendor_name for r in rows if clean(r.vendor_name)]

def import_existing_dashboard_vendor_names():
    vendors = set()
    for row in ImportDashboardRow.query.filter(ImportDashboardRow.source_key.in_(_IMPORT_VISIBLE_SOURCE_KEYS)).all():
        v = clean(row.vendor_name)
        if v: vendors.add(v)
        try: data = json.loads(row.data_json or '{}')
        except: data = {}
        for field in ('vendor_name', 'vendor'):
            v = clean(data.get(field))
            if v: vendors.add(v)
    return sorted(vendors, key=lambda s: s.lower())

def import_vendor_filter_names():
    uploaded = import_uploaded_vendor_names()
    if uploaded: return uploaded, 'vendor_import'
    existing = import_existing_dashboard_vendor_names()
    if existing: return existing, 'existing_import_rows'
    return [], 'none'

def import_vendor_names(force_default=False):
    uploaded = import_uploaded_vendor_names()
    return uploaded or import_default_vendors_from_layout(force=force_default)


def import_vendor_attrs_map():
    """Build a {vendor_name_lower: {'origin':..., 'top':..., 'non_ski':...}}
    map from the ImportVendor table. Used to inject Origin/TOP into import
    rows by matching the row's vendor name."""
    out = {}
    try:
        for v in db.session.query(ImportVendor).all():
            key = (v.vendor_name or '').strip().lower()
            if not key: continue
            out[key] = {
                'origin': (v.origin or '').strip(),
                'top': (v.top or '').strip(),
                'non_ski': (v.non_ski or '').strip(),
            }
    except Exception:
        pass
    return out

def import_detect_data_start(df, header_idx=None):
    """Deteksi baris pertama yang berisi DATA (bukan header/kosong).

    FIX V9: Sebelumnya hardcode kolom 7, 12, 16 untuk cek 'item name', 'qty',
    'vendor'. Itu salah untuk sheet yang layout-nya beda. Plus fallback `return 3`
    adalah hardcode — kalau header di row 5, data tidak mungkin mulai row 3.

    Sekarang: scan dari header_idx+1, cari row pertama yang punya ≥3 cell
    non-empty. Tidak peduli kolom mana — asal row tersebut berisi data.
    """
    if df is None or getattr(df, 'empty', True): return 0
    if header_idx is None:
        # Kalau header_idx tidak diketahui, scan dari atas sampai ketemu header
        header_idx = import_detect_header_row(df)
    # Scan dari header_idx + 1, cari row dengan ≥3 cell non-empty
    for idx in range(header_idx + 1, min(len(df), header_idx + 20)):
        try:
            row_values = [clean(v) for v in df.iloc[idx].tolist()]
        except: continue
        non_empty = sum(1 for v in row_values if v)
        if non_empty >= 3:
            return idx
    # Fallback: header_idx + 1 (data biasanya langsung setelah header)
    return header_idx + 1 if header_idx is not None else 3

def import_detect_header_row(df):
    """Deteksi baris yang berisi HEADER tabel.

    FIX V9: Dinamis — scan sampai 60 baris pertama, tidak peduli di row berapa
    header berada (bisa row 1, 3, 4, dst). Pakai scoring berdasarkan jumlah
    header yang match alias.

    Contoh:
    - Sheet 1: header di row 4 (1-indexed) → deteksi sebagai idx 3 (0-indexed)
    - Sheet 2: header di row 3 (1-indexed) → deteksi sebagai idx 2 (0-indexed)
    - Kalau user edit sheet dan pindah header ke row 1, masih bisa deteksi.
    """
    if df is None or getattr(df, 'empty', True): return 0
    # Fast-path: cek 12 baris pertama untuk kombinasi header yang sangat spesifik
    for idx in range(min(len(df), 12)):
        try: labels = [import_header_key(v) for v in df.iloc[idx].tolist()]
        except: continue
        if 'poyupi' in labels and 'reqdlvdate' in labels: return idx
        if 'itemname' in labels and 'ordqty' in labels and 'vendorname' in labels: return idx
    # Fallback: scoring-based detection (scan sampai 60 baris)
    alias_keys = set()
    for values in IMPORT_COLUMN_ALIASES.values():
        alias_keys.update(import_header_key(v) for v in values if v)
    alias_keys.update(import_header_key(c.get('label')) for c in import_all_mapping_columns(import_layout_columns()))
    alias_keys.update(import_header_key(c.get('field')) for c in import_all_mapping_columns(import_layout_columns()))
    alias_keys.discard('')
    best_idx, best_score = None, -1
    max_scan = min(len(df), 60)
    for idx in range(max_scan):
        labels = [import_header_key(v) for v in df.iloc[idx].tolist()]
        score = sum(1 for label in labels if label in alias_keys)
        if 'poyupi' in labels: score += 4
        if 'reqdlvdate' in labels: score += 4
        if 'itemname' in labels: score += 3
        if 'vendorname' in labels or 'vendor' in labels: score += 3
        if 'posementara' in labels: score += 2
        if score > best_score:
            best_idx = idx; best_score = score
    if best_idx is not None and best_score >= 4: return best_idx
    # Last resort: pakai import_detect_data_start - 1
    return max(import_detect_data_start(df) - 1, 0)

def import_source_header_score(df):
    if df is None or getattr(df, 'empty', True): return -1
    idx = import_detect_header_row(df)
    try: labels = [import_header_key(v) for v in df.iloc[idx].tolist()]
    except: labels = []
    weights = {'poyupi': 10, 'reqdlvdate': 10, 'itemyupi': 6, 'itemname': 6, 'vendorname': 6, 'posementara': 4, 'ordqty': 4, 'unitprice': 4, 'purchaseprice': 4, 'purchaseamount': 4, 'noso': 4, 'podatebyemail': 3}
    return sum(weights.get(label, 0) for label in labels)

def import_source_candidate_titles(source):
    titles = []
    try:
        preferred = import_sheet_title_for_gid(source['spreadsheet_id'], source.get('gid') or '0')
        if preferred: titles.append(preferred)
    except: pass
    try:
        metadata = google_sheets_metadata(source['spreadsheet_id'])
        for sheet in metadata.get('sheets', []):
            props = sheet.get('properties', {})
            title = props.get('title')
            if title and title not in titles: titles.append(title)
    except: pass
    return titles

def import_source_header_preview(source, force=False):
    cache_key = ('import_source_header_preview_v2', source.get('spreadsheet_id'), source.get('gid'))
    if not force:
        cached = runtime_cache_get(cache_key)
        if cached is not None:
            title, values = cached
            return title, pd.DataFrame(values).fillna('') if values else pd.DataFrame()
    best_title, best_values, best_score = None, [], -1
    for title in import_source_candidate_titles(source):
        try:
            values = google_sheets_values_get(source['spreadsheet_id'], f"'{title}'!A1:ZZ60", value_render_option='FORMATTED_VALUE').get('values', [])
            df = pd.DataFrame(values).fillna('') if values else pd.DataFrame()
            score = import_source_header_score(df)
            if score > best_score:
                best_title = title; best_values = values; best_score = score
            if score >= 35: break
        except: continue
    if best_title:
        runtime_cache_set(cache_key, (best_title, best_values), ttl_seconds=3600)
        return best_title, pd.DataFrame(best_values).fillna('') if best_values else pd.DataFrame()
    title = import_sheet_title_for_gid(source['spreadsheet_id'], source.get('gid') or '0')
    return title, pd.DataFrame()

def import_source_column_map(df, columns):
    """Build mapping {field_name: column_index_0_based} berdasarkan HEADER NAME detection.

    FIX V9: Per-sheet header detection. Kedua sheet source punya layout kolom
    yang BERBEDA (vendor di P untuk sheet 1, di Q untuk sheet 2; header row di 1
    untuk sheet 1, di 3 untuk sheet 2). Karena itu kita detect header by NAME,
    bukan by hardcoded column letter.

    Logika:
    1. Cari header row (bisa di row 1, 2, 3, dst — tergantung sheet)
    2. Untuk setiap field, cari kolom yang header-nya match alias (case-insensitive,
       hapus punctuation/spasi)
    3. Kalau tidak match, fallback ke hardcoded letter (last resort)
    """
    header_idx = import_detect_header_row(df)
    header_values = list(df.iloc[header_idx]) if len(df) else []
    source_fallbacks = import_source_fallback_columns(df, header_idx)
    by_key = {}
    for idx, raw in enumerate(header_values):
        key = import_header_key(raw)
        if key and key not in by_key: by_key[key] = idx
    source_map = {}
    for col in columns:
        field = col.get('field')
        if field in IMPORT_DASHBOARD_LOCAL_FIELDS: continue
        keys = []
        keys.extend(IMPORT_COLUMN_ALIASES.get(field, []))
        keys.extend([import_header_key(col.get('label')), import_header_key(field)])
        seen_keys = []
        for key in keys:
            if key and key not in seen_keys: seen_keys.append(key)
        source_idx = next((by_key[key] for key in seen_keys if key in by_key), None)
        fallback_sheet_col = source_fallbacks.get(field) or col.get('source_sheet_col')
        if not fallback_sheet_col and col.get('source_only'):
            fallback_sheet_col = source_fallbacks.get(field) or col.get('sheet_col')
        if source_idx is None and fallback_sheet_col:
            try: source_idx = column_index_from_letter(str(fallback_sheet_col)) - 1
            except: source_idx = None
        if source_idx is not None: source_map[field] = source_idx
    # FIX V9: log hasil column detection untuk debugging
    # (hanya field kritis yang sering bermasalah: vendor_name, po_yupi, po_date_by_email)
    try:
        import sys
        debug_fields = ('vendor_name', 'vendor', 'po_yupi', 'yupi_po', 'po_date_by_email', 'item_name')
        debug_map = {f: source_map.get(f) for f in debug_fields if f in source_map}
        # Convert 0-indexed ke letter untuk log yang readable
        debug_letters = {}
        for f, idx in debug_map.items():
            if idx is not None:
                try: debug_letters[f] = column_letter_from_index(idx + 1)
                except: debug_letters[f] = f'?{idx}'
        print(f"[import_source_column_map] header_idx={header_idx}, detected: {debug_letters}", file=sys.stderr)
    except Exception:
        pass
    if 'po_yupi' in source_map and 'yupi_po' not in source_map: source_map['yupi_po'] = source_map['po_yupi']
    if 'yupi_po' in source_map and 'po_yupi' not in source_map: source_map['po_yupi'] = source_map['yupi_po']
    if 'source_req_dlv_date' in source_map and 'req_dlv_date' not in source_map: source_map['req_dlv_date'] = source_map['source_req_dlv_date']
    if 'req_dlv_date' in source_map and 'source_req_dlv_date' not in source_map: source_map['source_req_dlv_date'] = source_map['req_dlv_date']
    _so_col_kind = import_source_kind_from_header(df, header_idx)
    _so_col_fallback = source_fallbacks.get('so')
    if _so_col_fallback and _so_col_kind in ('rm', 'sp'):
        try: source_map['so'] = column_index_from_letter(_so_col_fallback) - 1
        except: pass
    return source_map

def import_row_vendor_candidates(values, source_map, columns):
    """Kumpulkan kandidat nilai vendor dari sebuah row.

    FIX V9: Sebelumnya pakai IMPORT_FALLBACK_SOURCE_VENDOR_COLUMNS = (16,) yang
    di-interpretasi sebagai 0-indexed → kolom Q (Vendor Address di sheet 1).
    Itu bug — kolom Q berisi alamat, bukan nama vendor.

    Sekarang: pakai IMPORT_FALLBACK_SOURCE_VENDOR_LETTERS = ('P', 'Q') yang
    dikonversi ke 0-indexed. P (sheet 1: Vendor Name) dan Q (sheet 2: Vendor Name).
    Code pemakai akan iterasi semua kandidat dan ambil yang tidak kosong.
    """
    candidates = []
    for field in ('vendor_name', 'vendor'):
        col_idx = source_map.get(field)
        if col_idx is not None and col_idx < len(values): candidates.append(values[col_idx])
    # FIX V9: konversi letter ke 0-indexed column index
    for letter in IMPORT_FALLBACK_SOURCE_VENDOR_LETTERS:
        try:
            col_idx = column_index_from_letter(letter) - 1  # 1-indexed → 0-indexed
            if 0 <= col_idx < len(values): candidates.append(values[col_idx])
        except Exception:
            pass
    return [clean(v) for v in candidates if clean(v)]

def import_vendor_match(vendor_candidate, vendor_set):
    """Cek apakah vendor_candidate cocok dengan salah satu vendor di vendor_set.

    FIX V9: Sebelumnya pakai exact match case-insensitive:
        v.strip().lower() in vendor_set

    Masalah: Vendor di ImportVendor table sering ditulis singkat, misal
    "CURT GEORGI GMBH & CO", tapi di sheet source tertulis lengkap
    "CURT GEORGI GMBH & CO. KG" atau "CURT GEORGI GMBH & CO., LTD".
    Exact match → tidak match → row di-skip → PO 410100301 tidak masuk dashboard.

    Solusi: 3-tier matching:
    1. Exact match (case-insensitive)
    2. Prefix match: vendor_set adalah prefix dari candidate (atau sebaliknya)
    3. Contains match: candidate mengandung vendor_set (atau sebaliknya)
       — dipakai kalau keduanya minimal 8 karakter untuk hindari false positive

    Return True kalau ada match.
    """
    if not vendor_candidate or not vendor_set:
        return False
    v = vendor_candidate.strip().lower()
    if not v:
        return False

    # Tier 1: Exact match
    if v in vendor_set:
        return True

    # Tier 2 & 3: prefix/contains match
    # Normalisasi: hapus punctuation berlebih (titik, koma, &, dll) untuk matching
    def normalize(s):
        return re.sub(r'[^a-z0-9]+', '', s)

    v_norm = normalize(v)
    if not v_norm or len(v_norm) < 5:
        return False

    for target in vendor_set:
        if not target:
            continue
        t_norm = normalize(target)
        if not t_norm or len(t_norm) < 5:
            continue  # skip terlalu pendek (false positive risk)
        # Tier 2: prefix match
        if v_norm.startswith(t_norm) or t_norm.startswith(v_norm):
            return True
        # Tier 3: contains match (min 8 chars untuk hindari false positive)
        if len(t_norm) >= 8 and len(v_norm) >= 8 and (t_norm in v_norm or v_norm in t_norm):
            return True

    return False


def import_source_rows_fast(source, columns, vendor_set):
    mapping_columns = import_all_mapping_columns(columns)
    try:
        sheet_title, header_df = import_source_header_preview(source, force=True)
        if header_df.empty: return []
        source_map = import_source_column_map(header_df, mapping_columns)
        header_idx = import_detect_header_row(header_df)
        # FIX V9: deteksi data_start_row secara dinamis (bukan hardcode header_idx + 2).
        # Sebelumnya +2 asumsi data mulai 2 baris setelah header (untuk skip sub-header
        # atau blank row). Tapi itu tidak selalu benar — bisa saja data langsung di
        # baris setelah header.
        # Sekarang: pakai import_detect_data_start() yang scan dari header_idx+1
        # untuk cari row pertama dengan ≥3 cell non-empty.
        data_start_idx = import_detect_data_start(header_df, header_idx)
        # Convert 0-indexed → 1-indexed untuk Google Sheets API
        data_start_row = data_start_idx + 1
        if data_start_row <= header_idx + 1:
            # Safety: data tidak boleh mulai sebelum/sama dengan header
            data_start_row = header_idx + 2
        try:
            import sys
            print(f"[import_source_rows_fast] source={source['key']}, header_idx={header_idx} (row {header_idx+1} 1-indexed), data_start_row={data_start_row}", file=sys.stderr)
        except Exception: pass
        source_fallbacks = import_source_fallback_columns(header_df, header_idx)
        for field in ('po_yupi', 'yupi_po', 'source_req_dlv_date', 'req_dlv_date', 'po_date_by_email', 'site', 'po_sementara', 'item_yupi', 'item_name', 'spec', 'remark_yupi', 'reschedule', 'ord_qty', 'unit', 'unit_price', 'amount', 'vendor_name', 'vendor', 'purchase_price', 'currency', 'purchase_amount', 'so'):
            if field not in source_map and source_fallbacks.get(field):
                source_map[field] = column_index_from_letter(source_fallbacks[field]) - 1
        _so_kind = import_source_kind_from_header(header_df, header_idx)
        so_fallback_col = source_fallbacks.get('so')
        if so_fallback_col and _so_kind in ('rm', 'sp'):
            source_map['so'] = column_index_from_letter(so_fallback_col) - 1
        # FIX V9: konversi IMPORT_FALLBACK_SOURCE_VENDOR_LETTERS ke 0-indexed indexes
        _vendor_fallback_indexes = set()
        for letter in IMPORT_FALLBACK_SOURCE_VENDOR_LETTERS:
            try: _vendor_fallback_indexes.add(column_index_from_letter(letter) - 1)
            except Exception: pass
        needed_col_indexes = set(source_map.values()) | _vendor_fallback_indexes
        if not needed_col_indexes: return []
        needed_col_indexes = sorted(i for i in needed_col_indexes if i is not None and i >= 0)
        ranges = []
        for col_idx in needed_col_indexes:
            letter = column_letter_from_index(col_idx + 1)
            ranges.append(f"'{sheet_title}'!{letter}{data_start_row}:{letter}")
        batch = google_sheets_values_batch_get(source['spreadsheet_id'], ranges, value_render_option='FORMATTED_VALUE', major_dimension='COLUMNS')
        value_ranges = batch.get('valueRanges', [])
        if len(value_ranges) != len(needed_col_indexes):
            raise ValueError("batchGet column count mismatch")
        columns_by_idx = {}
        max_len = 0
        for col_idx, value_range in zip(needed_col_indexes, value_ranges):
            values = value_range.get('values') or []
            col_values = values[0] if values and isinstance(values[0], list) else []
            col_values = [clean(v) or '' for v in col_values]
            columns_by_idx[col_idx] = col_values
            max_len = max(max_len, len(col_values))
        rows = []
        # FIX V9: carry-over vendor name untuk row yang vendor-nya kosong.
        # Di sheet source, vendor name sering hanya diisi di row pertama untuk
        # group yang sama, row berikutnya dikosongkan (merged cell visual).
        # Tanpa carry-over, row berikutnya akan di-skip karena vendor kosong.
        last_known_vendor = ''
        for offset in range(max_len):
            values_by_idx = {col_idx: (vals[offset] if offset < len(vals) else '') for col_idx, vals in columns_by_idx.items()}
            vendor_candidates = []
            for field in ('vendor_name', 'vendor'):
                col_idx = source_map.get(field)
                if col_idx is not None: vendor_candidates.append(values_by_idx.get(col_idx, ''))
            # FIX V9: konversi letter ke 0-indexed, bukan pakai hardcoded tuple angka
            for letter in IMPORT_FALLBACK_SOURCE_VENDOR_LETTERS:
                try:
                    col_idx = column_index_from_letter(letter) - 1
                    vendor_candidates.append(values_by_idx.get(col_idx, ''))
                except Exception: pass
            vendor_candidates = [clean(v) for v in vendor_candidates if clean(v)]
            row_vendor = next((v for v in vendor_candidates if v), '')

            # FIX V9: carry-over vendor name dari row sebelumnya kalau row ini kosong.
            # Tapi hanya kalau row ini punya data PO/item (bukan row blank/subtotal).
            if not row_vendor and last_known_vendor:
                # Cek apakah row ini punya konten lain (PO Yupi / item_name)
                po_yupi_val = clean(values_by_idx.get(source_map.get('po_yupi', -1), '')) if 'po_yupi' in source_map else ''
                item_name_val = clean(values_by_idx.get(source_map.get('item_name', -1), '')) if 'item_name' in source_map else ''
                if po_yupi_val or item_name_val:
                    row_vendor = last_known_vendor
                    vendor_candidates = [row_vendor]

            if row_vendor:
                last_known_vendor = row_vendor

            # FIX V9: pakai import_vendor_match() (fuzzy matching) bukan exact match
            # supaya vendor "CURT GEORGI GMBH & CO" di ImportVendor match dengan
            # "CURT GEORGI GMBH & CO. KG" di sheet source.
            if vendor_set and not any(import_vendor_match(v, vendor_set) for v in vendor_candidates if v): continue
            row = {'_row_key': f"{source['key']}:{data_start_row + offset}", '_source_key': source['key'], '_source_label': source['label'], '_spreadsheet_id': source['spreadsheet_id'], '_gid': source['gid'], '_sheet_row': data_start_row + offset, '_vendor_name': row_vendor}
            for col in mapping_columns:
                col_idx = source_map.get(col['field'])
                row[col['field']] = values_by_idx.get(col_idx, '') if col_idx is not None else ''
            if not any(row.get(col['field']) for col in mapping_columns): continue
            rows.append(row)
        return rows
    except Exception as exc:
        print(f"[import_source_rows_fast] fast path failed: {exc}")
        df = read_public_sheet_csv(source['spreadsheet_id'], source['gid'])
        mapping_columns = import_all_mapping_columns(columns)
        source_map = import_source_column_map(df, mapping_columns)
        # FIX V9: pakai import_detect_data_start untuk dynamic start detection
        header_idx_fallback = import_detect_header_row(df)
        start_idx = import_detect_data_start(df, header_idx_fallback)
        rows = []
        for idx in range(start_idx, len(df)):
            values = [clean(v) or '' for v in df.iloc[idx].tolist()]
            vendor_candidates = import_row_vendor_candidates(values, source_map, columns)
            row_vendor = next((v for v in vendor_candidates if v), '')
            # FIX V9: pakai import_vendor_match() (fuzzy matching) bukan exact match
            if vendor_set and not any(import_vendor_match(v, vendor_set) for v in vendor_candidates if v): continue
            row = {'_row_key': f"{source['key']}:{idx + 1}", '_source_key': source['key'], '_source_label': source['label'], '_spreadsheet_id': source['spreadsheet_id'], '_gid': source['gid'], '_sheet_row': idx + 1, '_vendor_name': row_vendor}
            for col in mapping_columns:
                col_idx = source_map.get(col['field'])
                row[col['field']] = values[col_idx] if col_idx is not None and col_idx < len(values) else ''
            if not any(row.get(col['field']) for col in mapping_columns): continue
            rows.append(row)
        return rows

def import_sheet_rows(force_metadata=False):
    columns = import_layout_columns(force=force_metadata)
    filter_vendors, vendor_source = import_vendor_filter_names()
    vendor_set = {v.strip().lower() for v in filter_vendors if v and v.strip()}
    if not vendor_set: return columns, []
    rows = []
    for source in IMPORT_SOURCE_SHEETS:
        rows.extend(import_source_rows_fast(source, columns, vendor_set))
    import_meta_set('last_import_vendor_filter', {'source': vendor_source, 'count': len(vendor_set)})
    return columns, rows

def import_truthy_checkbox_value(value):
    s = (clean(value) or '').strip().lower()
    return s in ('true', '1', 'yes', 'ya', 'y', 'checked', 'done', 'ok', '✓', '✅')

def import_normalize_checkbox(value):
    raw = clean(value)
    if not raw: return ''
    return 'TRUE' if import_truthy_checkbox_value(raw) else 'FALSE'

def import_date_from_value(value):
    raw = clean(value)
    if not raw: return None
    s = str(raw).strip()
    if re.match(r'^\d+(\.0+)?$', s):
        try:
            serial = float(s)
            if 25000 <= serial <= 60000: return (datetime(1899, 12, 30) + timedelta(days=serial)).date()
        except: pass
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%y', '%d-%b-%Y', '%d-%b-%y', '%d %b %Y', '%d %b %y'):
        try: return datetime.strptime(s, fmt).date()
        except: pass
    # FIX V9: Yearless date formats (e.g. "10-Jun", "25 Dec") — sheet tidak punya
    # kolom tahun, hanya dd-mmm.
    #
    # Logika LAMA (bug): selalu pakai tahun ini, kalau tanggal < today → naikkan
    # ke tahun depan. Akibatnya tanggal 10-Jun diisi hari ini (25-Jun) jadi
    # 10-Jun-2027 (tahun depan), padahal user maksudnya 10-Jun-2026 (yang baru
    # lewat 15 hari).
    #
    # Logika BARU: pakai tahun berjalan (today.year). Kalau tanggal hasilnya
    # lebih dari 6 bulan ke depan dari hari ini, asumsikan user maksud tahun lalu.
    # Kenapa 6 bulan? Karena PO Date biasanya diisi untuk PO yang dikirim
    # baru-baru ini (mundur maksimal ±6 bulan), bukan PO masa depan.
    yearless_formats = ('%d %b', '%d-%b', '%d %B', '%d-%B', '%b %d', '%b-%d', '%B %d', '%B-%d')
    today = date.today()
    for fmt in yearless_formats:
        try: d = datetime.strptime(s, fmt).date()
        except: continue
        try: d = d.replace(year=today.year)
        except: d = d.replace(year=today.year + 1)
        # Kalau tanggal hasil > 6 bulan (183 hari) ke depan, asumsikan tahun lalu
        # Contoh: today=25-Jun-2026, input "10-Dec" → 10-Dec-2026 (5.5 bln ke depan) → keep
        #         today=25-Jun-2026, input "25-Feb" → 25-Feb-2026 (4 bln lalu) → keep
        #         today=25-Jun-2026, input "10-Jan" → 10-Jan-2026 (5.5 bln lalu) → keep
        #         today=25-Jun-2026, input "10-Aug" → 10-Aug-2026 (1.5 bln depan) → keep (wajar)
        #         today=25-Jun-2026, input "10-May" → 10-May-2026 (1.5 bln lalu) → keep (wajar)
        # Kalau tanggal hasil > 183 hari ke depan → pindah ke tahun lalu
        if (d - today).days > 183:
            try: d = d.replace(year=today.year - 1)
            except: pass
        return d
    return parse_date(raw)

def import_date_output(value, fallback=None):
    raw = clean(value)
    if raw: return str(raw)
    if fallback: return fallback.isoformat()
    return ''

def import_float_value(value):
    raw = clean(value)
    if not raw: return None
    s = str(raw).replace(',', '').replace('Rp', '').replace('IDR', '').strip()
    try: return float(s)
    except: return None

def import_format_number(value):
    if value is None: return ''
    try: f = float(value)
    except: return str(value)
    if abs(f - round(f)) < 0.000001: return str(int(round(f)))
    return f'{f:.2f}'.rstrip('0').rstrip('.')

def apply_import_formula_columns(row):
    if not isinstance(row, dict): return row
    for field in IMPORT_CHECKBOX_FIELDS:
        if field in row: row[field] = import_normalize_checkbox(row.get(field))
    has_business_data = any(clean(row.get(f)) for f in ('po_yupi', 'yupi_po', 'po_sementara', 'item_name', 'vendor_name', 'vendor', 'so'))
    po_send_date_raw = clean(row.get('po_send_date'))
    po_date_email = clean(row.get('po_date_by_email'))
    po_send_manual = clean(row.get('_po_send_date_manual')) == '1'
    if po_send_date_raw and not po_send_manual:
        if not po_date_email:
            row['po_date_by_email'] = po_send_date_raw; po_date_email = po_send_date_raw
        if clean(row.get('po_date_by_email')) == po_send_date_raw:
            row['po_send_date'] = ''; po_send_date_raw = ''
    po_send_date = po_send_date_raw
    status = (clean(row.get('status')) or '').upper()
    if has_business_data and not po_send_date: status = 'NEW'
    elif has_business_data and (not status or status == 'NEW'): status = 'ON PROCESS'
    if status:
        status = next((opt for opt in IMPORT_STATUS_OPTIONS if opt == status), status)
        row['status'] = status
    row['site'] = import_nonblank(row.get('site'))
    po_yupi = import_nonblank(row.get('po_yupi')) or import_nonblank(row.get('yupi_po'))
    row['po_yupi'] = po_yupi; row['yupi_po'] = po_yupi
    row['vendor'] = import_nonblank(row.get('vendor')) or import_nonblank(row.get('vendor_name'))
    try:
        req_raw = import_nonblank(row.get('source_req_dlv_date')) or import_nonblank(row.get('req_dlv_date'))
        row['source_req_dlv_date'] = req_raw
        if req_raw:
            req_parsed = import_date_from_value(req_raw)
            if req_parsed: row['req_dlv_date'] = req_parsed.isoformat()
        for _date_field in ('po_date_by_email', 'etd', 'eta', 'reschedule'):
            _raw = import_nonblank(row.get(_date_field))
            if not _raw: continue
            _parsed = import_date_from_value(_raw)
            if _parsed: row[_date_field] = _parsed.isoformat()
    except: pass
    etd_date = import_date_from_value(row.get('etd'))
    eta_date = import_date_from_value(row.get('eta'))
    req_date = req_parsed if 'req_parsed' in locals() else None
    status_upper = (clean(row.get('status')) or '').upper()
    if not has_business_data: row['days_left'] = ''
    elif status_upper == 'DELIVERED': row['days_left'] = '✅'
    elif status_upper == 'CANCELED': row['days_left'] = '❌'
    elif req_date: row['days_left'] = str((req_date - date.today()).days)
    else: row['days_left'] = ''
    if row.get('days_left') == '✅': row['arrival_check'] = '⚪'
    elif not eta_date or not req_date: row['arrival_check'] = ''
    elif eta_date <= req_date: row['arrival_check'] = '🟢 On Schedule'
    else: row['arrival_check'] = f'🔴 Delay ({(eta_date - req_date).days}D)'
    price = import_float_value(row.get('purchase_price'))
    qty = import_float_value(row.get('ord_qty'))
    if price is not None and qty is not None: row['purchase_amount'] = import_format_number(price * qty)
    if etd_date and eta_date: row['lt_days'] = str((eta_date - etd_date).days)
    else: row['lt_days'] = ''
    # NEW: Payment Date "Overdue" logic — if today > ETA + TOP days AND payment
    # is not DONE AND the user hasn't entered a payment_date yet, mark
    # payment_date as "Overdue" (red). The frontend reads this sentinel value
    # and renders it in red. Actual user-entered payment_date values are
    # preserved as-is — they take precedence over the auto-computed "Overdue".
    payment_value = (clean(row.get('payment')) or '').upper()
    payment_date_value = clean(row.get('payment_date'))
    # Only auto-mark "Overdue" if payment_date is currently empty or already
    # holds the sentinel (so we re-evaluate it each time as ETA/TOP changes).
    # A real date string (YYYY-MM-DD) is preserved.
    is_sentinel_or_empty = (not payment_date_value) or (payment_date_value == 'Overdue')
    # FIX V10: Kalau payment sudah DONE, JANGAN tampilkan 'Overdue' di payment_date.
    # Clear sentinel 'Overdue' supaya field kosong (atau pakai real date kalau ada).
    if payment_value == 'DONE':
        if payment_date_value == 'Overdue':
            row['payment_date'] = ''  # clear sentinel — payment sudah DONE, tidak overdue
    elif is_sentinel_or_empty and eta_date:
        try:
            top_days_str = clean(row.get('top')) or '0'
            # Strip non-digit chars (e.g. "30 days", "Net 30", "30D")
            top_match = re.search(r'\d+', str(top_days_str))
            top_days = int(top_match.group(0)) if top_match else 0
            due_date = eta_date + timedelta(days=top_days)
            if date.today() > due_date:
                row['payment_date'] = 'Overdue'
            else:
                # Not overdue anymore — clear the sentinel if present.
                if payment_date_value == 'Overdue':
                    row['payment_date'] = ''
        except Exception:
            pass
    return row

def import_row_payload(row, columns):
    payload = {col['field']: '' if row.get(col['field']) is None else str(row.get(col['field'])) for col in import_all_mapping_columns(columns)}
    return apply_import_formula_columns(payload)

def import_row_identity_detail_fingerprint(row):
    parts = [import_nonblank(row.get('item_name')), import_nonblank(row.get('spec')), import_nonblank(row.get('ord_qty')), import_nonblank(row.get('unit_price')), import_nonblank(row.get('unit'))]
    fingerprint = '|'.join((p or '').strip().upper() for p in parts)
    return fingerprint if any(parts) else None

def import_row_identity_payload(row):
    # FIX V12: Primary identity key kini mengikutkan po_sementara + req_dlv_date
    # (dari source_req_dlv_date, yaitu kolom K sheet asli — SEBELUM reschedule
    # diterapkan ke req_dlv_date dashboard).
    #
    # Masalah sebelumnya:
    #   Key = {po_yupi, item_yupi} — tidak membedakan 2 baris yang punya
    #   po_yupi & item_yupi sama tapi req_dlv_date beda (mis. SVOI...-01 vs -02,
    #   atau 1 PO yang belum di-reschedule vs yang sudah).
    #   Akibatnya:
    #   1. Saat sync, 2 baris sheet di-match ke 1 row DB yang sama → salah satu
    #      baris tidak tersimpan.
    #   2. Kalau req_dlv_date di dashboard di-update karena reschedule, hash
    #      primary berubah → baris baru di-insert duplikat.
    #
    # Solusi: sertakan po_sementara + source_req_dlv_date dalam primary key.
    # source_req_dlv_date = kolom K sheet asli = tanggal ASLI sebelum reschedule,
    # sehingga reschedule (yg hanya mengubah req_dlv_date dashboard, bukan kolom K)
    # TIDAK mengubah identity hash.
    po_yupi = (import_nonblank(row.get('po_yupi')) or import_nonblank(row.get('yupi_po')) or '').strip().upper()
    item_yupi = (import_nonblank(row.get('item_yupi')) or '').strip().upper()
    po_sementara = (import_nonblank(row.get('po_sementara')) or '').strip().upper()
    # Pakai source_req_dlv_date (kolom K asli) sebagai bagian identity.
    # Fallback ke req_dlv_date kalau source_req_dlv_date kosong.
    req_dlv = (import_nonblank(row.get('source_req_dlv_date')) or import_nonblank(row.get('req_dlv_date')) or '').strip().upper()
    detail_fp = import_row_identity_detail_fingerprint(row)
    if po_yupi and item_yupi and po_sementara:
        return {'po_yupi': po_yupi, 'item_yupi': item_yupi, 'po_sementara': po_sementara, 'req_dlv_date': req_dlv}
    if po_yupi and item_yupi:
        return {'po_yupi': po_yupi, 'item_yupi': item_yupi, 'req_dlv_date': req_dlv}
    if po_yupi:
        return {'po_yupi': po_yupi, 'item_yupi': '(none)', 'req_dlv_date': req_dlv, 'detail': detail_fp or '(blank)'}
    if po_sementara and item_yupi:
        return {'po_sementara': po_sementara, 'item_yupi': item_yupi, 'req_dlv_date': req_dlv}
    if po_sementara:
        return {'po_sementara': po_sementara, 'item_yupi': '(none)', 'req_dlv_date': req_dlv, 'detail': detail_fp or '(blank)'}
    return {'source': clean(row.get('_source_key')) or '', 'sheet_row': str(row.get('_sheet_row') or '')}

def import_row_identity_secondary(row):
    # FIX V12: Secondary identity (fallback lookup) juga sertakan req_dlv_date
    # supaya 2 baris dengan po_sementara + item_yupi sama tapi tanggal beda
    # tidak di-merge ke 1 row yang sama lewat jalur secondary lookup.
    po_sementara = (import_nonblank(row.get('po_sementara')) or '').strip().upper()
    item_yupi = (import_nonblank(row.get('item_yupi')) or '').strip().upper()
    req_dlv = (import_nonblank(row.get('source_req_dlv_date')) or import_nonblank(row.get('req_dlv_date')) or '').strip().upper()
    if not po_sementara: return None
    return {'po_sementara': po_sementara, 'item_yupi': item_yupi or '(none)', 'req_dlv_date': req_dlv}

def _identity_to_uid(payload):
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(',', ':'))
    return hashlib.sha1(raw.encode('utf-8')).hexdigest()

def import_row_source_uid(row, columns):
    return _identity_to_uid(import_row_identity_payload(row))

def import_row_secondary_uid(row):
    sec = import_row_identity_secondary(row)
    return _identity_to_uid(sec) if sec else None

def merge_import_existing_payload(existing_payload, sheet_payload):
    merged = dict(sheet_payload or {})
    existing_payload = existing_payload or {}
    row_exists_in_db = bool(existing_payload)

    # Req Dlv Date selalu mengikuti sheet (source-managed, behavior normal).
    # Kalau ada reschedule, row di-highlight kuning di frontend — tidak ada
    # proteksi khusus di backend untuk req_dlv_date.
    for field in IMPORT_LOCAL_EDIT_FIELDS:
        old_value = existing_payload.get(field)
        new_value = merged.get(field)
        if field in IMPORT_USER_LOCAL_ONLY_FIELDS and row_exists_in_db:
            # Hanya restore old_value kalau memang ada isinya.
            # Kalau old_value None/blank (field belum pernah diisi user),
            # biarkan new_value dari sheet (jangan overwrite dengan None).
            if not import_blankish(old_value):
                merged[field] = old_value
            continue
        if field in IMPORT_SOURCE_MANAGED_FIELDS:
            # Behavior normal: kalau sheet ada nilai, pakai sheet value
            if not import_blankish(new_value): continue
            # Kalau sheet kosong tapi existing ada, keep existing
            if not import_blankish(old_value): merged[field] = old_value
            continue
        if not import_blankish(old_value): merged[field] = old_value
    return apply_import_formula_columns(merged)

def import_dashboard_row_to_dict(row, columns, vendor_attrs_map=None, pre_parsed_data=None):
    """Convert an ImportDashboardRow to a dict matching the column schema.

    vendor_attrs_map (optional): {vendor_name_lower: {'origin':..., 'top':..., 'non_ski':...}}
    When provided, Origin and TOP are injected from this map (matched by
    the row's vendor name). Falls back to the row's stored data_json values
    if no match is found.

    pre_parsed_data (optional): FIX V10 — kalau sudah ada data yang di-parse
    + apply_import_formula_columns sebelumnya (mis. di get_import_data),
    lewatkan di sini supaya tidak panggil apply_import_formula_columns lagi
    (function berat — parse tanggal, lookup PIC, dll).
    """
    if pre_parsed_data is not None:
        data = dict(pre_parsed_data)
    else:
        try: data = json.loads(row.data_json or '{}')
        except: data = {}
        data = apply_import_formula_columns(dict(data))
    # Inject vendor attributes (Origin, TOP, Non SKI) from the ImportVendor
    # table. These are looked up by the row's vendor name (case-insensitive).
    if vendor_attrs_map is not None:
        row_vendor = import_nonblank(data.get('vendor')) or import_nonblank(data.get('vendor_name')) or import_nonblank(row.vendor_name)
        if row_vendor:
            attrs = vendor_attrs_map.get(str(row_vendor).strip().lower())
            if attrs:
                # Only inject if not already present in data (data wins for
                # backward-compat — though normally these fields are empty
                # in data_json because they come from the vendor master).
                if not import_nonblank(data.get('origin')):
                    data['origin'] = attrs.get('origin') or ''
                if not import_nonblank(data.get('top')):
                    data['top'] = attrs.get('top') or ''
                # Non SKI — injected from the vendor master so the rightmost
                # column reflects what was uploaded in the Vendor Import
                # template (col D "Non SKI").
                if not import_nonblank(data.get('non_ski')):
                    data['non_ski'] = attrs.get('non_ski') or ''
    out = {}
    for col in columns:
        field = col.get('field')
        out[field] = '' if data.get(field) is None else data.get(field, '')
    out.update({'_row_key': row.row_key, '_source_key': row.source_key, '_source_label': row.source_label, '_source_uid': row.source_uid, '_sheet_row': row.sheet_row, '_vendor_name': row.vendor_name, '_dashboard_id': row.id, '_first_seen_at': utc_isoformat(row.first_seen_at) if row.first_seen_at else '', '_last_seen_at': utc_isoformat(row.last_seen_at) if row.last_seen_at else '', '_updated_at': utc_isoformat(row.updated_at) if row.updated_at else ''})
    return out

def import_layout_tracker_visible_rows(columns=None):
    columns = import_layout_columns() if columns is None else columns
    sheet_title = import_layout_target_sheet_title()
    resp = google_sheets_values_get(IMPORT_LAYOUT_SHEET_ID, f"'{sheet_title}'!A1:DI", value_render_option='FORMATTED_VALUE')
    values = resp.get('values') or []
    if not values: return []
    try:
        alias_keys = set()
        for vs in IMPORT_COLUMN_ALIASES.values(): alias_keys.update(import_header_key(v) for v in vs if v)
        alias_keys.update(import_header_key(c.get('label')) for c in import_all_mapping_columns(columns))
        alias_keys.update(import_header_key(c.get('field')) for c in import_all_mapping_columns(columns))
        alias_keys.discard('')
        critical_labels = ('poyupi', 'reqdlvdate', 'posementara', 'itemname', 'itemyupi')
        best_idx, best_score = 0, -1
        max_scan = min(len(values), 12)
        for idx in range(max_scan):
            labels = [import_header_key(v) for v in values[idx]]
            score = sum(1 for label in labels if label in alias_keys)
            for crit in critical_labels:
                if crit in labels: score += 5
            if score > best_score: best_idx = idx; best_score = score
        header_idx = best_idx if best_score >= 4 else 0
    except: header_idx = 0
    header = values[header_idx]
    by_header = {}
    for idx, raw in enumerate(header):
        key = import_header_key(raw)
        if key and key not in by_header: by_header[key] = idx
    mapping_columns = import_all_mapping_columns(columns)
    field_to_idx = {}
    for col in mapping_columns:
        field = col.get('field')
        if not field: continue
        keys = []
        keys.extend(IMPORT_COLUMN_ALIASES.get(field, []))
        keys.extend([import_header_key(col.get('label')), import_header_key(field)])
        idx = next((by_header[k] for k in keys if k and k in by_header), None)
        field_to_idx[field] = idx
    rows = []
    last_po_yupi = ''
    last_po_sementara_prefix = ''
    for row_offset, row_values in enumerate(values[header_idx + 1:], start=header_idx + 2):
        payload = {'_source_key': IMPORT_LAYOUT_SOURCE_KEY, '_source_label': 'Import Tracker', '_sheet_row': row_offset}
        any_value = False
        for field, idx in field_to_idx.items():
            val = row_values[idx] if idx is not None and idx < len(row_values) else ''
            payload[field] = val
            if import_nonblank(val): any_value = True
        if not any_value: continue
        po_yupi_now = import_nonblank(payload.get('po_yupi')) or import_nonblank(payload.get('yupi_po'))
        po_sementara_now = import_nonblank(payload.get('po_sementara'))
        if po_yupi_now:
            last_po_yupi = po_yupi_now
            last_po_sementara_prefix = _extract_po_yupi_from_po_sementara(po_sementara_now) or ''
        else:
            extracted = _extract_po_yupi_from_po_sementara(po_sementara_now)
            if extracted:
                payload['po_yupi'] = extracted; payload['yupi_po'] = extracted
                last_po_yupi = extracted; last_po_sementara_prefix = extracted
            elif last_po_yupi:
                cur_prefix = _extract_po_yupi_from_po_sementara(po_sementara_now) or ''
                if cur_prefix and cur_prefix == last_po_sementara_prefix:
                    payload['po_yupi'] = last_po_yupi; payload['yupi_po'] = last_po_yupi
        py = import_nonblank(payload.get('po_yupi')) or import_nonblank(payload.get('yupi_po'))
        if py: payload['po_yupi'] = py; payload['yupi_po'] = py
        if not import_blankish(payload.get('po_send_date')): payload['_po_send_date_manual'] = '1'
        if any(import_nonblank(payload.get(f)) for f in ('po_send_date', 'status', 'po_yupi', 'yupi_po', 'po_sementara', 'item_yupi', 'item_name', 'vendor_name', 'vendor', 'so')):
            rows.append(apply_import_formula_columns(payload))
    return rows

def _extract_po_yupi_from_po_sementara(po_sementara):
    raw = clean(po_sementara)
    if not raw: return ''
    s = str(raw).strip().upper()
    m = re.match(r'^SVO[KI]?(\d+)-\d+$', s)
    if m: return m.group(1)
    m = re.match(r'^[A-Z]+(\d{5,})-\d+$', s)
    if m: return m.group(1)
    return ''

IMPORT_TRACKER_AUTHORITATIVE_FIELDS = {
    # FIX V9: Hanya field-field yang MEMANG diisi di tracker sheet yang authoritative.
    # Field-field user-edit manual (status, incoterm, forwarder, checklist, dll)
    # TIDAK boleh di sini — kalau ada, akan di-overwrite saat sync dari tracker.
    #
    # SEBELUMNYA (BUG): field-field ini ada di authoritative → saat sync, user edit
    # status/incoterm/forwarder/checklist hilang karena di-overwrite dengan nilai dari
    # tracker (yang biasanya kosong atau nilai lama).
    #
    # SESUDAH FIX: hanya field yang memang dikelola tracker sheet:
    'po_send_date', '_po_send_date_manual', 'etd', 'eta', 'import_remarks',
    # Field-field di bawah DIPINDAH ke USER_LOCAL_ONLY (tidak di-overwrite sync):
    # 'status', 'incoterm', 'forwarder', 'bl_number', 'inv_no', 'non_ski',
    # 'sap_input', 'bl_awb', 'invoice', 'pl', 'hc', 'msds', 'coa', 'coo', 'soft_copy_doc',
}

def merge_import_tracker_payload(existing_payload, tracker_payload):
    merged = dict(existing_payload or {})
    tracker_payload = tracker_payload or {}
    for field, tracker_value in tracker_payload.items():
        if field.startswith('_') and field != '_po_send_date_manual': continue
        if import_blankish(tracker_value): continue
        if field in IMPORT_TRACKER_AUTHORITATIVE_FIELDS:
            merged[field] = tracker_value; continue
        if import_blankish(merged.get(field)): merged[field] = tracker_value
    if not import_blankish(merged.get('po_yupi')) and import_blankish(merged.get('yupi_po')): merged['yupi_po'] = merged.get('po_yupi')
    if not import_blankish(merged.get('yupi_po')) and import_blankish(merged.get('po_yupi')): merged['po_yupi'] = merged.get('yupi_po')
    if not import_blankish(merged.get('source_req_dlv_date')) and import_blankish(merged.get('req_dlv_date')): merged['req_dlv_date'] = merged.get('source_req_dlv_date')
    if not import_blankish(merged.get('req_dlv_date')) and import_blankish(merged.get('source_req_dlv_date')): merged['source_req_dlv_date'] = merged.get('req_dlv_date')
    return apply_import_formula_columns(merged)

def sync_import_tracker_to_dashboard(columns=None):
    columns = import_layout_columns() if columns is None else columns
    tracker_rows = import_layout_tracker_visible_rows(columns)
    if not tracker_rows: return {'rows': 0, 'seen': 0, 'added': 0, 'skipped': 0}
    existing_rows = ImportDashboardRow.query.filter(ImportDashboardRow.source_key != 'import_tracker').all()
    existing_by_key = {}
    for row in existing_rows:
        try: data = json.loads(row.data_json or '{}')
        except: data = {}
        data = apply_import_formula_columns(dict(data))
        for key in import_layout_target_candidate_keys(data): existing_by_key.setdefault(key, row)
    now = datetime.utcnow()
    seen, added, skipped = 0, 0, 0
    duplicate_counts = {}
    for tracker_payload in tracker_rows:
        keys = import_layout_target_candidate_keys(tracker_payload)
        if not keys: skipped += 1; continue
        current = next((existing_by_key[k] for k in keys if k in existing_by_key), None)
        if current:
            try: existing_payload = json.loads(current.data_json or '{}')
            except: existing_payload = {}
            current.data_json = json.dumps(merge_import_tracker_payload(existing_payload, tracker_payload), ensure_ascii=False)
            current.updated_at = now
            if current.source_key == 'import_tracker':
                current.last_seen_at = now
                current.sheet_row = tracker_payload.get('_sheet_row') or current.sheet_row
            seen += 1; continue
        source_uid_raw = '|'.join(keys)
        source_uid = hashlib.sha1(source_uid_raw.encode('utf-8')).hexdigest()
        duplicate_counts[source_uid] = duplicate_counts.get(source_uid, 0) + 1
        row_key = f"import_tracker:{source_uid}:{duplicate_counts[source_uid]}"
        data = apply_import_formula_columns(dict(tracker_payload))
        row = ImportDashboardRow(row_key=row_key, source_key='import_tracker', source_label='Import Tracker', source_uid=source_uid, sheet_row=tracker_payload.get('_sheet_row'), vendor_name=import_nonblank(data.get('vendor_name')) or import_nonblank(data.get('vendor')), data_json=json.dumps(data, ensure_ascii=False), first_seen_at=now, last_seen_at=now, updated_at=now)
        db.session.add(row)
        added += 1
        for key in keys: existing_by_key.setdefault(key, row)
    return {'rows': len(tracker_rows), 'seen': seen, 'added': added, 'skipped': skipped}

def _migrate_import_source_uid_v12():
    """FIX V12 Migration: Re-compute source_uid semua row import yang ada di DB
    supaya hash-nya menggunakan skema baru (po_yupi + item_yupi + po_sementara +
    req_dlv_date). Row yang masih pakai hash skema lama (tanpa req_dlv_date) akan
    gagal di-match saat copy sheet → dianggap baru → duplikat.

    Fungsi ini dipanggil SEKALI di awal sync_import_sheet_to_dashboard.
    Setelah semua row punya source_uid baru, copy sheet berjalan normal.
    Idempoten: aman dipanggil berkali-kali (re-hash tidak mengubah hasil kalau
    skema sudah baru).
    """
    try:
        all_rows = ImportDashboardRow.query.filter(
            ImportDashboardRow.source_key.in_(_IMPORT_VISIBLE_SOURCE_KEYS)
        ).all()
        migrated = 0
        for row in all_rows:
            try:
                data = json.loads(row.data_json or '{}')
            except Exception:
                data = {}
            new_uid = import_row_source_uid(data, [])
            if row.source_uid != new_uid:
                row.source_uid = new_uid
                migrated += 1
        if migrated:
            db.session.commit()
            print(f"[_migrate_import_source_uid_v12] Re-hashed {migrated} row(s) ke skema V12 (po_sementara + req_dlv_date dalam key)")
    except Exception as exc:
        try: db.session.rollback()
        except: pass
        print(f"[_migrate_import_source_uid_v12] Gagal: {exc}")

def sync_import_sheet_to_dashboard():
    purged_legacy = 0
    sheet_rows = []
    columns = import_layout_columns(force=True)
    filter_vendors, vendor_source = [], 'none'
    vendor_count = len(import_uploaded_vendor_names())

    # FIX V12: Jalankan migrasi hash sebelum apapun, supaya row lama yang
    # punya source_uid skema lama (tanpa req_dlv_date) di-update terlebih
    # dahulu dan bisa di-match dengan benar di loop upsert di bawah.
    _migrate_import_source_uid_v12()

    try:
        stale = ImportDashboardRow.query.filter(ImportDashboardRow.source_key.in_(_LEGACY_IMPORT_SOURCE_KEYS)).all()
        purged_legacy = len(stale)
        for row in stale: db.session.delete(row)
        if purged_legacy: db.session.commit()
    except: db.session.rollback(); purged_legacy = 0
    try:
        sheet_rows = import_layout_tracker_visible_rows(columns)
        # Use UPLOADED vendor names (from ImportVendor table) — NOT the
        # import_vendor_filter_names() fallback that includes existing
        # dashboard vendor names. The fallback would defeat the purpose
        # of the vendor filter for both source-sheet sync and purge.
        filter_vendors = import_uploaded_vendor_names()
        vendor_source = 'vendor_import' if filter_vendors else 'none'
    except:
        import traceback; traceback.print_exc()
        return {'added': 0, 'updated': 0, 'seen': 0, 'sheet_rows': 0, 'vendor_count': vendor_count, 'vendor_filter_count': 0, 'vendor_filter_source': 'none', 'purged_legacy': purged_legacy, 'copy_only': True, 'columns': columns, 'error': 'Failed to read the live Import tracker sheet.', 'source_sheet_url': f'https://docs.google.com/spreadsheets/d/{IMPORT_LAYOUT_SHEET_ID}/edit#gid={IMPORT_LAYOUT_GID}'}

    # ── ALSO pull directly from the 2 source sheets (RM=source_1, SP=source_2)
    # The consolidated layout sheet may not contain all rows from both
    # sources — SP rows in particular are often missing. By pulling directly
    # from each source sheet we guarantee both RM and SP data make it into
    # the dashboard. Source rows are appended to `sheet_rows` so they go
    # through the same upsert logic below.
    #
    # IMPORTANT: Only copy rows whose vendor matches the uploaded Vendor
    # Import list. Rows with vendors NOT in the uploaded list are skipped
    # at the source level (vendor_set filter) so they never enter the
    # dashboard. If no vendors have been uploaded yet, we skip source-sheet
    # sync entirely (no point pulling thousands of unfiltered rows).
    #
    # NOTE: We use import_uploaded_vendor_names() directly (NOT
    # import_vendor_filter_names()) because the latter falls back to
    # existing dashboard vendor names when nothing is uploaded — which
    # would defeat the purpose of filtering.
    uploaded_vendor_names = import_uploaded_vendor_names()
    uploaded_vendor_set = {v.strip().lower() for v in uploaded_vendor_names if v and v.strip()}
    source_rows_added = 0
    source_errors = []
    if uploaded_vendor_set:
        for source in IMPORT_SOURCE_SHEETS:
            try:
                # vendor_set filter → only rows with matching vendors are returned
                src_rows = import_source_rows_fast(source, columns, uploaded_vendor_set)
                sheet_rows.extend(src_rows)
                source_rows_added += len(src_rows)
            except Exception as src_exc:
                import traceback; traceback.print_exc()
                source_errors.append({'source': source.get('key'), 'label': source.get('label'), 'error': str(src_exc)})
    else:
        # No uploaded vendors → don't pull from source sheets at all (would
        # import the entire sheet unfiltered, which is not what the user wants).
        source_errors.append({'source': 'all', 'label': 'All sources', 'error': 'No uploaded vendor list — source sheet sync skipped. Upload vendors via Vendor Import first.'})

    # ── FIX V13: Deduplikasi sheet_rows berdasarkan source_uid SEBELUM upsert loop.
    #
    # MASALAH:
    #   sheet_rows bisa berisi baris yang sama berkali-kali karena:
    #   1. Dual-source pull: row yang sama datang dari layout tracker sheet
    #      DAN dari source sheet (RM/SP). Kedua sumber biasanya berbeda di
    #      kolom `so` (layout sheet sering kosong, source sheet terisi).
    #   2. Akibatnya, 1 baris sheet asli → 2 (atau lebih) baris di sheet_rows.
    #
    #   Loop upsert di bawah menggunakan counter `consumed_per_uid` yang
    #   AKAN BERTAMBAH setiap kali baris dengan source_uid yang sama di-proses.
    #   Jika ada 2 sheet_rows dengan source_uid sama, sheet_row kedua tidak
    #   menemukan slot kandidat (karena candidates hanya 1) → DIANGGAP BARU
    #   → di-insert jadi row baru di DB → duplikat.
    #
    # SOLUSI:
    #   Group sheet_rows by source_uid. Untuk setiap group, simpan hanya 1 row
    #   — pilih row dengan field paling lengkap (paling banyak field non-blank,
    #   priority ke `so` terisi karena itu field yang paling sering beda antara
    #   layout vs source sheet). Kalau ada row A yang punya `so` dan row B yang
    #   tidak, simpan row A; field-field lain di-merge dari keduanya.
    #
    # EFEK:
    #   - Duplikat hilang: 1 baris sheet asli → 1 source_uid → 1 DB row.
    #   - Penambahan baris tetap bekerja: kalau sheet bertambah dari 2 → 3 baris
    #     untuk PO yang sama (dengan req_dlv_date beda), ketiganya punya
    #     source_uid berbeda → 3 DB row.
    try:
        _dedup_before_upsert = {}
        _dedup_seen_payloads = {}
        for _sr in sheet_rows:
            _sp = import_row_payload(_sr, columns)
            _uid = import_row_source_uid(_sp, columns)
            if not _uid:
                # Tanpa uid, tidak bisa dedup → simpan apa adanya
                _dedup_before_upsert.setdefault('__no_uid__', []).append(_sr)
                continue
            # Hitung richness = jumlah field non-blank (makin banyak makin baik)
            _rich = sum(1 for _v in _sp.values() if not import_blankish(_v))
            _has_so = 1 if not import_blankish(_sp.get('so')) else 0
            _has_reschedule = 1 if not import_blankish(_sp.get('reschedule')) else 0
            _has_poso = 1 if not import_blankish(_sp.get('po_sementara')) else 0
            _score = (_has_so, _has_reschedule, _has_poso, _rich)
            _prev = _dedup_seen_payloads.get(_uid)
            if _prev is None:
                _dedup_seen_payloads[_uid] = (_score, _sr, _sp)
            else:
                _prev_score, _prev_sr, _prev_sp = _prev
                if _score > _prev_score:
                    # Row baru lebih lengkap → merge field-field yang ada di prev
                    # tapi tidak di row baru, lalu simpan row baru.
                    _merged_sp = dict(_sp)
                    for _k, _v in _prev_sp.items():
                        if import_blankish(_merged_sp.get(_k)) and not import_blankish(_v):
                            _merged_sp[_k] = _v
                    # Update _sr in-place supaya perubahan tersimpan ke DB
                    for _col in columns:
                        _field = _col.get('field')
                        if _field and _field in _merged_sp:
                            _sr[_field] = _merged_sp[_field]
                    _dedup_seen_payloads[_uid] = (_score, _sr, _merged_sp)
                else:
                    # Row lama lebih lengkap → merge field-field dari row baru
                    # yang tidak ada di row lama, lalu update row lama.
                    _prev_sp_updated = dict(_prev_sp)
                    for _k, _v in _sp.items():
                        if import_blankish(_prev_sp_updated.get(_k)) and not import_blankish(_v):
                            _prev_sp_updated[_k] = _v
                    for _col in columns:
                        _field = _col.get('field')
                        if _field and _field in _prev_sp_updated:
                            _prev_sr[_field] = _prev_sp_updated[_field]
                    _dedup_seen_payloads[_uid] = (_prev_score, _prev_sr, _prev_sp_updated)
        # Bangun sheet_rows baru: 1 row per source_uid + rows tanpa uid
        _new_sheet_rows = [_entry[1] for _entry in _dedup_seen_payloads.values()]
        _new_sheet_rows.extend(_dedup_before_upsert.get('__no_uid__', []))
        _dup_removed = len(sheet_rows) - len(_new_sheet_rows)
        if _dup_removed > 0:
            print(f"[sync_import_sheet_to_dashboard] FIX V13: {_dup_removed} duplikat "
                  f"sheet_rows dihapus sebelum upsert "
                  f"({len(sheet_rows)} → {len(_new_sheet_rows)})")
        sheet_rows = _new_sheet_rows
    except Exception as _dedup_exc:
        import traceback; traceback.print_exc()
        # Kalau dedup gagal, lanjutkan dengan sheet_rows asli — jangan blok sync
        print(f"[sync_import_sheet_to_dashboard] FIX V13 dedup gagal: {_dedup_exc}")

    existing_rows = ImportDashboardRow.query.filter(
        db.or_(
            ImportDashboardRow.source_key == IMPORT_LAYOUT_SOURCE_KEY,
            ImportDashboardRow.source_key == 'import_tracker'
        )
    ).order_by(ImportDashboardRow.id.asc()).all()

    # ── CLEANUP: Remove already-copied rows whose vendor is NOT in the
    # uploaded Vendor Import list. This handles the case where vendors were
    # previously uploaded (so their rows got copied), but the vendor list
    # has since been updated to a smaller set — the old vendor rows should
    # be purged so they don't linger in the dashboard forever.
    #
    # IMPORTANT: We use uploaded_vendor_set (from ImportVendor table only)
    # — NOT import_vendor_filter_names() — because the latter falls back to
    # existing dashboard vendor names, which would mean NOTHING gets purged
    # (since the filter set would include every vendor already in the DB).
    #
    # We only purge when there IS an uploaded vendor list. If no vendors are
    # uploaded, we skip purge so the user doesn't accidentally lose all data.
    purged_non_vendor = 0
    if uploaded_vendor_set:
        rows_to_purge = []
        for ex_row in existing_rows:
            try: ex_payload = json.loads(ex_row.data_json or '{}')
            except: ex_payload = {}
            ex_vendor = import_nonblank(ex_payload.get('vendor')) or import_nonblank(ex_payload.get('vendor_name')) or import_nonblank(ex_row.vendor_name)
            if not ex_vendor:
                # No vendor on the row — can't match. These rows are stale
                # leftovers from before the vendor filter existed. PURGE them
                # too so the dashboard only contains vendor-attributed rows.
                # (If the user actually wants blank-vendor rows, they can
                # re-add the vendor name in the source sheet and re-sync.)
                rows_to_purge.append(ex_row)
                continue
            # FIX V9: pakai import_vendor_match() (fuzzy matching) bukan exact match
            # supaya vendor yang sedikit beda penulisan (mis. "CURT GEORGI GMBH & CO"
            # vs "CURT GEORGI GMBH & CO. KG") tidak ikut ter-purge.
            if not import_vendor_match(ex_vendor, uploaded_vendor_set):
                rows_to_purge.append(ex_row)
        for row in rows_to_purge:
            db.session.delete(row)
            purged_non_vendor += 1
        if purged_non_vendor:
            try: db.session.commit()
            except: db.session.rollback(); purged_non_vendor = 0
            # Refresh existing_rows to exclude the purged ones
            existing_rows = [r for r in existing_rows if r not in rows_to_purge]
    existing_by_uid = {}
    existing_by_sec_uid = {}
    existing_row_keys = set()
    for existing_row in existing_rows:
        existing_row_keys.add(existing_row.row_key)
        try: existing_payload = json.loads(existing_row.data_json or '{}')
        except: existing_payload = {}
        uid = existing_row.source_uid
        if not uid: uid = import_row_source_uid(existing_payload, columns) if existing_payload else None
        if uid: existing_by_uid.setdefault(uid, []).append(existing_row)
        sec_uid = import_row_secondary_uid(existing_payload)
        # FIX V11: SELALU daftarkan row ke index secondary (po_sementara + item_yupi),
        # bahkan kalau sec_uid == uid (primary). Sebelumnya hanya didaftarkan kalau
        # berbeda, dengan asumsi itu redundant. Tapi itu salah: row yang BARU dibuat
        # saat po_yupi belum ada punya primary uid == secondary uid (karena identity
        # payload-nya sama-sama berbasis po_sementara). Begitu po_yupi terisi belakangan
        # di sheet sumber, primary uid sheet row yang baru BERUBAH (sekarang berbasis
        # po_yupi), sehingga lookup primary gagal match row lama. Fallback ke secondary
        # lookup pun gagal kalau row lama tidak terdaftar di index ini — akibatnya row
        # lama dianggap row baru sama sekali, dan semua field user-edit (status,
        # checklist, incoterm, dll) hilang/balik ke default. Mendaftarkan tanpa
        # pengecualian membuat fallback secondary selalu bisa menemukan row lama.
        if sec_uid: existing_by_sec_uid.setdefault(sec_uid, []).append(existing_row)
    now = datetime.utcnow()
    added, updated, seen = 0, 0, 0
    duplicate_counts = {}
    consumed_per_uid = {}
    # FIX V13b: Track row yang BENAR-BENAR di-touch di upsert loop (update atau insert).
    # Sebelumnya, orphan purge pakai existing_by_uid.values() yang terlalu luas —
    # termasuk row yang sudah ada di DB tapi TIDAK di-update (duplikat source_uid
    # yang kalah di candidates[used]). Akibatnya row duplikat lama tidak pernah
    # di-purge karena dianggap "touched". Sekarang kita track explicit.
    touched_in_upsert_ids = set()     # id row existing yang di-update
    newly_inserted_rows = []          # objek row baru yang di-insert (id di-assign setelah commit)
    for sheet_row in sheet_rows:
        sheet_payload = import_row_payload(sheet_row, columns)
        source_uid = import_row_source_uid(sheet_payload, columns)
        duplicate_counts[source_uid] = duplicate_counts.get(source_uid, 0) + 1
        suffix = duplicate_counts[source_uid]
        candidates = existing_by_uid.get(source_uid) or []
        used_key = source_uid
        if not candidates:
            sec_uid_new = import_row_secondary_uid(sheet_payload)
            if sec_uid_new:
                candidates = existing_by_sec_uid.get(sec_uid_new) or []
                if candidates: used_key = sec_uid_new
        used = consumed_per_uid.get(used_key, 0)
        target_row = candidates[used] if used < len(candidates) else None
        if target_row is not None:
            consumed_per_uid[used_key] = used + 1
            # FIX V13b: tandai row ini sebagai benar-benar di-touch
            if target_row.id:
                touched_in_upsert_ids.add(target_row.id)
            try: existing_payload = json.loads(target_row.data_json or '{}')
            except: existing_payload = {}
            merged_payload = merge_import_existing_payload(existing_payload, sheet_payload)
            def _source_diff(old, new):
                for f in IMPORT_SOURCE_MANAGED_FIELDS:
                    old_v = (old.get(f) or '').strip(); new_v = (new.get(f) or '').strip()
                    if new_v and new_v != old_v: return True
                return False
            source_changed = _source_diff(existing_payload, merged_payload)
            full_old = json.dumps(existing_payload, ensure_ascii=False, sort_keys=True)
            full_new = json.dumps(merged_payload, ensure_ascii=False, sort_keys=True)
            any_change = source_changed or (full_old != full_new)
            if any_change:
                target_row.data_json = full_new; target_row.updated_at = now; updated += 1
            else: seen += 1
            target_row.last_seen_at = now
            target_row.sheet_row = sheet_row.get('_sheet_row') or target_row.sheet_row
            target_row.vendor_name = sheet_row.get('_vendor_name') or target_row.vendor_name
            target_row.source_key = IMPORT_LAYOUT_SOURCE_KEY
            target_row.source_label = sheet_row.get('_source_label') or 'Import Tracker'
            target_row.source_uid = source_uid
            existing_by_uid.setdefault(source_uid, [target_row])
            continue
        row_key = f"import:{source_uid}" if suffix == 1 else f"import:{source_uid}:{suffix}"
        while row_key in existing_row_keys:
            suffix += 1; row_key = f"import:{source_uid}:{suffix}"
        new_row = ImportDashboardRow(row_key=row_key, source_key=IMPORT_LAYOUT_SOURCE_KEY, source_label=sheet_row.get('_source_label') or '', source_uid=source_uid, sheet_row=sheet_row.get('_sheet_row'), vendor_name=sheet_row.get('_vendor_name') or '', data_json=json.dumps(sheet_payload, ensure_ascii=False), first_seen_at=now, last_seen_at=now, updated_at=now)
        db.session.add(new_row)
        added += 1
        # FIX V13b: simpan objek row baru supaya id-nya bisa di-track setelah commit
        newly_inserted_rows.append(new_row)
        existing_row_keys.add(row_key)
        existing_by_uid.setdefault(source_uid, []).append(new_row)
        consumed_per_uid[source_uid] = consumed_per_uid.get(source_uid, 0) + 1
    try: db.session.commit()
    except: db.session.rollback(); import traceback; traceback.print_exc()
    try:
        wib_now = datetime.utcnow() + timedelta(hours=7)
        import_meta_set('last_copy_at', wib_now.strftime('%Y-%m-%d %H:%M'))
    except: pass
    clear_runtime_caches()

    # ORPHAN PURGE: Hapus row dari vendor yang ada di uploaded_vendor_set yang
    # TIDAK ditemukan di sheet saat sync ini (last_seen_at tidak di-update = tidak ada di sheet).
    # Menangani kasus row lama yang tertinggal setelah sheet diubah (reschedule dihapus,
    # item tidak ada lagi di sheet, dll).
    #
    # Safety guards:
    # 1. Hanya berlaku kalau ada uploaded_vendor_set (vendor filter aktif)
    # 2. Hanya hapus row dari vendor yang ADA di uploaded_vendor_set
    # 3. Hanya kalau sheet_rows > 0 (bukan gagal baca sheet -> sheet kosong palsu)
    # 4. Toleransi 5 detik untuk menghindari false positive akibat clock skew
    purged_orphan = 0
    if uploaded_vendor_set and len(sheet_rows) > 0:
        try:
            sync_threshold = now - timedelta(seconds=5)
            # FIX V13b: Hanya row yang BENAR-BENAR di-touch di upsert loop yang
            # masuk touched_ids. Sebelumnya, kode pakai existing_by_uid.values()
            # yang juga meng-include row existing yang TIDAK di-update (duplikat
            # source_uid yang kalah di candidates[used]). Akibatnya duplikat lama
            # dianggap "touched" dan tidak pernah di-purge.
            touched_ids = set(touched_in_upsert_ids)
            # Tambahkan row baru yang baru saja di-insert (id di-assign setelah commit)
            for r in newly_inserted_rows:
                if r.id:
                    touched_ids.add(r.id)
            all_vendor_rows = ImportDashboardRow.query.filter(
                ImportDashboardRow.source_key.in_(_IMPORT_VISIBLE_SOURCE_KEYS)
            ).all()
            orphan_rows = []
            for ex_row in all_vendor_rows:
                # Skip row yang di-touch di sync ini
                if ex_row.id in touched_ids:
                    continue
                # Skip row yang last_seen_at NULL (row lama sebelum fitur ini ada —
                # tidak bisa dipastikan apakah ada di sheet atau tidak, jadi biarkan)
                if ex_row.last_seen_at is None:
                    continue
                # Hanya purge row dari vendor yang ada di uploaded_vendor_set
                try: ex_data = json.loads(ex_row.data_json or '{}')
                except: ex_data = {}
                ex_vendor = import_nonblank(ex_data.get('vendor')) or import_nonblank(ex_data.get('vendor_name')) or import_nonblank(ex_row.vendor_name)
                if not ex_vendor:
                    continue
                if not import_vendor_match(ex_vendor, uploaded_vendor_set):
                    continue
                orphan_rows.append(ex_row)
            for row in orphan_rows:
                db.session.delete(row)
                purged_orphan += 1
            if purged_orphan > 0:
                db.session.commit()
                clear_runtime_caches()
                print(f"[sync_import_sheet_to_dashboard] Orphan purge: {purged_orphan} row tidak ada di sheet dihapus")
        except Exception as orphan_exc:
            try: db.session.rollback()
            except: pass
            print(f"[sync_import_sheet_to_dashboard] Orphan purge gagal: {orphan_exc}")


    # ── POST-SYNC CLEANUP: Sesuaikan jumlah row di DB dengan jumlah baris di sheet.
    #
    # Masalah: karena sheet_rows bisa datang dari 2 sumber (layout sheet + source sheet),
    # row yang sama bisa masuk dua kali ke loop upsert → DB menumpuk lebih banyak row
    # dari yang seharusnya. Rescue logic ini membandingkan jumlah row di DB vs jumlah
    # unik di sheet per "business group" (po_yupi + item_yupi) lalu membuang kelebihan.
    #
    # PRINSIP UTAMA:
    # 1. "Expected count" = jumlah baris UNIK di sheet untuk group ini
    #    (deduplikasi berdasarkan source fingerprint dari IMPORT_SOURCE_MANAGED_FIELDS).
    # 2. Kalau DB punya LEBIH dari expected → hapus kelebihan, dengan prioritas:
    #    - Pertahankan row dengan USER RICHNESS tertinggi (paling banyak isian user)
    #    - Tiebreak: pertahankan row paling lama (id terkecil)
    # 3. Kalau DB punya KURANG dari expected → row baru sudah di-insert di upsert loop
    #    di atas, jadi tidak perlu action tambahan di sini.
    # 4. JANGAN hapus row dari vendor yang tidak di uploaded_vendor_set.
    duplicates_cleaned = 0
    excess_purged = 0
    try:
        all_db_rows = ImportDashboardRow.query.filter(
            ImportDashboardRow.source_key.in_(_IMPORT_VISIBLE_SOURCE_KEYS)
        ).order_by(ImportDashboardRow.id.asc()).all()

        # Cache semua parsed data sekaligus
        db_row_data = {}
        for db_row in all_db_rows:
            try: db_row_data[db_row.id] = json.loads(db_row.data_json or '{}')
            except: db_row_data[db_row.id] = {}

        # User richness = jumlah field user-only yang terisi (makin banyak = makin berharga)
        def _user_richness(data):
            return sum(1 for f in IMPORT_USER_LOCAL_ONLY_FIELDS if not import_blankish(data.get(f)))

        # Source fingerprint = hash dari field-field yang berasal dari sheet (bukan user-input).
        # Dua row dengan source fingerprint sama = berasal dari baris sheet yang SAMA.
        #
        # FIX V13: EXCLUDE field `so` dari fingerprint. Alasannya:
        #   - Layout tracker sheet dan source sheet (RM/SP) sering berbeda di kolom `so`
        #     (layout sering kosong, source terisi — atau sebaliknya).
        #   - Akibatnya, 2 baris di DB yang berasal dari baris sheet yang SAMA bisa
        #     punya fingerprint berbeda → tidak di-dedup oleh FASE 1 → duplikat tetap ada.
        #   - Setelah exclude `so`, kedua baris tersebut akan dianggap identik dan
        #     salah satunya akan dihapus (yang lebih rendah richness-nya).
        #   - Field `so` sendiri tetap disimpan di DB (tidak hilang) — kita hanya
        #     tidak menggunakannya sebagai pembeda identitas.
        # Field lain yang juga sering drift antara layout vs source sheet bisa
        # ditambahkan ke set ini kalau ditemukan kasus serupa di masa depan.
        _SRC_FP_EXCLUDE_FIELDS = {'so'}
        def _src_fingerprint(data):
            src = {k: (data.get(k) or '').strip().upper()
                   for k in IMPORT_SOURCE_MANAGED_FIELDS
                   if k not in _SRC_FP_EXCLUDE_FIELDS and not import_blankish(data.get(k))}
            return json.dumps(src, ensure_ascii=False, sort_keys=True, separators=(',', ':'))

        # Business group key = (po_yupi, item_yupi).
        # Ini yang menentukan "berapa baris seharusnya ada untuk PO + item ini".
        # TIDAK menyertakan req_dlv_date/po_sementara supaya semua varian tanggal
        # dari PO yang sama masuk ke group yang sama dan bisa di-compare sebagai satu kesatuan.
        def _biz_group_key(data):
            po = (import_nonblank(data.get('po_yupi')) or import_nonblank(data.get('yupi_po')) or '').strip().upper()
            item = (import_nonblank(data.get('item_yupi')) or '').strip().upper()
            pos = (import_nonblank(data.get('po_sementara')) or '').strip().upper()
            if po: return (po, item)
            if pos: return (pos, item)
            return None

        # ── Hitung expected count dari sheet_rows.
        #
        # sheet_rows bisa berisi duplikat karena datang dari 2 sumber (layout + source).
        # Deduplikasi dengan source fingerprint: 2 sheet rows yang punya fingerprint sama
        # = baris sheet yang sama, hanya dihitung 1x.
        seen_src_fps = set()
        sheet_group_count = {}  # biz_group_key → jumlah baris unik di sheet
        for sr in sheet_rows:
            sp = import_row_payload(sr, columns)
            fp = _src_fingerprint(sp)
            if fp in seen_src_fps:
                continue
            seen_src_fps.add(fp)
            gk = _biz_group_key(sp)
            if gk:
                sheet_group_count[gk] = sheet_group_count.get(gk, 0) + 1

        # ── Group DB rows per business group, juga deduplikasi per source fingerprint.
        #
        # Kalau 2 DB rows punya source fingerprint sama → salah satu pasti duplikat.
        # Di sini kita kumpulkan semua DB rows per group dan per src_fp supaya
        # FASE 1 bisa hapus yang benar-benar identik dari sisi source data.
        db_by_group = {}           # biz_group_key → [db_row, ...]
        db_by_src_fp = {}          # src_fingerprint → [db_row, ...]
        for db_row in all_db_rows:
            data = db_row_data.get(db_row.id, {})
            gk = _biz_group_key(data)
            if gk:
                db_by_group.setdefault(gk, []).append(db_row)
            fp = _src_fingerprint(data)
            db_by_src_fp.setdefault(fp, []).append(db_row)

        to_delete = set()  # set of db_row.id yang akan dihapus

        # ── FASE 0: Hapus DB rows yang punya source_uid sama.
        #
        # FIX V13c: FASE 1 dedup by source_fingerprint tidak menangkap kasus di mana
        # qty / unit_price / field source LAINNYA berubah di sheet (mis. qty 62.5 → 63).
        # Source fingerprint berubah karena field-field itu ikut di-hash, padahal
        # source_uid (identity key) TIDAK berubah karena tidak meng-include field
        # tersebut. Akibatnya 2 row DB — satu dengan qty lama (62.5, dari sync lampau)
        # dan satu dengan qty baru (63, dari sync terbaru) — punya source_uid SAMA tapi
        # source_fingerprint BEDA → FASE 1 lewat, duplikat tetap nangkring.
        #
        # Prinsip FASE 0: source_uid = identity. Kalau 2 row DB punya source_uid sama,
        # itu PASTI duplikat (atau row lama yang seharusnya sudah di-update). Simpan
        # yang paling up-to-date (last_seen_at terbaru — yaitu row yang baru saja
        # di-touch di sync ini). Kalau tie, ambil id terbesar (asumsi: id lebih besar
        # = row lebih baru di-insert). Sebelum hapus, MERGE field user-edit dari row
        # loser ke row winner supaya data user (status, checklist, incoterm, dll) tidak
        # hilang.
        #
        # Tujuan khusus: membersihkan residu duplikat dari sync-sync LAMPAU yang
        # menggunakan logika buggy (sebelum FIX V13). Sync-sync itu meninggalkan
        # row duplikat di DB yang tidak pernah di-purge.
        db_by_uid = {}  # source_uid → [db_row, ...]
        for db_row in all_db_rows:
            uid = (db_row.source_uid or '').strip()
            if uid:
                db_by_uid.setdefault(uid, []).append(db_row)

        for uid, uid_rows in db_by_uid.items():
            if len(uid_rows) <= 1:
                continue
            # Sort: winner = last_seen_at terbaru, tiebreak id terbesar
            uid_rows.sort(key=lambda r: (
                r.last_seen_at or datetime.min,
                r.id
            ), reverse=True)
            winner = uid_rows[0]
            winner_data = dict(db_row_data.get(winner.id, {}))
            for loser in uid_rows[1:]:
                if loser.id in to_delete:
                    continue
                loser_data = db_row_data.get(loser.id, {})
                # MERGE field user-edit dari loser ke winner (jangan overwrite yang
                # sudah terisi di winner, hanya isi yang masih kosong).
                for field in IMPORT_LOCAL_EDIT_FIELDS:
                    loser_v = loser_data.get(field)
                    if import_blankish(loser_v):
                        continue
                    if import_blankish(winner_data.get(field)):
                        winner_data[field] = loser_v
                # Cek apakah winner_data berubah → perlu persist
                if winner_data != db_row_data.get(winner.id, {}):
                    winner.data_json = json.dumps(winner_data, ensure_ascii=False)
                    winner.updated_at = now
                    db_row_data[winner.id] = winner_data
                to_delete.add(loser.id)
                duplicates_cleaned += 1
                print(f"[sync] FASE0 source_uid dup purge: id={loser.id} "
                      f"uid={uid[:12]}... richness_loser={_user_richness(loser_data)} "
                      f"→ merged ke winner id={winner.id}")

        # ── FASE 1: Hapus DB rows yang punya source fingerprint sama (= baris sheet sama).
        #
        # Di antara rows dengan fp sama, pertahankan yang paling banyak user-data-nya
        # (richness tertinggi). Tiebreak: id terkecil.
        for fp, fp_rows in db_by_src_fp.items():
            if len(fp_rows) <= 1:
                continue
            fp_rows.sort(key=lambda r: (-_user_richness(db_row_data.get(r.id, {})), r.id))
            for loser in fp_rows[1:]:
                if loser.id not in to_delete:
                    to_delete.add(loser.id)
                    duplicates_cleaned += 1
                    print(f"[sync] FASE1 src-identical purge: id={loser.id} "
                          f"richness={_user_richness(db_row_data.get(loser.id, {}))} dihapus")

        # ── FASE 2: Excess count purge — setelah FASE 1, cek masih ada kelebihan?
        #
        # Rekonstruksi db_by_group tanpa row yang sudah ditandai untuk dihapus di FASE 1.
        for gk, db_rows_in_group in db_by_group.items():
            # Skip row yang sudah ditandai hapus di FASE 1
            surviving = [r for r in db_rows_in_group if r.id not in to_delete]
            expected = sheet_group_count.get(gk, 0)
            if expected == 0 or len(surviving) <= expected:
                continue  # pas atau kurang — tidak perlu hapus lagi

            # Masih kelebihan setelah FASE 1 → hapus yang paling rendah richness-nya
            surviving.sort(key=lambda r: (-_user_richness(db_row_data.get(r.id, {})), r.id))
            losers = surviving[expected:]
            for loser in losers:
                if loser.id in to_delete:
                    continue
                loser_data = db_row_data.get(loser.id, {})
                loser_vendor = (import_nonblank(loser_data.get('vendor')) or
                                import_nonblank(loser_data.get('vendor_name')) or
                                import_nonblank(loser.vendor_name))
                if uploaded_vendor_set and loser_vendor:
                    if not import_vendor_match(loser_vendor, uploaded_vendor_set):
                        continue
                to_delete.add(loser.id)
                excess_purged += 1
                print(f"[sync] FASE2 excess purge: id={loser.id} group={gk} "
                      f"richness={_user_richness(loser_data)} dihapus "
                      f"(expected={expected}, surviving={len(surviving)})")

        # ── Eksekusi delete
        if to_delete:
            rows_to_del = [r for r in all_db_rows if r.id in to_delete]
            for r in rows_to_del:
                db.session.delete(r)
            db.session.commit()
            db.session.execute(text('PRAGMA wal_checkpoint(TRUNCATE)'))
            db.session.commit()
            clear_runtime_caches()
            print(f"[sync_import_sheet_to_dashboard] Cleanup selesai: "
                  f"{duplicates_cleaned} src-identical + {excess_purged} excess dihapus")

    except Exception as cleanup_exc:
        try: db.session.rollback()
        except: pass
        print(f"[sync_import_sheet_to_dashboard] Cleanup gagal: {cleanup_exc}")

    return {
        'added': added, 'updated': updated, 'seen': seen,
        'sheet_rows': len(sheet_rows),
        'source_rows_added': source_rows_added,
        'source_errors': source_errors,
        'purged_non_vendor': purged_non_vendor,
        'purged_orphan': purged_orphan,
        'duplicates_cleaned': duplicates_cleaned,
        'excess_purged': excess_purged,
        'vendor_count': vendor_count,
        'vendor_filter_count': len(filter_vendors),
        'vendor_filter_source': vendor_source,
        'purged_legacy': purged_legacy,
        'copy_only': True,
        'columns': columns,
        'source_sheet_url': f'https://docs.google.com/spreadsheets/d/{IMPORT_LAYOUT_SHEET_ID}/edit#gid={IMPORT_LAYOUT_GID}',
    }

RFQ_DASHBOARD_ONLY_FIELDS = {'private_remarks_1', 'private_remarks_2'}

def rfq_label(field): return dict(RFQ_TEMPLATE_COLUMNS).get(field, field)

def parse_rfq_number(value):
    raw = clean(value)
    if not raw: return None
    s = re.sub(r'[^0-9.\-]', '', str(raw))
    if not s or s in ('-', '.', '-.'): return None
    try: return float(s)
    except: return None

def fmt_rfq_amount(value):
    if value is None: return None
    if abs(value - round(value)) < 0.000001: return f'{int(round(value)):,}'
    return f'{value:,.2f}'

def rfq_days_left(closing_date):
    raw = clean(closing_date)
    if not raw: return None
    d = None
    for fmt in ('%d/%m/%Y', '%Y/%m/%d', '%Y-%m-%d'):
        try: d = datetime.strptime(str(raw).strip(), fmt).date(); break
        except: pass
    if d is None: d = parse_date(raw)
    if not d: return None
    if d < date.today(): return None
    return workdays_until(d)

def parse_rfq_closing_date_value(value):
    raw = clean(value)
    if not raw: return None
    for fmt in ('%d/%m/%Y', '%Y/%m/%d', '%Y-%m-%d'):
        try: return datetime.strptime(str(raw).strip(), fmt).date()
        except: pass
    return parse_date(raw)

def parse_rfq_date_value(value):
    raw = clean(value)
    if not raw: return None
    if not re.search(r'\d{4}', str(raw)) and not re.match(r'^\d{8}(\.0)?$', str(raw).strip()): return None
    for fmt in ('%d/%m/%Y', '%Y/%m/%d', '%Y-%m-%d'):
        try: return datetime.strptime(str(raw).strip(), fmt).date()
        except: pass
    return parse_date(raw)

def sort_rfq_rows(rows, sort_order='newest'):
    newest = sort_order != 'oldest'
    def key(row):
        d = parse_rfq_date_value(row.get('rfq_date'))
        ordinal = d.toordinal() if d else 0
        sheet_row = int(row.get('sheet_row') or 0)
        return (d is None, -ordinal if newest else ordinal, sheet_row)
    rows.sort(key=key)
    return rows

RFQ_SEARCH_FIELDS = ('rfq_code', 'request_number', 'item_name', 'detail_spec')

def rfq_multiline_search_terms(value):
    terms = []; seen = set()
    for raw in re.split(r'[\r\n]+', str(value or '')):
        term = raw.strip().lower()
        if term and term not in seen: seen.add(term); terms.append(term)
    return terms

def filter_rfq_rows_by_multiline_search(rows, value):
    terms = rfq_multiline_search_terms(value)
    if not terms: return rows
    filtered = []
    for row in rows:
        searchable_values = [str(row.get(field) or '').lower() for field in RFQ_SEARCH_FIELDS]
        if any(term in field_value for term in terms for field_value in searchable_values): filtered.append(row)
    return filtered

def rfq_check_value(item):
    if clean_product_id(item.get('product_id')): return 'complete'
    if 'reject' in (clean(item.get('sheet_status')) or '').lower(): return 'reject'
    closing_date = parse_rfq_closing_date_value(item.get('closing_date'))
    if closing_date and closing_date < date.today(): return 'closed'
    return 'open'

def rfq_check_label(value):
    return {'complete': 'Complete', 'reject': 'Reject', 'closed': 'Closed', 'open': 'Open'}.get(value or '', 'Open')

def apply_rfq_computed_fields(item):
    item['category_name'] = (clean(item.get('category_name')) or '').split('>')[0].strip() or None
    qty = parse_rfq_number(item.get('qty'))
    unit_price = parse_rfq_number(item.get('unit_price_idr'))
    item['amt_idr'] = fmt_rfq_amount(qty * unit_price) if qty is not None and unit_price is not None else None
    item['days_left'] = rfq_days_left(item.get('closing_date'))
    item['unit_price_missing'] = unit_price is None
    item['status'] = bool(clean_product_id(item.get('product_id')))
    item['check'] = rfq_check_value(item)
    return item

def rfq_cell(row, idx):
    try: return clean(row.iloc[idx])
    except: return None

def rfq_row_key(data, sheet_row):
    code = clean(data.get('source_code'))
    if code: return code
    parts = [data.get('no'), data.get('client_name'), data.get('rfq_date'), data.get('item_name')]
    key = '|'.join(str(clean(x) or '') for x in parts).strip('|')
    return key or f'row-{sheet_row}'

def fetch_rfq_rows(force=False):
    now = datetime.utcnow()
    if not force and RFQ_CACHE['expires_at'] and RFQ_CACHE['expires_at'] > now: return RFQ_CACHE['rows'], RFQ_CACHE['fetched_at']
    from urllib.parse import quote
    url = f'https://docs.google.com/spreadsheets/d/{RFQ_SHEET_ID}/gviz/tq?tqx=out:csv&sheet={quote(RFQ_SHEET_NAME)}'
    df = pd.read_csv(url, header=None, dtype=str, keep_default_na=False)
    rows = []
    for idx in range(3, len(df)):
        src = df.iloc[idx]
        product_id = clean_product_id(rfq_cell(src, 16))
        request_number = clean_request_number(rfq_cell(src, 17))
        data = {
            'sheet_row': idx + 1, 'no': rfq_cell(src, 1), 'client_name': rfq_cell(src, 2), 'rfq_date': rfq_cell(src, 4),
            'closing_date': rfq_cell(src, 5), 'sales_pic': rfq_cell(src, 6), 'rfq_code': rfq_cell(src, 7),
            'item_name': rfq_cell(src, 8), 'detail_spec': rfq_cell(src, 9), 'brand_manufacturer': rfq_cell(src, 10),
            'qty': rfq_cell(src, 11), 'unit': rfq_cell(src, 12), 'remark': rfq_cell(src, 13), 'category_id': rfq_cell(src, 14),
            'category_name': rfq_cell(src, 15), 'product_id': product_id, 'sheet_status': rfq_cell(src, 0),
            'request_number': request_number, 'purchase_pic': rfq_cell(src, 18), 'same_replacement': rfq_cell(src, 21),
            'vendor_name': rfq_cell(src, 22), 'unit_price_idr': rfq_cell(src, 23), 'amt_idr': rfq_cell(src, 24),
            'quoted_item_name': rfq_cell(src, 25), 'quoted_spec': rfq_cell(src, 26), 'quoted_brand': rfq_cell(src, 27),
            'quoted_unit': rfq_cell(src, 28), 'moq': rfq_cell(src, 29), 'lead_time_days': rfq_cell(src, 30),
            'valid_period': rfq_cell(src, 31), 'photo_url': rfq_cell(src, 32), 'remarks': rfq_cell(src, 33),
            'private_remarks_1': '', 'private_remarks_2': '', 'source_code': rfq_cell(src, 38),
        }
        data['purchase_pic'] = canonical_rfq_pic(data)
        if not any(data.get(field) for field, _ in RFQ_TEMPLATE_COLUMNS if field != 'check'): continue
        data['row_key'] = rfq_row_key(data, idx + 1)
        apply_rfq_computed_fields(data)
        rows.append(data)
    fetched_at = datetime.utcnow()
    RFQ_CACHE.update({'rows': rows, 'fetched_at': fetched_at, 'expires_at': fetched_at + timedelta(seconds=RFQ_CACHE_TTL_SECONDS)})
    return rows, fetched_at

def rfq_json_load(value, fallback):
    try: return json.loads(value or '')
    except: return fallback

def rfq_dashboard_payload(row):
    payload = dict(row or {})
    payload['row_key'] = clean(payload.get('row_key')) or rfq_row_key(payload, payload.get('sheet_row') or 0)
    try: payload['sheet_row'] = int(payload.get('sheet_row') or 0) or None
    except: payload['sheet_row'] = None
    apply_rfq_computed_fields(payload)
    return payload

def rfq_dashboard_row_to_dict(row):
    data = rfq_json_load(row.data_json, {})
    data['row_key'] = row.row_key
    data['sheet_row'] = row.sheet_row
    return data

def load_rfq_dashboard_rows():
    db_rows = RFQDashboardRow.query.options(load_only(RFQDashboardRow.id, RFQDashboardRow.row_key, RFQDashboardRow.sheet_row, RFQDashboardRow.data_json, RFQDashboardRow.last_seen_at)).order_by(RFQDashboardRow.sheet_row.is_(None), RFQDashboardRow.sheet_row.asc(), RFQDashboardRow.id.asc()).all()
    rows = [rfq_dashboard_row_to_dict(row) for row in db_rows]
    fetched_at = max((row.last_seen_at for row in db_rows if row.last_seen_at), default=None)
    return rows, fetched_at

def set_rfq_runtime_rows(rows, fetched_at):
    now = datetime.utcnow()
    RFQ_CACHE.update({'rows': [dict(row) for row in rows], 'fetched_at': fetched_at or now, 'expires_at': now + timedelta(seconds=RFQ_CACHE_TTL_SECONDS)})

def sync_rfq_sheet_to_dashboard():
    sheet_rows, fetched_at = fetch_rfq_rows(force=True)
    existing = {row.row_key: row for row in RFQDashboardRow.query.all()}
    duplicate_counts = {}
    now = datetime.utcnow()
    added, updated = 0, 0
    for sheet_row in sheet_rows:
        base_key = clean(sheet_row.get('row_key'))
        if not base_key: continue
        duplicate_counts[base_key] = duplicate_counts.get(base_key, 0) + 1
        row_key = base_key if duplicate_counts[base_key] == 1 else f"{base_key}#{duplicate_counts[base_key]}"
        sheet_row = dict(sheet_row)
        sheet_row['row_key'] = row_key
        incoming = rfq_dashboard_payload(sheet_row)
        current = existing.get(row_key)
        if not current:
            db.session.add(RFQDashboardRow(row_key=row_key, sheet_row=incoming.get('sheet_row'), data_json=json.dumps(incoming, ensure_ascii=False), dirty_fields_json='[]', first_seen_at=now, last_seen_at=fetched_at or now, updated_at=now))
            added += 1; continue
        local = rfq_json_load(current.data_json, {})
        dirty_fields = set(rfq_json_load(current.dirty_fields_json, []))
        for field, value in incoming.items():
            if field in dirty_fields and field in RFQ_EDITABLE_FIELDS: continue
            local[field] = value
        local['row_key'] = row_key
        local['sheet_row'] = incoming.get('sheet_row')
        apply_rfq_computed_fields(local)
        current.sheet_row = incoming.get('sheet_row')
        current.data_json = json.dumps(local, ensure_ascii=False)
        current.last_seen_at = fetched_at or now
        current.updated_at = now
        updated += 1
    db.session.commit()
    rows, loaded_at = load_rfq_dashboard_rows()
    clear_runtime_caches()
    set_rfq_runtime_rows(rows, loaded_at or fetched_at)
    return {'added': added, 'updated': updated, 'sheet_rows': len(sheet_rows), 'fetched_at': loaded_at or fetched_at}

def set_rfq_dashboard_cell(row_key, field, value, dirty=True, commit=True):
    row = RFQDashboardRow.query.filter_by(row_key=row_key).first()
    if not row: return False
    data = rfq_json_load(row.data_json, {})
    dirty_fields = set(rfq_json_load(row.dirty_fields_json, []))
    data[field] = value
    data['row_key'] = row.row_key
    data['sheet_row'] = row.sheet_row
    apply_rfq_computed_fields(data)
    if dirty: dirty_fields.add(field)
    else: dirty_fields.discard(field)
    row.data_json = json.dumps(data, ensure_ascii=False)
    row.dirty_fields_json = json.dumps(sorted(dirty_fields), ensure_ascii=False)
    row.updated_at = datetime.utcnow()
    if commit:
        db.session.commit()
        RFQ_CACHE['expires_at'] = None
        clear_runtime_caches()
    return True

def clear_rfq_dashboard_dirty_fields(updates, commit=True):
    grouped = {}
    for item in updates or []:
        row_key = clean(item.get('row_key')); field = clean(item.get('field'))
        if row_key and field: grouped.setdefault(row_key, set()).add(field)
    if not grouped: return
    for row in RFQDashboardRow.query.filter(RFQDashboardRow.row_key.in_(grouped.keys())).all():
        dirty_fields = set(rfq_json_load(row.dirty_fields_json, []))
        dirty_fields.difference_update(grouped.get(row.row_key, set()))
        row.dirty_fields_json = json.dumps(sorted(dirty_fields), ensure_ascii=False)
        row.updated_at = datetime.utcnow()
    if commit:
        db.session.commit()
        RFQ_CACHE['expires_at'] = None
        clear_runtime_caches()

def cleanup_rfq_sheet_backed_edits(commit=False):
    try:
        deleted = RFQCellEdit.query.filter(~RFQCellEdit.field.in_(list(RFQ_DASHBOARD_ONLY_FIELDS))).delete(synchronize_session=False)
        if commit and deleted: db.session.commit()
        return deleted or 0
    except:
        if commit: db.session.rollback()
        return 0

def rfq_rows_with_edits(force=False, prefer_stale_cache=False):
    now = datetime.utcnow()
    if force:
        sync_rfq_sheet_to_dashboard()
        rows, fetched_at = load_rfq_dashboard_rows()
    elif prefer_stale_cache and RFQ_CACHE.get('rows'):
        rows, fetched_at = RFQ_CACHE['rows'], RFQ_CACHE['fetched_at']
    elif RFQ_CACHE.get('expires_at') and RFQ_CACHE['expires_at'] > now and RFQ_CACHE.get('rows'):
        rows, fetched_at = RFQ_CACHE['rows'], RFQ_CACHE['fetched_at']
    else:
        has_dashboard_rows = RFQDashboardRow.query.with_entities(RFQDashboardRow.id).first() is not None
        if not has_dashboard_rows: sync_rfq_sheet_to_dashboard()
        rows, fetched_at = load_rfq_dashboard_rows()
        set_rfq_runtime_rows(rows, fetched_at)
    edits = RFQCellEdit.query.options(load_only(RFQCellEdit.row_key, RFQCellEdit.field, RFQCellEdit.value)).filter(RFQCellEdit.field.in_(list(RFQ_DASHBOARD_ONLY_FIELDS))).all()
    edit_map = {}
    for edit in edits: edit_map.setdefault(edit.row_key, {})[edit.field] = edit.value
    merged = []
    for row in rows:
        item = dict(row)
        for field, value in edit_map.get(item['row_key'], {}).items(): item[field] = value
        merged.append(item)
    return merged, fetched_at

def _candidate_registered_items_for_rfq_similarity(row, limit=1200):
    name_token = _similarity_token(row.get('item_name'))
    spec_token = _similarity_token(row.get('detail_spec'))
    if not clean(row.get('unit')) or not clean(row.get('item_name')) or not clean(row.get('detail_spec')): return []
    q = ProductIDDB.query.filter(ProductIDDB.product_id.isnot(None), ProductIDDB.product_id != '', db.or_(ProductIDDB.product_status.is_(None), ProductIDDB.product_status == '', func.lower(ProductIDDB.product_status) == 'use'))
    token_filters = []
    if name_token: token_filters.append(ProductIDDB.product_name.ilike(f'%{name_token}%'))
    if spec_token: token_filters.append(ProductIDDB.specification.ilike(f'%{spec_token}%'))
    if token_filters: q = q.filter(db.or_(*token_filters))
    return q.limit(limit).all()

def find_similar_rfq_registered_items(row):
    try:
        if (clean(row.get('check')) or '').lower() != 'open': return None
        if clean_product_id(row.get('product_id')): return None
        key_fields = [row.get('item_name'), row.get('detail_spec'), row.get('unit')]
        if not all(clean(v) for v in key_fields): return None
        current_prod_id = clean_product_id(row.get('product_id'))
        cache_key = '|'.join(['rfq_similar_v5', clean(row.get('row_key')) or '', current_prod_id, (clean(row.get('item_name')) or '').lower(), (clean(row.get('detail_spec')) or '').lower(), (clean(row.get('unit')) or '').lower()])
        if cache_key in _SIMILARITY_CACHE: return _SIMILARITY_CACHE[cache_key]
        similar_items = []
        for reg in _candidate_registered_items_for_rfq_similarity(row):
            reg_prod_id = clean_product_id(reg.product_id)
            if not reg_prod_id or (current_prod_id and reg_prod_id == current_prod_id): continue
            if not (clean(reg.product_name) and clean(reg.specification) and clean(reg.order_unit)): continue
            item_score = calculate_similarity(row.get('item_name'), reg.product_name)
            spec_score = calculate_similarity(row.get('detail_spec'), reg.specification)
            unit_score = calculate_similarity(row.get('unit'), reg.order_unit)
            if item_score >= 70.0 and spec_score >= 70.0 and unit_score >= 70.0:
                total_sim = (item_score + spec_score + unit_score) / 3
                similar_items.append({'product_id': reg_prod_id, 'product_name': reg.product_name or '', 'specification': reg.specification or '', 'manufacturer_name': reg.manufacturer_name or '', 'order_unit': reg.order_unit or '', 'similarity': round(total_sim, 1)})
        similar_items.sort(key=lambda x: (-x['similarity'], x['product_id']))
        if not similar_items: result = None
        else:
            result = {'product_ids': '\n'.join(x['product_id'] for x in similar_items), 'product_name': '\n'.join(x['product_name'] or '-' for x in similar_items), 'specification': '\n'.join(x['specification'] or '-' for x in similar_items), 'manufacturer_name': '\n'.join(x['manufacturer_name'] or '-' for x in similar_items), 'order_unit': '\n'.join(x['order_unit'] or '-' for x in similar_items), 'similarity': '\n'.join(f"{x['similarity']:.0f}%" for x in similar_items), 'count': len(similar_items)}
        _SIMILARITY_CACHE[cache_key] = result
        return result
    except Exception as e:
        print(f"Error finding RFQ similar items: {e}")
        return None

def apply_rfq_similarity(row):
    if (clean(row.get('check')) or '').lower() != 'open':
        row['similar_prod_ids'] = ''; row['similar_prod_name'] = ''; row['similar_spec'] = ''; row['similar_mfr_name'] = ''; row['similar_odr_unit'] = ''; row['similar_score'] = None
        return row
    similar = find_similar_rfq_registered_items(row)
    row['similar_prod_ids'] = (similar or {}).get('product_ids', '') if clean_product_id(row.get('product_id')) else (similar or {}).get('product_ids', 'No Similar Item')
    row['similar_prod_name'] = (similar or {}).get('product_name', '')
    row['similar_spec'] = (similar or {}).get('specification', '')
    row['similar_mfr_name'] = (similar or {}).get('manufacturer_name', '')
    row['similar_odr_unit'] = (similar or {}).get('order_unit', '')
    row['similar_score'] = (similar or {}).get('similarity', None)
    return row

def rfq_sheet_sync_credentials():
    raw_json = os.environ.get('GOOGLE_SERVICE_ACCOUNT_JSON') or os.environ.get('GOOGLE_SHEETS_SERVICE_ACCOUNT_JSON')
    raw_file = os.environ.get('GOOGLE_SERVICE_ACCOUNT_FILE') or os.environ.get('GOOGLE_APPLICATION_CREDENTIALS')
    if raw_json:
        try: return json.loads(raw_json)
        except json.JSONDecodeError as e: raise RuntimeError(f'Invalid GOOGLE_SERVICE_ACCOUNT_JSON: {e}')
    if raw_file and os.path.exists(raw_file):
        with open(raw_file, 'r', encoding='utf-8') as f: return json.load(f)
    return None

GOOGLE_SHEETS_SCOPE = ['https://www.googleapis.com/auth/spreadsheets']

def google_sheets_access_token():
    credentials_info = rfq_sheet_sync_credentials()
    if not credentials_info: raise RuntimeError('Google service account credential is not configured')
    try:
        from google.oauth2.service_account import Credentials
        from google.auth.transport.requests import Request
    except ImportError as e: raise RuntimeError('google-auth and requests are required for Google Sheets access') from e
    creds = Credentials.from_service_account_info(credentials_info, scopes=GOOGLE_SHEETS_SCOPE)
    creds.refresh(Request())
    return creds.token

def _summarize_non_json_response(response):
    """Turn a non-JSON (typically HTML) error body into a short, readable
    diagnostic instead of dumping raw HTML/JS into the error message.

    On PythonAnywhere free accounts, outbound requests to non-whitelisted
    hosts (or requests that don't go through the platform's proxy) can be
    intercepted and answered with an HTML block page instead of reaching
    Google's servers at all. That HTML looks nothing like a Google API error
    and is unreadable when shown verbatim in a toast, so detect it here.
    """
    text = (response.text or '').strip()
    looks_like_html = text[:15].lower().lstrip().startswith(('<!doctype', '<html', '<head', '<script'))
    if looks_like_html:
        return ('Got an HTML page instead of a response from Google Sheets API — this usually means the '
                'request never reached Google (blocked or intercepted by a network proxy/firewall before '
                'reaching sheets.googleapis.com). If this server is on PythonAnywhere\'s free tier, outbound '
                'requests must go through their proxy for *.googleapis.com to be reachable — check that the '
                'HTTP_PROXY/HTTPS_PROXY environment variables are set for this web app.')
    return text[:500]

def google_sheets_request(method, spreadsheet_id, path, params=None, body=None):
    try:
        import requests
        from urllib.parse import quote
    except ImportError as e: raise RuntimeError('requests is required for Google Sheets access') from e
    token = google_sheets_access_token()
    encoded_path = '/'.join(quote(str(part), safe='') for part in path)
    url = f'https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/{encoded_path}'
    headers = {'Authorization': f'Bearer {token}', 'Accept': 'application/json'}
    if body is not None: headers['Content-Type'] = 'application/json'
    proxies = {}
    if os.environ.get('HTTPS_PROXY'): proxies['https'] = os.environ.get('HTTPS_PROXY')
    if os.environ.get('HTTP_PROXY'): proxies['http'] = os.environ.get('HTTP_PROXY')
    kwargs = {'headers': headers, 'params': params or {}, 'timeout': 60}
    if body is not None: kwargs['json'] = body
    if proxies: kwargs['proxies'] = proxies
    try:
        response = requests.request(method, url, **kwargs)
    except requests.exceptions.RequestException as e:
        raise RuntimeError(f'Could not reach Google Sheets API ({method} {"/".join(path)}): {e}') from e
    if not response.ok:
        detail = _summarize_non_json_response(response)
        raise RuntimeError(f'Google Sheets API {method} {"/".join(path)} failed: {response.status_code} {detail}')
    if not response.text:
        return {}
    try:
        return response.json()
    except ValueError as e:
        detail = _summarize_non_json_response(response)
        raise RuntimeError(f'Google Sheets API {method} {"/".join(path)} returned a non-JSON response (HTTP {response.status_code}): {detail}') from e

def google_sheets_metadata(spreadsheet_id):
    try: import requests
    except ImportError as e: raise RuntimeError('requests is required for Google Sheets access') from e
    token = google_sheets_access_token()
    url = f'https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}'
    headers = {'Authorization': f'Bearer {token}', 'Accept': 'application/json'}
    proxies = {}
    if os.environ.get('HTTPS_PROXY'): proxies['https'] = os.environ.get('HTTPS_PROXY')
    if os.environ.get('HTTP_PROXY'): proxies['http'] = os.environ.get('HTTP_PROXY')
    kwargs = {'headers': headers, 'timeout': 60}
    if proxies: kwargs['proxies'] = proxies
    try:
        response = requests.get(url, **kwargs)
    except requests.exceptions.RequestException as e:
        raise RuntimeError(f'Could not reach Google Sheets API (metadata): {e}') from e
    if not response.ok:
        raise RuntimeError(f'Google Sheets metadata failed: {response.status_code} {_summarize_non_json_response(response)}')
    try:
        return response.json()
    except ValueError as e:
        raise RuntimeError(f'Google Sheets metadata returned a non-JSON response (HTTP {response.status_code}): {_summarize_non_json_response(response)}') from e

def google_sheets_values_get(spreadsheet_id, range_name, value_render_option='UNFORMATTED_VALUE'):
    return google_sheets_request('GET', spreadsheet_id, ['values', range_name], params={'valueRenderOption': value_render_option})

def google_sheets_values_batch_get(spreadsheet_id, ranges, value_render_option='FORMATTED_VALUE', major_dimension='COLUMNS'):
    return google_sheets_request('GET', spreadsheet_id, ['values:batchGet'], params={'ranges': list(ranges or []), 'valueRenderOption': value_render_option, 'majorDimension': major_dimension})

def google_sheets_values_update(spreadsheet_id, range_name, values):
    return google_sheets_request('PUT', spreadsheet_id, ['values', range_name], params={'valueInputOption': 'USER_ENTERED'}, body={'values': values})

def google_sheets_values_batch_update(spreadsheet_id, ranges):
    return google_sheets_request('POST', spreadsheet_id, ['values:batchUpdate'], body={'valueInputOption': 'USER_ENTERED', 'data': ranges})

def sync_rfq_cell_to_google_sheet(row, field, value):
    column = RFQ_SHEET_COLUMN_BY_FIELD.get(field)
    if field in RFQ_DASHBOARD_ONLY_FIELDS: return {'synced': False, 'local_only': True, 'reason': 'Dashboard-only field'}
    if not column: return {'synced': False, 'reason': 'Field is not mapped to RFQ sheet column'}
    sheet_row = row.get('sheet_row')
    if not sheet_row: return {'synced': False, 'reason': 'RFQ sheet row is missing'}
    range_name = f"'{RFQ_SHEET_NAME}'!{column}{sheet_row}"
    google_sheets_values_update(RFQ_SHEET_ID, range_name, [[value or '']])
    RFQ_CACHE['expires_at'] = None
    return {'synced': True, 'range': range_name}

def sync_rfq_cells_to_google_sheet(updates):
    ranges = []
    local_only_count = 0
    for item in updates:
        row = item.get('row') or {}; field = item.get('field'); value = item.get('value')
        if field in RFQ_DASHBOARD_ONLY_FIELDS: local_only_count += 1; continue
        column = RFQ_SHEET_COLUMN_BY_FIELD.get(field)
        sheet_row = row.get('sheet_row')
        if column and sheet_row: ranges.append({'range': f"'{RFQ_SHEET_NAME}'!{column}{sheet_row}", 'values': [[value or '']]})
    if not ranges:
        if local_only_count: return {'synced': False, 'local_only': True, 'reason': 'Dashboard-only fields'}
        return {'synced': False, 'reason': 'No mapped RFQ sheet cells to sync'}
    google_sheets_values_batch_update(RFQ_SHEET_ID, ranges)
    RFQ_CACHE['expires_at'] = None
    return {'synced': True, 'ranges': len(ranges), 'local_only': local_only_count}

def column_letter_from_index(index):
    result = ''
    while index > 0:
        index, rem = divmod(index - 1, 26)
        result = chr(65 + rem) + result
    return result

def column_index_from_letter(letter):
    result = 0
    for ch in str(letter or '').strip().upper():
        if not ('A' <= ch <= 'Z'): continue
        result = result * 26 + (ord(ch) - 64)
    return result

def vendor_control_sheet_name():
    if VENDOR_CONTROL_CACHE.get('sheet_name'): return VENDOR_CONTROL_CACHE['sheet_name']
    meta = google_sheets_metadata(VENDOR_CONTROL_SHEET_ID)
    for sheet in meta.get('sheets', []):
        props = sheet.get('properties', {})
        if props.get('sheetId') == VENDOR_CONTROL_SHEET_GID:
            VENDOR_CONTROL_CACHE['sheet_name'] = props.get('title')
            return VENDOR_CONTROL_CACHE['sheet_name']
    sheets = meta.get('sheets', [])
    if sheets:
        VENDOR_CONTROL_CACHE['sheet_name'] = sheets[0].get('properties', {}).get('title')
        return VENDOR_CONTROL_CACHE['sheet_name']
    raise RuntimeError('Vendor Control sheet not found')

def normalized_header(value): return re.sub(r'[^a-z0-9]+', '', str(value or '').lower())

def find_vendor_control_columns(headers):
    normalized = {}
    for idx, header in enumerate(headers or []):
        key = normalized_header(header)
        if key and key not in normalized: normalized[key] = idx + 1
    def pick(names):
        for name in names:
            idx = normalized.get(normalized_header(name))
            if idx: return idx
        return None
    return {'vendor_name': pick(['Vendor Name', 'Vendor Nm', 'Vendor', 'Supplier Name', 'Supplier']), 'vendor_id': pick(['Vendor ID', 'Vendor Id', 'VendorID', 'ID', 'User ID']), 'password': pick(['Password', 'Pass', 'PWD', 'Pwd'])}

def vendor_control_rows(force=False):
    now = datetime.utcnow()
    if (not force and VENDOR_CONTROL_CACHE.get('expires_at') and VENDOR_CONTROL_CACHE['expires_at'] > now and VENDOR_CONTROL_CACHE.get('rows')): return VENDOR_CONTROL_CACHE['rows'], VENDOR_CONTROL_CACHE.get('fetched_at')
    sheet_name = vendor_control_sheet_name()
    result = google_sheets_values_get(VENDOR_CONTROL_SHEET_ID, f"'{sheet_name}'!A:Z")
    values = result.get('values', [])
    if not values:
        rows = []; fetched_at = datetime.utcnow()
        VENDOR_CONTROL_CACHE.update({'rows': rows, 'fetched_at': fetched_at, 'expires_at': fetched_at + timedelta(seconds=VENDOR_CONTROL_CACHE_TTL_SECONDS), 'columns': {}})
        return rows, fetched_at
    header_index = 0; columns = {}
    for idx, candidate_headers in enumerate(values[:20]):
        candidate_columns = find_vendor_control_columns(candidate_headers)
        if all(candidate_columns.get(name) for name in ('vendor_name', 'vendor_id', 'password')):
            header_index = idx; columns = candidate_columns; break
    missing = [name for name in ('vendor_name', 'vendor_id', 'password') if not columns.get(name)]
    if missing: raise RuntimeError(f"Vendor Control sheet missing required columns: {', '.join(missing)}")
    def cell(row, col_index):
        idx = col_index - 1
        return clean(row[idx]) if idx < len(row) else ''
    rows = []
    for sheet_row, raw in enumerate(values[header_index + 1:], start=header_index + 2):
        vendor_name = cell(raw, columns['vendor_name'])
        vendor_id = cell(raw, columns['vendor_id'])
        password = cell(raw, columns['password'])
        if not (vendor_name and vendor_id and password): continue
        if re.fullmatch(r'\d+(?:\.0+)?', str(vendor_name).strip()): continue
        rows.append({'row_key': str(sheet_row), 'sheet_row': sheet_row, 'vendor_name': vendor_name, 'vendor_id': vendor_id, 'password': password})
    fetched_at = datetime.utcnow()
    VENDOR_CONTROL_CACHE.update({'rows': rows, 'fetched_at': fetched_at, 'expires_at': fetched_at + timedelta(seconds=VENDOR_CONTROL_CACHE_TTL_SECONDS), 'columns': columns})
    return rows, fetched_at

def sync_vendor_control_cell(sheet_row, field, value):
    if field not in ('vendor_id', 'password'): return {'synced': False, 'reason': 'Field is not editable'}
    sheet_name = vendor_control_sheet_name()
    columns = VENDOR_CONTROL_CACHE.get('columns') or {}
    if not columns.get(field):
        vendor_control_rows(force=True)
        columns = VENDOR_CONTROL_CACHE.get('columns') or {}
    column_index = columns.get(field)
    if not column_index: return {'synced': False, 'reason': f'Sheet column for {field} was not found'}
    range_name = f"'{sheet_name}'!{column_letter_from_index(column_index)}{sheet_row}"
    google_sheets_values_update(VENDOR_CONTROL_SHEET_ID, range_name, [[value or '']])
    VENDOR_CONTROL_CACHE['expires_at'] = None
    return {'synced': True, 'range': range_name}

def parse_date(val):
    if val is None: return None
    try:
        if pd.isna(val): return None
    except: pass
    raw = str(val).strip()
    if not raw or raw.lower() in ('nan', 'none', 'nat', '-', '#n/a', 'n/a'): return None
    if re.match(r'^\d{8}(\.0)?$', raw):
        try: return datetime.strptime(raw[:8], '%Y%m%d').date()
        except: pass
    try:
        parsed = pd.to_datetime(val, errors='coerce')
        if pd.isna(parsed): return None
        return parsed.date()
    except: return None

def safe_float(val, default=0.0):
    try:
        if pd.isna(val): return default
    except: pass
    try: return float(val)
    except: return default

def find_column(df, names):
    low = {c.lower().strip(): c for c in df.columns}
    for n in names:
        if n.lower().strip() in low: return low[n.lower().strip()]
    return None

def uploaded_files():
    files = []
    for key in ('file', 'files'): files.extend(request.files.getlist(key))
    return [f for f in files if f and f.filename]

def read_upload_excel(file):
    raw = file.read()
    file.seek(0)
    filename = (file.filename or '').lower()
    is_xls_format = raw[:4] == b'\xd0\xcf\x11\xe0'
    engine = 'xlrd' if is_xls_format or filename.endswith('.xls') else 'openpyxl'
    # Read ALL columns as strings so leading zeros in IDs (e.g. Category ID
    # "000200030000100007") are preserved. Pandas otherwise auto-infers
    # numeric columns and drops leading zeros — that destroyed Category ID
    # data on every upload. `keep_default_na=False` keeps empty cells as ''
    # instead of NaN so downstream `clean()` works uniformly.
    try:
        return pd.read_excel(file, sheet_name=0, engine=engine, dtype=str, keep_default_na=False)
    except Exception:
        # Fallback: some sheets cannot honor dtype=str (mixed-type quirks).
        # In that case read normally and let `clean()` normalize per cell.
        file.seek(0)
        return pd.read_excel(file, sheet_name=0, engine=engine)

def _json_rows_to_dataframe(rows, columns=None):
    if rows is None: rows = []
    if not isinstance(rows, list): raise ValueError('JSON rows/data must be a list')
    if columns: return pd.DataFrame(rows, columns=[str(c).strip() for c in columns])
    if not rows: return pd.DataFrame()
    if all(isinstance(r, dict) for r in rows): return pd.DataFrame(rows)
    return pd.DataFrame(rows)

def _json_payload_to_uploads(payload, default_filename='json_upload'):
    if payload is None: raise ValueError('Invalid or empty JSON body')
    def one(obj, index=1):
        if isinstance(obj, dict):
            filename = clean(obj.get('filename')) or clean(obj.get('name')) or f'{default_filename}_{index}.json'
            columns = obj.get('columns')
            rows = obj.get('rows') if 'rows' in obj else obj.get('data') if 'data' in obj else obj.get('records') if 'records' in obj else obj.get('items') if 'items' in obj else None
            if rows is None:
                row = {k: v for k, v in obj.items() if k not in ('filename', 'name', 'columns')}
                rows = [row] if row else []
            df = _json_rows_to_dataframe(rows, columns=columns)
            df.columns = [str(c).strip() for c in df.columns]
            return {'filename': filename, 'df': df}
        if isinstance(obj, list):
            filename = f'{default_filename}_{index}.json'
            df = _json_rows_to_dataframe(obj)
            df.columns = [str(c).strip() for c in df.columns]
            return {'filename': filename, 'df': df}
        raise ValueError('Each JSON upload must be an object or list')
    uploads = []
    if isinstance(payload, dict) and isinstance(payload.get('files'), list):
        for idx, item in enumerate(payload.get('files') or [], start=1): uploads.append(one(item, idx))
    else: uploads.append(one(payload, 1))
    return [u for u in uploads if u['df'] is not None]

def request_upload_dataframes(default_filename='upload'):
    content_type = (request.content_type or '').lower()
    if request.is_json or 'application/json' in content_type:
        payload = request.get_json(silent=True)
        uploads = _json_payload_to_uploads(payload, default_filename=default_filename)
        return uploads, 'json'
    files = uploaded_files()
    uploads = []
    for file in files:
        df = read_upload_excel(file)
        df.columns = [str(c).strip() for c in df.columns]
        uploads.append({'filename': file.filename, 'df': df})
    return uploads, 'excel'


def request_upload_all_sheets(default_filename='upload'):
    """Like request_upload_dataframes, but reads EVERY sheet from each
    uploaded Excel file (not just the first one). Used by Master PIC upload
    so the 4-sheet template (By Category / By Client ID / By Vendor / By Bid
    Type) can be processed in a single upload.

    Each returned upload dict has shape:
        {'filename': str, 'sheet_name': str, 'df': DataFrame}
    where `sheet_name` is the original sheet title from the workbook.
    """
    content_type = (request.content_type or '').lower()
    if request.is_json or 'application/json' in content_type:
        # JSON uploads have no concept of multiple sheets — fall through to
        # the regular single-df parser.
        payload = request.get_json(silent=True)
        uploads = _json_payload_to_uploads(payload, default_filename=default_filename)
        for u in uploads: u['sheet_name'] = ''
        return uploads, 'json'
    files = uploaded_files()
    uploads = []
    for file in files:
        raw = file.read()
        file.seek(0)
        filename = (file.filename or '').lower()
        is_xls_format = raw[:4] == b'\xd0\xcf\x11\xe0'
        engine = 'xlrd' if is_xls_format or filename.endswith('.xls') else 'openpyxl'
        try:
            sheet_to_df = pd.read_excel(io.BytesIO(raw), sheet_name=None, engine=engine, dtype=str, keep_default_na=False)
        except Exception:
            file.seek(0)
            try:
                sheet_to_df = pd.read_excel(file, sheet_name=None, engine=engine)
            except Exception:
                # Fallback: read just the first sheet
                file.seek(0)
                df = read_upload_excel(file)
                df.columns = [str(c).strip() for c in df.columns]
                uploads.append({'filename': file.filename, 'sheet_name': '', 'df': df})
                continue
        for sheet_name, df in sheet_to_df.items():
            df.columns = [str(c).strip() for c in df.columns]
            uploads.append({'filename': file.filename, 'sheet_name': str(sheet_name), 'df': df})
    return uploads, 'excel'

def upload_replace_mode():
    raw = request.args.get('replace') or request.args.get('snapshot') or ''
    if not raw and request.is_json:
        payload = request.get_json(silent=True) or {}
        if isinstance(payload, dict): raw = payload.get('replace') or payload.get('snapshot') or ''
    return str(raw).strip().lower() in ('1', 'true', 'yes', 'replace', 'snapshot')

def validate_upload_columns(filename, label, col_map, expected, required, max_missing=3):
    missing_expected = [display for key, display in expected if not col_map.get(key)]
    if len(missing_expected) > max_missing:
        raise ValueError(f'Struktur kolom tidak cocok untuk {label}: lebih dari {max_missing} kolom penting tidak ditemukan ({", ".join(missing_expected)}). Pastikan file yang diupload benar.')
    missing_required = [display for key, display in required if not col_map.get(key)]
    if missing_required:
        raise ValueError(f'Struktur kolom tidak cocok untuk {label}: kolom wajib tidak ditemukan: {", ".join(missing_required)}.')

def _product_id_columns(df):
    return {
        'product_id': find_column(df, ['Product ID', 'Prod. ID', 'Prod ID']),
        'category_id': find_column(df, ['Category ID', 'Category Id', 'CategoryID', 'Cat. ID', 'Cat. ID.']),
        'category_name': find_column(df, ['Category Name', 'Category Nm.', 'Cat. Nm.', 'Cat. Nm']),
        'product_name': find_column(df, ['Product Name', 'Prod. Nm.', 'Prod. Nm', 'Product Name(EN)']),
        'product_status': find_column(df, ['Product Status', 'Prod. Status', 'Prod Status']),
        'specification': find_column(df, ['Specification', 'Spec.', 'Spec']),
        'manufacturer_name': find_column(df, ['Manufacturer Name', 'Mfr. Nm.', 'Mfr. Nm', 'Maker Nm.']),
        # vendor_name intentionally omitted — source Excel has no Vendor column.
        'order_unit': find_column(df, ['Order Unit', 'Odr. Unit', 'Odr. Unit.']),
        'hub_handling_check': find_column(df, ['HUB Handling Check', 'HUB Handling Chk.', 'HUB Handling Chk']),
        'tax_type': find_column(df, ['Purchasing Price Tax Type', 'Tax Type', 'Tax Type.', 'Tax']),
        'registration_date': find_column(df, ['Registration Date', 'Prod. Reg. Date', 'Product Registration Date', 'Product Reg. Date', 'Reg. Date']),
        'product_registry_pic': find_column(df, ['Product Registy PIC(Name)', 'Product Registry PIC(Name)', 'Product Registy PIC', 'Product Registry PIC', 'Product Registered by(Name)', 'Prod. Reg. PIC Nm.', 'Prod. Reg. PIC Nm', 'Prod. Reg. PIC', 'Product Registry PIC Name']),
    }

def _master_pic_columns(df):
    return {
        'category_id': find_column(df, ['Category ID', 'Category Id', 'CategoryID', 'Cat. ID', 'Cat. ID.']),
        'category_name': find_column(df, ['Category Name', 'Category Nm.', 'Cat. Nm.', 'Cat. Nm']),
        'pic': find_column(df, ['PIC', 'PIC Name', 'Pur. PIC', 'Purchase PIC', 'Current PIC', 'Nama PIC']),
        'pic_update': find_column(df, ['Update New PIC', 'New PIC', 'Update PIC', 'PIC Baru', 'New PIC Name']),
    }


def _master_client_pic_columns(df):
    """Column detector for the 'By Client ID' sheet of the Master PIC template."""
    return {
        'client_id': find_column(df, ['Client ID', 'Client Id', 'ClientID', 'Client Cd.', 'Client Cd', 'Client Code']),
        'client_name': find_column(df, ['Client Name', 'Client Nm.', 'Client Nm', 'Client', 'Operation Unit Name', 'Op Unit']),
        'pic': find_column(df, ['PIC', 'PIC Name', 'Pur. PIC', 'Purchase PIC', 'Current PIC', 'Nama PIC']),
    }


def _master_vendor_pic_columns(df):
    """Column detector for the 'By Vendor' sheet of the Master PIC template."""
    return {
        'vendor_id': find_column(df, ['Vendor ID', 'Vendor Id', 'VendorID', 'Vendor Code', 'Supplier ID', 'Supplier Code']),
        'vendor_name': find_column(df, ['Vendor Name', 'Vendor Nm.', 'Vendor Nm', 'Vendor', 'Supplier Name', 'Supplier']),
        'pic': find_column(df, ['PIC', 'PIC Name', 'Pur. PIC', 'Purchase PIC', 'Current PIC', 'Nama PIC']),
    }


def _master_bid_type_pic_columns(df):
    """Column detector for the 'By Bid Type' sheet of the Master PIC template."""
    return {
        'bid_type': find_column(df, ['Bid Type', 'Bid Except Type', 'Bid Type Name']),
        'pic': find_column(df, ['PIC', 'PIC Name', 'Pur. PIC', 'Purchase PIC', 'Current PIC', 'Nama PIC']),
    }


def _detect_master_pic_sheet_kind(sheet_name, df):
    """Auto-detect which Master PIC sheet kind a DataFrame represents, based
    on its column headers. Returns one of: 'category', 'client', 'vendor',
    'bid_type', or None if no match.

    Sheet name is also used as a hint (e.g. 'By Category', 'By Client ID',
    'By Vendor', 'By Bid Type') but column detection is the primary signal
    so the function works even if the user renames the sheet."""
    name_lower = str(sheet_name or '').lower().strip()
    # Try column-based detection first — most reliable.
    has_category = bool(find_column(df, ['Category Name', 'Category Nm.', 'Cat. Nm.', 'Cat. Nm']))
    has_client_id = bool(find_column(df, ['Client ID', 'Client Id', 'ClientID', 'Client Cd.', 'Client Cd', 'Client Code']))
    has_vendor_id = bool(find_column(df, ['Vendor ID', 'Vendor Id', 'VendorID', 'Vendor Code', 'Supplier ID', 'Supplier Code']))
    has_bid_type = bool(find_column(df, ['Bid Type', 'Bid Except Type', 'Bid Type Name']))
    if has_category and not has_client_id and not has_vendor_id and not has_bid_type:
        return 'category'
    if has_client_id:
        return 'client'
    if has_vendor_id:
        return 'vendor'
    if has_bid_type:
        return 'bid_type'
    # Fall back to sheet name hints.
    if 'client' in name_lower: return 'client'
    if 'vendor' in name_lower: return 'vendor'
    if 'bid' in name_lower: return 'bid_type'
    if 'categ' in name_lower: return 'category'
    return None

def selected_clients(args=None):
    args = args if args is not None else request.args
    return [c.strip() for c in args.getlist('client') if c and c.strip()]

def selected_pics(args=None):
    args = args if args is not None else request.args
    return [p.strip() for p in args.getlist('pic') if p and p.strip()]

def matches_selected_client(value, clients):
    if not clients: return True
    v = (value or '').strip().lower()
    return any(v == c.lower() for c in clients)

def apply_so_client_filter(query, clients):
    if clients: return query.filter(SOData.operation_unit_name.in_(clients))
    return query

def apply_so_pic_filter(query, pics):
    if not pics: return query
    if '__NONE_PLACEHOLDER__' in pics: return query.filter(SOData.id.is_(None))
    if '(Kosong)' in pics:
        others = [p for p in pics if p != '(Kosong)']
        empty_pic = db.or_(SOData.pic_name.is_(None), SOData.pic_name == '')
        if others:
            return query.filter(db.or_(SOData.pic_name.in_(others), empty_pic))
        return query.filter(empty_pic)
    return query.filter(SOData.pic_name.in_(pics))

def canonical_pending_pic(pic, client_or_op_unit=None):
    # PIC is now fully controlled by Master PIC (by category, client ID, and
    # vendor ID). The old "Yupi → Andre" auto-assignment is removed.
    return pic or 'Unassigned'

def canonical_rfq_pic(row):
    return canonical_pending_pic(clean(row.get('purchase_pic')), row.get('client_name'))

def sort_pic_kpis(rows):
    return sorted(rows, key=lambda x: (-x.get('count', 0), x.get('pic') or ''))

def apply_item_registration_pic_filter(query, pics):
    if not pics: return query
    # IMPORTANT: ItemRegistration.pic column may be stale. The PIC shown in
    # table and KPI is resolved in real-time via resolve_item_registration_pic
    # which checks MasterClientPIC, MasterBidTypePIC, and MasterPIC.
    # Filter must match against RESOLVED PIC, not stored column.
    _client_pic_cache = {normalize_client_id(m.client_id): clean(m.pic_name) for m in db.session.query(MasterClientPIC).all() if clean(m.pic_name)}
    _bid_type_pic_cache = {clean(m.bid_type): clean(m.pic_name) for m in db.session.query(MasterBidTypePIC).all() if clean(m.pic_name)}
    _cat_by_id, _cat_by_name = master_pic_maps()
    all_rows = query.with_entities(
        ItemRegistration.id, ItemRegistration.client_id, ItemRegistration.bid_except_type,
        ItemRegistration.category_id, ItemRegistration.category, ItemRegistration.pic,
        ItemRegistration.client_name
    ).all()
    matching_ids = []
    for row_id, cid_val, bt_val, cat_id_val, cat_val, pic_val, client_val in all_rows:
        pic = None
        cid = normalize_client_id(cid_val) if cid_val else ''
        if cid and cid in _client_pic_cache: pic = _client_pic_cache[cid]
        if not pic and bt_val:
            bt = clean(bt_val)
            if bt and bt in _bid_type_pic_cache: pic = _bid_type_pic_cache[bt]
        if not pic:
            cat_id = normalize_category_id(cat_id_val)
            cat_name = normalize_category_name(cat_val)
            if cat_id and cat_id in _cat_by_id: pic = _cat_by_id[cat_id]
            elif cat_name and cat_name in _cat_by_name: pic = _cat_by_name[cat_name]
        pic = pic or clean(pic_val) or ''
        pic = canonical_pending_pic(pic, client_val)
        if pic in pics:
            matching_ids.append(row_id)
    if not matching_ids:
        return query.filter(ItemRegistration.id == -1)
    return query.filter(ItemRegistration.id.in_(matching_ids))

def item_registration_dict(row, registered_items=None, include_similarity=True):
    pic = resolve_item_registration_pic(row)
    similar_items = find_similar_registered_items(row, registered_items) if include_similarity else None
    # Normalize vendor_id: strip leading zeros + .0 suffix. Also try to look
    # up vendor_id from Vendor Control data by matching vendor_name (for rows
    # that don't have vendor_id stored directly).
    raw_vid = clean(getattr(row, 'vendor_id', None))
    vid = normalize_vendor_id(raw_vid) if raw_vid else ''
    if not vid and row.vendor_name:
        # Fallback: look up vendor_id from Vendor Control by vendor_name match
        try:
            vc_rows, _ = vendor_control_rows(force=False)
            for vc in vc_rows:
                if clean(vc.get('vendor_name')) and clean(vc.get('vendor_name')).strip().lower() == clean(row.vendor_name).strip().lower():
                    vid = normalize_vendor_id(vc.get('vendor_id')) or ''
                    break
        except Exception:
            pass
    return {
        'id': row.id, 'proc_status': row.proc_status or '', 'req_date': row.req_date.isoformat() if row.req_date else '',
        'existing_owner': row.existing_owner or '', 'client_name': row.client_name or '',
        'operation_unit_name': row.operation_unit_name or '',
        'category': source_category_level1(row.category),
        'pic': pic, 'req_no': row.req_no or '', 'prod_id': row.prod_id or '', 'batch_grp_no': row.batch_grp_no or '',
        'prod_name': row.prod_name or '', 'spec': row.spec or '', 'mfr_name': row.mfr_name or '', 'odr_unit': row.odr_unit or '',
        'bid_except_type': row.bid_except_type or '',
        'vendor_name': row.vendor_name or '', 'vendor_id': vid,
        'prod_price': row.prod_price or 0, 'curr': row.curr or '', 'remarks': row.remarks or '',
        'uploaded_at': utc_isoformat(row.uploaded_at), 'similar_items': similar_items,
        'similar_prod_ids': (similar_items or {}).get('product_ids', ''), 'similar_prod_name': (similar_items or {}).get('product_name', ''),
        'similar_spec': (similar_items or {}).get('specification', ''), 'similar_mfr_name': (similar_items or {}).get('manufacturer_name', ''),
        'similar_odr_unit': (similar_items or {}).get('order_unit', ''), 'similar_score': (similar_items or {}).get('similarity', None),
        'similar_count': (similar_items or {}).get('count', 0),
    }

def product_category_level1(product_id):
    if not product_id: return ''
    prod = db.session.query(ProductIDDB).filter_by(product_id=str(product_id).strip()).first()
    if not prod or not prod.category_name: return ''
    full_category = prod.category_name.strip()
    return full_category.split('>')[0].strip() if '>' in full_category else full_category

def source_category_level1(category_value):
    category = clean(category_value)
    if not category: return ''
    if '>' in category: return category.split('>', 1)[0].strip()
    return category.strip()

def normalize_category_id(value):
    """Normalize a Category ID for storage / lookup.

    IMPORTANT: never strip leading zeros. Category IDs in the source Excel
    look like "000200030000100007" and the leading zeros are significant
    (they identify the category hierarchy depth). Pandas may still convert
    a numeric string column to float on read, producing values like
    "2.00030000100007e+17" or "200030000100007.0" — we reverse that here.
    """
    cat_id = clean(value)
    if not cat_id: return ''
    # Reverse pandas float coercion: "123.0" -> "123" (only when the entire
    # string is digits + ".0" suffix, so we don't accidentally touch decimal
    # values).
    if re.match(r'^\d+\.0+$', cat_id): return cat_id.split('.', 1)[0]
    # Reverse scientific notation: "2.0003e+17" -> "200030000000000000"
    # Pandas produces this for big numbers when dtype wasn't honored.
    sci = re.match(r'^(\d+)\.(\d+)e\+(\d+)$', cat_id)
    if sci:
        mantissa_digits = sci.group(1) + sci.group(2)
        exponent = int(sci.group(3))
        # Pad with zeros so the total digit count equals exponent+1
        target_len = exponent + 1
        if len(mantissa_digits) <= target_len:
            mantissa_digits = mantissa_digits + '0' * (target_len - len(mantissa_digits))
        # We cannot reconstruct leading zeros from scientific notation —
        # the source must be re-uploaded with dtype=str. At least restore
        # the magnitude correctly.
        return mantissa_digits
    return cat_id.strip()

def normalize_vendor_id(value):
    vid = clean(value)
    if not vid: return ''
    if re.match(r'^\d+\.0+$', vid): vid = vid.split('.', 1)[0]
    vid = vid.lstrip('0') or '0'
    return vid

def normalize_client_id(value):
    cid = clean(value)
    if not cid: return ''
    if re.match(r'^\d+\.0+$', cid): cid = cid.split('.', 1)[0]
    return cid.strip()

def normalize_category_name(value):
    category = source_category_level1(value)
    if not category: return ''
    return re.sub(r'\s+', ' ', category).strip().lower()

def master_pic_category_key(category_name):
    norm = normalize_category_name(category_name)
    if not norm: return ''
    return f"CATNAME_{hashlib.sha1(norm.encode('utf-8')).hexdigest()[:16]}"

def find_master_pic_by_category_name(category_name):
    norm = normalize_category_name(category_name)
    if not norm: return None
    generated_key = master_pic_category_key(category_name)
    if generated_key:
        existing = db.session.query(MasterPIC).filter_by(category_id=generated_key).first()
        if existing: return existing
    for item in db.session.query(MasterPIC).order_by(MasterPIC.updated_at.desc()).all():
        if normalize_category_name(item.category_name) == norm: return item
    return None

def master_pic_unique_category_count():
    return len({normalize_category_name(m.category_name) for m in db.session.query(MasterPIC.category_name).all() if normalize_category_name(m.category_name)})

def invalidate_master_pic_cache():
    _MASTER_PIC_CACHE['signature'] = None
    _MASTER_PIC_CACHE['by_id'] = {}
    _MASTER_PIC_CACHE['by_name'] = {}

def master_pic_maps():
    if _MASTER_PIC_CACHE.get('signature') is not None: return _MASTER_PIC_CACHE['by_id'], _MASTER_PIC_CACHE['by_name']
    signature = db.session.query(func.count(MasterPIC.id), func.max(MasterPIC.updated_at)).one()
    signature = tuple(signature)
    by_id = {}; by_name = {}
    for m in MasterPIC.query.with_entities(MasterPIC.category_id, MasterPIC.category_name, MasterPIC.pic_name).order_by(MasterPIC.updated_at.desc()).all():
        pic = clean(m.pic_name)
        if not pic: continue
        cat_id = normalize_category_id(m.category_id)
        if cat_id and cat_id not in by_id: by_id[cat_id] = pic
        cat_name = normalize_category_name(m.category_name)
        if cat_name and cat_name not in by_name: by_name[cat_name] = pic
    _MASTER_PIC_CACHE['signature'] = signature
    _MASTER_PIC_CACHE['by_id'] = by_id
    _MASTER_PIC_CACHE['by_name'] = by_name
    return by_id, by_name

def _lookup_pic_by_category(category_id=None, category_name=None):
    by_id, by_name = master_pic_maps()
    cat_id = normalize_category_id(category_id)
    if cat_id and cat_id in by_id: return by_id[cat_id]
    cat_name = normalize_category_name(category_name)
    if cat_name and cat_name in by_name: return by_name[cat_name]
    return None

def resolve_pic_with_overrides(category_id=None, category_name=None, client_id=None, vendor_id=None, bid_except_type=None):
    if client_id:
        cid = normalize_client_id(client_id)
        if cid:
            r = db.session.query(MasterClientPIC.pic_name).filter(MasterClientPIC.client_id == cid).first()
            if r and clean(r[0]): return clean(r[0])
    if bid_except_type:
        bt = clean(bid_except_type)
        if bt:
            r = db.session.query(MasterBidTypePIC.pic_name).filter(MasterBidTypePIC.bid_type == bt).first()
            if r and clean(r[0]): return clean(r[0])
    if vendor_id:
        vid = normalize_vendor_id(vendor_id)
        if vid:
            r = db.session.query(MasterVendorPIC.pic_name).filter(MasterVendorPIC.vendor_id == vid).first()
            if r and clean(r[0]): return clean(r[0])
    return _lookup_pic_by_category(category_id, category_name)


def resolve_item_registration_pic(row):
    client_id = getattr(row, 'client_id', None) or (row.get('client_id') if isinstance(row, dict) else None)
    bid_except_type = getattr(row, 'bid_except_type', None) or (row.get('bid_except_type') if isinstance(row, dict) else None)
    pic = resolve_pic_with_overrides(
        category_id=row.category_id if hasattr(row, 'category_id') else row.get('category_id'),
        category_name=row.category if hasattr(row, 'category') else row.get('category'),
        client_id=client_id, vendor_id=None, bid_except_type=bid_except_type,
    )
    return canonical_pending_pic(pic or (row.pic if hasattr(row, 'pic') else row.get('pic')) or '', row.client_name if hasattr(row, 'client_name') else row.get('client_name'))

def is_existing_owner_pur_pic(value): return (clean(value) or '').strip().lower() == 'pur. pic'

ITEM_REG_KPI_EXCLUDED_STATUSES = {'sales pic terminate(pur. pic)', 'purchase exception termination', 'sales pic confirmation req.(pur. pic)', 'pre-reg. prod. proc.(pur.)'}

def item_registration_kpi_status_expr(): return func.lower(func.trim(func.coalesce(ItemRegistration.proc_status, '')))

def apply_item_registration_kpi_status_filter(query):
    status_expr = item_registration_kpi_status_expr()
    return query.filter(~status_expr.in_(list(ITEM_REG_KPI_EXCLUDED_STATUSES)), ~status_expr.like('%sales%'))

def apply_item_registration_visible_status_filter(query): return query.filter(~item_registration_kpi_status_expr().like('%sales%'))

def refresh_item_registration_mappings():
    rows = ItemRegistration.query.all()
    changed = False
    for row in rows:
        category = source_category_level1(row.category)
        normalized_cat_id = normalize_category_id(row.category_id)
        if row.category_id != normalized_cat_id: row.category_id = normalized_cat_id; changed = True
        if row.category != category: row.category = category; changed = True
        pic = resolve_pic_with_overrides(
            category_id=normalized_cat_id, category_name=category,
            client_id=row.client_id, vendor_id=None, bid_except_type=row.bid_except_type,
        ) or ''
        if row.pic != pic: row.pic = pic; changed = True
    if changed: db.session.commit()

def _item_registration_columns(df):
    return {
        'proc_status': find_column(df, ['Proc. Status', 'Proc Status', 'Process Status']),
        'req_date': find_column(df, ['Req. Date', 'Req Date', 'Request Date']),
        'existing_owner': find_column(df, ['Existing Owner', 'Existing Owner.', 'Owner']),
        'client_id': find_column(df, ['Client ID', 'Client Id', 'ClientID', 'Client Cd.', 'Client Cd', 'Client Code']),
        'client_name': find_column(df, ['Client Nm.', 'Client Nm', 'Client Name']),
        'operation_unit_name': find_column(df, ['Op. Unit Nm.', 'Op. Unit Nm', 'Op Unit Nm', 'Operation Unit Nm.', 'Operation Unit Nm', 'Operation Unit Name', 'Op. Unit Name', 'Op Unit Name']),
        'category': find_column(df, ['Cat. Nm.', 'Cat. Nm', 'Category', 'Cate. Nm.', 'Category Name']),
        'category_id': find_column(df, ['Cat. ID', 'Cat. ID.', 'Category ID', 'Category Id', 'CategoryID']),
        'pic': find_column(df, ['PIC', 'Pur. PIC', 'Purchase PIC']),
        'req_no': find_column(df, ['Req. No', 'Req. No.', 'Request No', 'Request Number']),
        'prod_id': find_column(df, ['Prod. ID', 'Prod ID', 'Product ID']),
        'product_status': find_column(df, ['Product Status', 'Prod. Status', 'Prod Status']),
        'batch_grp_no': find_column(df, ['Batch Grp. No.', 'Batch Grp. No', 'Batch Group No']),
        'prod_name': find_column(df, ['Prod. Nm.', 'Prod. Nm', 'Product Name', 'Prod. Nm.(Eng.)']),
        'spec': find_column(df, ['Spec.', 'Spec', 'Specification']),
        'mfr_name': find_column(df, ['Mfr. Nm.', 'Mfr. Nm', 'Manufacturer Name', 'Maker Nm.']),
        'odr_unit': find_column(df, ['Odr. Unit', 'Odr. Unit.', 'Order Unit']),
        'bid_except_type': find_column(df, ['Bid Except Type', 'Bid/Except Type', 'Bid Type', 'Bid/Except. Type']),
        'vendor_name': find_column(df, ['Vendor Nm.', 'Vendor Nm', 'Vendor Name']),
        'vendor_id': find_column(df, ['Vendor ID', 'Vendor Id', 'VendorID', 'Vendor Code', 'Supplier ID']),
        'prod_price': find_column(df, ['Prod. Price', 'Product Price', 'Price']),
        'curr': find_column(df, ['Curr.', 'Curr', 'Currency']),
        'hub_handling_check': find_column(df, ['HUB Handling Chk.', 'HUB Handling Chk', 'HUB Handling Check', 'Hub Handling Check', 'Hub Handling Chk.']),
        'tax_type': find_column(df, ['Tax Type', 'Tax Type.', 'Tax']),
        'registration_date': find_column(df, ['Prod. Reg. Date', 'Product Reg. Date', 'Product Registration Date', 'Registration Date', 'Reg. Date']),
        'product_registry_pic': find_column(df, ['Prod. Reg. PIC Nm.', 'Prod. Reg. PIC Nm', 'Prod. Reg. PIC', 'Product Registry PIC', 'Product Registration PIC', 'Product Reg. PIC']),
    }

def validate_item_registration_source_file(df, filename='Item Registration'):
    marker_cols = {str(c).strip().lower() for c in df.columns}
    process_markers = {'unified vendor', 'bid/quo.', 'multi. bidding required', 'bid no.', 'deadline', 'pur. info. proc. compl. date', 'vendor confirm req. detail', 'vendor confirm proc. detail'}
    prod_reg_markers = {'register request', 'prod. req. skip reason', 'prod. reg. req. compl. date', 'prod. reg. req. reject date'}
    matched_process = sorted(process_markers & marker_cols)
    matched_prod_reg = sorted(prod_reg_markers & marker_cols)
    if matched_prod_reg and not matched_process:
        raise ValueError('Struktur kolom tidak cocok untuk Item Registration. Upload yang benar adalah struktur SAP Process Pur. Info. Reg., bukan Prod. Reg. Status.')
    if not matched_process:
        raise ValueError('Struktur kolom tidak terlihat seperti SAP Process Pur. Info. Reg. Kolom marker wajib tidak ditemukan: Unified Vendor / Bid/Quo. / Multi. Bidding Required / Bid No. / Deadline.')

def import_item_registration_dataframe(df, filename='Item Registration'):
    df.columns = [str(c).strip() for c in df.columns]
    validate_item_registration_source_file(df, filename)
    col = _item_registration_columns(df)
    expected = [('proc_status', 'Proc. Status'), ('client_name', 'Client Nm.'), ('category', 'Cat. Nm.'), ('category_id', 'Category ID'), ('req_no', 'Req. No'), ('prod_name', 'Product Name'), ('spec', 'Specification'), ('mfr_name', 'Manufacturer Name'), ('odr_unit', 'Order Unit'), ('vendor_name', 'Vendor Name'), ('prod_price', 'Prod. Price'), ('curr', 'Curr.')]
    required = [('proc_status', 'Proc. Status'), ('client_name', 'Client Nm.'), ('category', 'Cat. Nm.'), ('category_id', 'Category ID'), ('req_no', 'Req. No'), ('prod_name', 'Product Name')]
    validate_upload_columns(filename, 'Item Registration', col, expected, required)
    incoming = {}
    for _, row in df.iterrows():
        req_no = clean(df_val(row, col['req_no']))
        prod_id = clean_product_id(df_val(row, col['prod_id']))
        prod_name = clean(df_val(row, col['prod_name']))
        if not req_no: continue
        category_id = normalize_category_id(df_val(row, col['category_id']))
        category = source_category_level1(df_val(row, col['category']))
        incoming[req_no] = {
            'proc_status': clean(df_val(row, col['proc_status'])), 'req_date': parse_date(df_val(row, col['req_date'])),
            'existing_owner': clean(df_val(row, col['existing_owner'])),
            'client_id': clean(df_val(row, col['client_id'])),
            'client_name': clean(df_val(row, col['client_name'])),
            'operation_unit_name': clean(df_val(row, col['operation_unit_name'])),
            'category': category, 'category_id': category_id, 'pic': _lookup_pic_by_category(category_id, category) or '',
            'req_no': req_no, 'prod_id': prod_id, 'product_status': clean(df_val(row, col['product_status'])),
            'batch_grp_no': clean(df_val(row, col['batch_grp_no'])), 'prod_name': prod_name, 'spec': clean(df_val(row, col['spec'])),
            'mfr_name': clean(df_val(row, col['mfr_name'])), 'odr_unit': clean(df_val(row, col['odr_unit'])),
            'bid_except_type': clean(df_val(row, col['bid_except_type'])),
            'vendor_name': clean(df_val(row, col['vendor_name'])), 'vendor_id': normalize_vendor_id(clean(df_val(row, col['vendor_id'])) or '') if col.get('vendor_id') else '', 'prod_price': safe_float(df_val(row, col['prod_price'])),
            'curr': clean(df_val(row, col['curr'])), 'hub_handling_check': clean(df_val(row, col['hub_handling_check'])),
            'tax_type': clean(df_val(row, col['tax_type'])), 'registration_date': parse_date(df_val(row, col['registration_date'])),
            'product_registry_pic': clean(df_val(row, col['product_registry_pic'])), 'uploaded_at': datetime.utcnow(),
        }
    req_numbers = list(incoming.keys())
    existing_map = {}; duplicate_rows = []
    if req_numbers:
        existing_rows = ItemRegistration.query.filter(ItemRegistration.req_no.in_(req_numbers)).order_by(ItemRegistration.id.asc()).all()
        for existing in existing_rows:
            if existing.req_no in existing_map: duplicate_rows.append(existing)
            else: existing_map[existing.req_no] = existing
    added = updated = removed_duplicates = 0
    for dup in duplicate_rows:
        db.session.delete(dup); removed_duplicates += 1
    for req_no, payload in incoming.items():
        existing = existing_map.get(req_no)
        if existing:
            for key, value in payload.items(): setattr(existing, key, value)
            updated += 1
        else:
            db.session.add(ItemRegistration(**payload)); added += 1
    db.session.add(UploadLog(file_type='ITEM_REG', filename=filename, records_count=len(incoming)))
    return {'processed': len(incoming), 'added': added, 'updated': updated, 'removed_duplicates': removed_duplicates, 'keys': list(incoming.keys())}

def ensure_default_item_registration_loaded():
    # Disabled: Membaca file Excel besar saat page load menyebabkan loading 10+ detik.
    # Data Item Registration harus di-upload manual via tombol Upload.
    return

def df_val(row, col): return row.get(col) if col else None

def get_aging_label(workday_count):
    if workday_count is None: return '180+'
    if workday_count >= 180: return '180+'
    if workday_count >= 90: return '90-180'
    if workday_count >= 30: return '30-90'
    return '0-30'

def so_dict(s):
    today = date.today()
    age_days = workdays_since(s.so_create_date, today)
    category_name = _pid_category_lookup(s.product_id) if s.product_id else ''
    # Pre-compute IDR-converted purchase price & amount for the frontend
    # so margin can be calculated correctly regardless of currency.
    pp_idr = purchase_price_idr(s)
    pa_idr = purchase_amount_idr_for_margin(s)
    return {
        'id': s.id, 'so_number': s.so_number, 'so_item': s.so_item, 'so_status': s.so_status,
        'operation_unit_name': s.operation_unit_name,
        'vendor_id': normalize_vendor_id(s.vendor_id) if s.vendor_id else '',
        'client_id': normalize_client_id(s.client_id) if s.client_id else '',
        'vendor_name': s.vendor_name,
        'customer_po_number': s.customer_po_number, 'delivery_memo': s.delivery_memo, 'product_name': s.product_name,
        'specification': s.specification, 'manufacturer_name': s.manufacturer_name or '', 'product_id': s.product_id,
        'category_name': category_name, 'svo_po': s.matched_po_number or '', 'so_qty': s.so_qty, 'sales_unit': s.sales_unit or '',
        'sales_price': s.sales_price, 'sales_amount': s.sales_amount, 'currency': s.currency or '',
        'purchasing_price': s.purchasing_price, 'purchasing_amount': s.purchasing_amount, 'purchasing_currency': s.purchasing_currency,
        'purchase_price_idr': pp_idr, 'purchase_amount_idr': pa_idr,
        'so_create_date': s.so_create_date.isoformat() if s.so_create_date else '',
        'delivery_possible_date': s.delivery_possible_date.isoformat() if s.delivery_possible_date else '',
        'delivery_plan_date': s.delivery_plan_date.isoformat() if s.delivery_plan_date else '',
        'remarks': s.remarks or '', 'pic_name': canonical_pending_pic(s.pic_name, s.operation_unit_name),
        'aging_days': age_days, 'aging_label': get_aging_label(age_days)
    }

def get_hidden_so_items(): return set()

def calculate_similarity(str1, str2):
    if not str1 or not str2: return 0.0
    s1 = str(str1).lower().strip(); s2 = str(str2).lower().strip()
    if s1 == s2: return 100.0
    tokens1 = set(s1.split()); tokens2 = set(s2.split())
    if not tokens1 or not tokens2: return 0.0
    intersection = tokens1.intersection(tokens2); union = tokens1.union(tokens2)
    jaccard = len(intersection) / len(union) * 100
    substring_bonus = 20.0 if (s1 in s2 or s2 in s1) else 0.0
    return min(100.0, jaccard + substring_bonus)

def _similarity_token(value):
    text_value = (clean(value) or '').lower()
    tokens = [t for t in re.split(r'[^a-z0-9]+', text_value) if len(t) >= 3]
    return max(tokens, key=len) if tokens else ''

def _candidate_registered_items_for_similarity(item, registered_items=None, limit=1200):
    unit = (clean(item.odr_unit) or '').lower()
    mfr_token = _similarity_token(item.mfr_name); name_token = _similarity_token(item.prod_name)
    if registered_items is not None:
        candidates = []
        for reg in registered_items:
            if unit and (clean(reg.order_unit) or '').lower() != unit: continue
            reg_mfr = (clean(reg.manufacturer_name) or '').lower(); reg_name = (clean(reg.product_name) or '').lower()
            token_matches = []
            if mfr_token: token_matches.append(mfr_token in reg_mfr)
            if name_token: token_matches.append(name_token in reg_name)
            if token_matches and not any(token_matches): continue
            candidates.append(reg)
            if len(candidates) >= limit: break
        return candidates
    q = ProductIDDB.query.filter(ProductIDDB.product_id.isnot(None), ProductIDDB.product_id != '')
    if unit: q = q.filter(func.lower(ProductIDDB.order_unit) == unit)
    token_filters = []
    if mfr_token: token_filters.append(ProductIDDB.manufacturer_name.ilike(f'%{mfr_token}%'))
    if name_token: token_filters.append(ProductIDDB.product_name.ilike(f'%{name_token}%'))
    if token_filters: q = q.filter(db.or_(*token_filters))
    elif not unit: return []
    return q.limit(limit).all()

def _similarity_score(values):
    scores = []
    for left, right in values:
        if clean(left) and clean(right): scores.append(calculate_similarity(left, right))
    if not scores: return 0.0
    return sum(scores) / len(scores)

def find_similar_registered_items(item, registered_items=None):
    try:
        key_fields = [item.prod_name, item.spec, item.mfr_name, item.odr_unit]
        if not any(clean(v) for v in key_fields): return None
        current_prod_id = clean_product_id(item.prod_id)
        cache_key = '|'.join(['similar_v4', clean(item.req_no), current_prod_id, clean(item.prod_name).lower(), clean(item.spec).lower(), clean(item.mfr_name).lower(), clean(item.odr_unit).lower()])
        if cache_key in _SIMILARITY_CACHE: return _SIMILARITY_CACHE[cache_key]
        registered_items = _candidate_registered_items_for_similarity(item, registered_items)
        similar_items = []
        for reg in registered_items:
            reg_prod_id = clean_product_id(reg.product_id)
            if not reg_prod_id or (current_prod_id and reg_prod_id == current_prod_id): continue
            has_descriptive_pair = any(clean(left) and clean(right) for left, right in [(item.prod_name, reg.product_name), (item.spec, reg.specification), (item.mfr_name, reg.manufacturer_name)])
            if not has_descriptive_pair: continue
            total_sim = _similarity_score([(item.prod_name, reg.product_name), (item.spec, reg.specification), (item.mfr_name, reg.manufacturer_name), (item.odr_unit, reg.order_unit)])
            if total_sim > 80.0:
                similar_items.append({'product_id': reg_prod_id, 'product_name': reg.product_name or '', 'specification': reg.specification or '', 'manufacturer_name': reg.manufacturer_name or '', 'order_unit': reg.order_unit or '', 'similarity': round(total_sim, 1)})
        similar_items.sort(key=lambda x: (-x['similarity'], x['product_id']))
        if not similar_items: result = None
        else:
            best = similar_items[0]
            result = {'product_ids': ', '.join(x['product_id'] for x in similar_items), 'product_name': best['product_name'], 'specification': best['specification'], 'manufacturer_name': best['manufacturer_name'], 'order_unit': best['order_unit'], 'similarity': best['similarity'], 'count': len(similar_items)}
        _SIMILARITY_CACHE[cache_key] = result
        return result
    except Exception as e:
        print(f"Error finding similar items: {e}")
        import traceback
        traceback.print_exc()
        return None

# BAGIAN 1 SELESAI DI SINI. SILAKAN BALAS "lanjut" UNTUK MENDAPATKAN BAGIAN 2.
@app.route('/api/dashboard/stats', methods=['GET'])
def get_dashboard_stats():
    try:
        cache_key = runtime_cache_key('dashboard_stats_v2_sql')
        cached = runtime_cache_get(cache_key)
        if cached is not None: return jsonify(cached)
        date_year, date_from, date_to = parse_so_date_args()
        clients = selected_clients()
        pics = selected_pics()
        is_sqlite = 'sqlite' in app.config.get('SQLALCHEMY_DATABASE_URI', '')
        def base_open_q(apply_client=True, apply_pic=True):
            q = db.session.query(SOData).filter(open_so_filter(), so_countable_sql_filter())
            if apply_client: q = apply_so_client_filter(q, clients)
            if apply_pic: q = apply_so_pic_filter(q, pics)
            return apply_so_create_date_filter(q, date_year, date_from, date_to, is_sqlite=is_sqlite)
        q = base_open_q()
        sales_expr = func.coalesce(SOData.sales_amount, 0.0)
        purchase_expr = dashboard_purchase_sql_expr()
        month_expr = func.strftime('%Y-%m', SOData.so_create_date) if is_sqlite else func.to_char(func.date_trunc('month', SOData.so_create_date), 'YYYY-MM')
        month_sort_expr = month_expr
        status_label = func.coalesce(func.nullif(func.trim(SOData.so_status), ''), 'Unknown')
        total_row = q.with_entities(func.count(SOData.id), func.coalesce(func.sum(sales_expr), 0.0)).first()
        total_so_count = int(total_row[0] or 0) if total_row else 0
        total_open_so_amount = float(total_row[1] or 0) if total_row else 0.0
        total_open_for_pct = total_so_count or 1
        monthly_rows = q.filter(SOData.so_create_date.isnot(None)).with_entities(month_expr.label('month'), month_sort_expr.label('month_sort'), func.count(SOData.id).label('so_count'), func.coalesce(func.sum(sales_expr), 0.0).label('amount'), func.coalesce(func.sum(purchase_expr), 0.0).label('purchase_amount')).group_by(month_sort_expr, month_expr).order_by(month_sort_expr).all()
        monthly_trend = [{'month': row.month, 'so_count': int(row.so_count or 0), 'amount': round(float(row.amount or 0) / 1_000_000, 2), 'purchase_amount': round(float(row.purchase_amount or 0) / 1_000_000, 2)} for row in monthly_rows]
        def top_group(label_expr, out_key, amount_expr, limit):
            rows = q.with_entities(label_expr.label(out_key), func.count(SOData.id).label('so_count'), func.coalesce(func.sum(amount_expr), 0.0).label('total_amount')).group_by(label_expr).order_by(desc(func.coalesce(func.sum(amount_expr), 0.0))).limit(limit).all()
            return [{out_key: getattr(row, out_key) or 'Unknown', 'so_count': int(row.so_count or 0), 'total_amount': round(float(row.total_amount or 0), 2)} for row in rows]
        vendor_label = func.coalesce(func.nullif(func.trim(SOData.vendor_name), ''), 'Unknown')
        op_unit_label = func.coalesce(func.nullif(func.trim(SOData.operation_unit_name), ''), 'Unknown')
        top_vendors = top_group(vendor_label, 'vendor', sales_expr, 5)
        top_op_units = top_group(op_unit_label, 'op_unit', sales_expr, 10)
        status_rows = q.with_entities(status_label.label('name'), func.count(SOData.id).label('value'), func.coalesce(func.sum(sales_expr), 0.0).label('amount')).group_by(status_label).order_by(desc(func.count(SOData.id))).all()
        so_status = [{'name': row.name or 'Unknown', 'value': int(row.value or 0), 'percentage': round((int(row.value or 0) / total_open_for_pct) * 100, 1), 'amount': round(float(row.amount or 0), 2)} for row in status_rows]
        monthly_status_rows = q.filter(SOData.so_create_date.isnot(None)).with_entities(status_label.label('name'), month_expr.label('month'), month_sort_expr.label('month_sort'), func.count(SOData.id).label('count'), func.coalesce(func.sum(sales_expr), 0.0).label('amount')).group_by(status_label, month_sort_expr, month_expr).order_by(month_sort_expr).all()
        status_months = []; status_month_sort = {}; status_acc = {}
        for row in monthly_status_rows:
            name = row.name or 'Unknown'; month = row.month
            if month not in status_month_sort: status_month_sort[month] = row.month_sort; status_months.append(month)
            item = status_acc.setdefault(name, {'monthly': {}, 'total': 0, 'amount': 0.0})
            c = int(row.count or 0); item['monthly'][month] = c; item['total'] += c; item['amount'] += float(row.amount or 0)
        status_months = sorted(status_months, key=lambda m: status_month_sort.get(m, m))
        so_status_monthly = sorted([{'name': name, 'monthly': data['monthly'], 'total': data['total'], 'percentage': round((data['total'] / total_open_for_pct) * 100, 1), 'amount': round(data['amount'], 2)} for name, data in status_acc.items()], key=lambda x: x['total'], reverse=True)
        # "Pending Item Registration" must mirror the EXACT same definition used by
        # the Item Registration page's "Total Pending Regist." KPI: status not in the
        # excluded KPI statuses AND prod_id is empty/blank/'-' (i.e. not yet registered
        # with a Product ID). Previously this only applied the status filter, so rows
        # that already had a Product ID (and are therefore NOT pending) were still
        # counted in these charts.
        item_reg_base_q = apply_item_registration_kpi_status_filter(db.session.query(ItemRegistration))
        item_reg_base_q = item_reg_base_q.filter(db.or_(ItemRegistration.prod_id.is_(None), ItemRegistration.prod_id == '', ItemRegistration.prod_id == '-'))
        if clients: item_reg_base_q = item_reg_base_q.filter(ItemRegistration.client_name.in_(clients))
        # These charts use the Date Filter control same as the rest of the dashboard,
        # but Item Registration rows are dated by req_date, not so_create_date.
        item_reg_base_q = apply_item_registration_date_filter(item_reg_base_q, date_year, date_from, date_to)
        def item_registration_distribution(column, limit=None):
            label_expr = func.coalesce(func.nullif(func.trim(column), ''), '(Kosong)')
            rows = item_reg_base_q.with_entities(label_expr.label('name'), func.count(ItemRegistration.id).label('value')).group_by(label_expr).order_by(func.count(ItemRegistration.id).desc(), label_expr.asc())
            if limit: rows = rows.limit(limit)
            return [{'name': name or '(Kosong)', 'value': int(value or 0)} for name, value in rows.all()]
        item_registration_proc_status = item_registration_distribution(ItemRegistration.proc_status)
        item_registration_clients = item_registration_distribution(ItemRegistration.client_name, limit=10)
        # PIC pie chart — PIC is resolved dynamically (Master Client PIC / Bid Type PIC /
        # Category PIC overrides), same as the per-PIC KPI cards on the Item Registration
        # page, so it can't be a simple SQL GROUP BY on a stored column. Build the
        # override caches once up front (same pattern as apply_item_registration_pic_filter)
        # instead of resolving per-row, to avoid N+1 queries on large tables.
        _ir_client_pic_cache = {normalize_client_id(m.client_id): clean(m.pic_name) for m in db.session.query(MasterClientPIC).all() if clean(m.pic_name)}
        _ir_bid_type_pic_cache = {clean(m.bid_type): clean(m.pic_name) for m in db.session.query(MasterBidTypePIC).all() if clean(m.pic_name)}
        _ir_cat_by_id, _ir_cat_by_name = master_pic_maps()
        item_reg_pic_counts = {}
        pic_rows = item_reg_base_q.with_entities(
            ItemRegistration.client_id, ItemRegistration.bid_except_type,
            ItemRegistration.category_id, ItemRegistration.category,
            ItemRegistration.pic, ItemRegistration.client_name
        ).all()
        for cid_val, bt_val, cat_id_val, cat_val, pic_val, client_val in pic_rows:
            pic = None
            cid = normalize_client_id(cid_val) if cid_val else ''
            if cid and cid in _ir_client_pic_cache: pic = _ir_client_pic_cache[cid]
            if not pic and bt_val:
                bt = clean(bt_val)
                if bt and bt in _ir_bid_type_pic_cache: pic = _ir_bid_type_pic_cache[bt]
            if not pic:
                cat_id = normalize_category_id(cat_id_val)
                cat_name = normalize_category_name(cat_val)
                if cat_id and cat_id in _ir_cat_by_id: pic = _ir_cat_by_id[cat_id]
                elif cat_name and cat_name in _ir_cat_by_name: pic = _ir_cat_by_name[cat_name]
            pic = pic or clean(pic_val) or ''
            pic = canonical_pending_pic(pic, client_val)
            item_reg_pic_counts[pic] = item_reg_pic_counts.get(pic, 0) + 1
        item_registration_pics = sorted(
            [{'name': name, 'value': value} for name, value in item_reg_pic_counts.items()],
            key=lambda x: (-x['value'], x['name'])
        )
        option_q = base_open_q(apply_client=False, apply_pic=False)
        client_options = [r[0] for r in option_q.with_entities(SOData.operation_unit_name).filter(SOData.operation_unit_name.isnot(None), SOData.operation_unit_name != '').distinct().order_by(SOData.operation_unit_name).all()]
        raw_pic_options = [r[0] for r in option_q.with_entities(SOData.pic_name).filter(SOData.pic_name.isnot(None), SOData.pic_name != '').distinct().order_by(SOData.pic_name).all()]
        pic_options = []; seen_pics = set()
        for p in raw_pic_options:
            label = canonical_pending_pic(p, '')
            if label and label not in seen_pics: seen_pics.add(label); pic_options.append(label)
        
        last_uploads = db.session.query(UploadLog.file_type, func.max(UploadLog.uploaded_at)).group_by(UploadLog.file_type).all()
        last_upload_map = {ft: dt for ft, dt in last_uploads}
        last_upload = max(last_upload_map.values()) if last_upload_map else None
        last_so_upload = last_upload_map.get('SO') or last_upload_map.get('SMRO')
        last_item_reg_upload = last_upload_map.get('ITEM_REG') or last_upload_map.get('ItemRegistration')
        last_po_upload = None
        rfq_fetched_at = RFQ_CACHE.get('fetched_at')
        
        so_date_range = db.session.query(func.min(SOData.so_create_date), func.max(SOData.so_create_date)).first()
        po_date_range = (None, None)
        if is_sqlite:
            covered_rows = db.session.query(func.strftime('%Y', SOData.so_create_date).label('yr'), func.strftime('%m', SOData.so_create_date).label('mo')).filter(SOData.so_create_date.isnot(None)).distinct().all()
        else:
            covered_rows = db.session.query(func.extract('year', SOData.so_create_date).label('yr'), func.extract('month', SOData.so_create_date).label('mo')).filter(SOData.so_create_date.isnot(None)).distinct().all()
        _MONTH_NAMES = ['January','February','March','April','May','June','July','August','September','October','November','December']
        so_covered_months = {}
        for yr, mo in covered_rows:
            if yr is None or mo is None: continue
            yr_s = str(int(yr)) if not isinstance(yr, str) else yr; mo_i = int(mo)
            so_covered_months.setdefault(yr_s, []).append((mo_i, _MONTH_NAMES[mo_i - 1]))
        so_covered_months = {yr: [name for _, name in sorted(months)] for yr, months in sorted(so_covered_months.items())}
        wib_today = (datetime.utcnow() + timedelta(hours=7)).date()
        today_start_utc = datetime.combine(wib_today, datetime.min.time()) - timedelta(hours=7)
        tomorrow_start_utc = today_start_utc + timedelta(days=1)
        # "SO months updated today" — show only months from the LATEST upload
        # batch, not all rows uploaded today. Find max(uploaded_at) and use
        # a 5-minute window to capture only the latest batch.
        latest_upload_time = db.session.query(func.max(SOData.uploaded_at)).scalar()
        so_updated_months_today = {}
        if latest_upload_time:
            latest_wib = latest_upload_time + timedelta(hours=7)
            if latest_wib.date() == wib_today:
                upload_window_start = latest_upload_time - timedelta(minutes=5)
                updated_today_filters = (
                    SOData.so_create_date.isnot(None),
                    SOData.uploaded_at.isnot(None),
                    SOData.uploaded_at >= upload_window_start,
                    SOData.uploaded_at <= latest_upload_time + timedelta(seconds=1),
                )
                if is_sqlite:
                    updated_month_rows = db.session.query(func.strftime('%Y', SOData.so_create_date).label('yr'), func.strftime('%m', SOData.so_create_date).label('mo')).filter(*updated_today_filters).distinct().all()
                else:
                    updated_month_rows = db.session.query(func.extract('year', SOData.so_create_date).label('yr'), func.extract('month', SOData.so_create_date).label('mo')).filter(*updated_today_filters).distinct().all()
                for yr, mo in updated_month_rows:
                    if yr is None or mo is None: continue
                    yr_s = str(int(yr)) if not isinstance(yr, str) else yr; mo_i = int(mo)
                    so_updated_months_today.setdefault(yr_s, []).append((mo_i, _MONTH_NAMES[mo_i - 1]))
                so_updated_months_today = {yr: [name for _, name in sorted(months)] for yr, months in sorted(so_updated_months_today.items())}
        payload = {
            'po_without_so': 0, 'so_without_po': total_so_count, 'total_po_count': 0, 'total_po_line_count': 0, 'total_po_amount': 0.0,
            'total_so_count': total_so_count, 'total_open_so_amount': total_open_so_amount, 'monthly_trend': monthly_trend,
            'top_vendors': top_vendors, 'top_op_units': top_op_units, 'so_status': so_status, 'so_status_monthly': so_status_monthly,
            'status_months': status_months, 'item_registration_proc_status': item_registration_proc_status, 'item_registration_clients': item_registration_clients,
            'item_registration_pics': item_registration_pics,
            'filters': {'clients': client_options, 'pics': pic_options}, 'last_updated': utc_isoformat(last_upload), 'last_updated_po': utc_isoformat(last_po_upload),
            'last_updated_smro': utc_isoformat(last_so_upload), 'last_updated_item_registration': utc_isoformat(last_item_reg_upload), 'last_updated_rfq': utc_isoformat(rfq_fetched_at),
            'so_covered_months': so_covered_months, 'so_updated_months_today': so_updated_months_today, 'so_updated_months_today_date': wib_today.isoformat(),
            'po_date_range': {'min': None, 'max': None},
            'so_date_range': {'min': so_date_range[0].isoformat() if so_date_range and so_date_range[0] else None, 'max': so_date_range[1].isoformat() if so_date_range and so_date_range[1] else None},
        }
        runtime_cache_set(cache_key, payload, ttl_seconds=300)
        return jsonify(payload)
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/debug/so-fields', methods=['GET'])
def debug_so_fields():
    try:
        total = db.session.query(func.count(SOData.id)).scalar() or 0
        has_spec = db.session.query(func.count(SOData.id)).filter(SOData.specification.isnot(None), SOData.specification != '').scalar() or 0
        has_pid = db.session.query(func.count(SOData.id)).filter(SOData.product_id.isnot(None), SOData.product_id != '').scalar() or 0
        samples = db.session.query(SOData.so_item, SOData.product_name, SOData.specification, SOData.product_id).limit(10).all()
        return jsonify({
            'total_so_records': total, 'records_with_specification': has_spec, 'records_with_product_id': has_pid,
            'spec_fill_pct': round(has_spec / total * 100, 1) if total else 0, 'pid_fill_pct': round(has_pid / total * 100, 1) if total else 0,
            'sample_rows': [{'so_item': r[0], 'product_name': r[1], 'specification': r[2], 'product_id': r[3]} for r in samples]
        })
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/debug/smro-columns', methods=['POST'])
def debug_smro_columns():
    try:
        if 'file' not in request.files: return jsonify({'error': 'No file uploaded'}), 400
        file = request.files['file']
        df = pd.read_excel(file, engine='openpyxl', nrows=3)
        df.columns = [str(c).strip() for c in df.columns]
        all_cols = df.columns.tolist()
        detected = {
            'col_so_item': find_column(df, ['SO Item', 'SO Item No', 'SO Line', 'Item No', 'Line']),
            'col_so_number': find_column(df, ['SO Number','SO No','SO No.','SO','Sales Order Number','No SO','Nomor SO']),
            'col_spec': find_column(df, ['Specification','Spec','Specifications','Product Specification','Material Description','Material Desc','Short Text']),
            'col_pid': find_column(df, ['Product ID','Product Id','Product Code','Material','Material No','Material Number','Material Code','SKU','Article','Article Number']),
            'col_prod': find_column(df, ['Product Name','Item Name','Description','Product']),
            'col_status': find_column(df, ['SO Status','Status','Order Status']),
            'col_vendor': find_column(df, ['Vendor Name','Vendor','Supplier']),
            'col_sodate': find_column(df, ['SO Create Date','Order Date','SO Date','Create Date']),
        }
        col_primary = detected['col_so_item'] or detected['col_so_number']
        missing_critical = []
        if not col_primary: missing_critical.append('col_so_item / col_so_number')
        for k in ('col_spec', 'col_pid'):
            if not detected[k]: missing_critical.append(k)
        return jsonify({
            'total_columns': len(all_cols), 'all_columns': all_cols, 'detected': detected,
            'primary_key_column': col_primary, 'missing_critical': missing_critical,
            'diagnosis': ('col_spec and/or col_pid NOT detected — column names in this file do not match any known alias. Check "all_columns" list and update backend aliases.' if missing_critical else 'SO Item key, col_spec, and col_pid all detected — upload should work correctly.')
        })
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/data/aging', methods=['GET'])
def get_aging_data():
    try:
        cache_key = runtime_cache_key('aging')
        cached = runtime_cache_get(cache_key)
        if cached is not None: return jsonify(cached)
        today = date.today()
        hidden_so = get_hidden_so_items()
        clients = selected_clients()
        pics = selected_pics()
        date_year, date_from, date_to = parse_so_date_args()
        vendors = {}
        q = db.session.query(SOData).filter(open_so_filter())
        q = apply_so_client_filter(q, clients)
        q = apply_so_pic_filter(q, pics)
        q = apply_so_create_date_filter(q, date_year, date_from, date_to)
        aging_fields = (SOData.so_number, SOData.so_item, SOData.vendor_name, SOData.customer_po_number, SOData.delivery_memo, SOData.so_create_date, SOData.sales_amount)
        for s in q.options(load_only(*aging_fields)).all():
            if s.so_item in hidden_so or s.so_number in hidden_so: continue
            if not so_is_countable(s.so_item, customer_po_number=s.customer_po_number, delivery_memo=s.delivery_memo): continue
            v = s.vendor_name or 'Unknown'
            if v not in vendors: vendors[v] = {'vendor': v, 'less_30': 0, 'days_30_90': 0, 'days_90_180': 0, 'more_180': 0, 'total_open': 0, 'sales_amount': 0.0}
            age = workdays_since(s.so_create_date, today) if s.so_create_date else None
            if age is None: vendors[v]['more_180'] += 1
            elif age < 30: vendors[v]['less_30'] += 1
            elif age < 90: vendors[v]['days_30_90'] += 1
            elif age < 180: vendors[v]['days_90_180'] += 1
            else: vendors[v]['more_180'] += 1
            vendors[v]['total_open'] += 1
            vendors[v]['sales_amount'] += float(s.sales_amount or 0)
        payload = sorted(vendors.values(), key=lambda x: x['total_open'], reverse=True)
        runtime_cache_set(cache_key, payload, ttl_seconds=180)
        return jsonify(payload)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/data/aging-detail/<path:vendor_name>', methods=['GET'])
def get_aging_detail(vendor_name):
    try:
        bucket = request.args.get('bucket')
        today = date.today()
        hidden_so = get_hidden_so_items()
        date_year, date_from, date_to = parse_so_date_args()
        clients = selected_clients()
        pics = selected_pics()
        q = db.session.query(SOData).filter(open_so_filter(), SOData.vendor_name == vendor_name)
        q = apply_so_client_filter(q, clients)
        q = apply_so_pic_filter(q, pics)
        q = apply_so_create_date_filter(q, date_year, date_from, date_to)
        sos = q.order_by(SOData.so_create_date.asc()).all()
        sos = [s for s in sos if s.so_item not in hidden_so and s.so_number not in hidden_so and so_is_countable(s.so_item, customer_po_number=s.customer_po_number, delivery_memo=s.delivery_memo)]
        if bucket:
            bucket = bucket.strip().replace(' ', '+')
            sos = [s for s in sos if get_aging_label(workdays_since(s.so_create_date, today) if s.so_create_date else None) == bucket]
        return jsonify([so_dict(s) for s in sos])
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/data/aging-detail-all', methods=['GET'])
def get_aging_detail_all():
    try:
        bucket = request.args.get('bucket')
        if bucket: bucket = bucket.strip().replace(' ', '+')
        today = date.today()
        hidden_so = get_hidden_so_items()
        date_year, date_from, date_to = parse_so_date_args()
        clients = selected_clients()
        pics = selected_pics()
        q = db.session.query(SOData).filter(open_so_filter())
        q = apply_so_client_filter(q, clients)
        q = apply_so_pic_filter(q, pics)
        q = apply_so_create_date_filter(q, date_year, date_from, date_to)
        sos = q.order_by(SOData.vendor_name.asc(), SOData.so_create_date.asc()).all()
        sos = [s for s in sos if s.so_item not in hidden_so and s.so_number not in hidden_so and so_is_countable(s.so_item, customer_po_number=s.customer_po_number, delivery_memo=s.delivery_memo)]
        if bucket:
            sos = [s for s in sos if get_aging_label(workdays_since(s.so_create_date, today) if s.so_create_date else None) == bucket]
        return jsonify([so_dict(s) for s in sos])
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard/pending-total', methods=['GET'])
def get_dashboard_pending_total():
    try:
        cache_key = runtime_cache_key('dashboard_pending_total')
        cached = runtime_cache_get(cache_key)
        if cached is not None: return jsonify(cached)
        date_year, date_from, date_to = parse_so_date_args()
        clients = selected_clients()
        pics = selected_pics()
        q = db.session.query(func.count(SOData.id)).filter(open_so_filter(), so_countable_sql_filter())
        q = apply_so_client_filter(q, clients)
        q = apply_so_pic_filter(q, pics)
        q = apply_so_create_date_filter(q, date_year, date_from, date_to)
        total = q.scalar() or 0
        payload = {'total': total}
        runtime_cache_set(cache_key, payload, ttl_seconds=60)
        return jsonify(payload)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/data/all-so', methods=['GET'])
def get_all_so():
    try:
        cache_key = runtime_cache_key('all_so')
        cached = runtime_cache_get(cache_key)
        if cached is not None: return jsonify(cached)
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 10))
        op_units = request.args.getlist('op_unit')
        vendors = request.args.getlist('vendor')
        manufacturers = request.args.getlist('manufacturer')
        statuses = request.args.getlist('status')
        aging_list = request.args.getlist('aging')
        so_items = request.args.getlist('so_item')
        pics = request.args.getlist('pic')
        kpi_pic = (request.args.get('kpi_pic') or '').strip()
        global_pics = request.args.getlist('global_pic')
        clients = selected_clients()
        margin_filter = request.args.get('margin_filter', 'all')
        sort_order = request.args.get('sort_order', 'oldest')
        date_year, date_from, date_to = parse_so_date_args()

        q = SOData.query.filter(open_so_filter(), so_countable_sql_filter())
        q = apply_so_client_filter(q, clients)
        q = apply_so_pic_filter(q, global_pics)
        if op_units: q = q.filter(SOData.operation_unit_name.in_(op_units))
        if vendors: q = q.filter(SOData.vendor_name.in_(vendors))
        if manufacturers: q = q.filter(SOData.manufacturer_name.in_(manufacturers))
        if statuses: q = q.filter(SOData.so_status.in_(statuses))
        if so_items: q = q.filter(SOData.so_item.in_(so_items))

        q = apply_so_create_date_filter(q, date_year, date_from, date_to)

        # Push PIC filter into SQL when possible — avoids loading all SOs
        # into Python just to filter by pic_name. canonical_pending_pic()
        # is now `pic or 'Unassigned'`, so 'Unassigned' = empty/null pic.
        if pics:
            pic_set = set(pics)
            non_empty_pics = [p for p in pic_set if p and p != 'Unassigned']
            wants_unassigned = 'Unassigned' in pic_set or '(Kosong)' in pic_set
            if non_empty_pics and wants_unassigned:
                q = q.filter(db.or_(SOData.pic_name.in_(non_empty_pics), SOData.pic_name.is_(None), SOData.pic_name == ''))
            elif non_empty_pics:
                q = q.filter(SOData.pic_name.in_(non_empty_pics))
            elif wants_unassigned:
                q = q.filter(db.or_(SOData.pic_name.is_(None), SOData.pic_name == ''))

        if kpi_pic and kpi_pic != 'Unassigned':
            q = q.filter(SOData.pic_name == kpi_pic)
        elif kpi_pic == 'Unassigned':
            q = q.filter(db.or_(SOData.pic_name.is_(None), SOData.pic_name == ''))

        raw_purchase_expr = case(
            (func.coalesce(SOData.purchasing_amount, 0) != 0, func.coalesce(SOData.purchasing_amount, 0)),
            else_=func.coalesce(SOData.purchasing_price, 0) * func.coalesce(SOData.so_qty, 0)
        )
        # Only consider rows with valid (positive) purchase amount — empty/
        # zero/negative purchase → margin is invalid, excluded from filter.
        valid_purchase_filter = (raw_purchase_expr > 0)
        if margin_filter == 'negative':
            q = q.filter(valid_purchase_filter, func.coalesce(SOData.sales_amount, 0) < raw_purchase_expr)
        elif margin_filter == 'positive':
            q = q.filter(valid_purchase_filter, func.coalesce(SOData.sales_amount, 0) >= raw_purchase_expr)

        so_list_fields = (
            SOData.id, SOData.so_number, SOData.so_item, SOData.so_status,
            SOData.operation_unit_name, SOData.vendor_name, SOData.manufacturer_name,
            SOData.customer_po_number, SOData.delivery_memo, SOData.so_create_date,
            SOData.pic_name, SOData.sales_amount, SOData.purchasing_amount,
            SOData.purchasing_price, SOData.so_qty, SOData.purchasing_currency,
            SOData.purchasing_amount_idr, SOData.product_id, SOData.product_name,
            SOData.specification, SOData.matched_po_number, SOData.remarks,
            SOData.delivery_plan_date, SOData.sales_price, SOData.sales_unit,
            SOData.vendor_id, SOData.currency, SOData.client_id,
        )

        if sort_order == 'oldest':
            all_sos = q.options(load_only(*so_list_fields)).order_by(SOData.so_create_date.asc(), SOData.so_item.asc()).all()
        else:
            all_sos = q.options(load_only(*so_list_fields)).order_by(SOData.so_create_date.desc(), SOData.so_item.asc()).all()

        # Prefetch exchange rates for USD/EUR rows so purchase_price_idr /
        # purchase_amount_idr_for_margin can convert correctly.
        prefetch_convertible_exchange_rates(all_sos, fetch_missing=False)

        if aging_list:
            today = date.today()
            def matches_aging(s):
                age = workdays_since(s.so_create_date, today)
                return get_aging_label(age) in aging_list
            all_sos = [s for s in all_sos if matches_aging(s)]

        approval_statuses = {'Approval Apply', 'Approval Reject'}
        approval_q = SOData.query.filter(SOData.so_status.in_(list(approval_statuses)), so_countable_sql_filter())
        approval_q = apply_so_client_filter(approval_q, clients)
        approval_q = apply_so_pic_filter(approval_q, global_pics)
        if op_units: approval_q = approval_q.filter(SOData.operation_unit_name.in_(op_units))
        if vendors: approval_q = approval_q.filter(SOData.vendor_name.in_(vendors))
        if manufacturers: approval_q = approval_q.filter(SOData.manufacturer_name.in_(manufacturers))
        if statuses: approval_q = approval_q.filter(SOData.so_status.in_(statuses))
        if so_items: approval_q = approval_q.filter(SOData.so_item.in_(so_items))
        approval_q = apply_so_create_date_filter(approval_q, date_year, date_from, date_to)
        # Apply the same SQL-side pic filters to approval_q so we don't have to
        # load every approval row into Python just to filter by pic_name.
        if pics:
            pic_set = set(pics)
            non_empty_pics = [p for p in pic_set if p and p != 'Unassigned']
            wants_unassigned = 'Unassigned' in pic_set or '(Kosong)' in pic_set
            if non_empty_pics and wants_unassigned:
                approval_q = approval_q.filter(db.or_(SOData.pic_name.in_(non_empty_pics), SOData.pic_name.is_(None), SOData.pic_name == ''))
            elif non_empty_pics:
                approval_q = approval_q.filter(SOData.pic_name.in_(non_empty_pics))
            elif wants_unassigned:
                approval_q = approval_q.filter(db.or_(SOData.pic_name.is_(None), SOData.pic_name == ''))
        if kpi_pic and kpi_pic != 'Unassigned':
            approval_q = approval_q.filter(SOData.pic_name == kpi_pic)
        elif kpi_pic == 'Unassigned':
            approval_q = approval_q.filter(db.or_(SOData.pic_name.is_(None), SOData.pic_name == ''))
        if sort_order == 'oldest':
            approval_sos = approval_q.options(load_only(*so_list_fields)).order_by(SOData.so_create_date.asc(), SOData.so_item.asc()).all()
        else:
            approval_sos = approval_q.options(load_only(*so_list_fields)).order_by(SOData.so_create_date.desc(), SOData.so_item.asc()).all()

        kpi_source_sos = list(all_sos)

        # PIC filter and kpi_pic filter are now applied in SQL above, so the
        # Python-side filters below are kept as a safety net (cheap no-op now).

        total = len(all_sos)
        subtotal_amount = sum(float(s.sales_amount or 0) for s in all_sos)
        paged = all_sos[(page-1)*per_page : page*per_page]

        option_source_sos = all_sos
        op_units_opts = sorted({s.operation_unit_name for s in option_source_sos if s.operation_unit_name})
        vendors_opts  = sorted({s.vendor_name for s in option_source_sos if s.vendor_name})
        manufacturers_opts = sorted({s.manufacturer_name for s in option_source_sos if s.manufacturer_name})
        statuses_opts = sorted({s.so_status for s in option_source_sos if s.so_status})

        _all_cat_by_id, _all_cat_by_name = master_pic_maps()
        # FIX V3: _all_known_so_pics DIPINDAHKAN ke sini (sebelum baris pakainya).
        # Sebelumnya definisi ada di baris 3199 (di bawah pemakaian di baris 3196),
        # menyebabkan "UnboundLocalError: local variable '_all_known_so_pics'
        # referenced before assignment" saat get_all_so dipanggil.
        _all_known_so_pics = set()
        _all_known_so_pics.update(_all_cat_by_id.values())
        _all_known_so_pics.update(_all_cat_by_name.values())
        _all_known_so_pics.update(clean(m.pic_name) for m in db.session.query(MasterClientPIC).all() if clean(m.pic_name))
        _all_known_so_pics.update(clean(m.pic_name) for m in db.session.query(MasterVendorPIC).all() if clean(m.pic_name))
        _all_known_so_pics.discard('')
        _all_known_so_pics.discard(None)
        _all_known_so_pics.update(canonical_pending_pic(s.pic_name, s.operation_unit_name) for s in kpi_source_sos if canonical_pending_pic(s.pic_name, s.operation_unit_name) != 'Unassigned')
        pics_opts     = sorted(_all_known_so_pics)

        pic_aggregations = {pic: {'pic': pic, 'count': 0, 'amount': 0.0} for pic in _all_known_so_pics}
        for s in kpi_source_sos:
            pic = canonical_pending_pic(s.pic_name, s.operation_unit_name)
            if not pic or pic == 'Unassigned': continue
            if pic not in pic_aggregations: pic_aggregations[pic] = {'pic': pic, 'count': 0, 'amount': 0.0}
            pic_aggregations[pic]['count'] += 1
            pic_aggregations[pic]['amount'] += float(s.sales_amount or 0)
        
        pic_aggs_list = sort_pic_kpis(list(pic_aggregations.values()))

        payload = {
            'data': [so_dict(s) for s in paged],
            'approval_data': [so_dict(s) for s in approval_sos],
            'total': total, 'subtotal_amount': round(subtotal_amount, 2), 'page': page, 'per_page': per_page,
            'filters': {'op_units': list(op_units_opts), 'vendors': list(vendors_opts), 'manufacturers': list(manufacturers_opts), 'statuses': list(statuses_opts), 'pics': list(pics_opts)},
            'pic_aggregations': pic_aggs_list
        }
        runtime_cache_set(cache_key, payload, ttl_seconds=60)
        return jsonify(payload)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/data/so-status-detail/<path:status>', methods=['GET'])
def get_so_status_detail(status):
    try:
        month = request.args.get('month')
        hidden_so = get_hidden_so_items()
        date_year, date_from, date_to = parse_so_date_args()
        clients = selected_clients()
        pics = selected_pics()
        q = SOData.query.filter_by(so_status=status)
        q = apply_so_client_filter(q, clients)
        q = apply_so_pic_filter(q, pics)
        q = apply_so_create_date_filter(q, date_year, date_from, date_to)
        sos = q.all()
        if month:
            sos = [s for s in sos if s.so_create_date and s.so_create_date.strftime('%b %Y') == month]
        sos = [s for s in sos if s.so_item not in hidden_so and s.so_number not in hidden_so and so_is_countable(s.so_item, customer_po_number=s.customer_po_number, delivery_memo=s.delivery_memo)]
        return jsonify([so_dict(s) for s in sos])
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/data/so-status-detail-all', methods=['GET'])
def get_so_status_detail_all():
    try:
        month = request.args.get('month')
        hidden_so = get_hidden_so_items()
        date_year, date_from, date_to = parse_so_date_args()
        clients = selected_clients()
        pics = selected_pics()
        q = SOData.query.filter(open_so_filter())
        q = apply_so_client_filter(q, clients)
        q = apply_so_pic_filter(q, pics)
        q = apply_so_create_date_filter(q, date_year, date_from, date_to)
        sos = q.order_by(SOData.so_create_date.desc()).all()
        if month:
            sos = [s for s in sos if s.so_create_date and s.so_create_date.strftime('%b %Y') == month]
        sos = [s for s in sos if s.so_item not in hidden_so and s.so_number not in hidden_so and so_is_countable(s.so_item, customer_po_number=s.customer_po_number, delivery_memo=s.delivery_memo)]
        return jsonify([so_dict(s) for s in sos])
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/data/top-vendor-detail/<path:vendor_name>', methods=['GET'])
def get_top_vendor_detail(vendor_name):
    try:
        hidden_so = get_hidden_so_items()
        date_year, date_from, date_to = parse_so_date_args()
        clients = selected_clients()
        pics = selected_pics()
        q = db.session.query(SOData).filter(open_so_filter(), SOData.vendor_name == vendor_name)
        q = apply_so_client_filter(q, clients)
        q = apply_so_pic_filter(q, pics)
        q = apply_so_create_date_filter(q, date_year, date_from, date_to)
        sos = q.all()
        sos = [s for s in sos if s.so_item not in hidden_so and s.so_number not in hidden_so and so_is_countable(s.so_item, customer_po_number=s.customer_po_number, delivery_memo=s.delivery_memo)]
        return jsonify([so_dict(s) for s in sos])
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/exchange-rate', methods=['GET'])
def list_exchange_rates():
    try:
        rates = ExchangeRate.query.order_by(ExchangeRate.rate_date.desc()).limit(120).all()
        return jsonify([{'id': r.id, 'date': r.rate_date.isoformat(), 'usd_to_idr': r.usd_to_idr, 'source': r.source} for r in rates])
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/exchange-rate', methods=['POST'])
def upsert_exchange_rate():
    try:
        data = request.json
        d = parse_date(data.get('date'))
        rate = float(data.get('usd_to_idr', 0))
        if not d: return jsonify({'error': 'Invalid date'}), 400
        if rate <= 0: return jsonify({'error': 'Rate must be > 0'}), 400
        rec = ExchangeRate.query.filter_by(rate_date=d).first()
        if rec:
            rec.usd_to_idr = rate; rec.source = 'manual'
        else:
            rec = ExchangeRate(rate_date=d, usd_to_idr=rate, source='manual')
            db.session.add(rec)
        db.session.commit()
        _RATE_CACHE.pop(d, None)
        return jsonify({'success': True, 'date': d.isoformat(), 'usd_to_idr': rate})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/exchange-rate/fetch', methods=['POST'])
def fetch_exchange_rates_bulk():
    try:
        is_sqlite = 'sqlite' in app.config['SQLALCHEMY_DATABASE_URI']
        usd_rows = db.session.query(SOData.so_create_date).filter(SOData.purchasing_currency == 'USD', SOData.so_create_date.isnot(None)).distinct().all()
        dates_needed = {r[0] for r in usd_rows}
        existing_dates = {r[0] for r in db.session.query(ExchangeRate.rate_date).all()}
        to_fetch = sorted(dates_needed - existing_dates)
        fetched = 0; failed = []
        # Use parallel fetching — much faster than sequential per-date HTTP calls.
        rates = _fetch_rates_parallel(to_fetch, 'USD')
        for d in to_fetch:
            rate = rates.get(d)
            if rate:
                try:
                    db.session.add(ExchangeRate(rate_date=d, usd_to_idr=rate, source='frankfurter'))
                    db.session.flush(); _RATE_CACHE[d] = rate; fetched += 1
                except Exception: db.session.rollback()
            else: failed.append(d.isoformat())
        db.session.commit()
        pending_fx_rows = SOData.query.filter(SOData.purchasing_amount_idr.is_(None), func.upper(func.coalesce(SOData.purchasing_currency, '')).in_(['USD', 'EUR'])).all()
        converted_rows = ensure_purchase_amount_idr_cache(pending_fx_rows, fetch_missing=True)
        return jsonify({
            'dates_needed': len(dates_needed), 'already_stored': len(existing_dates & dates_needed), 'fetched': fetched,
            'converted_rows': converted_rows, 'failed': failed,
            'message': f'{fetched} kurs USD berhasil di-fetch dan {converted_rows} transaksi USD/EUR dikonversi.' + (f' {len(failed)} tanggal gagal: {", ".join(failed[:5])}' if failed else '')
        })
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/exchange-rate/preview', methods=['GET'])
def preview_exchange_rate():
    try:
        d = parse_date(request.args.get('date', ''))
        if not d: return jsonify({'error': 'Provide ?date=YYYY-MM-DD'}), 400
        rate = get_usd_to_idr(d)
        rec = ExchangeRate.query.filter_by(rate_date=d).first()
        return jsonify({'date': d.isoformat(), 'usd_to_idr': rate, 'source': rec.source if rec else 'fallback/nearest', 'stored_exact': rec is not None})
    except Exception as e: return jsonify({'error': str(e)}), 500

CHUNK_SIZE = 200

def _norm_key(v): return str(v or '').strip()

def _latest_row(rows, timestamp_fields=('uploaded_at', 'updated_at')):
    def score(row):
        ts = None
        for field in timestamp_fields:
            val = getattr(row, field, None)
            if val is not None: ts = val; break
        return (ts or datetime.min, getattr(row, 'id', 0) or 0)
    return max(rows, key=score)

def _latest_nonblank_value(rows, field, timestamp_fields=('uploaded_at', 'updated_at')):
    candidates = [r for r in rows if str(getattr(r, field, '') or '').strip()]
    if not candidates: return None
    return getattr(_latest_row(candidates, timestamp_fields), field)

def cleanup_source_table_snapshot(model, key_attr, valid_keys=None, *, manual_fields=(), timestamp_fields=('uploaded_at', 'updated_at'), delete_blank=True, key_normalizer=_norm_key):
    valid_set = None if valid_keys is None else {key_normalizer(k) for k in valid_keys if key_normalizer(k)}
    groups = {}; blank_rows = []
    for row in db.session.query(model).order_by(model.id.asc()).all():
        key = key_normalizer(getattr(row, key_attr, None))
        if not key: blank_rows.append(row); continue
        groups.setdefault(key, []).append(row)
    removed_duplicates = removed_stale = removed_blank = 0
    if delete_blank:
        for row in blank_rows: db.session.delete(row); removed_blank += 1
    for key, rows in groups.items():
        if valid_set is not None and key not in valid_set:
            for row in rows: db.session.delete(row); removed_stale += 1
            continue
        if len(rows) <= 1: continue
        winner = _latest_row(rows, timestamp_fields)
        for field in manual_fields or ():
            val = _latest_nonblank_value(rows, field, timestamp_fields)
            if val is not None: setattr(winner, field, val)
        for row in rows:
            if row is winner: continue
            db.session.delete(row); removed_duplicates += 1
    return {'removed_duplicates': removed_duplicates, 'removed_stale': removed_stale, 'removed_blank': removed_blank}

def cleanup_master_pic_by_category_name(valid_category_names=None):
    valid_set = None if valid_category_names is None else {normalize_category_name(x) for x in valid_category_names if normalize_category_name(x)}
    groups = {}; blank_rows = []
    for row in db.session.query(MasterPIC).order_by(MasterPIC.id.asc()).all():
        key = normalize_category_name(row.category_name)
        if not key: blank_rows.append(row); continue
        groups.setdefault(key, []).append(row)
    removed_duplicates = removed_stale = removed_blank = 0
    for row in blank_rows: db.session.delete(row); removed_blank += 1
    for key, rows in groups.items():
        if valid_set is not None and key not in valid_set:
            for row in rows: db.session.delete(row); removed_stale += 1
            continue
        if len(rows) <= 1: continue
        winner = _latest_row(rows, ('updated_at', 'uploaded_at'))
        pic = _latest_nonblank_value(rows, 'pic_name', ('updated_at', 'uploaded_at'))
        if pic is not None: winner.pic_name = pic
        winner.category_name = source_category_level1(winner.category_name)
        if str(winner.category_id or '').startswith('CATNAME_'): winner.category_id = master_pic_category_key(winner.category_name)
        for row in rows:
            if row is winner: continue
            db.session.delete(row); removed_duplicates += 1
    return {'removed_duplicates': removed_duplicates, 'removed_stale': removed_stale, 'removed_blank': removed_blank}

def cleanup_item_registration_duplicates_only(): return cleanup_source_table_snapshot(ItemRegistration, 'req_no', None, timestamp_fields=('uploaded_at',), delete_blank=True)


def _build_pic_lookup_cache():
    """Build a {product_id: pic_name} map for all ProductIDDB rows.
    Used to batch PIC lookups during upload — avoids N+1 SQL queries."""
    try:
        prod_rows = db.session.query(ProductIDDB.product_id, ProductIDDB.category_id, ProductIDDB.category_name).all()
    except Exception:
        return {}
    by_id, by_name = master_pic_maps()
    cache = {}
    for pid, cat_id, cat_name in prod_rows:
        if not pid: continue
        pid_str = str(pid).strip()
        pic = _lookup_pic_by_category(cat_id, cat_name)
        if pic:
            cache[pid_str] = pic
    return cache


def _batch_lookup_pic(product_id_str, cache):
    """Look up PIC for a product_id using the preloaded cache.
    Falls back to _lookup_pic (single SQL query) on cache miss."""
    if not product_id_str:
        return None
    pid = str(product_id_str).strip()
    if pid in cache:
        return cache[pid]
    # Cache miss — fall back to single lookup and store back to cache
    pic = _lookup_pic(pid)
    if pic:
        cache[pid] = pic
    return pic


def _spawn_post_upload_fx_worker(so_items_scope):
    """Spawn a background thread to fetch missing USD/EUR exchange rates for
    the rows just uploaded. Returns immediately; the worker runs in the
    background and commits rate cache updates as they arrive.

    This used to be a blocking call inside upload_smro: a single upload with
    30+ unique USD/EUR dates could lock the upload endpoint for several
    minutes while frankfurter.dev was called sequentially with a 6s timeout
    per call. Now we return success immediately and let the worker catch
    up; the dashboard uses cached rates (or 0 if none yet) meanwhile."""
    if not so_items_scope:
        return
    items_list = [s for s in so_items_scope if s]
    if not items_list:
        return

    def _worker():
        try:
            with app.app_context():
                pending_fx_rows = SOData.query.filter(
                    SOData.so_item.in_(items_list),
                    SOData.purchasing_amount_idr.is_(None),
                    func.upper(func.coalesce(SOData.purchasing_currency, '')).in_(['USD', 'EUR']),
                ).all()
                if not pending_fx_rows:
                    return
                converted = ensure_purchase_amount_idr_cache(pending_fx_rows, fetch_missing=True)
                if converted:
                    clear_runtime_caches()
                    print(f'[fx-worker] Converted {converted} FX rows in background.')
        except Exception as exc:
            try:
                db.session.rollback()
            except Exception:
                pass
            print(f'[fx-worker] Background FX fetch failed: {exc}')

    threading.Thread(target=_worker, daemon=True, name='post-upload-fx-worker').start()


def _refresh_so_pic_names(so_items_scope=None, product_ids_scope=None):
    """Helper function (BUKAN route handler!) — refresh pic_name untuk SO rows.
    Me-return jumlah baris yang di-refresh (int).

    Optimization: by default only refresh rows whose so_item is in
    `so_items_scope` (e.g., the items just uploaded) OR whose product_id is
    in `product_ids_scope`. This avoids iterating ALL SO rows on every upload.
    If both are None, falls back to refreshing all rows (used by master-pic uploads).

    FIX V3: sebelumnya function ini disertai decorator @app.route('/api/upload/smro', ...)
    yang membuat Flask mendaftarkannya sebagai view function untuk endpoint upload.
    Akibatnya, saat client POST /api/upload/smro, Flask memanggil function ini dan
    mendapat return value int → throw TypeError → HTTP 500.

    Decorator @app.route sekarang DIPINDAHKAN ke upload_smro() yang benar.
    """
    prod_map = {str(p.product_id).strip(): (p.category_id, p.category_name) for p in db.session.query(ProductIDDB).all() if p.product_id}
    client_pic_cache = {normalize_client_id(m.client_id): clean(m.pic_name) for m in db.session.query(MasterClientPIC).all() if clean(m.pic_name)}
    vendor_pic_cache = {normalize_vendor_id(m.vendor_id): clean(m.pic_name) for m in db.session.query(MasterVendorPIC).all() if clean(m.pic_name)}
    cat_pic_cache = {}
    # Scope the refresh to rows that actually need it — much faster than
    # iterating every SO row in the table on every upload.
    q = db.session.query(SOData).filter(SOData.product_id.isnot(None), SOData.product_id != '')
    if so_items_scope is not None:
        items_list = [s for s in so_items_scope if s]
        if not items_list:
            return 0
        q = q.filter(SOData.so_item.in_(items_list))
    elif product_ids_scope is not None:
        pids_list = [str(p).strip() for p in product_ids_scope if p]
        if not pids_list:
            return 0
        q = q.filter(SOData.product_id.in_(pids_list))
    so_rows = q.all()
    refreshed = 0
    for s in so_rows:
        cat_id, cat_name = prod_map.get(str(s.product_id).strip(), (None, None))
        ck = (normalize_category_id(cat_id), normalize_category_name(cat_name))
        if ck not in cat_pic_cache: cat_pic_cache[ck] = _lookup_pic_by_category(cat_id, cat_name)
        new_pic = None
        cid = normalize_client_id(s.client_id) if s.client_id else ''
        if cid and cid in client_pic_cache: new_pic = client_pic_cache[cid]
        if not new_pic:
            vid = normalize_vendor_id(s.vendor_id) if s.vendor_id else ''
            if vid and vid in vendor_pic_cache: new_pic = vendor_pic_cache[vid]
        if not new_pic: new_pic = cat_pic_cache[ck]
        if s.pic_name != new_pic: s.pic_name = new_pic; refreshed += 1
    if refreshed:
        db.session.commit()
    clear_runtime_caches()
    return refreshed


# FIX V3: decorator @app.route untuk endpoint /api/upload/smro dan /api/upload/scor
# DIPINDAHKAN dari _refresh_so_pic_names() (yang return int) ke upload_smro() (yang return jsonify).
# Sebelumnya Flask memanggil _refresh_so_pic_names() saat POST /api/upload/smro,
# mendapat int, lalu throw "TypeError: ... but it was a int" → HTTP 500.
@app.route('/api/upload/scor-json', methods=['POST'])
@app.route('/api/upload/scor', methods=['POST'])
@app.route('/api/upload/smro-json', methods=['POST'])
@app.route('/api/upload/smro', methods=['POST'])
def upload_smro():
    try:
        uploads, upload_mode = request_upload_dataframes('smro')
        if not uploads: return jsonify({'error': 'No file uploaded or JSON rows supplied'}), 400
        replace_existing = upload_replace_mode()
        required_smro_cols = {
            'SO Item': ['SO Item', 'SO Item No', 'SO Number', 'SO No', 'SO No.', 'SO', 'Sales Order', 'Sales Order Number', 'No SO', 'Nomor SO'],
            'SO Status': ['SO Status', 'Status', 'Order Status'],
            'Operation Unit': ['Operation Unit Name', 'Op Unit', 'Client Name', 'Client', 'Operation Unit'],
            'Vendor Name': ['Vendor Name', 'Vendor', 'Supplier'],
            'Customer PO': ['Customer PO number', 'Customer PO Number', 'Customer PO', 'PO Ref', 'PO Reference'],
            'Sales Amount': ['Sales Amount(Exclude Tax)', 'Sales Amount', 'Amount', 'Total'],
            'SO Create Date': ['SO Create Date', 'Order Date', 'SO Date', 'Create Date', 'Create Sales Order Date'],
        }
        cleanup_pre = cleanup_source_table_snapshot(SOData, 'so_item', None, manual_fields=('delivery_plan_date', 'remarks'), timestamp_fields=('uploaded_at',), delete_blank=True)
        db.session.flush()
        
        all_so_items_in_upload = set()
        for upload in uploads:
            df = upload['df']
            df.columns = [str(c).strip() for c in df.columns]
            col_soitem = find_column(df, ['SO Item', 'SO Item No', 'SO Line', 'Item No', 'Line'])
            col_so = find_column(df, ['SO Number', 'SO No', 'SO No.', 'SO', 'Sales Order', 'Sales Order Number', 'No SO', 'Nomor SO'])
            col_primary = col_soitem or col_so
            if col_primary:
                for _, row in df.iterrows():
                    val = clean(row.get(col_primary))
                    if val: all_so_items_in_upload.add(val)
        
        existing_so = {}
        if all_so_items_in_upload:
            existing_so_rows = SOData.query.filter(SOData.so_item.in_(list(all_so_items_in_upload))).all()
            existing_so = {s.so_item: s for s in existing_so_rows if s.so_item}

        # Preload product_id → pic_name map ONCE for the whole upload batch.
        # Previously the loop called _lookup_pic(product_id) per row, which
        # issued a separate SQL query for each row (N+1 query problem —
        # 1,000 rows = 1,000 SQL queries just for PIC lookup).
        _pic_lookup_cache = _build_pic_lookup_cache()

        total_count = total_updated = total_inserted = 0
        total_removed_duplicates = cleanup_pre.get('removed_duplicates', 0)
        total_removed_stale = cleanup_pre.get('removed_stale', 0)
        total_removed_blank = cleanup_pre.get('removed_blank', 0)
        latest_so_items = set()
        diagnostics_by_file = []
        for upload in uploads:
            filename = upload['filename']; df = upload['df']
            df.columns = [str(c).strip() for c in df.columns]
            has_primary_key = find_column(df, ['SO Item', 'SO Item No', 'SO Number', 'SO No', 'SO No.', 'SO', 'Sales Order', 'Sales Order Number', 'No SO', 'Nomor SO'])
            if not has_primary_key:
                return jsonify({'error': f'Invalid file "{filename}" - kolom SO Item / SO Number tidak ditemukan. Available columns: {df.columns.tolist()}'}), 400
            missing_required = [name for name, aliases in required_smro_cols.items() if not find_column(df, aliases)]
            if upload_mode == 'excel' and len(missing_required) > 4:
                return jsonify({'error': f'Invalid file "{filename}" - {len(missing_required)} required columns not found: {", ".join(missing_required)}. Please make sure you are uploading the correct SMRO file.'}), 400
            col_soitem = find_column(df, ['SO Item', 'SO Item No', 'SO Line', 'Item No', 'Line'])
            col_so = find_column(df, ['SO Number', 'SO No', 'SO No.', 'SO', 'Sales Order', 'Sales Order Number', 'No SO', 'Nomor SO'])
            col_primary = col_soitem or col_so
            if not col_primary: return jsonify({'error': f'SO Item / SO Number column not found in "{filename}". Available columns: {df.columns.tolist()}'}), 400
            col_status   = find_column(df, ['SO Status', 'Status', 'Order Status', 'SO Status Code'])
            col_opunit   = find_column(df, ['Operation Unit Name', 'Op Unit', 'Client Name', 'Client', 'Operation Unit'])
            col_client_id = find_column(df, ['Client ID', 'Client Id', 'ClientID', 'Client Cd.', 'Client Cd', 'Client Code'])
            col_vendor_id = find_column(df, ['Vendor ID', 'Vendor Id', 'Vendor Code', 'Supplier ID', 'Supplier Code'])
            col_vendor   = find_column(df, ['Vendor Name', 'Vendor', 'Supplier'])
            col_custpo   = find_column(df, ['Customer PO number', 'Customer PO Number', 'Customer PO', 'PO Ref', 'PO Reference'])
            col_memo     = find_column(df, ['Delivery Memo', 'Memo', 'Delivery Note'])
            col_prod     = find_column(df, ['Product Name', 'Item Name', 'Description', 'Product'])
            col_spec     = find_column(df, ['Specification', 'Spec', 'Specifications', 'Product Specification', 'Material Description', 'Material Desc', 'Short Text'])
            col_mfr      = find_column(df, ['Manufacturer Name', 'Mfr. Nm.', 'Mfr. Nm', 'Maker Nm.', 'Manufacturer'])
            col_pid      = find_column(df, ['Product ID', 'Product Id', 'Product Code', 'Material', 'Material No', 'Material Number', 'Material Code', 'SKU', 'Article', 'Article Number'])
            col_qty      = find_column(df, ['SO Quantity', 'SO Qty', 'Qty', 'Quantity'])
            col_sunit    = find_column(df, ['Sales Unit', 'Unit', 'UOM'])
            col_sprice   = find_column(df, ['Sales Price(Exclude Tax)', 'Sales Price', 'Price', 'Unit Price'])
            col_samt     = find_column(df, ['Sales Amount(Exclude Tax)', 'Sales Amount', 'Amount', 'Total'])
            col_cur      = find_column(df, ['Currency', 'Curr'])
            col_pprice   = find_column(df, ['Purchasing Price', 'Purchase Price', 'PO Price'])
            col_pamt     = find_column(df, ['Purchasing Amount', 'Purchase Amount', 'PO Amount'])
            col_pcur     = find_column(df, ['Purchasing Currency', 'Purchase Currency', 'PO Currency', 'Purchasing Curr', 'Purchase Curr'])
            col_sodate   = find_column(df, ['SO Create Date', 'Order Date', 'SO Date', 'Create Date', 'Create Sales Order Date'])
            col_delposs  = find_column(df, ['Delivery Possible Date', 'Possible Delivery Date', 'Est Delivery'])
            col_matchpo  = find_column(df, ['Purchasing Order Number', 'Matched PO Number', 'Matched PO', 'PO HLI', 'PO HLI Number', 'PO Number'])
            count = updated = inserted = spec_filled = pid_filled = 0
            for _, row in df.iterrows():
                primary_val = clean(df_val(row, col_primary))
                if not primary_val: continue
                row_status = clean(df_val(row, col_status)) if col_status else None
                if row_status and row_status in DISCARDABLE_STATUSES:
                    if primary_val in existing_so: db.session.delete(existing_so.pop(primary_val))
                    continue
                if col_soitem:
                    so_item_val = primary_val
                    so_val = clean(df_val(row, col_so)) if col_so else None
                    if not so_val: so_val = so_item_val.rsplit('-', 1)[0] if '-' in so_item_val else so_item_val
                else:
                    so_val = primary_val; so_item_val = so_val
                if so_item_val: latest_so_items.add(so_item_val)
                spec_val = clean(df_val(row, col_spec)) if col_spec else None
                pid_val = clean(df_val(row, col_pid)) if col_pid else None
                if spec_val: spec_filled += 1
                if pid_val: pid_filled += 1
                new_data = {
                    'so_number': so_val, 'so_item': so_item_val, 'so_status': clean(df_val(row, col_status)),
                    'operation_unit_name': clean(df_val(row, col_opunit)),
                    'client_id': normalize_client_id(clean(df_val(row, col_client_id))) if col_client_id else None,
                    'vendor_id': normalize_vendor_id(clean(df_val(row, col_vendor_id)) or ''),
                    'vendor_name': clean(df_val(row, col_vendor)), 'customer_po_number': clean(df_val(row, col_custpo)),
                    'delivery_memo': clean(df_val(row, col_memo)), 'product_name': clean(df_val(row, col_prod)),
                    'specification': spec_val, 'manufacturer_name': clean(df_val(row, col_mfr)), 'product_id': pid_val,
                    'so_qty': safe_float(df_val(row, col_qty)), 'sales_unit': clean(df_val(row, col_sunit)),
                    'sales_price': safe_float(df_val(row, col_sprice)), 'sales_amount': safe_float(df_val(row, col_samt)),
                    'currency': clean(df_val(row, col_cur)) or 'IDR', 'purchasing_price': safe_float(df_val(row, col_pprice)),
                    'purchasing_amount': safe_float(df_val(row, col_pamt)), 'purchasing_currency': clean(df_val(row, col_pcur)) if col_pcur else None,
                    'purchasing_amount_idr': None, 'purchasing_amount_idr_cached_at': None,
                    'so_create_date': parse_date(df_val(row, col_sodate)), 'delivery_possible_date': parse_date(df_val(row, col_delposs)),
                    'matched_po_number': clean(df_val(row, col_matchpo)), 'uploaded_at': datetime.utcnow(),
                }
                if so_item_val and so_item_val in existing_so:
                    existing = existing_so[so_item_val]
                    preserved_remarks = existing.remarks; preserved_plan_date = existing.delivery_plan_date
                    preserved_spec = existing.specification; preserved_pid = existing.product_id
                    preserved_amount_idr = existing.purchasing_amount_idr; preserved_amount_idr_cached_at = existing.purchasing_amount_idr_cached_at
                    old_purchase_signature = (float(existing.purchasing_amount or 0), float(existing.purchasing_price or 0), float(existing.so_qty or 0), (existing.purchasing_currency or 'IDR').strip().upper(), existing.so_create_date)
                    new_purchase_signature = (float(new_data.get('purchasing_amount') or 0), float(new_data.get('purchasing_price') or 0), float(new_data.get('so_qty') or 0), (new_data.get('purchasing_currency') or 'IDR').strip().upper(), new_data.get('so_create_date'))
                    purchase_inputs_changed = old_purchase_signature != new_purchase_signature
                    for field, val in new_data.items(): setattr(existing, field, val)
                    existing.remarks = preserved_remarks; existing.delivery_plan_date = preserved_plan_date
                    if not purchase_inputs_changed:
                        existing.purchasing_amount_idr = preserved_amount_idr; existing.purchasing_amount_idr_cached_at = preserved_amount_idr_cached_at
                    if not col_spec or spec_val is None: existing.specification = preserved_spec
                    if not col_pid or pid_val is None: existing.product_id = preserved_pid
                    if existing.product_id: existing.pic_name = _batch_lookup_pic(existing.product_id, _pic_lookup_cache)
                    updated += 1
                else:
                    new_rec = SOData(**new_data)
                    if new_rec.product_id: new_rec.pic_name = _batch_lookup_pic(new_rec.product_id, _pic_lookup_cache)
                    db.session.add(new_rec)
                    if so_item_val: existing_so[so_item_val] = new_rec
                    inserted += 1
                count += 1
                if count % CHUNK_SIZE == 0: db.session.flush()
            db.session.add(UploadLog(file_type='SO', filename=filename, records_count=count))
            total_count += count; total_updated += updated; total_inserted += inserted
            diagnostics = {'filename': filename, 'columns_detected': {'so_item': col_primary, 'so_item_col': col_soitem, 'so_number_col': col_so, 'specification': col_spec, 'product_id': col_pid}, 'rows_with_specification': spec_filled, 'rows_with_product_id': pid_filled, 'all_file_columns': df.columns.tolist()}
            warnings = []
            if not col_spec and not col_pid: warnings.append("File ini tidak mengandung kolom 'Specification' maupun 'Product ID'. Spec/Product ID di DB tidak diubah.")
            else:
                if not col_spec: warnings.append("Kolom 'Specification' tidak ditemukan di file ini - Specification di DB dipertahankan.")
                elif spec_filled == 0: warnings.append(f"Kolom '{col_spec}' terdeteksi tapi semua baris kosong.")
                if not col_pid: warnings.append("Kolom 'Product ID' tidak ditemukan di file ini - Product ID di DB dipertahankan.")
                elif pid_filled == 0: warnings.append(f"Kolom '{col_pid}' terdeteksi tapi semua baris kosong.")
            if warnings: diagnostics['warning'] = ' '.join(warnings)
            diagnostics_by_file.append(diagnostics)
        db.session.flush()
        cleanup_post = cleanup_source_table_snapshot(SOData, 'so_item', latest_so_items if replace_existing else None, manual_fields=('delivery_plan_date', 'remarks'), timestamp_fields=('uploaded_at',), delete_blank=True)
        total_removed_duplicates += cleanup_post.get('removed_duplicates', 0)
        total_removed_stale += cleanup_post.get('removed_stale', 0)
        total_removed_blank += cleanup_post.get('removed_blank', 0)
        db.session.commit()
        try:
            db.session.execute(text('PRAGMA wal_checkpoint(TRUNCATE)'))
            db.session.commit()
        except: pass
        # Clear caches BEFORE returning so the dashboard refetches fresh data.
        clear_runtime_caches()

        # Scope PIC refresh to just the uploaded so_items — much faster than
        # iterating every SO row in the table.
        try:
            _refresh_so_pic_names(so_items_scope=latest_so_items)
        except Exception as pic_exc:
            import traceback; traceback.print_exc()

        # Defer FX rate fetching to a background thread. Previously this
        # blocked the upload response — every unique USD/EUR date that
        # lacked a cached rate triggered a 6-second HTTP call to
        # frankfurter.dev, and 30+ unique dates could lock the upload
        # endpoint for minutes. Now we return immediately and the rates
        # get cached asynchronously; the dashboard will use whatever
        # rates are already cached (or 0 if none) until the background
        # worker catches up.
        try:
            _spawn_post_upload_fx_worker(latest_so_items)
        except Exception as fx_spawn_exc:
            print(f'[upload_smro] FX worker spawn failed: {fx_spawn_exc}')

        diagnostics = diagnostics_by_file[-1] if diagnostics_by_file else {}
        if len(diagnostics_by_file) > 1: diagnostics = {**diagnostics, 'files': diagnostics_by_file}
        return jsonify({
            'message': f'Berhasil upload {len(uploads)} file: {total_inserted} SO baru ditambahkan, {total_updated} SO diperbarui, {total_removed_duplicates} duplicate lama dihapus, {total_removed_stale} SO lama dibuang.',
            'uploaded': total_count, 'files': len(uploads), 'mode': upload_mode, 'replace': replace_existing,
            'inserted': total_inserted, 'updated': total_updated, 'removed_duplicates': total_removed_duplicates,
            'removed_stale': total_removed_stale, 'removed_blank': total_removed_blank,
            'fx_converted': 0, 'fx_warning': None, 'fx_pending': True,
            'diagnostics': diagnostics,
        })
    except Exception as e:
        db.session.rollback(); import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/cleanup-discardable', methods=['POST'])
def admin_cleanup_discardable():
    try:
        deleted = db.session.query(SOData).filter(SOData.so_status.in_(list(DISCARDABLE_STATUSES))).delete(synchronize_session=False)
        db.session.commit()
        db.session.execute(text('PRAGMA wal_checkpoint(TRUNCATE)'))
        db.session.commit()
        clear_runtime_caches()
        return jsonify({'deleted': deleted, 'message': f'{deleted} SO rows dengan status discardable berhasil dihapus.'})
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/cleanup-item-registration', methods=['POST'])
def admin_cleanup_item_registration():
    """One-time cleanup untuk hapus duplikat di tabel ItemRegistration.

    FIX V8: Karena bug sebelumnya (replace_existing default False), data lama
    menumpuk dan record menjadi 2x lipat (8000 vs 4000 asli). Endpoint ini
    hapus duplikat req_no (keep yang terbaru berdasarkan uploaded_at), dan
    hapus juga req_no yang proc_status-nya sudah "selesai" (tidak ada di sistem
    source lagi).

    Hapus req_no dengan proc_status berikut (data sudah selesai di sistem source):
    - 'Registry Completed'
    - 'Vendor Approved'
    - 'Closed'
    - 'Cancelled'
    - 'Rejected'
    """
    try:
        # 1. Hapus duplikat req_no (keep row dengan uploaded_at terbaru)
        all_rows = db.session.query(ItemRegistration).order_by(ItemRegistration.uploaded_at.desc()).all()
        seen_req_no = set()
        duplicates_deleted = 0
        for row in all_rows:
            if row.req_no in seen_req_no:
                db.session.delete(row)
                duplicates_deleted += 1
            else:
                seen_req_no.add(row.req_no)
        db.session.flush()

        # 2. Hapus req_no dengan proc_status "selesai" (data sudah hilang dari sistem source)
        # User bilang: "di sistem jika sudah selesai proses data akan hilang"
        finished_statuses = ['Registry Completed', 'Vendor Approved', 'Closed', 'Cancelled', 'Rejected']
        stale_deleted = db.session.query(ItemRegistration).filter(
            ItemRegistration.proc_status.in_(finished_statuses)
        ).delete(synchronize_session=False)
        db.session.flush()

        db.session.commit()
        db.session.execute(text('PRAGMA wal_checkpoint(TRUNCATE)'))
        db.session.commit()
        clear_runtime_caches()

        total_remaining = db.session.query(ItemRegistration).count()
        return jsonify({
            'duplicates_deleted': duplicates_deleted,
            'stale_deleted': stale_deleted,
            'total_deleted': duplicates_deleted + stale_deleted,
            'total_remaining': total_remaining,
            'message': f'Cleanup selesai: {duplicates_deleted} duplikat + {stale_deleted} data stale dihapus. Sisa: {total_remaining} records.'
        })
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/upload/smro-backfill-spec-json', methods=['POST'])
@app.route('/api/upload/smro-backfill-spec', methods=['POST'])
def upload_smro_backfill_spec():
    try:
        uploads, upload_mode = request_upload_dataframes('smro_backfill_spec')
        if not uploads: return jsonify({'error': 'No file uploaded or JSON rows supplied'}), 400
        _ensure_extra_columns()
        all_so = SOData.query.all()
        by_soitem = {}; by_sonum = {}
        for s in all_so:
            if s.so_item: by_soitem[s.so_item] = s
            if s.so_number: by_sonum.setdefault(s.so_number, []).append(s)
        updated = 0; skipped_no_match = 0; skipped_no_data = 0; flush_counter = 0; diagnostics = []
        for upload in uploads:
            filename = upload['filename']; df = upload['df']
            df.columns = [str(c).strip() for c in df.columns]
            col_sonum  = find_column(df, ['SO Number', 'SO No', 'SO No.', 'SO'])
            col_soitem = find_column(df, ['SO Item', 'SO Item No', 'SO Line', 'Item No', 'Line'])
            col_spec   = find_column(df, ['Specification', 'Spec', 'Specifications', 'Product Specification'])
            col_pid    = find_column(df, ['Product ID', 'Product Id', 'Product Code', 'Material', 'Material No', 'Material Number', 'Material Code', 'SKU'])
            if not col_soitem and not col_sonum: return jsonify({'error': f'SO Item / SO Number column not found in "{filename}". Columns: {df.columns.tolist()}'}), 400
            if not col_spec and not col_pid: return jsonify({'error': f'Neither Specification nor Product ID column found in "{filename}".'}), 400
            file_updated = 0; file_skipped_no_match = 0; file_skipped_no_data = 0
            for _, row in df.iterrows():
                so_item_val = clean(df_val(row, col_soitem)) if col_soitem else None
                so_num_val  = clean(df_val(row, col_sonum))  if col_sonum  else None
                spec_val    = clean(df_val(row, col_spec))   if col_spec   else None
                pid_val     = clean(df_val(row, col_pid))    if col_pid    else None
                if spec_val is None and pid_val is None: skipped_no_data += 1; file_skipped_no_data += 1; continue
                matched_recs = []
                if so_item_val:
                    rec = by_soitem.get(so_item_val)
                    if rec: matched_recs = [rec]
                    else:
                        parts = so_item_val.rsplit('-', 1)
                        so_num_from_item = parts[0] if len(parts) == 2 else so_item_val
                        candidates = by_sonum.get(so_num_from_item, [])
                        if len(parts) == 2:
                            item_line = parts[1]
                            line_matched = [c for c in candidates if c.so_item and c.so_item.endswith(f'-{item_line}')]
                            matched_recs = line_matched or candidates
                        else: matched_recs = candidates
                if not matched_recs and so_num_val: matched_recs = by_sonum.get(so_num_val, [])
                if not matched_recs: skipped_no_match += 1; file_skipped_no_match += 1; continue
                for rec in matched_recs:
                    changed = False
                    if spec_val is not None and rec.specification != spec_val: rec.specification = spec_val; changed = True
                    if pid_val is not None and rec.product_id != pid_val: rec.product_id = pid_val; changed = True
                    if changed:
                        updated += 1; file_updated += 1; flush_counter += 1
                        if flush_counter % 300 == 0: db.session.flush()
            diagnostics.append({'filename': filename, 'updated': file_updated, 'skipped_no_match': file_skipped_no_match, 'skipped_no_data': file_skipped_no_data, 'spec_column_detected': col_spec, 'pid_column_detected': col_pid, 'soitem_column_detected': col_soitem, 'sonumber_column_detected': col_sonum})
        db.session.commit()
        clear_runtime_caches()
        return jsonify({
            'message': f'Backfill selesai: {updated} SO record diperbarui' + (f', {skipped_no_match} baris tidak cocok di DB' if skipped_no_match else '') + (f', {skipped_no_data} baris tidak ada data Spec/PID' if skipped_no_data else '') + '.',
            'mode': upload_mode, 'files': len(uploads), 'updated': updated, 'skipped_no_match': skipped_no_match, 'skipped_no_data': skipped_no_data,
            'diagnostics': diagnostics[-1] if len(diagnostics) == 1 else diagnostics,
        })
    except ValueError as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/data/so/<int:so_id>', methods=['PUT'])
def update_so(so_id):
    try:
        data = request.json
        so = db.session.get(SOData, so_id)
        if not so: return jsonify({'error': 'Not found'}), 404
        if 'delivery_plan_date' in data: so.delivery_plan_date = parse_date(data['delivery_plan_date'])
        if 'remarks' in data: so.remarks = data['remarks']
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback(); return jsonify({'error': str(e)}), 500

@app.route('/api/data/so/by-item/<path:so_item>', methods=['PUT'])
def update_so_by_item(so_item):
    try:
        data = request.json or {}
        so = SOData.query.filter_by(so_item=so_item).first()
        if not so: return jsonify({'error': 'Not found'}), 404
        if 'delivery_plan_date' in data: so.delivery_plan_date = parse_date(data['delivery_plan_date'])
        if 'remarks' in data: so.remarks = data['remarks']
        db.session.commit()
        return jsonify({'success': True, 'id': so.id, 'so_item': so.so_item})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/data/so/template', methods=['GET'])
def download_so_batch_template():
    try:
        op_units   = request.args.getlist('op_unit')
        vendors    = request.args.getlist('vendor')
        manufacturers = request.args.getlist('manufacturer')
        statuses   = request.args.getlist('status')
        aging_list = request.args.getlist('aging')
        so_items   = request.args.getlist('so_item')
        pics       = request.args.getlist('pic')
        kpi_pic = (request.args.get('kpi_pic') or '').strip()
        global_pics = request.args.getlist('global_pic')
        clients = selected_clients()
        margin_filter = request.args.get('margin_filter', 'all')
        date_year, date_from, date_to = parse_so_date_args()
        q = SOData.query.filter(open_so_filter())
        q = apply_so_client_filter(q, clients)
        q = apply_so_pic_filter(q, global_pics)
        if op_units: q = q.filter(SOData.operation_unit_name.in_(op_units))
        if vendors: q = q.filter(SOData.vendor_name.in_(vendors))
        if manufacturers: q = q.filter(SOData.manufacturer_name.in_(manufacturers))
        if statuses: q = q.filter(SOData.so_status.in_(statuses))
        if so_items: q = q.filter(SOData.so_item.in_(so_items))
        q = apply_so_pic_filter(q, pics)
        q = apply_so_create_date_filter(q, date_year, date_from, date_to)
        all_sos = q.order_by(SOData.so_create_date.asc()).all()
        if aging_list:
            today = date.today()
            def matches_aging(s): return get_aging_label(workdays_since(s.so_create_date, today)) in aging_list
            all_sos = [s for s in all_sos if matches_aging(s)]
        if margin_filter in ('positive', 'negative'):
            prefetch_convertible_exchange_rates(all_sos)
            def calc_margin(s):
                po_amt = purchase_amount_idr_for_margin(s)
                if po_amt <= 0: return None  # invalid purchase → no margin
                return float(s.sales_amount or 0) - po_amt
            if margin_filter == 'negative': all_sos = [s for s in all_sos if (lambda m: m is not None and m < 0)(calc_margin(s))]
            else: all_sos = [s for s in all_sos if (lambda m: m is not None and m >= 0)(calc_margin(s))]
        if kpi_pic: all_sos = [s for s in all_sos if canonical_pending_pic(s.pic_name, s.operation_unit_name) == kpi_pic]
        wb = Workbook(); ws = wb.active; ws.title = "SO Batch Upload"
        headers = ['SO Item', 'Delivery Plan Date', 'Remarks']
        ws.append(headers); ws.freeze_panes = 'A2'
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        col_widths   = [35, 25, 50]
        for i, cell in enumerate(ws[1], 1):
            cell.fill = header_fill; cell.font = Font(bold=True, color="000000"); cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(i)].width = col_widths[i - 1]
        grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        red_font  = Font(color="FF0000")
        ws.append(['example : 9008988017-10', 'example : 2025-12-31', 'example : Waiting for vendor confirmation'])
        for cell in ws[2]:
            cell.font = red_font; cell.fill = grey_fill
        for s in all_sos:
            if not s.so_item: continue
            plan = s.delivery_plan_date.isoformat() if s.delivery_plan_date else ''
            ws.append([s.so_item, plan, s.remarks or ''])
        output = io.BytesIO(); wb.save(output); output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f"Template_SO_BatchUpload_{datetime.now().strftime('%Y%m%d')}.xlsx")
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/data/so/batch-upload', methods=['POST'])
def batch_upload_so():
    try:
        if 'file' not in request.files: return jsonify({'error': 'No file'}), 400
        file = request.files['file']
        df = pd.read_excel(file, engine='openpyxl', skiprows=[1])
        df.columns = [str(c).strip() for c in df.columns]
        col_so_item = find_column(df, ['SO Item', 'SO Item No', 'SO Item Number'])
        col_plan    = find_column(df, ['Delivery Plan Date', 'Plan Date'])
        col_rem     = find_column(df, ['Remarks', 'Remark'])
        if not col_so_item: return jsonify({'error': f'Column "SO Item" not found. Available: {df.columns.tolist()}'}), 400
        updated = 0; not_found = 0
        for _, row in df.iterrows():
            so_item_val = clean(df_val(row, col_so_item)) if col_so_item else None
            if not so_item_val: continue
            so = SOData.query.filter_by(so_item=so_item_val).first()
            if so:
                if col_plan: so.delivery_plan_date = parse_date(df_val(row, col_plan))
                if col_rem: so.remarks = clean(df_val(row, col_rem)) or ''
                updated += 1
            else: not_found += 1
        db.session.commit()
        return jsonify({'updated': updated, 'not_found': not_found})
    except Exception as e:
        db.session.rollback(); import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

def _style_wb(ws, headers, num_cols=None):
    ws.append(headers); ws.freeze_panes = 'A2'
    fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for i, cell in enumerate(ws[1], 1):
        cell.fill = fill; cell.font = Font(bold=True, color="FFFFFF"); cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(i)].width = 20
    if num_cols:
        for row in ws.iter_rows(min_row=2):
            for ci in num_cols: row[ci-1].number_format = '#,##0.00'

@app.route('/api/export/all-so', methods=['GET'])
def export_all_so():
    try:
        q = SOData.query.filter(open_so_filter())
        op_units = request.args.getlist('op_unit')
        vendors  = request.args.getlist('vendor')
        manufacturers = request.args.getlist('manufacturer')
        statuses = request.args.getlist('status')
        aging_list = request.args.getlist('aging')
        so_items = request.args.getlist('so_item')
        pics = request.args.getlist('pic')
        kpi_pic = (request.args.get('kpi_pic') or '').strip()
        global_pics = request.args.getlist('global_pic')
        clients = selected_clients()
        margin_filter = request.args.get('margin_filter', 'all')
        sort_order = request.args.get('sort_order', 'oldest')
        date_year, date_from, date_to = parse_so_date_args()
        q = apply_so_client_filter(q, clients)
        q = apply_so_pic_filter(q, global_pics)
        if op_units: q = q.filter(SOData.operation_unit_name.in_(op_units))
        if vendors: q = q.filter(SOData.vendor_name.in_(vendors))
        if manufacturers: q = q.filter(SOData.manufacturer_name.in_(manufacturers))
        if statuses: q = q.filter(SOData.so_status.in_(statuses))
        if so_items: q = q.filter(SOData.so_item.in_(so_items))
        q = apply_so_pic_filter(q, pics)
        q = apply_so_create_date_filter(q, date_year, date_from, date_to)
        if sort_order == 'newest': sos = q.order_by(SOData.so_create_date.desc(), SOData.so_item.asc()).all()
        else: sos = q.order_by(SOData.so_create_date.asc(), SOData.so_item.asc()).all()
        today = date.today()
        hidden_so = get_hidden_so_items()
        sos = [s for s in sos if s.so_item not in hidden_so and s.so_number not in hidden_so and so_is_countable(s.so_item, customer_po_number=s.customer_po_number, delivery_memo=s.delivery_memo)]
        if aging_list: sos = [s for s in sos if get_aging_label(workdays_since(s.so_create_date, today)) in aging_list]
        if margin_filter in ('positive', 'negative'):
            def calc_margin(s):
                po_amt = purchase_amount_idr_for_margin(s)
                if po_amt <= 0: return None  # invalid purchase → no margin
                return float(s.sales_amount or 0) - po_amt
            if margin_filter == 'negative': sos = [s for s in sos if (lambda m: m is not None and m < 0)(calc_margin(s))]
            else: sos = [s for s in sos if (lambda m: m is not None and m >= 0)(calc_margin(s))]
        if kpi_pic: sos = [s for s in sos if canonical_pending_pic(s.pic_name, s.operation_unit_name) == kpi_pic]
        wb = Workbook(); ws = wb.active; ws.title = "SO List"
        headers = ['Aging', 'Day', 'SO Create Date', 'SO Item', 'PO No.', 'SO Status', 'Category', 'PIC', 'Product ID', 'Product Name', 'Specification', 'Manufacturer Name', 'SO Quantity', 'Sales Unit', 'Operation Unit Name', 'Vendor ID', 'Vendor Name', 'Currency', 'Sales Price(Exclude Tax)', 'Sales Amount(Exclude Tax)', 'Purchasing Currency', 'Purchasing Price', 'Margin', '%Margin', 'Delivery Memo', 'Plan Date', 'Remarks']
        _style_wb(ws, headers, num_cols=[2,13,19,20,22,23,24])
        widths = [14, 10, 16, 22, 22, 24, 22, 16, 18, 30, 44, 28, 14, 14, 30, 16, 28, 12, 22, 24, 20, 18, 18, 12, 30, 16, 70]
        for i, width in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = width
        for s in sos:
            day = workdays_since(s.so_create_date, today)
            # Use IDR-converted purchase amount for margin (handles USD/EUR).
            po_amount = purchase_amount_idr_for_margin(s)
            sales_amount = float(s.sales_amount or 0)
            # FIX V9: Margin valid only when purchase is present AND positive (in IDR)
            # AND sales > 0. Kalau PO kosong/0 atau sales kosong/0 → margin = None.
            purchase_valid = po_amount > 0
            sales_valid = sales_amount > 0
            margin = (sales_amount - po_amount) if (purchase_valid and sales_valid) else None
            # FIX V9: margin_pct hanya dihitung kalau po_amount > 0 (hindari division by zero)
            margin_pct = (margin / po_amount * 100) if (purchase_valid and sales_valid and po_amount) else None
            ws.append([
                get_aging_label(day), day if day is not None else '', s.so_create_date.isoformat() if s.so_create_date else '', s.so_item or '', s.matched_po_number or '', s.so_status or '',
                product_category_level1(s.product_id), canonical_pending_pic(s.pic_name, s.operation_unit_name), s.product_id or '', s.product_name or '', s.specification or '', s.manufacturer_name or '',
                s.so_qty or 0, s.sales_unit or '', s.operation_unit_name or '', s.vendor_id or '', s.vendor_name or '', s.currency or '', s.sales_price or 0, sales_amount, s.purchasing_currency or '', s.purchasing_price or 0,
                margin, margin_pct if margin_pct is not None else '', s.delivery_memo or '', s.delivery_plan_date.isoformat() if s.delivery_plan_date else '', s.remarks or '',
            ])
        output = io.BytesIO(); wb.save(output); output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f"SO_List_{datetime.now().strftime('%Y%m%d')}.xlsx")
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/completed/summary', methods=['GET'])
def completed_summary():
    try:
        year_filter = request.args.get('year', 'all')
        date_year   = request.args.get('date_year', '')
        date_from   = request.args.get('date_from', '')
        date_to     = request.args.get('date_to', '')
        yoy_base_year = request.args.get('yoy_base_year', '')
        mode = (request.args.get('mode') or '').strip().lower()
        light_mode = mode in ('dashboard', 'light', 'kpi')
        is_sqlite = 'sqlite' in app.config['SQLALCHEMY_DATABASE_URI']
        clients = selected_clients()
        pics = selected_pics()
        q = db.session.query(SOData).filter(SOData.so_status == 'Delivery Completed')
        q = apply_so_client_filter(q, clients)
        q = apply_so_pic_filter(q, pics)
        yoy_q = db.session.query(SOData).filter(SOData.so_status == 'Delivery Completed')
        yoy_q = apply_so_client_filter(yoy_q, clients)
        yoy_q = apply_so_pic_filter(yoy_q, pics)
        effective_year = date_year or (year_filter if year_filter and year_filter != 'all' else '')
        if effective_year:
            try:
                yr = int(effective_year)
                start_date = date(yr, 1, 1); end_date = date(yr, 12, 31)
                q = q.filter(SOData.so_create_date >= start_date, SOData.so_create_date <= end_date)
            except ValueError: pass
        else:
            if date_from: q = q.filter(SOData.so_create_date >= date_from)
            if date_to: q = q.filter(SOData.so_create_date <= date_to)
        q = q.filter(~SOData.operation_unit_name.in_(list(EXCLUDED_OP_UNITS)))
        yoy_q = yoy_q.filter(~SOData.operation_unit_name.in_(list(EXCLUDED_OP_UNITS)))
        cache_key = (_RUNTIME_CACHE_VERSION, year_filter or 'all', date_year or '', date_from or '', date_to or '', yoy_base_year or '', mode or '', tuple(sorted(clients)), tuple(sorted(pics)))
        now_ts = datetime.utcnow().timestamp()
        with _COMPLETED_CACHE_LOCK:
            cache_entry = _COMPLETED_SUMMARY_CACHE.get(cache_key)
            if cache_entry and now_ts - cache_entry.get('created_at', 0) < _COMPLETED_SUMMARY_CACHE_TTL_SECONDS: return jsonify(cache_entry['payload'])
        if light_mode:
            currency_expr = func.upper(func.trim(func.coalesce(SOData.purchasing_currency, '')))
            raw_purchase_expr = case(
                (func.coalesce(SOData.purchasing_amount, 0) != 0, func.coalesce(SOData.purchasing_amount, 0)),
                else_=func.coalesce(SOData.purchasing_price, 0) * func.coalesce(SOData.so_qty, 0)
            )
            purchase_expr = case(
                (SOData.purchasing_amount_idr.isnot(None), SOData.purchasing_amount_idr),
                (currency_expr.in_(['', 'IDR']), raw_purchase_expr),
                else_=0.0
            )
            sales_expr = func.coalesce(SOData.sales_amount, 0.0)
            # Margin valid only when purchase is present AND positive (not empty/zero/negative).
            # has_purchase_expr: purchase amount/price exists and is non-zero
            # purchase_positive_expr: the computed purchase amount is > 0
            has_purchase_expr = db.or_(db.and_(SOData.purchasing_amount.isnot(None), SOData.purchasing_amount != 0), db.and_(SOData.purchasing_price.isnot(None), SOData.purchasing_price != 0))
            purchase_positive_expr = (purchase_expr > 0)
            margin_expr = case((db.and_(has_purchase_expr, purchase_positive_expr), sales_expr - purchase_expr), else_=None)
            sum_purchase_expr = func.coalesce(func.sum(purchase_expr), 0.0)
            sum_sales_expr = func.coalesce(func.sum(sales_expr), 0.0)
            sum_margin_expr = func.coalesce(func.sum(func.coalesce(margin_expr, 0.0)), 0.0)
            count_expr = func.count(SOData.id)
            kpi_row = q.with_entities(count_expr, sum_sales_expr, sum_purchase_expr, func.coalesce(func.sum(case((margin_expr > 0, 1), else_=0)), 0), func.coalesce(func.sum(case((margin_expr < 0, 1), else_=0)), 0), func.coalesce(func.sum(case((margin_expr == 0, 1), else_=0)), 0)).first()
            total_count = int(kpi_row[0] or 0) if kpi_row else 0
            total_sales = float(kpi_row[1] or 0) if kpi_row else 0.0
            total_purchase = float(kpi_row[2] or 0) if kpi_row else 0.0
            pos = int(kpi_row[3] or 0) if kpi_row else 0
            neg = int(kpi_row[4] or 0) if kpi_row else 0
            zero = int(kpi_row[5] or 0) if kpi_row else 0
            month_expr = func.strftime('%Y-%m', SOData.so_create_date) if is_sqlite else func.to_char(func.date_trunc('month', SOData.so_create_date), 'YYYY-MM')
            monthly_trend = []
            month_rows = q.filter(SOData.so_create_date.isnot(None)).with_entities(month_expr.label('month'), count_expr.label('count'), sum_sales_expr.label('sales_amount'), sum_purchase_expr.label('purchase_amount')).group_by(month_expr).order_by(month_expr).all()
            for month, cnt, sales_amt, purchase_amt in month_rows:
                monthly_trend.append({'month': month, 'count': int(cnt or 0), 'sales_amount': float(sales_amt or 0), 'purchase_amount': float(purchase_amt or 0)})
            current_year = datetime.utcnow().year
            def _int_year(value):
                try: return int(str(value)[:4])
                except: return None
            base_year = _int_year(yoy_base_year) or _int_year(effective_year) or _int_year(date_from) or current_year
            latest_three_years = sorted({current_year, current_year - 1, current_year - 2})
            yoy_years = [year for year in latest_three_years if year != base_year]
            if len(yoy_years) > 2: yoy_years = yoy_years[-2:]
            yoy_fields = {year: f'purchase_{year}' for year in yoy_years}
            purchase_yoy_trend = []; purchase_yoy_by_month = {}
            for month_num in range(1, 13):
                row = {'month': month_num, 'month_label': datetime(current_year, month_num, 1).strftime('%B')}
                for field in yoy_fields.values(): row[field] = 0.0
                purchase_yoy_trend.append(row); purchase_yoy_by_month[month_num] = row
            if yoy_years:
                if is_sqlite:
                    yoy_year_expr = func.strftime('%Y', SOData.so_create_date)
                    yoy_month_expr = func.strftime('%m', SOData.so_create_date)
                    yoy_filter = yoy_year_expr.in_([str(y) for y in yoy_years])
                else:
                    yoy_year_expr = func.extract('year', SOData.so_create_date)
                    yoy_month_expr = func.extract('month', SOData.so_create_date)
                    yoy_filter = yoy_year_expr.in_(yoy_years)
                yoy_rows = yoy_q.filter(SOData.so_create_date.isnot(None), yoy_filter).with_entities(yoy_year_expr.label('yr'), yoy_month_expr.label('mo'), sum_purchase_expr.label('purchase_amount')).group_by(yoy_year_expr, yoy_month_expr).all()
                for yr, mo, purchase_amt in yoy_rows:
                    try: year_int = int(yr); month_int = int(mo)
                    except: continue
                    field = yoy_fields.get(year_int)
                    if field and month_int in purchase_yoy_by_month: purchase_yoy_by_month[month_int][field] = round(float(purchase_amt or 0), 2)
            def group_top(base_q, label_expr, label_key, value_key='purchase_amount', limit=5, extra_filter=None):
                gq = base_q
                if extra_filter is not None: gq = gq.filter(extra_filter)
                rows = gq.with_entities(label_expr.label(label_key), count_expr.label('count'), sum_sales_expr.label('sales_amount'), sum_purchase_expr.label('purchase_amount'), sum_margin_expr.label('margin')).group_by(label_expr).order_by(desc(value_key if isinstance(value_key, str) else value_key)).limit(limit).all()
                result = []
                for label, cnt, sales_amt, purchase_amt, margin_amt in rows:
                    result.append({label_key: label or 'Unknown', 'count': int(cnt or 0), 'sales_amount': float(sales_amt or 0), 'purchase_amount': float(purchase_amt or 0), 'margin': float(margin_amt or 0)})
                return result
            vendor_label = func.coalesce(func.nullif(func.trim(SOData.vendor_name), ''), 'Unknown')
            client_label = func.coalesce(func.nullif(func.trim(SOData.operation_unit_name), ''), 'Unknown')
            local_filter = currency_expr.in_(['', 'IDR'])
            import_filter = db.not_(currency_expr.in_(['', 'IDR']))
            top_vendors = group_top(q, vendor_label, 'vendor', value_key=sum_purchase_expr, limit=5)
            top_vendors_local = group_top(q, vendor_label, 'vendor', value_key=sum_purchase_expr, limit=5, extra_filter=local_filter)
            top_vendors_import = group_top(q, vendor_label, 'vendor', value_key=sum_purchase_expr, limit=5, extra_filter=import_filter)
            top_clients = group_top(q, client_label, 'client', value_key=sum_sales_expr, limit=5)
            missing_conversion_count = q.filter(SOData.purchasing_amount_idr.is_(None), db.not_(currency_expr.in_(['', 'IDR'])), raw_purchase_expr > 0).count()
            payload = {
                'total_count': total_count, 'total_sales': total_sales, 'total_purchase': total_purchase,
                # FIX V9: total_margin hanya dihitung kalau total_purchase > 0 DAN total_sales > 0.
                # Kalau semua PO kosong → total_margin = None (tidak menyesatkan dengan angka 0 atau negatif palsu).
                'total_margin': (total_sales - total_purchase) if (total_sales > 0 and total_purchase > 0) else None,
                'monthly_trend': monthly_trend, 'purchase_yoy_years': yoy_years, 'purchase_yoy_trend': purchase_yoy_trend,
                'top_vendors': top_vendors, 'top_vendors_local': top_vendors_local, 'top_vendors_import': top_vendors_import,
                'top_clients': top_clients, 'top_items': [], 'worst_margin_vendors': [], 'worst_margin_transactions': [],
                'margin_distribution': {'positive': pos, 'negative': neg, 'zero': zero},
                'conversion_status': {'checked': True, 'had_missing_cache': missing_conversion_count > 0, 'converted_count': 0, 'pending_count': int(missing_conversion_count or 0), 'message': 'Dashboard memakai cache currency yang sudah tersimpan. Backfill rate dijalankan terpisah.'}
            }
            with _COMPLETED_CACHE_LOCK: _COMPLETED_SUMMARY_CACHE[cache_key] = {'created_at': now_ts, 'payload': payload}
            return jsonify(payload)
        completed_summary_fields = (SOData.so_number, SOData.so_item, SOData.operation_unit_name, SOData.vendor_name, SOData.vendor_id, SOData.client_id, SOData.so_qty, SOData.sales_amount, SOData.purchasing_price, SOData.purchasing_amount, SOData.purchasing_currency, SOData.purchasing_amount_idr, SOData.so_create_date)
        if not light_mode: completed_summary_fields = completed_summary_fields + (SOData.product_name, SOData.specification, SOData.product_id)
        purchase_summary_fields = (SOData.so_qty, SOData.purchasing_price, SOData.purchasing_amount, SOData.purchasing_currency, SOData.purchasing_amount_idr, SOData.so_create_date)
        rows = q.options(load_only(*completed_summary_fields)).all()
        yoy_rows = yoy_q.options(load_only(*purchase_summary_fields)).all()
        missing_conversion_count = sum(1 for s in rows if s.purchasing_amount_idr is None and str(s.purchasing_currency or 'IDR').strip().upper() != 'IDR' and raw_purchase_amount(s) > 0)
        converted_count = ensure_purchase_amount_idr_cache(rows, fetch_missing=False)
        def po_amt_of(s): return purchase_amount_idr(s)
        enriched = []
        for s in rows:
            po_amt = po_amt_of(s); sales = float(s.sales_amount or 0)
            # FIX V9: Margin valid only when purchase (in IDR) is present AND positive
            # AND sales > 0. Kalau PO kosong/0 atau sales kosong/0 → margin = None.
            po_amt_margin = purchase_amount_idr_for_margin(s)
            margin = (sales - po_amt_margin) if (po_amt_margin > 0 and sales > 0) else None
            enriched.append((s, po_amt, sales, margin))
        monthly = {}
        for s, po_amt, sales, _m in enriched:
            if not s.so_create_date: continue
            key = s.so_create_date.strftime('%Y-%m')
            if key not in monthly: monthly[key] = {'month': key, 'count': 0, 'sales_amount': 0.0, 'purchase_amount': 0.0}
            monthly[key]['count'] += 1; monthly[key]['sales_amount'] += sales; monthly[key]['purchase_amount'] += po_amt
        monthly_trend = sorted(monthly.values(), key=lambda x: x['month'])
        current_year = datetime.utcnow().year
        def _int_year(value):
            try: return int(str(value)[:4])
            except: return None
        base_year = _int_year(yoy_base_year) or _int_year(effective_year) or _int_year(date_from) or current_year
        latest_three_years = sorted({current_year, current_year - 1, current_year - 2})
        yoy_years = [year for year in latest_three_years if year != base_year]
        if len(yoy_years) > 2: yoy_years = yoy_years[-2:]
        yoy_fields = {year: f'purchase_{year}' for year in yoy_years}
        purchase_yoy_trend = []; purchase_yoy_by_month = {}
        for month_num in range(1, 13):
            row = {'month': month_num, 'month_label': datetime(current_year, month_num, 1).strftime('%B')}
            for field in yoy_fields.values(): row[field] = 0.0
            purchase_yoy_trend.append(row); purchase_yoy_by_month[month_num] = row
        ensure_purchase_amount_idr_cache(yoy_rows, fetch_missing=False)
        for s in yoy_rows:
            if not s.so_create_date: continue
            year = s.so_create_date.year
            if year not in yoy_fields: continue
            purchase_yoy_by_month[s.so_create_date.month][yoy_fields[year]] += purchase_amount_idr(s)
        for row in purchase_yoy_trend:
            for field in yoy_fields.values(): row[field] = round(row[field], 2)
        def currency_bucket(s):
            cur = (s.purchasing_currency or 'IDR').strip().upper()
            return 'local' if cur in ('', 'IDR') else 'import'
        def add_vendor(target, s, po_amt, sales, m):
            v = s.vendor_name or 'Unknown'
            if v not in target: target[v] = {'vendor': v, 'count': 0, 'sales_amount': 0.0, 'purchase_amount': 0.0, 'margin': 0.0}
            target[v]['count'] += 1; target[v]['sales_amount'] += sales; target[v]['purchase_amount'] += po_amt
            if m is not None: target[v]['margin'] += m
        vendor_map = {}; vendor_local_map = {}; vendor_import_map = {}; client_map = {}
        for s, po_amt, sales, m in enriched:
            add_vendor(vendor_map, s, po_amt, sales, m)
            if currency_bucket(s) == 'local': add_vendor(vendor_local_map, s, po_amt, sales, m)
            else: add_vendor(vendor_import_map, s, po_amt, sales, m)
            client = s.operation_unit_name or 'Unknown'
            if client not in client_map: client_map[client] = {'client': client, 'count': 0, 'sales_amount': 0.0, 'purchase_amount': 0.0, 'margin': 0.0}
            client_map[client]['count'] += 1; client_map[client]['sales_amount'] += sales; client_map[client]['purchase_amount'] += po_amt
            if m is not None: client_map[client]['margin'] += m
        def top_purchase_vendors(mapping):
            return sorted((row for row in mapping.values() if float(row.get('purchase_amount') or 0) > 0), key=lambda x: x['purchase_amount'], reverse=True)[:5]
        top_vendors = top_purchase_vendors(vendor_map)
        top_vendors_local = top_purchase_vendors(vendor_local_map)
        top_vendors_import = top_purchase_vendors(vendor_import_map)
        top_clients = sorted(client_map.values(), key=lambda x: x['sales_amount'], reverse=True)[:5]
        pos = neg = zero = 0
        total_sales = 0.0; total_purchase = 0.0
        # FIX V9: Track total_margin secara terpisah (hanya dari row dengan valid margin).
        # Sebelumnya total_margin = total_sales - total_purchase, yang salah karena
        # total_sales dan total_purchase mencakup SEMUA row (termasuk yang PO-nya 0/kosong
        # dan tidak punya margin). Margin seharusnya hanya dihitung dari row dengan
        # po_amt > 0 DAN sales > 0.
        total_margin_sum = 0.0
        valid_margin_count = 0
        for _s, po_amt, sales, m in enriched:
            total_sales += sales; total_purchase += po_amt
            if m is not None:
                # FIX V9: hanya hitung margin kalau sales > 0 juga (hindari margin
                # dari row dengan sales kosong/0 yang menyesatkan).
                if sales > 0:
                    total_margin_sum += m
                    valid_margin_count += 1
                if m > 0: pos += 1
                elif m < 0: neg += 1
                else: zero += 1
        top_items = []; worst_margin_vendors = []; worst_margin_transactions = []
        if not light_mode:
            item_map = {}
            for s, po_amt, sales, m in enriched:
                pid = (s.product_id or '').strip(); label = s.product_name or s.so_item or 'Unknown'; key = pid or label
                if key not in item_map: item_map[key] = {'item': label, 'specification': s.specification or '', 'product_id': pid, 'count': 0, 'sales_amount': 0.0, 'purchase_amount': 0.0, 'margin': 0.0}
                agg = item_map[key]; agg['count'] += 1; agg['sales_amount'] += sales; agg['purchase_amount'] += po_amt
                if m is not None: agg['margin'] += m
                if not agg['specification'] and s.specification: agg['specification'] = s.specification
            top_items = sorted(item_map.values(), key=lambda x: x['sales_amount'], reverse=True)[:20]
            neg_vendor_map = {}
            for s, po_amt, sales, m in enriched:
                if m is None or m >= 0: continue
                v = s.vendor_name or 'Unknown'
                if v not in neg_vendor_map: neg_vendor_map[v] = {'vendor': v, 'margin': 0.0, 'count': 0, 'total_sales': 0.0, 'total_purchase': 0.0}
                neg_vendor_map[v]['margin'] += m; neg_vendor_map[v]['count'] += 1; neg_vendor_map[v]['total_sales'] += sales; neg_vendor_map[v]['total_purchase'] += po_amt
            worst_margin_vendors = sorted(neg_vendor_map.values(), key=lambda x: x['margin'])[:50]
            neg_txns = [(s, po_amt, sales, m) for s, po_amt, sales, m in enriched if m is not None and m < 0]
            neg_txns.sort(key=lambda x: x[3])
            for s, po_amt, sales, m in neg_txns[:30]:
                # FIX V9: margin_pct hanya dihitung kalau sales > 0 (hindari division by zero)
                pct = round(m / sales * 100, 1) if (m is not None and sales > 0) else None
                worst_margin_transactions.append({'so_item': s.so_item, 'so_number': s.so_number, 'item_code': (s.item_code if hasattr(s, 'item_code') and s.item_code else (s.so_item or '-')), 'product': s.product_name or '-', 'vendor': s.vendor_name or '-', 'sales_amount': sales, 'purchase_amount': po_amt, 'margin': m, 'margin_pct': pct, 'count': 1, 'date': s.so_create_date.isoformat() if s.so_create_date else None})
        payload = {
            'total_count': len(rows), 'total_sales': total_sales, 'total_purchase': total_purchase,
            # FIX V9: total_margin sekarang hanya dari row dengan valid margin (po > 0 dan sales > 0).
            # Sebelumnya: (total_sales - total_purchase) yang mencakup row dengan PO kosong → margin menyesatkan.
            'total_margin': round(total_margin_sum, 2) if valid_margin_count > 0 else None,
            'valid_margin_count': valid_margin_count,
            'monthly_trend': monthly_trend, 'purchase_yoy_years': yoy_years, 'purchase_yoy_trend': purchase_yoy_trend,
            'top_vendors': top_vendors, 'top_vendors_local': top_vendors_local, 'top_vendors_import': top_vendors_import,
            'top_clients': top_clients, 'top_items': top_items, 'worst_margin_vendors': worst_margin_vendors,
            'worst_margin_transactions': worst_margin_transactions, 'margin_distribution': {'positive': pos, 'negative': neg, 'zero': zero},
            'conversion_status': {'checked': True, 'had_missing_cache': missing_conversion_count > 0, 'converted_count': converted_count, 'pending_count': max(missing_conversion_count - converted_count, 0), 'message': (f'Konversi currency selesai dan disimpan untuk {converted_count} data baru.' if converted_count else 'Tidak ada data currency baru yang perlu dikonversi.')}
        }
        with _COMPLETED_CACHE_LOCK: _COMPLETED_SUMMARY_CACHE[cache_key] = {'created_at': now_ts, 'payload': payload}
        return jsonify(payload)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/completed/margin-detail', methods=['GET'])
def completed_margin_detail():
    try:
        category = request.args.get('category', 'positive')
        date_from = request.args.get('date_from', ''); date_to   = request.args.get('date_to', ''); date_year = request.args.get('date_year', '')
        is_sqlite = 'sqlite' in app.config['SQLALCHEMY_DATABASE_URI']
        clients = selected_clients(); pics = selected_pics()
        q = db.session.query(SOData).filter(SOData.so_status == 'Delivery Completed')
        q = apply_so_client_filter(q, clients); q = apply_so_pic_filter(q, pics)
        if date_year:
            try:
                yr = int(date_year); start_date = date(yr, 1, 1); end_date = date(yr, 12, 31)
                q = q.filter(SOData.so_create_date >= start_date, SOData.so_create_date <= end_date)
            except ValueError: pass
        elif date_from or date_to:
            if date_from: q = q.filter(SOData.so_create_date >= date_from)
            if date_to: q = q.filter(SOData.so_create_date <= date_to)
        rows = q.filter(~SOData.operation_unit_name.in_(list(EXCLUDED_OP_UNITS))).all()
        ensure_purchase_amount_idr_cache(rows)
        def get_po_amt(s): return purchase_amount_idr(s)
        result = []
        for s in rows:
            po_amt = get_po_amt(s)
            # Margin valid only when purchase (in IDR) is positive.
            # Use purchase_amount_idr_for_margin for USD/EUR conversion.
            po_amt_margin = purchase_amount_idr_for_margin(s)
            sales_val = float(s.sales_amount or 0)
            # FIX V9: margin hanya dihitung kalau PO amount > 0 DAN sales > 0.
            # Kalau PO kosong/0 atau sales kosong/0 → margin = None (tidak dihitung).
            m = (sales_val - po_amt_margin) if (po_amt_margin > 0 and sales_val > 0) else None
            if m is None and category in ('positive', 'negative'): continue
            if category == 'positive' and (m is None or m <= 0): continue
            elif category == 'negative' and (m is None or m >= 0): continue
            elif category == 'zero' and (m is None or m != 0): continue
            result.append({
                'id': s.id, 'so_item': s.so_item, 'so_number': s.so_number, 'product': s.product_name or '-', 'vendor': s.vendor_name or '-',
                'item_code': (s.item_code if hasattr(s, 'item_code') and s.item_code else '-'), 'sales_amount': sales_val, 'purchase_amount': po_amt,
                # FIX V9: margin_pct hanya dihitung kalau sales > 0 (hindari division by zero)
                'margin': m, 'margin_pct': round(m / sales_val * 100, 1) if (m is not None and sales_val > 0) else None, 'date': s.so_create_date.isoformat() if s.so_create_date else None,
                'so_status': s.so_status, 'pic_name': canonical_pending_pic(s.pic_name, s.operation_unit_name), 'operation_unit_name': s.operation_unit_name,
            })
        result.sort(key=lambda x: x['margin'])
        return jsonify(result)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/clients', methods=['GET'])
def get_clients():
    try:
        ensure_default_item_registration_loaded()
        clients = set()
        clients.update(c for (c,) in db.session.query(SOData.operation_unit_name).distinct().all() if c)
        clients.update(c for (c,) in db.session.query(ItemRegistration.client_name).distinct().all() if c)
        return jsonify(sorted(c for c in clients if c))
    except Exception as e: return jsonify({'error': str(e)}), 500

def load_similarity_cache():
    global _SIMILARITY_CACHE
    try:
        if os.path.exists(_SIMILARITY_CACHE_FILE):
            with open(_SIMILARITY_CACHE_FILE, 'r', encoding='utf-8') as f: _SIMILARITY_CACHE = json.load(f)
    except Exception as e:
        print(f"Error loading similarity cache: {e}"); _SIMILARITY_CACHE = {}

def save_similarity_cache():
    try:
        os.makedirs(os.path.dirname(_SIMILARITY_CACHE_FILE), exist_ok=True)
        with open(_SIMILARITY_CACHE_FILE, 'w', encoding='utf-8') as f: json.dump(_SIMILARITY_CACHE, f, ensure_ascii=False, separators=(',', ':'))
    except Exception as e:
        print(f"Error saving similarity cache: {e}")

@app.route('/api/item-registration/data', methods=['GET'])
def get_item_registration_data():
    try:
        cache_key = runtime_cache_key('item_registration_data')
        cached = runtime_cache_get(cache_key)
        if cached is not None: return jsonify(cached)
        
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 10))
        search = request.args.get('search', '').strip()
        req_numbers = [n.strip() for n in request.args.getlist('req_no') if n.strip()]
        date_year, date_from, date_to = parse_so_date_args()
        clients = selected_clients()
        global_pics = [p.strip() for p in request.args.getlist('global_pic') if p.strip()]
        item_clients = [c.strip() for c in request.args.getlist('item_client') if c.strip()]
        categories = [c.strip() for c in request.args.getlist('category') if c.strip()]
        pics = [p.strip() for p in request.args.getlist('pic') if p.strip()]
        kpi_pic = (request.args.get('kpi_pic') or '').strip()
        proc_statuses = [s.strip() for s in request.args.getlist('proc_status') if s.strip()]
        mfr_names = [s.strip() for s in request.args.getlist('mfr_name') if s.strip()]
        op_units = [s.strip() for s in request.args.getlist('op_unit') if s.strip()]
        bid_types = [s.strip() for s in request.args.getlist('bid_except_type') if s.strip()]
        
        q = apply_item_registration_visible_status_filter(apply_item_registration_date_filter(ItemRegistration.query, date_year, date_from, date_to))
        if clients: q = q.filter(ItemRegistration.client_name.in_(clients))
        q = apply_item_registration_pic_filter(q, global_pics)
        if item_clients: q = q.filter(ItemRegistration.client_name.in_(item_clients))
        if categories: q = q.filter(ItemRegistration.category.in_(categories))
        if proc_statuses: q = q.filter(ItemRegistration.proc_status.in_(proc_statuses))
        if mfr_names: q = q.filter(ItemRegistration.mfr_name.in_(mfr_names))
        if op_units: q = q.filter(ItemRegistration.operation_unit_name.in_(op_units))
        if bid_types: q = q.filter(ItemRegistration.bid_except_type.in_(bid_types))
        if req_numbers: q = q.filter(ItemRegistration.req_no.in_(req_numbers))
        if search:
            pattern = f'%{search}%'
            q = q.filter(db.or_(ItemRegistration.req_no.ilike(pattern), ItemRegistration.prod_id.ilike(pattern), ItemRegistration.prod_name.ilike(pattern), ItemRegistration.vendor_name.ilike(pattern), ItemRegistration.mfr_name.ilike(pattern), ItemRegistration.remarks.ilike(pattern)))

        missing_q = apply_item_registration_kpi_status_filter(q).filter(db.or_(ItemRegistration.prod_id.is_(None), ItemRegistration.prod_id == '', ItemRegistration.prod_id == '-'))
        _client_pic_cache = {normalize_client_id(m.client_id): clean(m.pic_name) for m in db.session.query(MasterClientPIC).all() if clean(m.pic_name)}
        _bid_type_pic_cache = {clean(m.bid_type): clean(m.pic_name) for m in db.session.query(MasterBidTypePIC).all() if clean(m.pic_name)}
        _cat_by_id, _cat_by_name = master_pic_maps()
        all_known_pics = set()
        all_known_pics.update(_cat_by_id.values())
        all_known_pics.update(_cat_by_name.values())
        all_known_pics.update(_client_pic_cache.values())
        all_known_pics.update(_bid_type_pic_cache.values())
        all_known_pics.discard('')
        all_known_pics.discard(None)
        missing_rows = missing_q.all()
        missing_by_pic = {pic: 0 for pic in all_known_pics}
        for row in missing_rows:
            pic = None
            cid = normalize_client_id(row.client_id) if row.client_id else ''
            if cid and cid in _client_pic_cache: pic = _client_pic_cache[cid]
            if not pic and row.bid_except_type:
                bt = clean(row.bid_except_type)
                if bt and bt in _bid_type_pic_cache: pic = _bid_type_pic_cache[bt]
            if not pic:
                cat_id = normalize_category_id(row.category_id)
                cat_name = normalize_category_name(row.category)
                if cat_id and cat_id in _cat_by_id: pic = _cat_by_id[cat_id]
                elif cat_name and cat_name in _cat_by_name: pic = _cat_by_name[cat_name]
            pic = pic or row.pic or ''
            pic = canonical_pending_pic(clean(pic), row.client_name)
            if pic and pic != 'Unassigned': missing_by_pic[pic] = missing_by_pic.get(pic, 0) + 1
        missing_prod_id_by_pic = [{'pic': pic, 'count': count} for pic, count in sorted(missing_by_pic.items(), key=lambda item: (-item[1], item[0]))]

        q = apply_item_registration_pic_filter(q, pics)
        if kpi_pic:
            q = apply_item_registration_pic_filter(q, [kpi_pic])
            # Apply the SAME KPI status filter used for counting, so the
            # table shows exactly the same rows as the KPI count.
            q = apply_item_registration_kpi_status_filter(q)
            q = q.filter(db.or_(ItemRegistration.prod_id.is_(None), ItemRegistration.prod_id == '', ItemRegistration.prod_id == '-'))

        total = q.count()
        rows = q.order_by(ItemRegistration.uploaded_at.desc(), ItemRegistration.id.asc()).offset((page-1)*per_page).limit(per_page).all()
        
        option_q = apply_item_registration_visible_status_filter(apply_item_registration_date_filter(ItemRegistration.query, date_year, date_from, date_to))
        if clients: option_q = option_q.filter(ItemRegistration.client_name.in_(clients))
        option_q = apply_item_registration_pic_filter(option_q, global_pics)
        if item_clients: option_q = option_q.filter(ItemRegistration.client_name.in_(item_clients))
        if op_units: option_q = option_q.filter(ItemRegistration.operation_unit_name.in_(op_units))
        if bid_types: option_q = option_q.filter(ItemRegistration.bid_except_type.in_(bid_types))
        if categories: option_q = option_q.filter(ItemRegistration.category.in_(categories))
        if proc_statuses: option_q = option_q.filter(ItemRegistration.proc_status.in_(proc_statuses))
        if mfr_names: option_q = option_q.filter(ItemRegistration.mfr_name.in_(mfr_names))
        if req_numbers: option_q = option_q.filter(ItemRegistration.req_no.in_(req_numbers))
        if search:
            pattern = f'%{search}%'
            option_q = option_q.filter(db.or_(ItemRegistration.req_no.ilike(pattern), ItemRegistration.prod_id.ilike(pattern), ItemRegistration.prod_name.ilike(pattern), ItemRegistration.vendor_name.ilike(pattern), ItemRegistration.mfr_name.ilike(pattern), ItemRegistration.remarks.ilike(pattern)))

        def distinct_options(query, column):
            return sorted({clean(value) for (value,) in query.with_entities(column).distinct().all() if clean(value)})

        all_clients = distinct_options(option_q, ItemRegistration.client_name)
        all_op_units = distinct_options(option_q, ItemRegistration.operation_unit_name)
        all_bid_types = distinct_options(option_q, ItemRegistration.bid_except_type)
        all_categories = distinct_options(option_q, ItemRegistration.category)
        all_proc_statuses = distinct_options(option_q, ItemRegistration.proc_status)
        all_mfr_names = distinct_options(option_q, ItemRegistration.mfr_name)
        
        all_pics = set(all_known_pics)
        pic_option_rows = option_q.all()
        for row in pic_option_rows:
            pic = None
            cid = normalize_client_id(row.client_id) if row.client_id else ''
            if cid and cid in _client_pic_cache: pic = _client_pic_cache[cid]
            if not pic and row.bid_except_type:
                bt = clean(row.bid_except_type)
                if bt and bt in _bid_type_pic_cache: pic = _bid_type_pic_cache[bt]
            if not pic:
                cat_id = normalize_category_id(row.category_id)
                cat_name = normalize_category_name(row.category)
                if cat_id and cat_id in _cat_by_id: pic = _cat_by_id[cat_id]
                elif cat_name and cat_name in _cat_by_name: pic = _cat_by_name[cat_name]
            pic = pic or row.pic or ''
            resolved = canonical_pending_pic(clean(pic), row.client_name)
            if resolved != 'Unassigned': all_pics.add(resolved)
        all_pics = sorted(list(all_pics))

        last_upload = db.session.query(func.max(UploadLog.uploaded_at)).filter(UploadLog.file_type == 'ITEM_REG').scalar()
        response_rows = [item_registration_dict(r, include_similarity=False) for r in rows]

        payload = {
            'data': response_rows, 'total': total, 'page': page, 'per_page': per_page,
            'client_options': all_clients, 'op_unit_options': all_op_units, 'bid_except_type_options': all_bid_types, 'category_options': all_categories, 'pic_options': all_pics,
            'proc_status_options': all_proc_statuses, 'mfr_name_options': all_mfr_names,
            'missing_prod_id_by_pic': missing_prod_id_by_pic, 'last_updated': utc_isoformat(last_upload),
        }
        runtime_cache_set(cache_key, payload, ttl_seconds=60)
        return jsonify(payload)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/upload/item-registration-json', methods=['POST'])
@app.route('/api/upload/item-registration', methods=['POST'])
def upload_item_registration():
    try:
        uploads, upload_mode = request_upload_dataframes('item_registration')
        if not uploads: return jsonify({'error': 'No file uploaded or JSON rows supplied'}), 400

        # FIX V8: Item Registration selalu pakai replace mode (forced).
        # Di sistem source, data yang sudah selesai proses akan hilang.
        # Jadi req_no lama yang tidak ada di upload terbaru → harus dihapus.
        # Sebelumnya: replace_existing default False → data lama menumpuk,
        # record menjadi 2x lipat (8000 vs 4000 asli).
        replace_existing = True  # forced, abaikan upload_replace_mode()

        summary = {'processed': 0, 'added': 0, 'updated': 0, 'removed_duplicates': 0, 'removed_stale': 0, 'removed_blank': 0}
        latest_req_numbers = set()
        for upload in uploads:
            df = upload['df']
            result = import_item_registration_dataframe(df, upload['filename'])
            latest_req_numbers.update(result.get('keys', []))
            for key in summary: summary[key] += result.get(key, 0)

        db.session.flush()
        # FIX V8: selalu pass latest_req_numbers sebagai valid_keys (bukan None),
        # supaya req_no lama yang tidak ada di upload terbaru dihapus.
        cleanup = cleanup_source_table_snapshot(ItemRegistration, 'req_no', latest_req_numbers, timestamp_fields=('uploaded_at',), delete_blank=True)
        for key, value in cleanup.items(): summary[key] = summary.get(key, 0) + value

        db.session.commit()
        clear_runtime_caches()
        return jsonify({
            'message': f'Berhasil upload {len(uploads)} file Item Registration: +{summary["added"]} added, {summary["updated"]} updated, {summary["removed_duplicates"]} duplicate lama dihapus, {summary["removed_stale"]} data lama dibuang',
            'uploaded': summary['processed'], 'files': len(uploads), 'mode': upload_mode, 'replace': replace_existing, **summary,
        })
    except ValueError as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/item-registration/<int:item_id>', methods=['PUT'])
def update_item_registration(item_id):
    try:
        data = request.json or {}
        item = db.session.get(ItemRegistration, item_id)
        if not item: return jsonify({'error': 'Not found'}), 404
        if 'remarks' in data: item.remarks = data['remarks'] or ''
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

def apply_item_registration_request_filters(query):
    search = request.args.get('search', '').strip()
    req_numbers = [n.strip() for n in request.args.getlist('req_no') if n.strip()]
    date_year, date_from, date_to = parse_so_date_args()
    clients = selected_clients()
    global_pics = [p.strip() for p in request.args.getlist('global_pic') if p.strip()]
    item_clients = [c.strip() for c in request.args.getlist('item_client') if c.strip()]
    categories = [c.strip() for c in request.args.getlist('category') if c.strip()]
    pics = [p.strip() for p in request.args.getlist('pic') if p.strip()]
    kpi_pic = (request.args.get('kpi_pic') or '').strip()
    proc_statuses = [s.strip() for s in request.args.getlist('proc_status') if s.strip()]
    mfr_names = [s.strip() for s in request.args.getlist('mfr_name') if s.strip()]
    op_units = [s.strip() for s in request.args.getlist('op_unit') if s.strip()]
    bid_types = [s.strip() for s in request.args.getlist('bid_except_type') if s.strip()]
    query = apply_item_registration_date_filter(query, date_year, date_from, date_to)
    if clients: query = query.filter(ItemRegistration.client_name.in_(clients))
    query = apply_item_registration_pic_filter(query, global_pics)
    if item_clients: query = query.filter(ItemRegistration.client_name.in_(item_clients))
    if op_units: query = query.filter(ItemRegistration.operation_unit_name.in_(op_units))
    if bid_types: query = query.filter(ItemRegistration.bid_except_type.in_(bid_types))
    if categories: query = query.filter(ItemRegistration.category.in_(categories))
    query = apply_item_registration_pic_filter(query, pics)
    if proc_statuses: query = query.filter(ItemRegistration.proc_status.in_(proc_statuses))
    if mfr_names: query = query.filter(ItemRegistration.mfr_name.in_(mfr_names))
    if req_numbers: query = query.filter(ItemRegistration.req_no.in_(req_numbers))
    if search:
        pattern = f'%{search}%'
        query = query.filter(db.or_(ItemRegistration.req_no.ilike(pattern), ItemRegistration.prod_id.ilike(pattern), ItemRegistration.prod_name.ilike(pattern), ItemRegistration.vendor_name.ilike(pattern), ItemRegistration.mfr_name.ilike(pattern), ItemRegistration.remarks.ilike(pattern)))
    if kpi_pic:
        query = apply_item_registration_pic_filter(query, [kpi_pic])
        query = apply_item_registration_kpi_status_filter(query)
        query = query.filter(db.or_(ItemRegistration.prod_id.is_(None), ItemRegistration.prod_id == '', ItemRegistration.prod_id == '-'))
    return query


@app.route('/api/item-registration/vendor-auto-approve', methods=['POST'])
def item_registration_vendor_auto_approve():
    """Vendor Auto Approve endpoint.

    Accepts a list of req_no values from the Item Registration table.
    Groups them by vendor_id (looked up from Vendor Control data by matching
    vendor_name). For each vendor (processed SEQUENTIALLY, not in parallel):
      1. Login to Serveone Mall using vendor credentials (V + vendor_id + password)
      2. Navigate to Batch Unit Price Agreement menu
      3. Find rows matching the req numbers in the Request Number column
      4. Check matching rows
      5. Click Agreement button
      6. Handle confirmation + success popups
    Returns a summary of which req numbers were processed vs not found.
    """
    try:
        data = request.get_json(silent=True) or {}
        req_numbers = data.get('req_numbers', [])
        if not req_numbers or not isinstance(req_numbers, list):
            return jsonify({'error': 'req_numbers (list) is required'}), 400

        # ── Step 1: Look up vendor info for each req_no ──────────────────
        # Get ItemRegistration rows matching the req_numbers
        req_set = {str(r).strip() for r in req_numbers if str(r).strip()}
        items = ItemRegistration.query.filter(ItemRegistration.req_no.in_(list(req_set))).all()

        # Build req_no → vendor_name map
        req_to_vendor_name = {}
        for item in items:
            vn = clean(item.vendor_name)
            if vn:
                req_to_vendor_name[str(item.req_no).strip()] = vn

        # Get Vendor Control rows (vendor_name → vendor_id + password)
        try:
            vc_rows, _ = vendor_control_rows(force=True)
        except Exception as vc_exc:
            return jsonify({'error': f'Failed to load Vendor Control data: {vc_exc}'}), 500

        # Build vendor_name_lower → {vendor_id, password, vendor_name} map
        vc_by_name = {}
        for vc in vc_rows:
            vname = clean(vc.get('vendor_name'))
            if vname:
                vc_by_name[vname.strip().lower()] = {
                    'vendor_id': clean(vc.get('vendor_id')) or '',
                    'password': clean(vc.get('password')) or '',
                    'vendor_name': vname,
                }

        # Group req_numbers by vendor_id
        # vendor_groups = {vendor_id: {'password': ..., 'vendor_name': ..., 'req_numbers': [...]}}
        vendor_groups = {}
        req_without_vendor = []
        req_without_credentials = []

        for req_no in req_set:
            vendor_name = req_to_vendor_name.get(req_no)
            if not vendor_name:
                req_without_vendor.append(req_no)
                continue
            vc_info = vc_by_name.get(vendor_name.strip().lower())
            if not vc_info or not vc_info.get('vendor_id') or not vc_info.get('password'):
                req_without_credentials.append(req_no)
                continue
            vid = vc_info['vendor_id']
            if vid not in vendor_groups:
                vendor_groups[vid] = {
                    'password': vc_info['password'],
                    'vendor_name': vc_info['vendor_name'],
                    'req_numbers': [],
                }
            vendor_groups[vid]['req_numbers'].append(req_no)

        # ── Step 2: Process each vendor SEQUENTIALLY ─────────────────────
        # Import the automation script
        import sys
        import os
        scripts_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'scripts')
        if scripts_dir not in sys.path:
            sys.path.insert(0, scripts_dir)
        try:
            from vendor_auto_approve import run_vendor_auto_approve
        except ImportError as imp_exc:
            return jsonify({'error': f'Failed to import automation module: {imp_exc}. Ensure /scripts/vendor_auto_approve.py exists.'}), 500

        all_results = []
        all_processed = []
        all_not_found = []
        all_errors = []

        for vid, info in vendor_groups.items():
            vendor_result = {
                'vendor_id': vid,
                'vendor_name': info['vendor_name'],
                'req_numbers': info['req_numbers'],
                'processed': [],
                'not_found': [],
                'error': None,
            }
            try:
                auto_result = run_vendor_auto_approve(
                    vendor_id=vid,
                    password=info['password'],
                    req_numbers=info['req_numbers'],
                    headless=True,
                )
                vendor_result['processed'] = auto_result.get('processed', [])
                vendor_result['not_found'] = auto_result.get('not_found', [])
                vendor_result['error'] = auto_result.get('error')
                vendor_result['log'] = auto_result.get('log', [])
                all_processed.extend(vendor_result['processed'])
                all_not_found.extend(vendor_result['not_found'])
                if vendor_result['error']:
                    all_errors.append(f"Vendor {info['vendor_name']} ({vid}): {vendor_result['error']}")
            except Exception as exc:
                vendor_result['error'] = str(exc)
                all_errors.append(f"Vendor {info['vendor_name']} ({vid}): {exc}")
            all_results.append(vendor_result)

        # ── Step 3: Build summary ────────────────────────────────────────
        return jsonify({
            'status': 'ok',
            'total_req_numbers': len(req_set),
            'vendors_processed': len(vendor_groups),
            'processed': all_processed,
            'not_found': all_not_found,
            'without_vendor': req_without_vendor,
            'without_credentials': req_without_credentials,
            'vendor_results': all_results,
            'errors': all_errors,
            'summary': {
                'total': len(req_set),
                'processed': len(all_processed),
                'not_found': len(all_not_found),
                'without_vendor': len(req_without_vendor),
                'without_credentials': len(req_without_credentials),
                'errors': len(all_errors),
            },
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/item-registration/template', methods=['GET'])
def download_item_registration_batch_template():
    try:
        ensure_default_item_registration_loaded()
        refresh_item_registration_mappings()
        rows = apply_item_registration_request_filters(ItemRegistration.query).order_by(ItemRegistration.uploaded_at.desc(), ItemRegistration.id.asc()).all()
        wb = Workbook(); ws = wb.active; ws.title = "Item Reg Batch Upload"
        headers = ['Req. No', 'Remarks']
        ws.append(headers); ws.freeze_panes = 'A2'
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        col_widths = [28, 70]
        for i, cell in enumerate(ws[1], 1):
            cell.fill = header_fill; cell.font = Font(bold=True, color="000000"); cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(i)].width = col_widths[i - 1]
        grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        red_font = Font(color="FF0000")
        ws.append(['example : 100010723616', 'example : Waiting for product registration'])
        for cell in ws[2]:
            cell.font = red_font; cell.fill = grey_fill
        seen = set()
        for row in rows:
            req_no = clean(row.req_no)
            if not req_no or req_no in seen: continue
            seen.add(req_no)
            ws.append([req_no, row.remarks or ''])
        output = io.BytesIO(); wb.save(output); output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f"Template_ItemRegistration_BatchUpload_{datetime.now().strftime('%Y%m%d')}.xlsx")
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/item-registration', methods=['GET'])
def export_item_registration():
    try:
        ensure_default_item_registration_loaded()
        refresh_item_registration_mappings()
        rows = apply_item_registration_request_filters(ItemRegistration.query).order_by(ItemRegistration.uploaded_at.desc(), ItemRegistration.id.asc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Item Registration"
        headers = [
            'Proc. Status', 'Client Nm.', 'Op. Unit Nm.', 'Category', 'PIC', 'Req. No', 'Prod. ID',
            'Prod. Nm.', 'Spec.', 'Mfr. Nm.', 'Odr. Unit', 'Bid Except Type', 'Prod. Price', 'Curr.', 'Remarks'
        ]
        _style_wb(ws, headers, num_cols=[12])
        widths = [26, 34, 34, 24, 16, 18, 18, 28, 48, 24, 14, 28, 16, 12, 60]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        for row in rows:
            ws.append([
                row.proc_status or '',
                row.client_name or '',
                row.operation_unit_name or '',
                source_category_level1(row.category),
                row.pic or '',
                row.req_no or '',
                row.prod_id or '',
                row.prod_name or '',
                row.spec or '',
                row.mfr_name or '',
                row.odr_unit or '',
                row.bid_except_type or '',
                row.prod_price or 0,
                row.curr or '',
                row.remarks or '',
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"Item_Registration_{datetime.now().strftime('%Y%m%d')}.xlsx")
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/item-registration/batch-upload', methods=['POST'])
def batch_upload_item_registration():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file'}), 400
        file = request.files['file']
        df = pd.read_excel(file, engine='openpyxl', skiprows=[1])
        df.columns = [str(c).strip() for c in df.columns]
        col_req = find_column(df, ['Req. No', 'Req. No.', 'Request No', 'Request Number'])
        col_rem = find_column(df, ['Remarks', 'Remark'])
        if not col_req:
            return jsonify({'error': f'Column "Req. No" not found. Available: {df.columns.tolist()}'}), 400
        if not col_rem:
            return jsonify({'error': f'Column "Remarks" not found. Available: {df.columns.tolist()}'}), 400

        updated = 0
        not_found = 0
        for _, row in df.iterrows():
            req_no = clean(df_val(row, col_req))
            if not req_no or req_no.lower().startswith('example'):
                continue
            req_no = req_no.replace('example :', '').replace('example:', '').strip()
            matches = ItemRegistration.query.filter_by(req_no=req_no).all()
            if not matches:
                not_found += 1
                continue
            remarks = clean(df_val(row, col_rem)) or ''
            for item in matches:
                item.remarks = remarks
                updated += 1
        db.session.commit()
        return jsonify({'updated': updated, 'not_found': not_found})
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def _lookup_pic_by_category_id(category_id):
    """Return PIC name for a Master PIC category id, or None if not found."""
    return _lookup_pic_by_category(category_id, None)


def _lookup_pic(product_id_str):
    """Return PIC name for a product_id string, or None if not found."""
    if not product_id_str:
        return None
    pid = str(product_id_str).strip()
    prod = db.session.query(ProductIDDB).filter_by(product_id=pid).first()
    if not prod:
        return None
    return _lookup_pic_by_category(prod.category_id, prod.category_name)


@app.route('/api/upload/product-id-json', methods=['POST'])
@app.route('/api/upload/product-id', methods=['POST'])
def upload_product_id():
    """Upload Prod_ID Excel from SAP. Upserts product_id → category_id mapping."""
    try:
        uploads, upload_mode = request_upload_dataframes('product_id')
        if not uploads:
            return jsonify({'error': 'No file uploaded or JSON rows supplied'}), 400
        replace_existing = upload_replace_mode()

        cleanup_pre = cleanup_source_table_snapshot(
            ProductIDDB,
            'product_id',
            None,
            timestamp_fields=('updated_at',),
            delete_blank=True,
        )
        db.session.flush()
        added = updated = 0
        removed_duplicates = cleanup_pre.get('removed_duplicates', 0)
        removed_stale = cleanup_pre.get('removed_stale', 0)
        removed_blank = cleanup_pre.get('removed_blank', 0)
        latest_product_ids = set()
        pic_cache = {}  # category_id → pic_name

        expected = [
            ('product_id', 'Product ID'), ('category_id', 'Category ID'),
            ('category_name', 'Category Name'), ('product_name', 'Product Name'),
            ('product_status', 'Product Status'), ('specification', 'Specification'),
            ('manufacturer_name', 'Manufacturer Name'),
            ('order_unit', 'Order Unit'),
            ('hub_handling_check', 'HUB Handling Check'), ('tax_type', 'Tax Type'),
            ('registration_date', 'Registration Date'), ('product_registry_pic', 'Product Registry PIC')
        ]
        required = [('product_id', 'Product ID')]

        all_pids_in_upload = set()
        for upload in uploads:
            df = upload['df']
            df.columns = [str(c).strip() for c in df.columns]
            col = _product_id_columns(df)
            for _, row in df.iterrows():
                pid = clean_product_id(df_val(row, col['product_id']))
                if pid: all_pids_in_upload.add(pid)
        
        existing_pid_map = {}
        if all_pids_in_upload:
            existing_pid_rows = db.session.query(ProductIDDB).filter(ProductIDDB.product_id.in_(list(all_pids_in_upload))).all()
            existing_pid_map = {p.product_id: p for p in existing_pid_rows}

        for upload in uploads:
            df = upload['df']
            df.columns = [str(c).strip() for c in df.columns]
            col = _product_id_columns(df)
            validate_upload_columns(upload['filename'], 'Prod ID', col, expected, required)

            for _, row in df.iterrows():
                pid = clean_product_id(df_val(row, col['product_id']))
                if not pid:
                    continue
                latest_product_ids.add(pid)
                cat_id = normalize_category_id(df_val(row, col['category_id']))
                payload = {
                    'category_id': cat_id,
                    'category_name': clean(df_val(row, col['category_name'])),
                    'product_name': clean(df_val(row, col['product_name'])),
                    'product_status': clean(df_val(row, col['product_status'])),
                    'specification': clean(df_val(row, col['specification'])),
                    'manufacturer_name': clean(df_val(row, col['manufacturer_name'])),
                    # vendor_name intentionally omitted — source Excel has no
                    # Vendor column for product master data.
                    'order_unit': clean(df_val(row, col['order_unit'])),
                    'hub_handling_check': clean(df_val(row, col['hub_handling_check'])),
                    'tax_type': clean(df_val(row, col['tax_type'])),
                    'registration_date': parse_date(df_val(row, col['registration_date'])),
                    'product_registry_pic': clean(df_val(row, col['product_registry_pic'])),
                    'updated_at': datetime.utcnow(),
                }

                existing = existing_pid_map.get(pid)
                if existing:
                    for key, value in payload.items():
                        setattr(existing, key, value)
                    updated += 1
                else:
                    db.session.add(ProductIDDB(product_id=pid, **payload))
                    added += 1

        db.session.flush()
        cleanup_post = cleanup_source_table_snapshot(
            ProductIDDB,
            'product_id',
            latest_product_ids if replace_existing else None,
            timestamp_fields=('updated_at',),
            delete_blank=True,
        )
        removed_duplicates += cleanup_post.get('removed_duplicates', 0)
        removed_stale += cleanup_post.get('removed_stale', 0)
        removed_blank += cleanup_post.get('removed_blank', 0)

        db.session.commit()
        _pid_category_cache_invalidate()
        clear_runtime_caches()

        global _SIMILARITY_CACHE
        _SIMILARITY_CACHE = {}

        # Refresh PIC only for SO rows whose product_id is in this upload
        # batch — uses the preloaded prod_map internally, no N+1 queries.
        # Previously this loop called db.session.query(ProductIDDB).filter_by
        # per SO row (1,000 SO rows = 1,000 SQL queries).
        refreshed = _refresh_so_pic_names(product_ids_scope=latest_product_ids)

        return jsonify({
            'status': 'ok',
            'files': len(uploads),
            'mode': upload_mode,
            'added': added, 'updated': updated,
            'removed_duplicates': removed_duplicates,
            'removed_stale': removed_stale,
            'removed_blank': removed_blank,
            'so_pic_refreshed': refreshed,
            'total_in_db': db.session.query(ProductIDDB).count()
        })
    except ValueError as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        import traceback; traceback.print_exc()
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/upload/master-pic-json', methods=['POST'])
@app.route('/api/upload/master-pic', methods=['POST'])
def upload_master_pic():
    """Upload Master PIC Excel. Supports the 4-sheet template
    (By Category / By Client ID / By Vendor / By Bid Type) and processes
    each sheet according to its kind. Also still accepts the legacy
    single-sheet "Category Name | PIC | Update New PIC" format for backwards
    compatibility. After upserting all 4 tables, refreshes SO pic_name and
    Item Registration mappings so the new PIC assignments take effect."""
    try:
        # Use the all-sheets reader so we get every sheet from the workbook.
        uploads, upload_mode = request_upload_all_sheets('master_pic')
        if not uploads:
            return jsonify({'error': 'No file uploaded or JSON rows supplied'}), 400
        replace_existing = upload_replace_mode()

        # ── Per-kind counters ─────────────────────────────────────────────
        # Each kind tracks its own added/updated/unchanged so the response
        # shows the user exactly what changed in each table.
        stats = {
            'category':   {'added': 0, 'updated': 0, 'unchanged': 0, 'removed_duplicates': 0, 'removed_stale': 0, 'removed_blank': 0},
            'client':     {'added': 0, 'updated': 0, 'unchanged': 0, 'removed_duplicates': 0, 'removed_stale': 0, 'removed_blank': 0},
            'vendor':     {'added': 0, 'updated': 0, 'unchanged': 0, 'removed_duplicates': 0, 'removed_stale': 0, 'removed_blank': 0},
            'bid_type':   {'added': 0, 'updated': 0, 'unchanged': 0, 'removed_duplicates': 0, 'removed_stale': 0, 'removed_blank': 0},
        }
        latest_category_names = set()
        latest_client_ids = set()
        latest_vendor_ids = set()
        latest_bid_types = set()
        seen_kinds = set()

        # ── Pre-load existing records for each table ──────────────────────
        # Avoids N+1 queries inside the row loop.
        all_master_pics = db.session.query(MasterPIC).all()
        master_pic_by_norm_name = {}
        master_pic_by_cat_id = {}
        for m in all_master_pics:
            norm = normalize_category_name(m.category_name)
            if norm and norm not in master_pic_by_norm_name:
                master_pic_by_norm_name[norm] = m
            if m.category_id:
                master_pic_by_cat_id[m.category_id] = m

        all_client_pics = db.session.query(MasterClientPIC).all()
        client_pic_by_id = {}
        for m in all_client_pics:
            cid = normalize_client_id(m.client_id)
            if cid and cid not in client_pic_by_id:
                client_pic_by_id[cid] = m

        all_vendor_pics = db.session.query(MasterVendorPIC).all()
        vendor_pic_by_id = {}
        for m in all_vendor_pics:
            vid = normalize_vendor_id(m.vendor_id)
            if vid and vid not in vendor_pic_by_id:
                vendor_pic_by_id[vid] = m

        all_bid_type_pics = db.session.query(MasterBidTypePIC).all()
        bid_type_pic_by_name = {}
        for m in all_bid_type_pics:
            bt = clean(m.bid_type)
            if bt and bt not in bid_type_pic_by_name:
                bid_type_pic_by_name[bt] = m

        # ── Pre-cleanup (replace mode only) ───────────────────────────────
        # When replace=true, we delete ALL existing rows in each table that's
        # present in the upload. Non-replace mode keeps existing rows and
        # just upserts.
        if replace_existing:
            # Determine which kinds are present in the upload so we only
            # wipe the corresponding table.
            upload_kinds = set()
            for upload in uploads:
                kind = _detect_master_pic_sheet_kind(upload.get('sheet_name'), upload['df'])
                if kind:
                    upload_kinds.add(kind)
            if 'category' in upload_kinds:
                stats['category']['removed_duplicates'] = cleanup_master_pic_by_category_name(None).get('removed_duplicates', 0)
            if 'client' in upload_kinds:
                stats['client']['removed_duplicates'] = db.session.query(MasterClientPIC).delete(synchronize_session=False) or 0
                client_pic_by_id = {}
            if 'vendor' in upload_kinds:
                stats['vendor']['removed_duplicates'] = db.session.query(MasterVendorPIC).delete(synchronize_session=False) or 0
                vendor_pic_by_id = {}
            if 'bid_type' in upload_kinds:
                stats['bid_type']['removed_duplicates'] = db.session.query(MasterBidTypePIC).delete(synchronize_session=False) or 0
                bid_type_pic_by_name = {}
            db.session.flush()

        # ── Process each uploaded sheet ───────────────────────────────────
        for upload in uploads:
            df = upload['df']
            df.columns = [str(c).strip() for c in df.columns]
            sheet_name = upload.get('sheet_name') or ''
            kind = _detect_master_pic_sheet_kind(sheet_name, df)
            if not kind:
                # Skip sheets we can't identify (e.g. an empty "Sheet1" the
                # user forgot to delete). Don't error — just skip silently.
                continue
            seen_kinds.add(kind)

            if kind == 'category':
                col = _master_pic_columns(df)
                if not col.get('category_name'):
                    continue  # invalid sheet, skip
                for _, row in df.iterrows():
                    cat_name = source_category_level1(df_val(row, col['category_name']))
                    if not cat_name:
                        continue
                    current_pic = clean(df_val(row, col['pic']))
                    update_pic = clean(df_val(row, col['pic_update']))
                    pic_name = update_pic or current_pic
                    if not pic_name:
                        continue
                    latest_category_names.add(cat_name)
                    norm_cat_name = normalize_category_name(cat_name)
                    generated_key = master_pic_category_key(cat_name)
                    existing = master_pic_by_cat_id.get(generated_key) or master_pic_by_norm_name.get(norm_cat_name)
                    if existing:
                        changed = (clean(existing.pic_name) != pic_name
                                   or normalize_category_name(existing.category_name) != norm_cat_name
                                   or (str(existing.category_id or '').startswith('CATNAME_') and existing.category_id != generated_key))
                        existing.category_name = cat_name
                        if str(existing.category_id or '').startswith('CATNAME_'):
                            existing.category_id = generated_key
                        existing.pic_name = pic_name
                        existing.updated_at = datetime.utcnow()
                        if changed: stats['category']['updated'] += 1
                        else: stats['category']['unchanged'] += 1
                    else:
                        new_rec = MasterPIC(category_id=generated_key, category_name=cat_name, pic_name=pic_name, updated_at=datetime.utcnow())
                        db.session.add(new_rec)
                        stats['category']['added'] += 1
                        master_pic_by_norm_name[norm_cat_name] = new_rec
                        master_pic_by_cat_id[generated_key] = new_rec

            elif kind == 'client':
                col = _master_client_pic_columns(df)
                if not col.get('client_id'):
                    continue
                for _, row in df.iterrows():
                    cid_raw = clean(df_val(row, col['client_id']))
                    if not cid_raw:
                        continue
                    cid = normalize_client_id(cid_raw)
                    cname = clean(df_val(row, col['client_name']))
                    pic_name = clean(df_val(row, col['pic']))
                    if not pic_name:
                        continue
                    latest_client_ids.add(cid)
                    existing = client_pic_by_id.get(cid)
                    if existing:
                        changed = (clean(existing.pic_name) != pic_name or clean(existing.client_name) != cname)
                        existing.client_name = cname
                        existing.pic_name = pic_name
                        existing.updated_at = datetime.utcnow()
                        if changed: stats['client']['updated'] += 1
                        else: stats['client']['unchanged'] += 1
                    else:
                        new_rec = MasterClientPIC(client_id=cid, client_name=cname, pic_name=pic_name, updated_at=datetime.utcnow())
                        db.session.add(new_rec)
                        stats['client']['added'] += 1
                        client_pic_by_id[cid] = new_rec

            elif kind == 'vendor':
                col = _master_vendor_pic_columns(df)
                if not col.get('vendor_id'):
                    continue
                for _, row in df.iterrows():
                    vid_raw = clean(df_val(row, col['vendor_id']))
                    if not vid_raw:
                        continue
                    vid = normalize_vendor_id(vid_raw)
                    vname = clean(df_val(row, col['vendor_name']))
                    pic_name = clean(df_val(row, col['pic']))
                    if not pic_name:
                        continue
                    latest_vendor_ids.add(vid)
                    existing = vendor_pic_by_id.get(vid)
                    if existing:
                        changed = (clean(existing.pic_name) != pic_name or clean(existing.vendor_name) != vname)
                        existing.vendor_name = vname
                        existing.pic_name = pic_name
                        existing.updated_at = datetime.utcnow()
                        if changed: stats['vendor']['updated'] += 1
                        else: stats['vendor']['unchanged'] += 1
                    else:
                        new_rec = MasterVendorPIC(vendor_id=vid, vendor_name=vname, pic_name=pic_name, updated_at=datetime.utcnow())
                        db.session.add(new_rec)
                        stats['vendor']['added'] += 1
                        vendor_pic_by_id[vid] = new_rec

            elif kind == 'bid_type':
                col = _master_bid_type_pic_columns(df)
                if not col.get('bid_type'):
                    continue
                for _, row in df.iterrows():
                    bt = clean(df_val(row, col['bid_type']))
                    if not bt:
                        continue
                    pic_name = clean(df_val(row, col['pic']))
                    if not pic_name:
                        continue
                    latest_bid_types.add(bt)
                    existing = bid_type_pic_by_name.get(bt)
                    if existing:
                        changed = (clean(existing.pic_name) != pic_name)
                        existing.pic_name = pic_name
                        existing.updated_at = datetime.utcnow()
                        if changed: stats['bid_type']['updated'] += 1
                        else: stats['bid_type']['unchanged'] += 1
                    else:
                        new_rec = MasterBidTypePIC(bid_type=bt, pic_name=pic_name, updated_at=datetime.utcnow())
                        db.session.add(new_rec)
                        stats['bid_type']['added'] += 1
                        bid_type_pic_by_name[bt] = new_rec

        db.session.flush()

        # ── Post-cleanup (replace mode only) ──────────────────────────────
        # Remove rows that weren't in the upload for each kind that was
        # actually uploaded (replace mode = "snapshot" of what's in the file).
        if replace_existing:
            if 'category' in seen_kinds and latest_category_names:
                # cleanup_master_pic_by_category_name removes rows whose
                # normalized category name is NOT in latest_category_names.
                post = cleanup_master_pic_by_category_name(latest_category_names)
                stats['category']['removed_stale'] = post.get('removed_stale', 0)
                stats['category']['removed_blank'] = post.get('removed_blank', 0)
            if 'client' in seen_kinds and latest_client_ids:
                # We already wiped MasterClientPIC above; remove rows in DB
                # whose client_id is not in the upload. Since we wiped and
                # re-inserted, this is effectively a no-op, but we run it
                # anyway for safety in case the wipe was rolled back.
                stale = db.session.query(MasterClientPIC).filter(~MasterClientPIC.client_id.in_(list(latest_client_ids))).delete(synchronize_session=False)
                stats['client']['removed_stale'] = int(stale or 0)
            if 'vendor' in seen_kinds and latest_vendor_ids:
                stale = db.session.query(MasterVendorPIC).filter(~MasterVendorPIC.vendor_id.in_(list(latest_vendor_ids))).delete(synchronize_session=False)
                stats['vendor']['removed_stale'] = int(stale or 0)
            if 'bid_type' in seen_kinds and latest_bid_types:
                stale = db.session.query(MasterBidTypePIC).filter(~MasterBidTypePIC.bid_type.in_(list(latest_bid_types))).delete(synchronize_session=False)
                stats['bid_type']['removed_stale'] = int(stale or 0)

        db.session.commit()
        invalidate_master_pic_cache()

        # Master PIC change can affect ANY SO row, so we do a full refresh.
        # Uses preloaded prod_map + master_pic_maps internally (no N+1).
        refreshed = _refresh_so_pic_names()
        refresh_item_registration_mappings()

        # Aggregate totals for the response.
        total_added = sum(s['added'] for s in stats.values())
        total_updated = sum(s['updated'] for s in stats.values())
        total_unchanged = sum(s['unchanged'] for s in stats.values())
        total_removed_dup = sum(s['removed_duplicates'] for s in stats.values())
        total_removed_stale = sum(s['removed_stale'] for s in stats.values())
        total_removed_blank = sum(s['removed_blank'] for s in stats.values())

        return jsonify({
            'status': 'ok',
            'files': len(uploads),
            'mode': upload_mode,
            'replace': replace_existing,
            'added': total_added,
            'updated': total_updated,
            'unchanged': total_unchanged,
            'removed_duplicates': total_removed_dup,
            'removed_stale': total_removed_stale,
            'removed_blank': total_removed_blank,
            'so_pic_refreshed': refreshed,
            'total_categories': master_pic_unique_category_count(),
            # Per-kind breakdown so the user can see exactly what changed.
            'by_kind': stats,
            'sheets_processed': sorted(seen_kinds),
        })
    except ValueError as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        import traceback; traceback.print_exc()
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/master-pic/status', methods=['GET'])
def master_pic_status():
    try:
        total_pid = db.session.query(ProductIDDB).count()
        last_pid = db.session.query(func.max(ProductIDDB.updated_at)).scalar()
        total_pic = master_pic_unique_category_count()
        last_pic = db.session.query(func.max(MasterPIC.updated_at)).scalar()
        return jsonify({
            'product_id_count': total_pid,
            # FIX V7: pakai utc_isoformat() supaya timestamp punya +00:00 suffix
            # Sebelumnya: last_pid.isoformat() → "2026-06-25T01:04:00" (tanpa tz)
            # Frontend interpret sebagai local time → jam tampil salah 7 jam
            # Sekarang: "2026-06-25T01:04:00+00:00" → frontend bisa konversi ke WIB
            'last_product_id_upload': utc_isoformat(last_pid),
            'master_pic_count': total_pic,
            'last_pic_update': utc_isoformat(last_pic),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/vendor-control/debug', methods=['GET'])
def vendor_control_debug():
    try:
        raw_file = os.environ.get('GOOGLE_SERVICE_ACCOUNT_FILE') or os.environ.get('GOOGLE_APPLICATION_CREDENTIALS')
        raw_json = bool(os.environ.get('GOOGLE_SERVICE_ACCOUNT_JSON') or os.environ.get('GOOGLE_SHEETS_SERVICE_ACCOUNT_JSON'))
        info = {
            'credential_file': raw_file or '',
            'credential_file_exists': bool(raw_file and os.path.exists(raw_file)),
            'credential_json_env_set': raw_json,
            'sheet_id': VENDOR_CONTROL_SHEET_ID,
            'sheet_gid': VENDOR_CONTROL_SHEET_GID,
        }
        sheet_name = vendor_control_sheet_name()
        info['sheet_name'] = sheet_name
        result = google_sheets_values_get(VENDOR_CONTROL_SHEET_ID, f"'{sheet_name}'!A1:Z20")
        values = result.get('values', [])
        info['sample_rows'] = len(values)
        info['header_candidates'] = []
        matched = None
        for idx, candidate_headers in enumerate(values[:20]):
            candidate_columns = find_vendor_control_columns(candidate_headers)
            looks_like_header = all(candidate_columns.get(name) for name in ('vendor_name', 'vendor_id', 'password'))
            info['header_candidates'].append({
                'row': idx + 1,
                'non_empty_cells': sum(1 for cell in candidate_headers if clean(cell)),
                'detected_columns': candidate_columns,
                'looks_like_header': looks_like_header,
            })
            if looks_like_header:
                matched = {'header_row': idx + 1, 'columns': candidate_columns, 'headers': candidate_headers}
                break
        info['matched_header'] = matched
        return jsonify(info)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/vendor-control/data', methods=['GET'])
def get_vendor_control_data():
    try:
        page = max(int(request.args.get('page', 1)), 1)
        per_page = min(max(int(request.args.get('per_page', 10)), 1), 500)
        search = (clean(request.args.get('search')) or '').lower()
        vendors = [clean(v) for v in request.args.getlist('vendor') if clean(v)]
        force = str(request.args.get('refresh', '')).lower() in ('1', 'true', 'yes')
        rows, fetched_at = vendor_control_rows(force=force)
        if vendors:
            vendor_needles = [v.lower() for v in vendors]
            rows = [row for row in rows if any(
                needle == str(row.get('vendor_name') or '').lower()
                or needle == str(row.get('vendor_id') or '').lower()
                or needle in str(row.get('vendor_name') or '').lower()
                or needle in str(row.get('vendor_id') or '').lower()
                for needle in vendor_needles
            )]
        if search:
            rows = [row for row in rows if search in str(row.get('vendor_name') or '').lower() or search in str(row.get('vendor_id') or '').lower()]
        rows = sorted(rows, key=lambda row: (str(row.get('vendor_name') or '').lower(), str(row.get('vendor_id') or '').lower()))
        total = len(rows)
        start = (page - 1) * per_page
        return jsonify({
            'data': rows[start:start + per_page],
            'total': total,
            'page': page,
            'per_page': per_page,
            'suggestions': [row.get('vendor_name') for row in rows[:20] if row.get('vendor_name')],
            'last_updated': utc_isoformat(fetched_at),
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({
            'error': str(e),
            'hint': 'Check PythonAnywhere WSGI credential path, google-api-python-client/google-auth installation, service account sheet permission, and Vendor Control sheet headers.'
        }), 500

@app.route('/api/vendor-control/<path:row_key>', methods=['PUT'])
def update_vendor_control(row_key):
    try:
        data = request.json or {}
        field = clean(data.get('field')) or ''
        value = clean(data.get('value')) or ''
        if field not in ('vendor_id', 'password'):
            return jsonify({'error': 'Only Vendor ID and Password can be edited'}), 400
        try:
            sheet_row = int(str(row_key).strip())
        except ValueError:
            return jsonify({'error': 'Invalid vendor row key'}), 400
        sync = sync_vendor_control_cell(sheet_row, field, value)
        if sync.get('synced'):
            for row in VENDOR_CONTROL_CACHE.get('rows') or []:
                if str(row.get('row_key')) == str(row_key):
                    row[field] = value
        return jsonify({'success': True, 'sheet_sync': sync})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/vendor-control/login/<path:row_key>', methods=['GET'])
def vendor_control_login(row_key):
    try:
        rows, _ = vendor_control_rows(force=False)
        row = next((item for item in rows if str(item.get('row_key')) == str(row_key)), None)
        if not row:
            return '<h3>Vendor credential was not found or incomplete.</h3>', 404
        vendor_id = row.get('vendor_id') or ''
        password = row.get('password') or ''
        action = 'https://mall.serveone.id/vendor/cmm/doLogin.dev?signData=noSign'
        if vendor_id.upper().startswith('FW'):
            action = 'https://mall.serveone.id/vendor/fwdr/fwdr/doChkFirstLogin.dev?mallType=FORWARDER'
        vendor_name = row.get('vendor_name') or vendor_id
        return f'''<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Vendor Login - {html.escape(str(vendor_name))}</title>
  <style>
    body {{ font-family: Arial, sans-serif; background: #f8fafc; color: #0f172a; display: grid; min-height: 100vh; place-items: center; margin: 0; }}
    .box {{ background: white; border: 1px solid #e2e8f0; border-radius: 14px; padding: 24px; width: min(420px, calc(100vw - 32px)); box-shadow: 0 20px 50px rgba(15,23,42,.12); }}
    h1 {{ font-size: 18px; margin: 0 0 6px; }}
    p {{ color: #475569; font-size: 13px; line-height: 1.5; margin: 0 0 18px; }}
    button {{ width: 100%; border: 0; border-radius: 10px; background: #2563eb; color: white; font-weight: 700; padding: 12px; cursor: pointer; }}
  </style>
</head>
<body>
  <div class="box">
    <h1>Logging in to {html.escape(str(vendor_name))}</h1>
    <p>This tab will submit the vendor login form automatically. If it does not continue, click the button below.</p>
    <form id="vendorLoginForm" method="post" action="{html.escape(action)}">
      <input type="hidden" name="cprtcpUsrId" value="{html.escape(str(vendor_id))}">
      <input type="hidden" name="cprtcpSectNo" value="{html.escape(str(password))}">
      <input type="hidden" name="agreType" value="">
      <input type="hidden" name="signData" value="noSign">
      <button type="submit">Log-In</button>
    </form>
  </div>
  <script>setTimeout(function(){{ document.getElementById('vendorLoginForm').submit(); }}, 250);</script>
</body>
</html>'''
    except Exception as e:
        import traceback; traceback.print_exc()
        return f'<h3>Vendor login failed to prepare.</h3><p>{html.escape(str(e))}</p>', 500

@app.route('/api/rfq/data', methods=['GET'])
def get_rfq_data():
    try:
        force = str(request.args.get('refresh', '')).lower() in ('1', 'true', 'yes')
        cache_key = runtime_cache_key('rfq_data')
        if not force:
            cached = runtime_cache_get(cache_key)
            if cached is not None:
                return jsonify(cached)

        page = max(int(request.args.get('page', 1)), 1)
        per_page = min(max(int(request.args.get('per_page', 10)), 1), 500)
        search = clean(request.args.get('search')) or ''
        pic = clean(request.args.get('pic')) or ''
        clients = [clean(v) for v in request.args.getlist('client_name') if clean(v)]
        rfq_numbers = [clean(v) for v in request.args.getlist('rfq_no') if clean(v)]
        brands = [clean(v) for v in request.args.getlist('brand_manufacturer') if clean(v)]
        purchase_pics = [clean(v) for v in request.args.getlist('purchase_pic') if clean(v)]
        vendors = [clean(v) for v in request.args.getlist('vendor_name') if clean(v)]
        checks = [clean(v).lower() for v in request.args.getlist('check') if clean(v)]
        include_similarity = str(request.args.get('similarity', '')).lower() in ('1', 'true', 'yes')

        rows, fetched_at = rfq_rows_with_edits(force=force)
        if search:
            rows = filter_rfq_rows_by_multiline_search(rows, search)

        search_rows = list(rows)

        def rfq_filter_rows(source_rows, exclude_field=None):
            excluded = set(exclude_field or []) if isinstance(exclude_field, (set, list, tuple)) else ({exclude_field} if exclude_field else set())
            result = list(source_rows)
            if 'clients' not in excluded and clients:
                result = [row for row in result if clean(row.get('client_name')) in clients]
            if 'rfq_numbers' not in excluded and rfq_numbers:
                result = [row for row in result if clean(row.get('rfq_code')) in rfq_numbers]
            if 'brands' not in excluded and brands:
                result = [row for row in result if clean(row.get('brand_manufacturer')) in brands]
            if 'vendors' not in excluded and vendors:
                result = [row for row in result if clean(row.get('vendor_name')) in vendors]
            if 'checks' not in excluded and checks:
                result = [row for row in result if clean(row.get('check')) and clean(row.get('check')).lower() in checks]
            if 'purchase_pics' not in excluded and purchase_pics:
                result = [row for row in result if clean(row.get('purchase_pic')) in purchase_pics]
            if 'pic' not in excluded and pic:
                result = [
                    row for row in result
                    if clean(row.get('purchase_pic')) == pic
                    and clean(row.get('check')) not in ('complete', 'reject')
                    and row.get('unit_price_missing')
                    and not clean_product_id(row.get('product_id'))
                ]
            return result

        filtered_cache = {}
        def filtered_for(exclude_field=None):
            if isinstance(exclude_field, (set, list, tuple)):
                key = tuple(sorted(str(x) for x in exclude_field))
            elif exclude_field:
                key = (str(exclude_field),)
            else:
                key = ()
            if key not in filtered_cache:
                filtered_cache[key] = rfq_filter_rows(search_rows, exclude_field)
            return filtered_cache[key]

        kpi_rows = filtered_for('pic')
        all_rfq_pics = set()
        for r in search_rows:
            p = clean(r.get('purchase_pic'))
            if p and p.lower() != 'unassigned': all_rfq_pics.add(p)
        pending_by_pic = {p: 0 for p in all_rfq_pics}
        for row in kpi_rows:
            if clean_product_id(row.get('product_id')):
                continue
            check_val = clean(row.get('check', '')).lower()
            if check_val in ('complete', 'reject'):
                continue
            if not row.get('unit_price_missing'):
                continue
            row_pic = clean(row.get('purchase_pic'))
            if not row_pic or row_pic.lower() == 'unassigned':
                continue
            pending_by_pic[row_pic] = pending_by_pic.get(row_pic, 0) + 1
        pic_kpis = [{'pic': key, 'count': val} for key, val in sorted(pending_by_pic.items(), key=lambda item: (-item[1], item[0]))]

        rows = filtered_for()

        total = len(rows)
        start = (page - 1) * per_page
        page_rows = [dict(row) for row in rows[start:start + per_page]]
        if include_similarity:
            page_rows = [apply_rfq_similarity(row) for row in page_rows]
            save_similarity_cache()

        clients_rows = filtered_for('clients')
        rfq_no_rows = filtered_for('rfq_numbers')
        brand_rows = filtered_for('brands')
        purchase_pic_rows = filtered_for('purchase_pics')
        vendor_rows = filtered_for('vendors')
        check_rows = filtered_for('checks')
        available_checks = {clean(row.get('check')).lower() for row in check_rows if clean(row.get('check'))}

        payload = {
            'data': page_rows,
            'total': total,
            'page': page,
            'per_page': per_page,
            'columns': [{'field': field, 'label': label} for field, label in RFQ_TEMPLATE_COLUMNS],
            'similarity_columns': [{'field': field, 'label': label} for field, label in RFQ_SIMILARITY_COLUMNS],
            'editable_fields': sorted(RFQ_EDITABLE_FIELDS),
            'pic_kpis': pic_kpis,
            'filters': {
                'clients': sorted({clean(row.get('client_name')) for row in clients_rows if clean(row.get('client_name'))}),
                'rfq_numbers': sorted({clean(row.get('rfq_code')) for row in rfq_no_rows if clean(row.get('rfq_code'))}),
                'brands': sorted({clean(row.get('brand_manufacturer')) for row in brand_rows if clean(row.get('brand_manufacturer'))}),
                'purchase_pics': sorted({clean(row.get('purchase_pic')) for row in purchase_pic_rows if clean(row.get('purchase_pic')) and clean(row.get('purchase_pic')).lower() != 'unassigned'}),
                'vendors': sorted({clean(row.get('vendor_name')) for row in vendor_rows if clean(row.get('vendor_name'))}),
                'checks': [rfq_check_label(key) for key in ['complete', 'reject', 'closed', 'open'] if key in available_checks],
            },
            'last_updated': utc_isoformat(fetched_at),
        }
        runtime_cache_set(cache_key, payload, ttl_seconds=180)
        return jsonify(payload)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

def rfq_filtered_rows_from_request(force=False):
    search = clean(request.args.get('search')) or ''
    pic = clean(request.args.get('pic')) or ''
    clients = [clean(v) for v in request.args.getlist('client_name') if clean(v)]
    rfq_numbers = [clean(v) for v in request.args.getlist('rfq_no') if clean(v)]
    brands = [clean(v) for v in request.args.getlist('brand_manufacturer') if clean(v)]
    purchase_pics = [clean(v) for v in request.args.getlist('purchase_pic') if clean(v)]
    vendors = [clean(v) for v in request.args.getlist('vendor_name') if clean(v)]
    checks = [clean(v).lower() for v in request.args.getlist('check') if clean(v)]
    rows, fetched_at = rfq_rows_with_edits(force=force)
    if search:
        rows = filter_rfq_rows_by_multiline_search(rows, search)
    if clients:
        rows = [row for row in rows if clean(row.get('client_name')) in clients]
    if rfq_numbers:
        rows = [row for row in rows if clean(row.get('rfq_code')) in rfq_numbers]
    if brands:
        rows = [row for row in rows if clean(row.get('brand_manufacturer')) in brands]
    if purchase_pics:
        rows = [row for row in rows if clean(row.get('purchase_pic')) in purchase_pics]
    if vendors:
        rows = [row for row in rows if clean(row.get('vendor_name')) in vendors]
    if checks:
        rows = [row for row in rows if clean(row.get('check')) and clean(row.get('check')).lower() in checks]
    if pic:
        rows = [row for row in rows
                if clean(row.get('purchase_pic')) == pic
                and clean(row.get('check')) not in ('complete', 'reject')
                and row.get('unit_price_missing')
                and not clean_product_id(row.get('product_id'))]
    return rows, fetched_at

@app.route('/api/rfq/template', methods=['GET'])
def download_rfq_batch_template():
    try:
        rows, _ = rfq_filtered_rows_from_request(force=False)
        wb = Workbook()
        ws = wb.active
        ws.title = 'RFQ Batch Upload'

        # Context fields (white columns – pre-filled, read-only reference)
        context_fields = ['client_name', 'purchase_pic', 'item_name', 'detail_spec', 'brand_manufacturer', 'qty', 'unit', 'remark']
        # Upload fields (blue columns – user fills in)
        upload_fields = ['same_replacement', 'vendor_name', 'unit_price_idr', 'quoted_item_name', 'quoted_spec', 'quoted_brand', 'quoted_unit', 'moq', 'lead_time_days', 'valid_period', 'photo_url', 'remarks']

        headers = ['No'] + [rfq_label(field) for field in context_fields] + [rfq_label(field) for field in upload_fields]
        num_context = 1 + len(context_fields)  # No + context fields
        num_upload = len(upload_fields)
        total_cols = num_context + num_upload

        thin_border = Border(
            left=Side(style='thin', color='D9E2EF'),
            right=Side(style='thin', color='D9E2EF'),
            top=Side(style='thin', color='D9E2EF'),
            bottom=Side(style='thin', color='D9E2EF'),
        )

        # --- Row 1: header row (MUST be row 1 so pandas reads column names correctly) ---
        for i, hdr in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=hdr)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            if i <= num_context:
                cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
                cell.font = Font(bold=True, color='000000')
            else:
                cell.fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')

        # Add instruction as a comment on the first blue header cell so pandas doesn't see it
        from openpyxl.comments import Comment
        first_blue_cell = ws.cell(row=1, column=num_context + 1)
        first_blue_cell.comment = Comment(
            'Silahkan isi penawaran di Kolom Biru / Kindly fill in your quotation in the blue columns',
            'System'
        )

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(total_cols)}1'

        # --- Data rows (starting row 2) ---
        ref_body_fill = PatternFill(start_color='EDEDED', end_color='EDEDED', fill_type='solid')
        blue_body_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')

        seen = set()
        for row in rows:
            no = clean(row.get('no'))
            if not no or no in seen:
                continue
            seen.add(no)
            data_row = [no] + [row.get(field) or '' for field in context_fields] + [row.get(field) or '' for field in upload_fields]
            ws.append(data_row)
            row_idx = ws.max_row
            for ci in range(1, total_cols + 1):
                cell = ws.cell(row=row_idx, column=ci)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = thin_border
                if ci <= num_context:
                    cell.fill = ref_body_fill
                else:
                    cell.fill = blue_body_fill

        # Column widths matching the image layout
        widths = [8, 18, 16, 28, 50, 20, 10, 10, 28, 20, 28, 18, 28, 18, 18, 10, 14, 18, 18, 28, 28]
        for i, width in enumerate(widths[:total_cols], 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 30
        ws.row_dimensions[1].height = 26

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Template_RFQ_BatchUpload_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# Fields that appear as blue/upload columns in the batch upload template.
# Only these fields are processed during batch upload – context (white) columns are ignored.
RFQ_TEMPLATE_UPLOAD_FIELDS = [
    'same_replacement', 'vendor_name', 'unit_price_idr', 'quoted_item_name',
    'quoted_spec', 'quoted_brand', 'quoted_unit', 'moq', 'lead_time_days',
    'valid_period', 'photo_url', 'remarks',
]

@app.route('/api/rfq/batch-upload-json', methods=['POST'])
@app.route('/api/rfq/batch-upload', methods=['POST'])
def batch_upload_rfq():
    try:
        uploads, upload_mode = request_upload_dataframes('rfq_batch')
        if not uploads:
            return jsonify({'error': 'No file uploaded or JSON rows supplied'}), 400
        rows, _ = rfq_rows_with_edits(force=False)
        no_map = {}
        row_by_key = {}
        for row in rows:
            row_by_key[row['row_key']] = row
            no = clean(row.get('no'))
            if no:
                no_map.setdefault(no, []).append(row['row_key'])

        updated = 0
        skipped_empty = 0
        not_found = 0
        sheet_updates = []
        local_updates = 0
        for upload in uploads:
            df = upload['df']
            df.columns = [str(c).strip() for c in df.columns]
            col_no = find_column(df, ['No'])
            if not col_no:
                return jsonify({'error': f'Column "No" not found. Available: {df.columns.tolist()}'}), 400
            # Only map blue/upload column fields – context columns are read-only
            col_map = {field: find_column(df, [rfq_label(field), field]) for field in RFQ_TEMPLATE_UPLOAD_FIELDS}
            for _, src in df.iterrows():
                no = clean(df_val(src, col_no))
                if not no or no.lower().startswith('example'):
                    continue
                keys = no_map.get(no, [])
                if not keys:
                    not_found += 1
                    continue
                for field, col in col_map.items():
                    if not col:
                        continue
                    value = clean(df_val(src, col))
                    # Skip empty values – don't overwrite existing data with blanks
                    if not value:
                        skipped_empty += 1
                        continue
                    for row_key in keys:
                        base_row = row_by_key.get(row_key)
                        if base_row:
                            set_rfq_dashboard_cell(row_key, field, str(value), dirty=True, commit=False)
                            sheet_updates.append({'row': base_row, 'field': field, 'value': str(value)})
                        updated += 1

        db.session.commit()

        try:
            sheet_sync = sync_rfq_cells_to_google_sheet(sheet_updates) if sheet_updates else {'synced': True, 'updated_ranges': 0}
            if sheet_sync.get('synced'):
                clear_rfq_dashboard_dirty_fields(sheet_updates)
        except Exception as sync_error:
            sheet_sync = {'synced': False, 'reason': str(sync_error)}

        clear_runtime_caches()
        return jsonify({
            'updated': updated,
            'sheet_updates': len(sheet_updates),
            'local_updates': local_updates,
            'skipped_empty': skipped_empty,
            'not_found': not_found,
            'files': len(uploads),
            'mode': upload_mode,
            'sheet_sync': sheet_sync,
        })
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

def _style_rfq_export_sheet(ws, headers, editable_start_col=19):
    last_col = len(headers)
    last_col_letter = get_column_letter(last_col)
    ref_end_col = editable_start_col - 1

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{last_col_letter}{ws.max_row}'

    ref_header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ref_body_fill = PatternFill(start_color='EDEDED', end_color='EDEDED', fill_type='solid')
    input_header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    note_font = Font(color='0070C0')
    ref_header_font = Font(bold=True, color='000000')
    input_header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin', color='D9E2EF'),
        right=Side(style='thin', color='D9E2EF'),
        top=Side(style='thin', color='D9E2EF'),
        bottom=Side(style='thin', color='D9E2EF'),
    )

    note_cell = ws.cell(row=1, column=editable_start_col)
    note_cell.value = 'Silahkan isi penawaran di Kolom Biru / Kindly fill in your quotation in the blue columns'
    note_cell.font = note_font

    for cell in ws[2]:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        if cell.column <= ref_end_col:
            cell.fill = ref_header_fill
            cell.font = ref_header_font
        else:
            cell.fill = input_header_fill
            cell.font = input_header_font

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = thin_border
            if cell.column <= ref_end_col:
                cell.fill = ref_body_fill

    widths = [
        10, 10, 10, 7, 55, 14, 18, 12, 28, 14, 24, 42, 20, 8, 8, 24, 18, 20,
        17, 18, 18, 18, 24, 42, 18, 12, 12, 18, 18, 28, 50, 32, 32
    ]
    for i, width in enumerate(widths[:last_col], 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for row_idx in range(3, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 30
    ws.row_dimensions[2].height = 26

@app.route('/api/export/rfq', methods=['GET'])
def export_rfq():
    try:
        rows, _ = rfq_filtered_rows_from_request(force=False)
        wb = Workbook()
        ws = wb.active
        ws.title = 'RFQ'
        headers = [label for _, label in RFQ_TEMPLATE_COLUMNS]

        ws.append([''] * len(headers))
        ws.append(headers)
        for row in rows:
            values = []
            for field, _label in RFQ_TEMPLATE_COLUMNS:
                if field == 'check':
                    values.append(rfq_check_label(row.get('check')))
                elif field == 'days_left':
                    values.append(row.get('days_left') if row.get('days_left') is not None else '-')
                else:
                    values.append(row.get(field) or '')
            ws.append(values)

        _style_rfq_export_sheet(ws, headers, editable_start_col=19)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'RFQ_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/rfq/batch-cells', methods=['PUT'])
def update_rfq_cells_batch():
    try:
        payload = request.get_json(silent=True) or {}
        updates = payload.get('updates') or []
        if not isinstance(updates, list) or not updates:
            return jsonify({'error': 'No RFQ cell updates supplied'}), 400
        if len(updates) > 1000:
            return jsonify({'error': 'Maximum 1000 cells can be updated at once'}), 400

        base_rows, _ = rfq_rows_with_edits(force=False, prefer_stale_cache=True)
        row_map = {row.get('row_key'): row for row in base_rows}
        sheet_updates = []
        updated = 0
        skipped = []

        for idx, item in enumerate(updates):
            row_key = clean(item.get('row_key'))
            field = clean(item.get('field'))
            value = item.get('value')
            if field not in RFQ_EDITABLE_FIELDS and field not in RFQ_DIRECT_UPDATE_FIELDS:
                skipped.append({'index': idx, 'reason': 'Field is not editable', 'field': field})
                continue
            base_row = row_map.get(row_key)
            if not base_row:
                skipped.append({'index': idx, 'reason': 'RFQ row not found', 'row_key': row_key})
                continue

            clean_value = clean_product_id(value) if field == 'product_id' else ('' if value is None else str(value))
            if field in RFQ_DASHBOARD_ONLY_FIELDS:
                edit = RFQCellEdit.query.filter_by(row_key=row_key, field=field).first()
                if not edit:
                    edit = RFQCellEdit(row_key=row_key, field=field)
                    db.session.add(edit)
                edit.value = clean_value
                edit.updated_at = datetime.utcnow()
                set_rfq_dashboard_cell(row_key, field, clean_value, dirty=False, commit=False)
            else:
                set_rfq_dashboard_cell(row_key, field, clean_value, dirty=True, commit=False)
                sheet_updates.append({'row': base_row, 'field': field, 'value': clean_value})
            updated += 1

        db.session.commit()
        clear_runtime_caches()
        try:
            sheet_sync = sync_rfq_cells_to_google_sheet(sheet_updates) if sheet_updates else {'synced': True, 'local_only': True}
            if sheet_sync.get('synced'):
                clear_rfq_dashboard_dirty_fields(sheet_updates)
        except Exception as sync_error:
            sheet_sync = {'synced': False, 'reason': str(sync_error)}
        return jsonify({'success': True, 'updated': updated, 'skipped': skipped, 'sheet_sync': sheet_sync})
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/rfq/<path:row_key>', methods=['PUT'])
def update_rfq_cell(row_key):
    try:
        payload = request.get_json(silent=True) or {}
        field = clean(payload.get('field'))
        value = payload.get('value')
        if field not in RFQ_EDITABLE_FIELDS and field not in RFQ_DIRECT_UPDATE_FIELDS:
            return jsonify({'error': 'Field is not editable'}), 400
        base_rows, _ = rfq_rows_with_edits(force=False, prefer_stale_cache=True)
        base_row = next((row for row in base_rows if row.get('row_key') == row_key), None)
        if not base_row:
            return jsonify({'error': 'RFQ row not found'}), 404
        clean_value = clean_product_id(value) if field == 'product_id' else ('' if value is None else str(value))
        if field in RFQ_DASHBOARD_ONLY_FIELDS:
            edit = RFQCellEdit.query.filter_by(row_key=row_key, field=field).first()
            if not edit:
                edit = RFQCellEdit(row_key=row_key, field=field)
                db.session.add(edit)
            edit.value = clean_value
            edit.updated_at = datetime.utcnow()
            set_rfq_dashboard_cell(row_key, field, clean_value, dirty=False, commit=False)
            db.session.commit()
            clear_runtime_caches()
            sheet_sync = {'synced': True, 'local_only': True}
        else:
            RFQCellEdit.query.filter_by(row_key=row_key, field=field).delete()
            db.session.commit()
            set_rfq_dashboard_cell(row_key, field, clean_value, dirty=True)
            try:
                sheet_sync = sync_rfq_cell_to_google_sheet(base_row, field, clean_value)
                if sheet_sync.get('synced'):
                    clear_rfq_dashboard_dirty_fields([{'row_key': row_key, 'field': field}])
            except Exception as sync_error:
                sheet_sync = {'synced': False, 'reason': str(sync_error)}
            clear_runtime_caches()
        return jsonify({'success': True, 'row_key': row_key, 'field': field, 'value': clean_value, 'sheet_sync': sheet_sync})
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

def import_sheet_title_for_gid(spreadsheet_id, gid):
    metadata = google_sheets_metadata(spreadsheet_id)
    gid_int = int(gid)
    for sheet in metadata.get('sheets', []):
        props = sheet.get('properties', {})
        if props.get('sheetId') == gid_int:
            return props.get('title')
    raise RuntimeError(f'Sheet gid {gid} not found')


def import_source_config(source_key):
    return next((s for s in IMPORT_SOURCE_SHEETS if s.get('key') == source_key), None)

def import_source_sheet_title(source):
    cache_key = ('import_source_sheet_title_v2', source.get('spreadsheet_id'), source.get('gid'))
    cached = runtime_cache_get(cache_key)
    if cached:
        return cached
    try:
        title, _preview_df = import_source_header_preview(source)
    except Exception:
        title = import_sheet_title_for_gid(source['spreadsheet_id'], source.get('gid') or '0')
    runtime_cache_set(cache_key, title, ttl_seconds=3600)
    return title

def import_source_map_for_sync(source):
    cache_key = ('import_source_map_for_sync_v2', source.get('spreadsheet_id'), source.get('gid'))
    cached = runtime_cache_get(cache_key)
    if cached is not None:
        return cached
    columns = import_all_mapping_columns(import_layout_columns())
    try:
        _sheet_title, df = import_source_header_preview(source)
    except Exception:
        df = read_public_sheet_csv(source['spreadsheet_id'], source.get('gid') or '0', nrows=60)
    source_map = import_source_column_map(df, columns)
    header_idx = import_detect_header_row(df)
    source_fallbacks = import_source_fallback_columns(df, header_idx)
    for field in ('po_yupi', 'yupi_po', 'source_req_dlv_date', 'req_dlv_date', 'po_date_by_email', 'site', 'po_sementara', 'item_yupi', 'item_name', 'spec', 'remark_yupi', 'reschedule', 'ord_qty', 'unit', 'unit_price', 'amount', 'vendor_name', 'vendor', 'purchase_price', 'currency', 'purchase_amount', 'so'):
        if field not in source_map and source_fallbacks.get(field):
            source_map[field] = column_index_from_letter(source_fallbacks[field]) - 1
    runtime_cache_set(cache_key, source_map, ttl_seconds=300)
    return source_map

def import_sheet_field_for_dashboard_field(field):
    return IMPORT_SYNC_FIELD_ALIASES.get(field, field)

def set_import_payload_field_aliases(data, field, value):
    data[field] = value
    if field == 'po_send_date':
        data['_po_send_date_manual'] = '1' if clean(value) else ''
    if field == 'yupi_po':
        data['po_yupi'] = value
    elif field == 'po_yupi':
        data['yupi_po'] = value
    elif field == 'req_dlv_date':
        data['source_req_dlv_date'] = value
    elif field == 'source_req_dlv_date':
        data['req_dlv_date'] = value
    return data

def sync_import_cells_to_source_sheets(items):
    grouped = {}
    skipped = 0
    for item in items:
        row = item.get('row')
        field = clean(item.get('field'))
        value = '' if item.get('value') is None else str(item.get('value'))
        if not row or not field:
            skipped += 1
            continue
        if field in IMPORT_DASHBOARD_LOCAL_FIELDS:
            skipped += 1
            continue
        source = import_source_config(row.source_key)
        if not source:
            skipped += 1
            continue
        sheet_field = import_sheet_field_for_dashboard_field(field)
        try:
            source_map = import_source_map_for_sync(source)
        except Exception:
            source_map = {}
        col_idx = source_map.get(sheet_field)
        if col_idx is None:
            skipped += 1
            continue
        try:
            sheet_title = import_source_sheet_title(source)
        except Exception:
            skipped += 1
            continue
        sheet_row = row.sheet_row
        if not sheet_row:
            skipped += 1
            continue
        col_letter = column_letter_from_index(col_idx + 1)
        spreadsheet_id = source['spreadsheet_id']
        grouped.setdefault(spreadsheet_id, []).append({
            'range': f"'{sheet_title}'!{col_letter}{sheet_row}",
            'values': [[value or '']],
        })
    if not grouped:
        return {'synced': False, 'reason': 'No mapped Import sheet cells to sync', 'skipped': skipped}
    total_ranges = 0
    for spreadsheet_id, ranges in grouped.items():
        google_sheets_values_batch_update(spreadsheet_id, ranges)
        total_ranges += len(ranges)
    return {'synced': True, 'ranges': total_ranges, 'spreadsheets': len(grouped), 'skipped': skipped}


IMPORT_LAYOUT_TARGET_FORMULA_FIELDS = {
    'days_left', 'site', 'yupi_po', 'vendor', 'req_dlv_date',
    'arrival_check', 'purchase_amount', 'lt_days',
}

def import_layout_target_sheet_title():
    cache_key = ('import_layout_target_sheet_title', IMPORT_LAYOUT_SHEET_ID, IMPORT_LAYOUT_GID)
    cached = runtime_cache_get(cache_key)
    if cached:
        return cached
    title = import_sheet_title_for_gid(IMPORT_LAYOUT_SHEET_ID, IMPORT_LAYOUT_GID)
    runtime_cache_set(cache_key, title, ttl_seconds=3600)
    return title

def import_layout_target_field_columns():
    mapping = {}
    for col in import_all_mapping_columns(import_layout_columns()):
        field = col.get('field')
        sheet_col = col.get('sheet_col')
        if field and sheet_col:
            mapping[field] = str(sheet_col).upper()
    return mapping

def import_layout_target_field_for_dashboard_field(field):
    if field == 'yupi_po':
        return 'po_yupi'
    if field == 'req_dlv_date':
        return 'source_req_dlv_date'
    return field

def import_layout_target_candidate_keys(data):
    data = data or {}
    po_sementara = clean(data.get('po_sementara'))
    so = clean(data.get('so'))
    item_yupi = clean(data.get('item_yupi'))
    po_yupi = clean(data.get('po_yupi')) or clean(data.get('yupi_po'))
    keys = []
    def add(prefix, *parts):
        parts = [clean(p) for p in parts]
        if all(parts):
            keys.append(prefix + ':' + '|'.join(p.strip().lower() for p in parts))
    add('po_sem_item_so', po_sementara, item_yupi, so)
    add('po_sem_item', po_sementara, item_yupi)
    add('po_yupi_item_so', po_yupi, item_yupi, so)
    add('po_yupi_item', po_yupi, item_yupi)
    add('so_item', so, item_yupi)
    add('po_sem', po_sementara)
    out = []
    for key in keys:
        if key not in out:
            out.append(key)
    return out

def import_layout_target_lookup(sheet_title):
    resp = google_sheets_values_get(
        IMPORT_LAYOUT_SHEET_ID,
        f"'{sheet_title}'!A2:S",
        value_render_option='FORMATTED_VALUE',
    )
    values = resp.get('values') or []
    by_key = {}
    max_row = 1
    def cell(row, one_based_idx):
        idx = one_based_idx - 1
        return row[idx] if idx < len(row) else ''
    for row_no, row_values in enumerate(values, start=2):
        if any(clean(v) for v in row_values):
            max_row = row_no
        row_data = {
            'yupi_po': cell(row_values, 5),
            'so': cell(row_values, 12),
            'po_sementara': cell(row_values, 17),
            'po_yupi': cell(row_values, 18),
            'item_yupi': cell(row_values, 19),
        }
        for key in import_layout_target_candidate_keys(row_data):
            by_key.setdefault(key, row_no)
    return {'by_key': by_key, 'max_row': max_row}

def sync_import_cells_to_layout_sheet(items):
    if not items:
        return {'synced': False, 'reason': 'No Import cells to sync'}

    field_columns = import_layout_target_field_columns()
    if not field_columns:
        return {'synced': False, 'reason': 'No Import tracker columns are mapped'}

    sheet_title = import_layout_target_sheet_title()
    lookup = import_layout_target_lookup(sheet_title)
    by_key = lookup['by_key']
    next_row = lookup['max_row'] + 1

    grouped = {}
    for item in items:
        row = item.get('row')
        field = clean(item.get('field'))
        if not row or not field:
            continue
        grouped.setdefault(row.row_key, {'row': row, 'fields': set()})['fields'].add(field)

    ranges = []
    skipped = 0
    appended = 0
    updated_rows = set()

    for group in grouped.values():
        row = group['row']
        try:
            data = json.loads(row.data_json or '{}')
        except (TypeError, json.JSONDecodeError):
            data = {}
        data = apply_import_formula_columns(dict(data))

        target_row = None
        is_new_target_row = False
        for key in import_layout_target_candidate_keys(data):
            if key in by_key:
                target_row = by_key[key]
                break
        if not target_row:
            target_row = next_row
            next_row += 1
            appended += 1
            is_new_target_row = True
            for key in import_layout_target_candidate_keys(data):
                by_key.setdefault(key, target_row)

        candidate_fields = set(data.keys()) if is_new_target_row else set(group['fields'])

        expanded_fields = set()
        for field in candidate_fields:
            mapped = import_layout_target_field_for_dashboard_field(field)
            expanded_fields.add(mapped)
            if field in ('yupi_po', 'po_yupi'):
                expanded_fields.add('po_yupi')
            if field in ('req_dlv_date', 'source_req_dlv_date'):
                expanded_fields.add('source_req_dlv_date')
            if field == 'po_send_date':
                expanded_fields.add('status')

        for field in sorted(expanded_fields):
            if field in IMPORT_LAYOUT_TARGET_FORMULA_FIELDS:
                continue
            col_letter = field_columns.get(field)
            if not col_letter:
                skipped += 1
                continue
            value = data.get(field, '')
            if field == 'po_yupi':
                value = clean(data.get('po_yupi')) or clean(data.get('yupi_po'))
            elif field == 'source_req_dlv_date':
                value = clean(data.get('source_req_dlv_date')) or clean(data.get('req_dlv_date'))
            ranges.append({
                'range': f"'{sheet_title}'!{col_letter}{target_row}",
                'values': [['' if value is None else str(value)]],
            })
            updated_rows.add(target_row)

    if not ranges:
        return {'synced': False, 'reason': 'No mapped Import tracker cells to sync', 'skipped': skipped}

    google_sheets_values_batch_update(IMPORT_LAYOUT_SHEET_ID, ranges)
    return {
        'synced': True,
        'ranges': len(ranges),
        'rows': len(updated_rows),
        'appended_rows': appended,
        'skipped': skipped,
        'spreadsheet_id': IMPORT_LAYOUT_SHEET_ID,
    }


def sync_import_cells_to_google_sheet(items):
    return {
        'synced': False,
        'source': {'synced': False, 'reason': 'Dashboard Import edits are local only'},
        'import_tracker': {'synced': False, 'reason': 'Dashboard Import edits are local only'},
    }

def sync_import_cell_to_google_sheet(row, field, value):
    result = sync_import_cells_to_google_sheet([{'row': row, 'field': field, 'value': value}])
    return result


def import_sort_date_value(value):
    raw = import_nonblank(value)
    if not raw:
        return None
    d = import_date_from_value(raw)
    if d:
        return d
    s = str(raw).strip().replace('.', '')
    for fmt in ('%d %b %Y', '%d-%b-%Y', '%d %B %Y', '%d-%B-%Y', '%Y/%m/%d', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    for fmt in ('%d %b', '%d-%b', '%d %B', '%d-%B'):
        try:
            parsed = datetime.strptime(s, fmt)
            return date(date.today().year, parsed.month, parsed.day)
        except ValueError:
            pass
    return None

@app.route('/api/import/data', methods=['GET'])
def get_import_data():
    try:
        force = str(request.args.get('refresh', '')).lower() in ('1', 'true', 'yes')
        page = max(int(request.args.get('page', 1)), 1)
        per_page = min(max(int(request.args.get('per_page', 25)), 1), 500)
        search = clean(request.args.get('search')) or ''
        selected_yupi_po = [clean(v) for v in request.args.getlist('yupi_po') if clean(v)]
        selected_vendors = [clean(v) for v in request.args.getlist('vendor_name') if clean(v)]
        # NEW: status filter (NEW / ON PROCESS / ON DELIVERY / DELIVERED / CANCELED)
        selected_statuses = [clean(v) for v in request.args.getlist('status') if clean(v)]
        # NEW: days_left color filter (red / yellow / green / today) — supports
        # multiple values via getlist so the user can filter by 2+ zones.
        selected_days_left = [clean(v).lower() for v in request.args.getlist('days_left') if clean(v)]
        req_dlv_sort = str(clean(request.args.get('req_dlv_sort')) or '').lower() or 'oldest'
        if req_dlv_sort not in ('oldest', 'newest'):
            req_dlv_sort = 'oldest'
        yupi_po_sort_val = clean(request.args.get('yupi_po_sort'))
        yupi_po_sort = str(yupi_po_sort_val).lower() if yupi_po_sort_val else ''
        if yupi_po_sort not in ('asc', 'desc'):
            yupi_po_sort = ''
        none_yupi_po = any(v == '__NONE_PLACEHOLDER__' for v in selected_yupi_po)
        none_vendor = any(v == '__NONE_PLACEHOLDER__' for v in selected_vendors)
        selected_yupi_po = {v.strip().lower() for v in selected_yupi_po if v != '__NONE_PLACEHOLDER__'}
        selected_vendors = {v.strip().lower() for v in selected_vendors if v != '__NONE_PLACEHOLDER__'}
        # Normalize selected statuses (uppercase) for case-insensitive comparison.
        selected_status_set = {str(v).strip().upper() for v in selected_statuses if v}
        sync_info = None

        if force:
            sync_info = sync_import_sheet_to_dashboard()

        columns = sync_info['columns'] if sync_info else import_layout_columns()
        vendor_count = sync_info.get('vendor_count') if sync_info else len(import_uploaded_vendor_names())

        q = ImportDashboardRow.query.filter(
            ImportDashboardRow.source_key.in_(_IMPORT_VISIBLE_SOURCE_KEYS)
        )
        if search:
            terms = [t.strip().lower() for t in re.split(r'[\n,]+', search) if t.strip()]
            for term in terms:
                pattern = f'%{term}%'
                q = q.filter(db.or_(
                    ImportDashboardRow.row_key.ilike(pattern),
                    ImportDashboardRow.source_label.ilike(pattern),
                    ImportDashboardRow.vendor_name.ilike(pattern),
                    ImportDashboardRow.data_json.ilike(pattern),
                ))

        candidate_rows = q.order_by(
            ImportDashboardRow.first_seen_at.desc(),
            ImportDashboardRow.id.desc(),
        ).all()

        parsed = []
        # Build the uploaded-vendor set ONCE. When the user has uploaded a
        # Vendor Import list, only rows whose vendor matches the uploaded
        # set are eligible for display — even if those rows still exist in
        # the DB from a previous (broader) sync. This is a safety net that
        # complements the purge logic in sync_import_sheet_to_dashboard():
        # if purge hasn't run yet (user hasn't clicked Copy Sheet), the
        # /data endpoint still hides non-matching rows so the user doesn't
        # see stale data.
        # FIX V10: Performance optimization — cache uploaded vendor list dan vendor attrs
        # supaya tidak query DB berulang kali. Sebelumnya, import_uploaded_vendor_names()
        # dan import_vendor_attrs_map() di-call terpisah, masing-masing query DB.
        _uploaded_vendor_names_cached = import_uploaded_vendor_names()
        uploaded_vendor_set_for_display = {v.strip().lower() for v in _uploaded_vendor_names_cached if v and v.strip()}
        _vendor_attrs_map_cached = import_vendor_attrs_map()

        # FIX V10: Pre-parse semua data_json sekali, cache hasil apply_import_formula_columns
        # supaya tidak panggil function berat berkali-kali untuk row yang sama.
        # Sebelumnya, apply_import_formula_columns dipanggil di:
        #   1. Loop parsed (line 7086)
        #   2. passes() untuk filter options (4x loop)
        #   3. import_dashboard_row_to_dict untuk page items
        # Sekarang cache hasil di parsed item, reuse di tempat lain.
        parsed = []
        for row in candidate_rows:
            try:
                data = json.loads(row.data_json or '{}')
            except (TypeError, json.JSONDecodeError):
                data = {}
            data = apply_import_formula_columns(dict(data))
            yupi = import_nonblank(data.get('yupi_po')) or import_nonblank(data.get('po_yupi'))
            vendor = import_nonblank(data.get('vendor_name')) or import_nonblank(data.get('vendor')) or import_nonblank(row.vendor_name)
            # Apply uploaded-vendor filter at the parse level — non-matching
            # rows never enter `parsed`, so they don't show up in the table
            # or in any filter dropdown options. Rows with no vendor are also
            # excluded when an uploaded vendor list exists.
            if uploaded_vendor_set_for_display:
                if not vendor or vendor.strip().lower() not in uploaded_vendor_set_for_display:
                    continue
            parsed.append({'row': row, 'data': data, 'yupi_po': yupi, 'vendor': vendor})

        def passes(item, ignore=None):
            if none_yupi_po or none_vendor:
                return False
            if ignore != 'yupi_po' and selected_yupi_po and str(item.get('yupi_po') or '').strip().lower() not in selected_yupi_po:
                return False
            if ignore != 'vendor' and selected_vendors and str(item.get('vendor') or '').strip().lower() not in selected_vendors:
                return False
            # NEW: status filter. Status is computed by apply_import_formula_columns
            # and stored at item['data']['status'].
            if ignore != 'status' and selected_status_set:
                row_status = str((item.get('data') or {}).get('status') or '').strip().upper()
                if row_status not in selected_status_set:
                    return False
            # NEW: days_left color filter. days_left is computed by
            # apply_import_formula_columns. Values: '✅', '❌', '', or a number
            # string (can be negative). Color zones match the frontend:
            #   red    → <= 7 or negative
            #   yellow → 8–29
            #   green  → >= 30
            #   today  → exactly 0
            # Supports MULTIPLE zones — row passes if it matches ANY selected
            # zone. empty selected_days_left → no filter.
            if ignore != 'days_left' and selected_days_left:
                raw = str((item.get('data') or {}).get('days_left') or '').strip()
                if raw in ('✅', '❌', '', '-'):
                    return False  # icon/empty rows don't match any color zone
                try:
                    dl = int(float(raw))
                except (ValueError, TypeError):
                    return False
                matched = False
                for zone in selected_days_left:
                    if zone == 'red' and (dl < 0 or dl <= 7):
                        matched = True; break
                    if zone == 'yellow' and 8 <= dl <= 29:
                        matched = True; break
                    if zone == 'green' and dl >= 30:
                        matched = True; break
                    if zone == 'today' and dl == 0:
                        matched = True; break
                if not matched:
                    return False
            return True

        filtered_items = [item for item in parsed if passes(item)]

        # FIX V10: Performance — cache hasil passes() per item supaya tidak
        # panggil passes() berulang kali (4x loop untuk filter options).
        # Sebelumnya, passes() dipanggil untuk setiap item di:
        #   1. filtered_items (1x per item)
        #   2. yupi_options loop (ignore='yupi_po')
        #   3. vendor_options loop (ignore='vendor')
        #   4. status_options loop (ignore='status')
        #   5. days_left_zone_options loop (ignore='days_left')
        # Total: 5x per item. Sekarang cache hasil passes() dengan setiap ignore variant.
        _passes_cache = {}
        def passes_cached(item, ignore=None):
            cache_key = (id(item), ignore)
            if cache_key not in _passes_cache:
                _passes_cache[cache_key] = passes(item, ignore=ignore)
            return _passes_cache[cache_key]

        # ── Interdependent filter options ────────────────────────────────
        # Each filter's option list is built from rows that pass ALL OTHER
        # filters (but not itself). This means:
        #   - If no rows have status "NEW", the status dropdown won't show "NEW"
        #   - If you filter by vendor A, the YUPI PO dropdown only shows POs
        #     from vendor A
        #   - If you filter by status "DELIVERED", the days_left dropdown only
        #     shows color zones that exist among delivered rows
        # This matches the user's expectation: "jika tidak ada status yang
        # new maka di drop down filternya tidak ada pilihan new".
        yupi_options = sorted({str(item.get('yupi_po') or '').strip() for item in parsed if str(item.get('yupi_po') or '').strip() and passes_cached(item, ignore='yupi_po')}, key=lambda s: s.lower())
        vendor_options = sorted({str(item.get('vendor') or '').strip() for item in parsed if str(item.get('vendor') or '').strip() and passes_cached(item, ignore='vendor')}, key=lambda s: s.lower())
        # Status options — built from rows that pass all filters EXCEPT status.
        # Only statuses that actually appear in the (otherwise-filtered) data
        # are listed. NEW is included only if a row's status computed to NEW.
        status_options = sorted({str((item.get('data') or {}).get('status') or '').strip().upper() for item in parsed if str((item.get('data') or {}).get('status') or '').strip() and passes_cached(item, ignore='status')}, key=lambda s: s.lower())
        # Days Left color options — built from rows that pass all filters
        # EXCEPT days_left. Only color zones that actually appear are listed.
        days_left_zone_options = []
        _dl_zones_seen = set()
        for item in parsed:
            if not passes_cached(item, ignore='days_left'): continue
            raw = str((item.get('data') or {}).get('days_left') or '').strip()
            if raw in ('✅', '❌', '', '-'): continue
            try: dl = int(float(raw))
            except (ValueError, TypeError): continue
            if dl < 0 or dl <= 7: zone = 'red'
            elif 8 <= dl <= 29: zone = 'yellow'
            elif dl >= 30: zone = 'green'
            elif dl == 0: zone = 'today'
            else: continue
            if zone not in _dl_zones_seen:
                _dl_zones_seen.add(zone)
                days_left_zone_options.append(zone)
        # Sort: red, yellow, green, today (consistent order)
        _zone_order = {'red': 0, 'yellow': 1, 'green': 2, 'today': 3}
        days_left_zone_options.sort(key=lambda z: _zone_order.get(z, 99))

        def _import_req_date_key(item):
            data = item.get('data') or {}
            d = import_sort_date_value(import_nonblank(data.get('req_dlv_date')) or import_nonblank(data.get('source_req_dlv_date')))
            try:
                if d is None or pd.isna(d):
                    return (1, 0)
                # Konversi ke date Python murni untuk hindari error NaT pandas
                if isinstance(d, datetime):
                    d = d.date()
                elif hasattr(d, 'to_pydatetime'):
                    d = d.to_pydatetime().date()
                ordinal = d.toordinal()
            except Exception:
                return (1, 0)
            return (0, ordinal if req_dlv_sort == 'oldest' else -ordinal)

        if yupi_po_sort:
            filtered_items.sort(key=_import_req_date_key)
            with_yupi = [item for item in filtered_items if str(item.get('yupi_po') or '').strip()]
            without_yupi = [item for item in filtered_items if not str(item.get('yupi_po') or '').strip()]
            with_yupi.sort(key=lambda item: str(item.get('yupi_po') or '').strip().lower(), reverse=(yupi_po_sort == 'desc'))
            filtered_items = with_yupi + without_yupi
        else:
            filtered_items.sort(key=_import_req_date_key)

        total = len(filtered_items)
        page_items = filtered_items[(page - 1) * per_page: page * per_page]
        # FIX V10: gunakan _vendor_attrs_map_cached + pre_parsed_data dari item
        # supaya import_dashboard_row_to_dict tidak panggil apply_import_formula_columns
        # lagi (data sudah di-parse di loop parsed atas).
        rows = [import_dashboard_row_to_dict(item['row'], columns,
                                              vendor_attrs_map=_vendor_attrs_map_cached,
                                              pre_parsed_data=item['data'])
                for item in page_items]

        # ── KPIs ───────────────────────────────────────────────────────────
        # Computed across ALL filtered rows (not just the current page):
        #  - Total PO           : count of distinct YUPI POs
        #  - This Week Arrival  : count of rows whose ETA falls within this week
        #    (Mon–Sun). Sub-KPI: of those, how many have SAP INPUT unchecked.
        #  - Sales Amount       : sum of AMOUNT column across filtered rows.
        #  - PO Amount (IDR)    : sum of PURCHASE AMOUNT converted to IDR using
        #    the exchange rate for the row's ETA date.
        #  - Gross Margin       : Sales Amount − PO Amount (IDR).
        try:
            today_wib = (datetime.utcnow() + timedelta(hours=7)).date()
            # ISO week: Monday is the start.
            week_start = today_wib - timedelta(days=today_wib.weekday())  # Monday 00:00
            week_end = week_start + timedelta(days=6)  # Sunday
            total_po = 0
            this_week_arrival = 0
            this_week_no_sap = 0
            sales_amount_total = 0.0
            po_amount_idr_total = 0.0
            seen_yupi_pos = set()
            # Pre-collect ETA dates per row for FX rate prefetching.
            eta_dates_usd_eur = set()
            for item in filtered_items:
                data = item.get('data') or {}
                yupi = import_nonblank(data.get('yupi_po')) or import_nonblank(data.get('po_yupi'))
                if yupi:
                    seen_yupi_pos.add(str(yupi).strip().lower())
                eta_date = import_date_from_value(data.get('eta'))
                if eta_date and week_start <= eta_date <= week_end:
                    this_week_arrival += 1
                    # SAP INPUT checkbox — if not checked, count as "no SAP input"
                    sap_val = import_nonblank(data.get('sap_input'))
                    if not import_normalize_checkbox(sap_val):
                        this_week_no_sap += 1
                # Sales Amount = AMOUNT column
                amt = import_float_value(data.get('amount'))
                if amt is not None:
                    sales_amount_total += amt
                # PO Amount (IDR) = PURCHASE AMOUNT converted using ETA-date FX rate
                pa = import_float_value(data.get('purchase_amount'))
                if pa is not None and pa > 0:
                    cur = (clean(data.get('currency')) or 'IDR').strip().upper()
                    if cur in ('', 'IDR'):
                        po_amount_idr_total += pa
                    elif cur in ('USD', 'EUR') and eta_date:
                        eta_dates_usd_eur.add((cur, eta_date))
            # Prefetch FX rates for the ETA dates we need (cache-only — no API
            # calls in the KPI path; if rates are missing, those rows contribute
            # 0 to the total rather than blocking the response).
            for cur in ('USD', 'EUR'):
                dates_for_cur = sorted({d for c, d in eta_dates_usd_eur if c == cur})
                if dates_for_cur:
                    prefetch_exchange_rates(dates_for_cur, fetch_missing=False, currency=cur)
            # Now compute PO Amount (IDR) per row using cached rates.
            for item in filtered_items:
                data = item.get('data') or {}
                pa = import_float_value(data.get('purchase_amount'))
                if pa is None or pa <= 0: continue
                cur = (clean(data.get('currency')) or 'IDR').strip().upper()
                eta_date = import_date_from_value(data.get('eta'))
                if cur in ('', 'IDR'):
                    continue  # already added above
                if cur in ('USD', 'EUR') and eta_date:
                    rate = get_currency_to_idr(cur, eta_date, cache_only=True)
                    po_amount_idr_total += pa * rate
            total_po = len(seen_yupi_pos)
            gross_margin = sales_amount_total - po_amount_idr_total
            kpis = {
                'total_po': total_po,
                'this_week_arrival': this_week_arrival,
                'this_week_no_sap': this_week_no_sap,
                'sales_amount': round(sales_amount_total, 2),
                'po_amount_idr': round(po_amount_idr_total, 2),
                'gross_margin': round(gross_margin, 2),
            }
        except Exception as kpi_exc:
            import traceback; traceback.print_exc()
            kpis = {
                'total_po': 0, 'this_week_arrival': 0, 'this_week_no_sap': 0,
                'sales_amount': 0.0, 'po_amount_idr': 0.0, 'gross_margin': 0.0,
            }

        last_copy_at = import_meta_get('last_copy_at') or ''

        return jsonify({
            'data': rows,
            'columns': columns,
            'total': total,
            'page': page,
            'per_page': per_page,
            'vendor_count': vendor_count,
            'last_copy_at': last_copy_at,
            'filters': {
                'yupi_po': yupi_options,
                'vendors': vendor_options,
                # Interdependent filter options — only show values that
                # actually exist in the (otherwise-filtered) data.
                'statuses': status_options,
                'days_left_zones': days_left_zone_options,
            },
            'req_dlv_sort': req_dlv_sort,
            'yupi_po_sort': yupi_po_sort,
            'sources': [{'key': s['key'], 'label': s['label']} for s in IMPORT_SOURCE_SHEETS],
            'sync': sync_info,
            'kpis': kpis,
        })
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/import/debug-source', methods=['GET'])
def import_debug_source():
    try:
        source_key = clean(request.args.get('source')) or 'source_1'
        vendor_filter = clean(request.args.get('vendor')) or ''
        source = next((s for s in IMPORT_SOURCE_SHEETS if s['key'] == source_key), None)
        if not source:
            return jsonify({
                'error': f"Unknown source key '{source_key}'",
                'available_sources': [s['key'] for s in IMPORT_SOURCE_SHEETS],
            }), 400

        columns = import_layout_columns()
        mapping_columns = import_all_mapping_columns(columns)
        sheet_title, header_df = import_source_header_preview(source, force=True)
        if header_df.empty:
            return jsonify({
                'error': 'Could not read a usable header preview for this source (empty result).',
                'sheet_title_tried': sheet_title,
                'spreadsheet_id': source['spreadsheet_id'],
            }), 500

        header_idx = import_detect_header_row(header_df)
        kind = import_source_kind_from_header(header_df, header_idx)
        source_map = import_source_column_map(header_df, mapping_columns)
        header_row_values = [clean(v) for v in header_df.iloc[header_idx].tolist()] if len(header_df) else []

        source_map_letters = {}
        for field, idx in source_map.items():
            try:
                source_map_letters[field] = column_letter_from_index(idx + 1)
            except Exception:
                source_map_letters[field] = f'(idx {idx})'

        if vendor_filter:
            vendor_set = {vendor_filter.strip().lower()}
        else:
            filter_vendors, _ = import_vendor_filter_names()
            vendor_set = {v.strip().lower() for v in filter_vendors}

        sample_rows = import_source_rows_fast(source, columns, vendor_set)[:5]
        sample_out = []
        for row in sample_rows:
            sample_out.append({
                'sheet_row': row.get('_sheet_row'),
                'vendor_name_detected': row.get('_vendor_name'),
                'po_yupi': row.get('po_yupi'),
                'yupi_po': row.get('yupi_po'),
                'po_sementara': row.get('po_sementara'),
                'req_dlv_date': row.get('req_dlv_date'),
                'po_date_by_email': row.get('po_date_by_email'),
                'etd': row.get('etd'),
                'eta': row.get('eta'),
                'so': row.get('so'),
                'group': row.get('group'),
                'item_name': row.get('item_name'),
            })

        return jsonify({
            'source_key': source['key'],
            'spreadsheet_id': source['spreadsheet_id'],
            'sheet_title_used': sheet_title,
            'detected_header_row_1based': header_idx + 1,
            'detected_kind': kind or '(none -> common fallback letters used)',
            'header_row_raw_values': header_row_values,
            'column_map_field_to_letter': source_map_letters,
            'vendor_fallback_letters': list(IMPORT_FALLBACK_SOURCE_VENDOR_LETTERS),
            'vendor_filter_used': sorted(vendor_set),
            'matched_row_count': len(sample_rows),
            'sample_rows': sample_out,
            # FIX V9: tambah hint troubleshooting kalau matched_row_count = 0
            'troubleshooting': (
                "Jika matched_row_count=0 padahal sheet ada data: "
                "(1) Cek vendor_filter_used — pastikan vendor yang diupload persis match "
                "atau jadi prefix dari vendor di sheet. "
                "(2) Cek header_row_raw_values — pastikan header row terdeteksi benar "
                "(mis. 'Vendor Name', 'PO YUPI', 'PO DATE By Email'). "
                "(3) Cek column_map_field_to_letter — vendor_name dan po_yupi harus terpetakan. "
                "(4) Fuzzy matching aktif: 'CURT GEORGI GMBH & CO' akan match dengan "
                "'CURT GEORGI GMBH & CO. KG' (prefix match)."
            ),
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/cleanup-import-duplicates', methods=['POST'])
def admin_cleanup_import_duplicates():
    """Auto-cleanup duplicate rows di Import Dashboard.

    FIX V10 (CORRECTED): HANYA hapus row yang 100% IDENTICAL (semua field bisnis
    sama, termasuk req_dlv_date, item_yupi, dll). JANGAN hapus row hanya karena
    po_yupi + item_yupi sama — 1 PO bisa punya multiple item dengan Req Dlv Date
    berbeda (reschedule history), itu entry valid yang harus dipertahankan.

    Logika:
    1. Build fingerprint dari semua field bisnis (IMPORT_LOCAL_EDIT_FIELDS) yang tidak kosong
    2. Group rows by fingerprint
    3. Kalau ada >1 row dengan fingerprint sama (100% identical):
       a. Keep row dengan id terkecil (paling lama, paling stabil)
       b. Hapus row lain
    4. Return summary

    Return:
    - total_rows_before: jumlah row sebelum cleanup
    - total_rows_after: jumlah row setelah cleanup
    - identical_duplicate_groups: jumlah group yang 100% identical
    - deleted_count: jumlah row yang dihapus
    """
    try:
        all_rows = ImportDashboardRow.query.filter(
            ImportDashboardRow.source_key.in_(_IMPORT_VISIBLE_SOURCE_KEYS)
        ).all()

        # Group by FULL data fingerprint (semua field bisnis sama = true identical duplicate)
        by_fingerprint = {}
        for row in all_rows:
            try: data = json.loads(row.data_json or '{}')
            except: data = {}
            # Build fingerprint dari semua field bisnis yang tidak kosong
            biz_fields = {k: v for k, v in data.items()
                          if k in IMPORT_LOCAL_EDIT_FIELDS and not import_blankish(v)}
            fingerprint = json.dumps(biz_fields, ensure_ascii=False, sort_keys=True, separators=(',', ':'))
            by_fingerprint.setdefault(fingerprint, []).append(row)

        # Find groups with identical duplicates
        duplicate_groups = 0
        deleted_count = 0
        details = []

        for fingerprint, rows in by_fingerprint.items():
            if len(rows) <= 1: continue  # no duplicate

            # Sort by id asc (keep paling lama = id terkecil)
            rows.sort(key=lambda x: x.id)
            winner = rows[0]
            losers = rows[1:]

            # Hapus losers (100% identical, tidak perlu merge karena datanya sama persis)
            deleted_row_keys = []
            for loser in losers:
                deleted_row_keys.append(loser.row_key)
                db.session.delete(loser)
                deleted_count += 1

            duplicate_groups += 1
            try: winner_data = json.loads(winner.data_json or '{}')
            except: winner_data = {}
            details.append({
                'po_yupi': clean(winner_data.get('po_yupi')) or clean(winner_data.get('yupi_po')) or '',
                'item_yupi': clean(winner_data.get('item_yupi')) or '',
                'req_dlv_date': clean(winner_data.get('req_dlv_date')) or '',
                'kept_row_key': winner.row_key,
                'deleted_row_keys': deleted_row_keys,
                'note': '100% identical — semua field bisnis sama',
            })

        # Commit changes
        if deleted_count > 0:
            db.session.commit()
            db.session.execute(text('PRAGMA wal_checkpoint(TRUNCATE)'))
            db.session.commit()
            clear_runtime_caches()

        total_rows_after = ImportDashboardRow.query.filter(
            ImportDashboardRow.source_key.in_(_IMPORT_VISIBLE_SOURCE_KEYS)
        ).count()

        return jsonify({
            'success': True,
            'total_rows_before': len(all_rows),
            'total_rows_after': total_rows_after,
            'identical_duplicate_groups': duplicate_groups,
            'deleted_count': deleted_count,
            'details': details[:50],
            'message': f'Cleanup selesai: {deleted_count} identical duplicate row dihapus (hanya row dengan 100% field sama). Sisa: {total_rows_after} rows.',
            'note': 'HANYA row yang 100% identical (semua field bisnis sama) yang dihapus. Row dengan po_yupi sama tapi req_dlv_date/item_yupi berbeda DIPERTAHANKAN.',
        })
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/import/debug-duplicates', methods=['GET'])
def import_debug_duplicates():
    """Cari row di dashboard yang punya po_yupi sama tapi row_key berbeda (duplicate).

    FIX V10: Endpoint ini untuk diagnose kenapa item yang di-reschedule muncul
    sebagai 2 baris terpisah di dashboard, padahal seharusnya 1 baris (update).

    Return:
    - duplicates: list of {po_yupi, rows: [{row_key, item_yupi, req_dlv_date, ...}]}
    - summary: total duplicates found
    """
    try:
        all_rows = ImportDashboardRow.query.filter(
            ImportDashboardRow.source_key.in_(_IMPORT_VISIBLE_SOURCE_KEYS)
        ).all()
        # Group by po_yupi (case-insensitive)
        by_po_yupi = {}
        for row in all_rows:
            try: data = json.loads(row.data_json or '{}')
            except: data = {}
            po_yupi = clean(data.get('po_yupi')) or clean(data.get('yupi_po')) or ''
            if not po_yupi: continue
            key = po_yupi.strip().lower()
            by_po_yupi.setdefault(key, []).append({
                'row_key': row.row_key,
                'po_yupi': po_yupi,
                'item_yupi': clean(data.get('item_yupi')) or '',
                'po_sementara': clean(data.get('po_sementara')) or '',
                'req_dlv_date': clean(data.get('req_dlv_date')) or '',
                'source_req_dlv_date': clean(data.get('source_req_dlv_date')) or '',
                'reschedule': clean(data.get('reschedule')) or '',
                'vendor_name': clean(data.get('vendor_name')) or row.vendor_name or '',
                'item_name': clean(data.get('item_name')) or '',
                'source_uid': row.source_uid or '',
                'sheet_row': row.sheet_row,
            })
        # Find duplicates (po_yupi with more than 1 row)
        duplicates = []
        for po_yupi_key, rows in by_po_yupi.items():
            if len(rows) > 1:
                duplicates.append({
                    'po_yupi': rows[0]['po_yupi'],
                    'row_count': len(rows),
                    'rows': rows,
                })
        duplicates.sort(key=lambda x: x['row_count'], reverse=True)
        return jsonify({
            'total_rows': len(all_rows),
            'unique_po_yupi_count': len(by_po_yupi),
            'duplicate_po_yupi_count': len(duplicates),
            'duplicates': duplicates[:20],  # top 20
            'troubleshooting': (
                "Kalau ada duplicate (po_yupi sama, row berbeda): "
                "(1) Cek item_yupi — kalau beda, UID akan beda → dianggap row berbeda. "
                "(2) Cek source_uid — kalau beda, berarti identity payload berbeda. "
                "(3) Cek req_dlv_date — kalau beda, kemungkinan item di-reschedule. "
                "Solusi: pastikan item_yupi konsisten di sheet source, atau update identity logic."
            ),
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/import/debug-find', methods=['GET'])
def import_debug_find():
    """Cari row di source sheet by PO Yupi number — tampilkan semua row yang cocok.

    FIX V9: Endpoint ini untuk diagnose kenapa PO tertentu (mis. 410100301)
    tidak tercopy. User bisa lihat persis di row berapa PO itu ada, dan apa
    vendor yang terbaca di row tersebut.

    Query params:
    - source: source_1 atau source_2
    - po: PO Yupi number (mis. 410100301)
    """
    try:
        source_key = clean(request.args.get('source')) or 'source_1'
        po_search = clean(request.args.get('po')) or ''
        source = next((s for s in IMPORT_SOURCE_SHEETS if s['key'] == source_key), None)
        if not source:
            return jsonify({'error': f"Unknown source '{source_key}'"}), 400
        if not po_search:
            return jsonify({'error': 'Param ?po= wajib (PO Yupi number, mis. 410100301)'}), 400

        columns = import_layout_columns()
        mapping_columns = import_all_mapping_columns(columns)
        sheet_title, header_df = import_source_header_preview(source, force=True)
        if header_df.empty:
            return jsonify({'error': 'Could not read header preview'}), 500

        header_idx = import_detect_header_row(header_df)
        source_map = import_source_column_map(header_df, mapping_columns)

        # Get uploaded vendor list
        uploaded_vendor_names = import_uploaded_vendor_names()
        uploaded_vendor_set = {v.strip().lower() for v in uploaded_vendor_names if v and v.strip()}

        # Scan ALL rows (empty set = ambil semua, tanpa filter vendor)
        all_rows = import_source_rows_fast(source, columns, set())

        # Cari row yang po_yupi-nya match
        po_search_norm = po_search.strip().lower()
        matching_rows = []
        for row in all_rows:
            row_po = clean(row.get('po_yupi')) or clean(row.get('yupi_po')) or ''
            if row_po and row_po.strip().lower() == po_search_norm:
                row_vendor = row.get('_vendor_name') or ''
                is_match = import_vendor_match(row_vendor, uploaded_vendor_set) if row_vendor else False
                matching_rows.append({
                    'sheet_row': row.get('_sheet_row'),
                    'po_yupi': row_po,
                    'po_sementara': row.get('po_sementara'),
                    'item_name': row.get('item_name'),
                    'vendor_name_detected': row_vendor,
                    'is_vendor_match': is_match,
                    'all_fields': {k: v for k, v in row.items() if not k.startswith('_') and v},
                })

        return jsonify({
            'source_key': source_key,
            'po_searched': po_search,
            'total_rows_scanned': len(all_rows),
            'matching_rows_found': len(matching_rows),
            'matched_rows_detail': matching_rows,
            'vendor_filter_used': sorted(uploaded_vendor_set),
            'troubleshooting': (
                "Jika matching_rows_found = 0: PO tidak ada di sheet source (cek sheet langsung). "
                "Jika matching_rows_found > 0 tapi is_vendor_match = false: "
                "(1) Cek vendor_name_detected — apakah persis sama atau prefix dari vendor di vendor_filter_used? "
                "(2) Kalau vendor_name_detected kosong, berika carry-over tidak aktif "
                "(mungkin row tidak punya po_yupi/item_name yang terbaca). "
                "(3) Kalau vendor_name_detected terisi tapi beda penulisan, "
                "update vendor di ImportVendor table atau tambahkan variant nama. "
                "Jika is_vendor_match = true tapi row belum tercopy di dashboard: "
                "trigger sync via POST /api/import/copy-sheet."
            ),
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/import/debug-row', methods=['GET'])
def import_debug_row():
    """Cek row spesifik di source sheet — tampilkan SEMUA kolom yang terbaca.

    FIX V9: Endpoint ini untuk diagnose kenapa row tertentu (mis. row 144 dengan
    PO 410100301) tidak tercopy. User bisa lihat persis nilai dari setiap kolom.

    Query params:
    - source: source_1 atau source_2
    - row: nomor row (1-indexed, sesuai tampilan Google Sheet)
    """
    try:
        source_key = clean(request.args.get('source')) or 'source_1'
        row_num = int(clean(request.args.get('row')) or 0)
        source = next((s for s in IMPORT_SOURCE_SHEETS if s['key'] == source_key), None)
        if not source:
            return jsonify({'error': f"Unknown source '{source_key}'"}), 400
        if not row_num or row_num < 1:
            return jsonify({'error': 'Param ?row= wajib (1-indexed row number)'}), 400

        columns = import_layout_columns()
        mapping_columns = import_all_mapping_columns(columns)
        sheet_title, header_df = import_source_header_preview(source, force=True)
        if header_df.empty:
            return jsonify({'error': 'Could not read header preview'}), 500

        header_idx = import_detect_header_row(header_df)
        source_map = import_source_column_map(header_df, mapping_columns)

        # Baca row spesifik langsung dari sheet
        # row_num adalah 1-indexed (sesuai Google Sheet)
        result = google_sheets_values_get(
            source['spreadsheet_id'],
            f"'{sheet_title}'!A{row_num}:ZZ{row_num}",
            value_render_option='FORMATTED_VALUE'
        )
        values = result.get('values', [])
        if not values or not values[0]:
            return jsonify({
                'source_key': source_key,
                'row_requested': row_num,
                'error': f'Row {row_num} kosong atau tidak terbaca',
            }), 404

        row_values = values[0]
        # Petakan ke field name berdasarkan source_map
        field_values = {}
        for field, col_idx in source_map.items():
            if col_idx is not None and col_idx < len(row_values):
                field_values[field] = row_values[col_idx]
            else:
                field_values[field] = ''

        # Header row untuk referensi
        header_row = [clean(v) for v in header_df.iloc[header_idx].tolist()] if len(header_df) else []

        return jsonify({
            'source_key': source_key,
            'sheet_title_used': sheet_title,
            'row_requested': row_num,
            'header_row_1based': header_idx + 1,
            'header_row_values': header_row[:25],
            'column_map_field_to_letter': {f: column_letter_from_index(idx + 1) for f, idx in source_map.items() if idx is not None},
            'raw_row_values': row_values[:25],  # 25 kolom pertama
            'field_values_detected': field_values,
            'vendor_value_at_Q': row_values[16] if len(row_values) > 16 else '(empty)',  # Q = index 16
            'amount_value_at_P': row_values[15] if len(row_values) > 15 else '(empty)',  # P = index 15
            'troubleshooting': (
                "Cek field_values_detected.vendor_name — kalau kosong padahal di sheet ada, "
                "kemungkinan: (1) vendor di-merge cell, (2) header detection salah kolom. "
                "Cek raw_row_values untuk lihat semua kolom. "
                "Kalau vendor ada di kolom lain (mis. R, S), berarti header detection salah — "
                "kirim screenshot ke developer untuk fix."
            ),
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/import/debug-scan', methods=['GET'])
def import_debug_scan():
    """Scan SEMUA rows di source sheet, return statistic match/no-match + sample row yang gagal.

    FIX V9: Endpoint ini untuk diagnose kenapa PO tertentu tidak tercopy.
    Return:
    - total_rows_scanned: jumlah row dengan data di sheet
    - matched_rows: row yang vendor-nya match dengan uploaded vendor list
    - unmatched_rows: row yang vendor-nya TIDAK match (dengan detail vendor yang terbaca)
    - matched_sample: 5 row pertama yang match
    - unmatched_sample: 5 row pertama yang tidak match (supaya user bisa lihat vendor apa yang tidak match)
    - vendor_filter_used: daftar vendor yang di-upload ke ImportVendor table
    """
    try:
        source_key = clean(request.args.get('source')) or 'source_1'
        source = next((s for s in IMPORT_SOURCE_SHEETS if s['key'] == source_key), None)
        if not source:
            return jsonify({
                'error': f"Unknown source key '{source_key}'",
                'available_sources': [s['key'] for s in IMPORT_SOURCE_SHEETS],
            }), 400

        columns = import_layout_columns()
        mapping_columns = import_all_mapping_columns(columns)
        sheet_title, header_df = import_source_header_preview(source, force=True)
        if header_df.empty:
            return jsonify({'error': 'Could not read header preview', 'source': source_key}), 500

        header_idx = import_detect_header_row(header_df)
        source_map = import_source_column_map(header_df, mapping_columns)
        header_row_values = [clean(v) for v in header_df.iloc[header_idx].tolist()] if len(header_df) else []

        source_map_letters = {}
        for field, idx in source_map.items():
            try: source_map_letters[field] = column_letter_from_index(idx + 1)
            except: source_map_letters[field] = f'(idx {idx})'

        # Get uploaded vendor list
        uploaded_vendor_names = import_uploaded_vendor_names()
        uploaded_vendor_set = {v.strip().lower() for v in uploaded_vendor_names if v and v.strip()}

        if not uploaded_vendor_set:
            return jsonify({
                'source_key': source_key,
                'error': 'No uploaded vendor list! Source sheet sync di-skip entirely ketika vendor list kosong.',
                'vendor_filter_used': [],
                'fix': 'Upload vendor list dulu via tombol Upload Vendor di dashboard, atau via /api/import/vendors/upload.',
            }), 200

        # Scan ALL rows (tidak hanya 5 sample) dengan vendor_set = kosong (ambil semua)
        # Lalu manual check match untuk setiap row
        all_rows = import_source_rows_fast(source, columns, set())  # empty set = ambil semua
        # Tunggu — empty set akan skip filter, tapi code cek `if vendor_set and not any(...)`.
        # Kalau vendor_set kosong, `if vendor_set` False → tidak filter → ambil semua. ✓

        matched_rows = []
        unmatched_rows = []
        for row in all_rows:
            row_vendor = row.get('_vendor_name') or ''
            is_match = import_vendor_match(row_vendor, uploaded_vendor_set) if row_vendor else False
            row_info = {
                'sheet_row': row.get('_sheet_row'),
                'vendor_name_detected': row_vendor,
                'po_yupi': row.get('po_yupi'),
                'po_sementara': row.get('po_sementara'),
                'item_name': row.get('item_name'),
                'is_match': is_match,
            }
            if is_match:
                matched_rows.append(row_info)
            else:
                unmatched_rows.append(row_info)

        # Statistic vendor yang tidak match (group by vendor name)
        unmatched_vendors = {}
        for r in unmatched_rows:
            v = r['vendor_name_detected'] or '(empty)'
            unmatched_vendors[v] = unmatched_vendors.get(v, 0) + 1

        return jsonify({
            'source_key': source_key,
            'sheet_title_used': sheet_title,
            'detected_header_row_1based': header_idx + 1,
            'column_map_field_to_letter': source_map_letters,
            'header_row_raw_values': header_row_values[:25],
            'vendor_fallback_letters': list(IMPORT_FALLBACK_SOURCE_VENDOR_LETTERS),
            'vendor_filter_used': sorted(uploaded_vendor_set),
            'vendor_filter_count': len(uploaded_vendor_set),
            'total_rows_scanned': len(all_rows),
            'matched_rows': len(matched_rows),
            'unmatched_rows': len(unmatched_rows),
            'unmatched_vendors_breakdown': unmatched_vendors,
            'matched_sample': matched_rows[:5],
            'unmatched_sample': unmatched_rows[:10],
            'troubleshooting': (
                "Jika PO yang dicari ada di unmatched_sample: "
                "(1) Cek vendor_name_detected — apakah persis sama atau prefix dari vendor di vendor_filter_used? "
                "(2) Kalau vendor belum di-upload, tambahkan via /api/import/vendors/upload. "
                "(3) Fuzzy match aktif: prefix match (min 5 char) atau contains match (min 8 char). "
                "Jika PO tidak ada di matched_sample DAN tidak ada di unmatched_sample: "
                "(a) Cek total_rows_scanned — kalau 0, berarti header detection gagal. "
                "(b) Cek header_row_raw_values — pastikan header terdeteksi benar. "
                "(c) Cek column_map_field_to_letter — vendor_name dan po_yupi harus terpetakan."
            ),
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/import/cell', methods=['PUT'])
def update_import_cell():
    try:
        payload = request.get_json(silent=True) or {}
        row_key = clean(payload.get('row_key'))
        field = clean(payload.get('field'))
        value = '' if payload.get('value') is None else str(payload.get('value'))
        if not row_key or not field:
            return jsonify({'error': 'row_key and field are required'}), 400
        columns = import_layout_columns()
        column = next((col for col in columns if col['field'] == field), None)
        if not column:
            return jsonify({'error': 'Unknown import column'}), 400
        row = ImportDashboardRow.query.filter_by(row_key=row_key).first()
        if not row:
            return jsonify({'error': 'Import dashboard row not found'}), 404
        try:
            data = json.loads(row.data_json or '{}')
        except (TypeError, json.JSONDecodeError):
            data = {}
        data = set_import_payload_field_aliases(data, field, value)
        data = apply_import_formula_columns(data)
        row.data_json = json.dumps(data, ensure_ascii=False)
        row.updated_at = datetime.utcnow()
        sheet_sync = {'synced': False, 'reason': 'Not attempted'}
        try:
            sync_items = [{'row': row, 'field': field, 'value': value}]
            if field in ('purchase_price', 'ord_qty') and clean(data.get('purchase_amount')):
                sync_items.append({'row': row, 'field': 'purchase_amount', 'value': data.get('purchase_amount')})
            sheet_sync = sync_import_cells_to_google_sheet(sync_items)
        except Exception as sync_exc:
            sheet_sync = {'synced': False, 'reason': str(sync_exc)}
        db.session.commit()
        clear_runtime_caches()
        columns = import_layout_columns()
        updated_row = import_dashboard_row_to_dict(row, columns, vendor_attrs_map=import_vendor_attrs_map())
        return jsonify({'success': True, 'row_key': row_key, 'field': field, 'value': value, 'row': updated_row, 'sheet_sync': sheet_sync})
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e), 'sheet_sync': {'synced': False, 'reason': str(e)}}), 500


@app.route('/api/import/export', methods=['GET'])
def export_import_data():
    try:
        search = clean(request.args.get('search')) or ''
        selected_yupi_po_raw = [clean(v) for v in request.args.getlist('yupi_po') if clean(v)]
        selected_vendors_raw = [clean(v) for v in request.args.getlist('vendor_name') if clean(v)]
        selected_statuses_raw = [clean(v) for v in request.args.getlist('status') if clean(v)]
        selected_days_left = [clean(v).lower() for v in request.args.getlist('days_left') if clean(v)]
        none_yupi_po = any(v == '__NONE_PLACEHOLDER__' for v in selected_yupi_po_raw)
        none_vendor = any(v == '__NONE_PLACEHOLDER__' for v in selected_vendors_raw)
        selected_yupi_po = {v.strip().lower() for v in selected_yupi_po_raw if v != '__NONE_PLACEHOLDER__'}
        selected_vendors = {v.strip().lower() for v in selected_vendors_raw if v != '__NONE_PLACEHOLDER__'}
        selected_status_set = {str(v).strip().upper() for v in selected_statuses_raw if v}

        columns = import_layout_columns()
        q = ImportDashboardRow.query.filter(
            ImportDashboardRow.source_key.in_(_IMPORT_VISIBLE_SOURCE_KEYS)
        )
        if search:
            terms = [t.strip().lower() for t in re.split(r'[\n,]+', search) if t.strip()]
            for term in terms:
                pattern = f'%{term}%'
                q = q.filter(db.or_(
                    ImportDashboardRow.row_key.ilike(pattern),
                    ImportDashboardRow.source_label.ilike(pattern),
                    ImportDashboardRow.vendor_name.ilike(pattern),
                    ImportDashboardRow.data_json.ilike(pattern),
                ))

        candidate_rows = q.order_by(
            ImportDashboardRow.first_seen_at.desc(),
            ImportDashboardRow.id.desc(),
        ).all()

        # Apply the uploaded-vendor filter to exports too — same safety net
        # as /api/import/data. When a Vendor Import list exists, only rows
        # with matching vendors are exported.
        uploaded_vendor_set_for_export = {v.strip().lower() for v in import_uploaded_vendor_names() if v and v.strip()}

        filtered = []
        for row in candidate_rows:
            try:
                data = json.loads(row.data_json or '{}')
            except (TypeError, json.JSONDecodeError):
                data = {}
            data = apply_import_formula_columns(dict(data))
            yupi = import_nonblank(data.get('yupi_po')) or import_nonblank(data.get('po_yupi'))
            vendor = import_nonblank(data.get('vendor_name')) or import_nonblank(data.get('vendor')) or import_nonblank(row.vendor_name)
            # Uploaded-vendor filter — skip rows whose vendor is not in the
            # uploaded list (or has no vendor at all).
            if uploaded_vendor_set_for_export:
                if not vendor or vendor.strip().lower() not in uploaded_vendor_set_for_export:
                    continue
            if none_yupi_po or none_vendor:
                continue
            if selected_yupi_po and str(yupi or '').strip().lower() not in selected_yupi_po:
                continue
            if selected_vendors and str(vendor or '').strip().lower() not in selected_vendors:
                continue
            # Status filter (case-insensitive, exact match)
            if selected_status_set and str(data.get('status') or '').strip().upper() not in selected_status_set:
                continue
            # Days Left color filter (supports multiple zones — row passes if
            # it matches ANY selected zone)
            if selected_days_left:
                raw_dl = str(data.get('days_left') or '').strip()
                if raw_dl in ('✅', '❌', '', '-'):
                    continue
                try:
                    dl = int(float(raw_dl))
                except (ValueError, TypeError):
                    continue
                matched = False
                for zone in selected_days_left:
                    if zone == 'red' and (dl < 0 or dl <= 7):
                        matched = True; break
                    if zone == 'yellow' and 8 <= dl <= 29:
                        matched = True; break
                    if zone == 'green' and dl >= 30:
                        matched = True; break
                    if zone == 'today' and dl == 0:
                        matched = True; break
                if not matched:
                    continue
            filtered.append((row, data))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Import Dashboard'

        visible_cols = [col for col in columns if not col.get('source_only')]
        header_labels = [col.get('label', col.get('field', '')).replace('\n', ' ') for col in visible_cols]
        ws.append(header_labels)

        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.row_dimensions[1].height = 32

        for i, col in enumerate(visible_cols, 1):
            width = min(max((col.get('width', 120)) / 7, 8), 50)
            ws.column_dimensions[get_column_letter(i)].width = width

        alt_fill = PatternFill(start_color='F0F4FF', end_color='F0F4FF', fill_type='solid')
        # Build vendor-attr map once for the export so Origin/TOP/Non SKI can be
        # injected into each row by matching the row's vendor name.
        _export_vendor_attrs = import_vendor_attrs_map()
        # FIX V10: Carry-over field yang di-merge di dashboard supaya di Excel
        # semua row di group dapat nilai yang sama.
        #
        # Frontend merge berdasarkan yupi_po + req_dlv_date sama (IMPORT_MERGE_KEY_FIELDS).
        # Field yang di-merge (bukan group_per_item): status, days_left, vendor,
        # po_send_date, etd, eta, dll — semua field EXCEPT yang group_per_item=True.
        # Saat user edit cell di row pertama group, hanya row pertama yang di-update
        # di DB. Row lain di group tetap kosong. Saat export, kita harus fill-down
        # nilai dari row pertama ke semua row di group yang sama.
        #
        # Logika:
        # 1. Group rows by (yupi_po + req_dlv_date) — sama seperti frontend
        # 2. Untuk setiap group, cari nilai non-empty untuk setiap field dari row manapun
        # 3. Apply nilai tersebut ke semua row di group
        # Field yang TIDAK di-carry-over: yang group_per_item=True (item_name, ord_qty,
        # unit_price, amount, spec, dll — ini unique per item)
        GROUP_PER_ITEM_FIELDS = {'item_name', 'item_yupi', 'spec', 'remark_yupi', 'ord_qty',
                                 'unit', 'unit_price', 'amount', 'purchase_price', 'purchase_amount',
                                 'currency', 'lt_days', 'reschedule', 'req_dlv_date', 'source_req_dlv_date',
                                 'po_sementara', 'days_left', 'arrival_check', 'po_date_by_email',
                                 'so', 'soft_copy_doc', 'bl_number', 'inv_no', 'incoterm', 'forwarder',
                                 'sap_input', 'bl_awb', 'invoice', 'pl', 'hc', 'msds', 'coa', 'coo',
                                 'payment', 'payment_date', 'import_remarks', 'etd', 'eta'}
        CARRYOVER_FIELDS = [col.get('field') for col in visible_cols
                           if col.get('field') and col.get('field') not in GROUP_PER_ITEM_FIELDS]

        # Group rows by (yupi_po + req_dlv_date)
        groups = []
        current_group = []
        current_key = None
        for db_row, data in filtered:
            yupi = (clean(data.get('yupi_po')) or clean(data.get('po_yupi')) or '').strip()
            req_dlv = clean(data.get('req_dlv_date')).strip()
            group_key = f"{yupi}|{req_dlv}"
            # Hanya group kalau yupi_po tidak kosong
            if not yupi:
                # Row tanpa yupi_po = group sendiri
                if current_group:
                    groups.append(current_group)
                groups.append([(db_row, data)])
                current_group = []
                current_key = None
                continue
            if current_key is None or group_key != current_key:
                if current_group:
                    groups.append(current_group)
                current_group = [(db_row, data)]
                current_key = group_key
            else:
                current_group.append((db_row, data))
        if current_group:
            groups.append(current_group)

        # Untuk setiap group, collect nilai non-empty per field, lalu apply ke semua row
        processed_rows = []
        for group in groups:
            if len(group) == 1:
                # Single row group, tidak perlu carry-over
                processed_rows.append(group[0])
                continue
            # Collect nilai non-empty per field dari semua row di group
            group_values = {}
            for db_row, data in group:
                for field in CARRYOVER_FIELDS:
                    val = data.get(field)
                    if not import_blankish(val) and field not in group_values:
                        group_values[field] = val
            # Apply nilai ke semua row di group
            for db_row, data in group:
                for field in CARRYOVER_FIELDS:
                    if import_blankish(data.get(field)) and field in group_values:
                        data[field] = group_values[field]
                processed_rows.append((db_row, data))

        for row_idx, (db_row, data) in enumerate(processed_rows, 2):
            # Inject vendor attributes (origin, top, non_ski) if missing in row data.
            row_vendor = import_nonblank(data.get('vendor')) or import_nonblank(data.get('vendor_name')) or import_nonblank(db_row.vendor_name)
            if row_vendor:
                attrs = _export_vendor_attrs.get(str(row_vendor).strip().lower())
                if attrs:
                    if not import_nonblank(data.get('origin')):
                        data['origin'] = attrs.get('origin') or ''
                    if not import_nonblank(data.get('top')):
                        data['top'] = attrs.get('top') or ''
                    if not import_nonblank(data.get('non_ski')):
                        data['non_ski'] = attrs.get('non_ski') or ''
            row_vals = []
            for col in visible_cols:
                field = col.get('field', '')
                val = data.get(field, '')
                if col.get('checkbox'):
                    val = 'YES' if import_truthy_checkbox_value(val) else ''
                row_vals.append(val if val is not None else '')
            ws.append(row_vals)
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = alt_fill

        ws.freeze_panes = 'A2'
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"Import_Dashboard_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/import/cells', methods=['PUT'])
def update_import_cells_batch():
    try:
        payload = request.get_json(silent=True) or {}
        updates = payload.get('updates') or []
        if not isinstance(updates, list) or not updates:
            return jsonify({'error': 'updates must be a non-empty list'}), 400
        columns = import_layout_columns()
        valid_fields = {col['field'] for col in import_all_mapping_columns(columns)}
        row_keys = [clean(item.get('row_key')) for item in updates if clean(item.get('row_key'))]
        rows = ImportDashboardRow.query.filter(ImportDashboardRow.row_key.in_(row_keys)).all() if row_keys else []
        row_by_key = {r.row_key: r for r in rows}
        sheet_items = []
        updated_keys = set()
        for item in updates:
            row_key = clean(item.get('row_key'))
            field = clean(item.get('field'))
            value = '' if item.get('value') is None else str(item.get('value'))
            if not row_key or not field or field not in valid_fields:
                continue
            row = row_by_key.get(row_key)
            if not row:
                continue
            try:
                data = json.loads(row.data_json or '{}')
            except (TypeError, json.JSONDecodeError):
                data = {}
            data = set_import_payload_field_aliases(data, field, value)
            data = apply_import_formula_columns(data)
            row.data_json = json.dumps(data, ensure_ascii=False)
            row.updated_at = datetime.utcnow()
            sheet_items.append({'row': row, 'field': field, 'value': value})
            if field in ('purchase_price', 'ord_qty') and clean(data.get('purchase_amount')):
                sheet_items.append({'row': row, 'field': 'purchase_amount', 'value': data.get('purchase_amount')})
            updated_keys.add(row_key)
        sheet_sync = {'synced': False, 'reason': 'No mapped Import sheet cells to sync'}
        if sheet_items:
            try:
                sheet_sync = sync_import_cells_to_google_sheet(sheet_items)
            except Exception as sync_exc:
                sheet_sync = {'synced': False, 'reason': str(sync_exc)}
        db.session.commit()
        clear_runtime_caches()
        _v_attrs_map = import_vendor_attrs_map()
        updated_rows = [import_dashboard_row_to_dict(row_by_key[k], columns, vendor_attrs_map=_v_attrs_map) for k in updated_keys if k in row_by_key]
        return jsonify({'success': True, 'updated': len(sheet_items), 'rows': updated_rows, 'sheet_sync': sheet_sync})
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e), 'sheet_sync': {'synced': False, 'reason': str(e)}}), 500

@app.route('/api/import/cleanup', methods=['POST'])
def import_cleanup_duplicates():
    try:
        columns = import_layout_columns()
        all_rows = ImportDashboardRow.query.filter(
            ImportDashboardRow.source_key.in_(_IMPORT_VISIBLE_SOURCE_KEYS)
        ).order_by(ImportDashboardRow.updated_at.desc(), ImportDashboardRow.id.desc()).all()

        groups = {}
        for row in all_rows:
            try:
                data = json.loads(row.data_json or '{}')
            except (TypeError, json.JSONDecodeError):
                data = {}
            po_yupi = import_nonblank(data.get('po_yupi')) or import_nonblank(data.get('yupi_po'))
            item_yupi = import_nonblank(data.get('item_yupi'))
            po_sementara = import_nonblank(data.get('po_sementara'))
            detail_fp = import_row_identity_detail_fingerprint(data) or '(blank)'

            if po_yupi and item_yupi:
                biz_key = f"poyupi:{po_yupi.strip().upper()}|item:{item_yupi.strip().upper()}"
            elif po_yupi:
                biz_key = f"poyupi:{po_yupi.strip().upper()}|item:(none)|detail:{detail_fp}"
            elif po_sementara and item_yupi:
                biz_key = f"posem:{po_sementara.strip().upper()}|item:{item_yupi.strip().upper()}"
            elif po_sementara:
                biz_key = f"posem:{po_sementara.strip().upper()}|item:(none)|detail:{detail_fp}"
            else:
                continue  # no key = skip

            if biz_key not in groups:
                groups[biz_key] = []
            groups[biz_key].append(row)

        deleted = 0
        merged = 0
        status_rank = {s: i for i, s in enumerate(IMPORT_STATUS_OPTIONS)}

        def _status_progress(data):
            return status_rank.get(str(data.get('status') or '').strip().upper(), -1)

        for biz_key, rows in groups.items():
            if len(rows) <= 1:
                continue

            def _row_data(r):
                try:
                    return json.loads(r.data_json or '{}')
                except Exception:
                    return {}

            rows_with_data = [(r, _row_data(r)) for r in rows]
            rows_with_data.sort(
                key=lambda rd: (_status_progress(rd[1]), rd[0].updated_at or datetime.min, rd[0].id),
                reverse=True,
            )
            winner, winner_data = rows_with_data[0]
            duplicates = [rd[0] for rd in rows_with_data[1:]]

            for dup in duplicates:
                dup_data = _row_data(dup)
                for field in IMPORT_LOCAL_EDIT_FIELDS:
                    if field in IMPORT_SOURCE_MANAGED_FIELDS:
                        continue
                    if field == 'status':
                        continue  # status already chosen via _status_progress above
                    if import_blankish(winner_data.get(field)) and not import_blankish(dup_data.get(field)):
                        winner_data[field] = dup_data[field]

            winner.data_json = json.dumps(apply_import_formula_columns(winner_data), ensure_ascii=False)
            winner.updated_at = datetime.utcnow()
            merged += 1

            for dup in duplicates:
                db.session.delete(dup)
                deleted += 1

        db.session.commit()
        clear_runtime_caches()
        return jsonify({
            'success': True,
            'deleted': deleted,
            'merged': merged,
            'message': f'{deleted} baris duplikat dihapus, {merged} baris dipertahankan dengan data tergabung.',
        })
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/import/vendors', methods=['GET'])
def list_import_vendors():
    """List semua vendor yang sudah di-upload ke ImportVendor table.

    FIX V9: Endpoint GET untuk list vendor (sebelumnya hanya POST upload).
    Return: daftar vendor dengan origin, top, non_ski, uploaded_at.
    """
    try:
        vendors = ImportVendor.query.order_by(ImportVendor.vendor_name.asc()).all()
        return jsonify({
            'count': len(vendors),
            'vendors': [
                {
                    'vendor_name': v.vendor_name,
                    'origin': v.origin or '',
                    'top': v.top or '',
                    'non_ski': v.non_ski or '',
                    'uploaded_at': utc_isoformat(v.uploaded_at) if v.uploaded_at else None,
                }
                for v in vendors
            ],
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/import/vendor-template', methods=['GET'])
def download_import_vendor_template():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Import Vendors'
        # Template now has 4 columns: Vendor Name, Origin, TOP, Non SKI.
        # Pre-populated with existing vendor records (including their last
        # saved origin/top/non_ski values) so the user can edit in place
        # rather than re-entering everything.
        ws.append(['Vendor Name', 'Origin', 'TOP', 'Non SKI'])
        existing = db.session.query(ImportVendor).order_by(ImportVendor.vendor_name.asc()).all()
        if existing:
            for v in existing:
                ws.append([v.vendor_name or '', v.origin or '', v.top or '', v.non_ski or ''])
        else:
            # Fall back to default vendor list (no attributes yet).
            for vendor in import_vendor_names():
                ws.append([vendor, '', '', ''])
        # Column widths tuned for content.
        ws.column_dimensions['A'].width = 42
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        # Style header row (bold white on blue).
        for col_letter in ('A', 'B', 'C', 'D'):
            cell = ws[f'{col_letter}1']
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='2563EB')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='Import_Vendor_Template.xlsx')
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/import/vendors/upload', methods=['POST'])
def upload_import_vendors():
    try:
        files = request.files.getlist('file') or request.files.getlist('files')
        if not files:
            return jsonify({'error': 'No file uploaded'}), 400
        # Collect vendor records as dicts so we can preserve origin/top/non_ski.
        # Last definition wins on duplicate vendor names (case-insensitive).
        vendors = {}
        for file in files:
            name = (file.filename or '').lower()
            if name.endswith('.csv'):
                df = pd.read_csv(file, dtype=str, keep_default_na=False)
            else:
                df = pd.read_excel(file, dtype=str, keep_default_na=False)
            if df.empty:
                continue
            df.columns = [str(c).strip() for c in df.columns]
            # Identify the vendor column (case-insensitive header match).
            vendor_col = next((c for c in df.columns if str(c).strip().lower() in ('vendor name', 'vendor', 'vendor_name')), df.columns[0] if len(df.columns) else None)
            if not vendor_col:
                continue
            # Identify optional Origin / TOP / Non SKI columns by header.
            origin_col = next((c for c in df.columns if str(c).strip().lower() in ('origin', 'orig', 'negara')), None)
            top_col = next((c for c in df.columns if str(c).strip().lower() in ('top', 'term of payment', 'term_of_payment', 'payment term')), None)
            non_ski_col = next((c for c in df.columns if str(c).strip().lower() in ('non ski', 'non-ski', 'nonski', 'non_ski')), None)
            for _, row in df.iterrows():
                vendor = clean(row.get(vendor_col))
                if not vendor or vendor.lower() in ('vendor', 'vendor name'):
                    continue
                vendors[vendor.lower()] = {
                    'vendor_name': vendor,
                    'origin': clean(row.get(origin_col)) if origin_col else '',
                    'top': clean(row.get(top_col)) if top_col else '',
                    'non_ski': clean(row.get(non_ski_col)) if non_ski_col else '',
                }
        ImportVendor.query.delete()
        now = datetime.utcnow()
        for key in sorted(vendors.keys()):
            rec = vendors[key]
            db.session.add(ImportVendor(
                vendor_name=rec['vendor_name'],
                origin=rec.get('origin') or '',
                top=rec.get('top') or '',
                non_ski=rec.get('non_ski') or '',
                uploaded_at=now,
            ))
        db.session.commit()
        clear_runtime_caches()
        return jsonify({
            'success': True,
            'count': len(vendors),
            'message': f'Import vendor list updated: {len(vendors)} vendors',
        })
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

def base_all_registered_items_query():
    return db.session.query(ProductIDDB).filter(db.or_(
        ProductIDDB.product_status.is_(None),
        ProductIDDB.product_status == '',
        func.lower(ProductIDDB.product_status) == 'use'
    ))


def apply_all_registered_items_filters(q, args, exclude_fields=None):
    exclude_fields = set(exclude_fields or [])
    search = (args.get('search', '') or '').strip()
    prod_ids = [clean_product_id(p) for p in args.getlist('prod_id') if clean_product_id(p)]
    date_filter = args.get('date_filter', 'all')
    date_from = args.get('date_from', '')
    date_to = args.get('date_to', '')
    # NOTE: `pic_name` is intentionally NOT applied here as a SQL filter.
    # The "PIC" shown in the table is the RESOLVED pic (looked up from
    # MasterPIC by category_id or category_name), not the raw
    # `product_registry_pic` column. Filtering by SQL on the raw column would
    # give wrong results. PIC filtering is done in Python after resolution
    # (see get_all_registered_items / export_all_registered_items). To support
    # `exclude_fields={'pic_name'}` semantics for the option-builder queries,
    # we simply accept the key here as a no-op.
    # pic_name = (args.get('pic_name', '') or '').strip()
    mfr_names = [clean(v) for v in args.getlist('mfr_name') if clean(v)]
    # NOTE: vendor_name filter removed — the source Excel has no Vendor column
    # for product master data, so the filter would never match anything.
    # vendor_names = [clean(v) for v in args.getlist('vendor_name') if clean(v)]

    if date_filter != 'all':
        today = datetime.now().date()
        if date_filter == 'today':
            q = q.filter(func.date(ProductIDDB.registration_date) == today)
        elif date_filter == 'week':
            week_start = today - timedelta(days=today.weekday())
            q = q.filter(ProductIDDB.registration_date >= week_start)
        elif date_filter == 'month':
            q = q.filter(ProductIDDB.registration_date >= today.replace(day=1))
        elif date_filter == 'year':
            q = q.filter(ProductIDDB.registration_date >= today.replace(month=1, day=1))
        elif date_filter == 'custom':
            if date_from:
                q = q.filter(ProductIDDB.registration_date >= date_from)
            if date_to:
                q = q.filter(ProductIDDB.registration_date <= date_to)

    if prod_ids:
        q = q.filter(ProductIDDB.product_id.in_(prod_ids))
    if 'mfr_name' not in exclude_fields and mfr_names:
        q = q.filter(ProductIDDB.manufacturer_name.in_(mfr_names))

    if search:
        terms = rfq_multiline_search_terms(search)
        term_filters = []
        for term in terms:
            pattern = f'%{term}%'
            term_filters.append(db.or_(
                ProductIDDB.product_id.ilike(pattern),
                ProductIDDB.product_name.ilike(pattern),
                ProductIDDB.specification.ilike(pattern),
                ProductIDDB.manufacturer_name.ilike(pattern),
                ProductIDDB.category_name.ilike(pattern),
            ))
        if term_filters:
            q = q.filter(db.or_(*term_filters))
    return q


def serialize_registered_product(row, pic_map=None, pic_by_name=None):
    """Serialize a ProductIDDB row for the All Registered Items API.

    PIC resolution mirrors Item Registration / Pending Delivery:
    1. Try category_id lookup (MasterPIC.category_id == row.category_id)
    2. Fall back to category_name lookup (MasterPIC.category_name == source_category_level1(row.category_name))
    This double lookup is required because some ProductIDDB rows don't have a
    category_id that matches any MasterPIC entry, but their category_name
    still matches a MasterPIC row.

    NOTE: vendor_name is intentionally NOT serialized — the source Excel does
    not contain a Vendor column for product master data, so the field would
    always be empty. The frontend no longer shows a Vendor column either.
    """
    pic_map = pic_map or {}
    pic_by_name = pic_by_name or {}
    cat_id = normalize_category_id(row.category_id)
    # Try by ID first, then by category name (level-1 only).
    pic = pic_map.get(cat_id)
    if not pic:
        cat_name_norm = normalize_category_name(row.category_name)
        if cat_name_norm:
            pic = pic_by_name.get(cat_name_norm)
    return {
        'id': row.id,
        'prod_id': clean_product_id(row.product_id),
        'category_id': cat_id or '',
        'category': source_category_level1(row.category_name),
        'pic': pic or '',
        'prod_name': row.product_name or '',
        'spec': row.specification or '',
        'mfr_name': row.manufacturer_name or '',
        'odr_unit': row.order_unit or '',
        'hub_handling_check': row.hub_handling_check or '',
        'tax_type': row.tax_type or '',
        'registration_date': row.registration_date.isoformat() if row.registration_date else '',
        'product_registry_pic': row.product_registry_pic or '',
        'client_name': '',
        'req_no': '',
        'proc_status': row.product_status or '',
        'prod_price': 0,
        'curr': '',
        'batch_grp_no': '',
    }


@app.route('/api/all-registered-items', methods=['GET'])
def get_all_registered_items():
    try:
        cache_key = runtime_cache_key('all_registered_items')
        cached = runtime_cache_get(cache_key)
        if cached is not None:
            return jsonify(cached)

        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 10))
        q = apply_all_registered_items_filters(base_all_registered_items_query(), request.args)
        total = q.count()

        # Build BOTH PIC lookup maps (by category_id and by normalized category
        # name) so we can resolve PIC for rows whose category_id doesn't have
        # a direct MasterPIC match but whose category_name does. Mirrors
        # `_lookup_pic_by_category` used by Item Registration.
        by_id, by_name = master_pic_maps()
        pic_map = by_id
        pic_by_name = by_name

        rows = q.order_by(ProductIDDB.registration_date.desc(), ProductIDDB.product_id.asc()).offset((page-1)*per_page).limit(per_page).all()
        data = [serialize_registered_product(row, pic_map, pic_by_name) for row in rows]

        # If user filtered by PIC, we must filter the *resolved* PIC (not the
        # raw product_registry_pic column). The resolver runs in Python, so we
        # need to re-resolve every candidate row and keep only matching ones.
        # This is a paginated re-query: fetch a larger candidate pool, filter,
        # then slice — acceptable because the table size is modest and the
        # filter is rare.
        pic_filter = (request.args.get('pic_name', '') or '').strip()
        if pic_filter:
            # Re-query WITHOUT pagination to get the full filtered set, then
            # paginate in Python so the resolved PIC filter actually narrows
            # the result instead of just slicing an unfiltered page.
            all_rows = q.order_by(ProductIDDB.registration_date.desc(), ProductIDDB.product_id.asc()).all()
            all_data = [serialize_registered_product(r, pic_map, pic_by_name) for r in all_rows]
            data = [d for d in all_data if d.get('pic') == pic_filter]
            total = len(data)
            start = (page - 1) * per_page
            data = data[start:start + per_page]

        # Build PIC options for the dropdown — collect distinct resolved PICs
        # across the (un-PIC-filtered) full set. Cheap because we already have
        # by_id / by_name maps; we just need to enumerate MasterPIC PIC names
        # that actually appear in the current filtered data.
        pic_option_q = apply_all_registered_items_filters(base_all_registered_items_query(), request.args, exclude_fields={'pic_name'})
        # Resolve PIC for every row in the option query (no pagination). Use
        # Python-side resolution since PIC is derived, not stored.
        option_rows = pic_option_q.all()
        pic_set = set()
        for r in option_rows:
            resolved = serialize_registered_product(r, pic_map, pic_by_name).get('pic') or ''
            if resolved:
                pic_set.add(resolved)
        pic_options = sorted(pic_set)

        mfr_option_q = apply_all_registered_items_filters(base_all_registered_items_query(), request.args, exclude_fields={'mfr_name'})
        # NOTE: vendor_names option removed — source has no Vendor column.
        payload = {
            'data': data,
            'total': total,
            'page': page,
            'per_page': per_page,
            'filters': {
                'mfr_names': sorted([r[0] for r in mfr_option_q.with_entities(ProductIDDB.manufacturer_name).distinct().all() if r[0]]),
                'pic_options': pic_options,
            }
        }
        runtime_cache_set(cache_key, payload, ttl_seconds=60)
        return jsonify(payload)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/all-registered-items', methods=['GET'])
def export_all_registered_items():
    try:
        q = apply_all_registered_items_filters(base_all_registered_items_query(), request.args)
        by_id, by_name = master_pic_maps()
        rows = q.order_by(ProductIDDB.registration_date.desc(), ProductIDDB.product_id.asc()).all()
        # If user filtered by PIC, narrow rows down by resolved PIC value
        # (the resolved PIC isn't a DB column, so we filter in Python).
        pic_filter = (request.args.get('pic_name', '') or '').strip()
        wb = Workbook()
        ws = wb.active
        ws.title = 'All Registered Items'
        headers = ['Product ID', 'Category ID', 'Category', 'PIC', 'Product Name', 'Specification', 'Manufacturer Name', 'Order Unit', 'Hub Handling Check', 'Tax Type', 'Registration Date', 'Registry PIC', 'Status']
        ws.append(headers)
        header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        for row in rows:
            item = serialize_registered_product(row, by_id, by_name)
            if pic_filter and (item.get('pic') or '') != pic_filter:
                continue
            ws.append([
                item['prod_id'], item['category_id'], item['category'], item['pic'], item['prod_name'], item['spec'],
                item['mfr_name'], item['odr_unit'], item['hub_handling_check'],
                item['tax_type'], item['registration_date'], item['product_registry_pic'], item['proc_status']
            ])
        widths = [18, 22, 28, 18, 35, 45, 28, 14, 20, 14, 18, 24, 16]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f"All_Registered_Items_{datetime.now().strftime('%Y%m%d')}.xlsx")
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500



@app.route('/api/dashboard/status-detail', methods=['GET'])
def get_dashboard_status_detail():
    try:
        status = request.args.get('status', '').strip()
        month = request.args.get('month', '').strip()
        hidden_so = get_hidden_so_items()
        date_year, date_from, date_to = parse_so_date_args()
        clients = selected_clients()
        pics = selected_pics()
        is_sqlite = 'sqlite' in app.config.get('SQLALCHEMY_DATABASE_URI', '')
        month_expr = func.strftime('%Y-%m', SOData.so_create_date) if is_sqlite else func.to_char(func.date_trunc('month', SOData.so_create_date), 'YYYY-MM')

        def so_q(*extra_filters):
            q = db.session.query(SOData).filter(*extra_filters) if extra_filters else db.session.query(SOData)
            q = apply_so_client_filter(q, clients)
            q = apply_so_pic_filter(q, pics)
            return apply_so_create_date_filter(q, date_year, date_from, date_to)

        if status:
            q = so_q(open_so_filter(), so_countable_sql_filter())
            if status == 'Unknown':
                # 'Unknown' is a display fallback (see status_label in get_dashboard_stats)
                # for rows where so_status is NULL or blank — match that same condition
                # here instead of comparing against the literal string 'Unknown'.
                q = q.filter(func.trim(func.coalesce(SOData.so_status, '')) == '')
            else:
                q = q.filter(func.trim(func.coalesce(SOData.so_status, '')) == status)
        else:
            q = so_q(open_so_filter(), so_countable_sql_filter())
        if month:
            # The frontend sends the raw aggregation key, which is 'YYYY-MM'
            # (see month_expr / status_months in get_dashboard_stats) — NOT a
            # 'Mon YYYY' label. Match that format directly. Keep a fallback
            # parse for older links/bookmarks that may still use 'Mon YYYY'
            # or 'Mon YY' so this stays backward compatible.
            month_key = None
            if re.match(r'^\d{4}-\d{2}$', month):
                month_key = month
            else:
                for fmt in ('%b %Y', '%b %y'):
                    try:
                        month_key = datetime.strptime(month, fmt).strftime('%Y-%m')
                        break
                    except ValueError:
                        continue
            if month_key:
                q = q.filter(month_expr == month_key)

        rows = q.all()
        result = []
        for s in rows:
            if s.so_item in hidden_so or s.so_number in hidden_so:
                continue
            if not so_is_countable(s.so_item, customer_po_number=s.customer_po_number, delivery_memo=s.delivery_memo):
                continue
            result.append({
                'so_item': s.so_item,
                'so_number': s.so_number,
                'so_status': s.so_status,
                'pic_name': canonical_pending_pic(s.pic_name, s.operation_unit_name),
                'operation_unit_name': s.operation_unit_name,
                'vendor_name': s.vendor_name,
                'product_name': s.product_name,
                'so_qty': s.so_qty,
                'sales_price': s.sales_price,
                'sales_amount': s.sales_amount,
                'customer_po_number': s.customer_po_number,
                'delivery_memo': s.delivery_memo,
                'so_create_date': s.so_create_date.isoformat() if s.so_create_date else None,
                'delivery_plan_date': s.delivery_plan_date.isoformat() if s.delivery_plan_date else None,
                'remarks': s.remarks,
            })
        return jsonify(result)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/data/pic-kpi', methods=['GET'])
def get_pic_kpi():
    try:
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        date_year = request.args.get('date_year', '')
        is_sqlite = 'sqlite' in app.config['SQLALCHEMY_DATABASE_URI']
        
        q = db.session.query(SOData).filter(
            SOData.so_status.notin_(['Delivery Completed', 'SO Cancel'])
        )
        
        if date_year:
            try:
                yr = int(date_year)
                start_date = date(yr, 1, 1)
                end_date = date(yr, 12, 31)
                q = q.filter(SOData.so_create_date >= start_date, SOData.so_create_date <= end_date)
            except ValueError:
                pass
        elif date_from or date_to:
            if date_from:
                q = q.filter(SOData.so_create_date >= date_from)
            if date_to:
                q = q.filter(SOData.so_create_date <= date_to)
        
        rows = q.all()
        
        pic_map = {}
        for s in rows:
            pic = s.pic_name or 'Unassigned'
            if pic not in pic_map:
                pic_map[pic] = {
                    'pic_name': pic,
                    'so_count': 0,
                    'total_sales': 0.0,
                    'total_purchase': 0.0,
                    'total_margin': 0.0,
                    'positive_margin_count': 0,
                    'negative_margin_count': 0,
                }
            
            pic_map[pic]['so_count'] += 1
            sales = float(s.sales_amount or 0)
            # Use IDR-converted purchase amount for margin (handles USD/EUR).
            po_amount = purchase_amount_idr_for_margin(s)
            # FIX V9: Margin valid only when purchase (in IDR) is positive AND sales > 0.
            # Kalau PO kosong/0 atau sales kosong/0 → margin = None, excluded dari KPI.
            purchase_valid = po_amount > 0
            sales_valid = sales > 0
            margin = (sales - po_amount) if (purchase_valid and sales_valid) else None

            pic_map[pic]['total_sales'] += sales
            pic_map[pic]['total_purchase'] += po_amount if purchase_valid else 0
            if margin is not None:
                pic_map[pic]['total_margin'] += margin
                if margin > 0:
                    pic_map[pic]['positive_margin_count'] += 1
                elif margin < 0:
                    pic_map[pic]['negative_margin_count'] += 1
        
        result = sorted(pic_map.values(), key=lambda x: x['so_count'], reverse=True)
        
        return jsonify(result)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/template/master-pic', methods=['GET'])
def download_master_pic_template():
    try:
        # ───────────────────────────────────────────────────────────────────
        # Master PIC Update Template — 4 sheets, matching the structure the
        # user uploaded on 2026-06-23:
        #
        #   Sheet 1 "By Category"   : Category Name | PIC
        #                            (gray header — reference data, edit PIC column)
        #   Sheet 2 "By Client ID"  : Client ID | Client Name | PIC
        #                            (gray ref headers on first two, blue input on PIC)
        #   Sheet 3 "By Vendor"     : Vendor ID | Vendor Name | PIC
        #                            (gray ref headers on first two, blue input on PIC)
        #   Sheet 4 "By Bid Type"   : Bid Type | PIC
        #                            (gray header — reference data, edit PIC column)
        #
        # All sheets are pre-populated with the current Master PIC data so
        # the user can edit PIC values in place rather than starting blank.
        # PIC resolution priority (also enforced at upload time):
        #   1. Client ID  (highest)
        #   2. Bid Type
        #   3. Vendor ID
        #   4. Category Name  (lowest)
        # ───────────────────────────────────────────────────────────────────
        wb = Workbook()

        ref_header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        input_header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
        input_font = Font(bold=True, color='FFFFFF', size=10)
        ref_font = Font(bold=True, color='000000', size=10)
        body_font = Font(size=10)
        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='center')

        def style_header(ws, ref_cols, input_cols):
            """Apply header styling to a sheet.
            ref_cols: list of column letters that are reference data (gray header).
            input_cols: list of column letters that are user-input (blue header)."""
            for col_letter in ref_cols + input_cols:
                cell = ws[f'{col_letter}1']
                cell.alignment = center
                if col_letter in ref_cols:
                    cell.fill = ref_header_fill
                    cell.font = ref_font
                else:
                    cell.fill = input_header_fill
                    cell.font = input_font

        def style_body(ws, max_col, left_cols=()):
            """Style all body cells (rows 2+). Numeric/text columns get text
            number format so vendor IDs like '443395' don't get auto-converted
            to numbers and lose leading zeros."""
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.font = body_font
                    cell.alignment = left if cell.column_letter in left_cols else center
                    cell.number_format = '@'

        def pad_to_min_rows(ws, min_rows, max_col):
            while ws.max_row < min_rows:
                ws.append([''] * max_col)

        # ── Sheet 1: By Category ───────────────────────────────────────────
        ws1 = wb.active
        ws1.title = 'By Category'
        ws1.append(['Category Name', 'PIC'])
        ws1.freeze_panes = 'A2'
        style_header(ws1, ref_cols=['A', 'B'], input_cols=[])

        category_rows = {}
        for m in db.session.query(MasterPIC).order_by(MasterPIC.category_name).all():
            cat_name = source_category_level1(m.category_name)
            norm = normalize_category_name(cat_name)
            if norm:
                category_rows[norm] = {'category_name': cat_name, 'pic': clean(m.pic_name) or ''}
        # Also include categories that appear in ProductIDDB but don't yet
        # have a MasterPIC entry — gives the user a complete list to fill in.
        for (cat_name_raw,) in db.session.query(ProductIDDB.category_name).filter(
            ProductIDDB.category_name.isnot(None), ProductIDDB.category_name != ''
        ).distinct().all():
            cat_name = source_category_level1(cat_name_raw)
            norm = normalize_category_name(cat_name)
            if norm and norm not in category_rows:
                category_rows[norm] = {'category_name': cat_name, 'pic': _lookup_pic_by_category(None, cat_name) or ''}

        for item in sorted(category_rows.values(), key=lambda x: x['category_name'].lower()):
            ws1.append([item['category_name'], item['pic']])
        pad_to_min_rows(ws1, max(20, len(category_rows) + 5), 2)
        style_body(ws1, max_col=2, left_cols=('A',))
        ws1.column_dimensions['A'].width = 32
        ws1.column_dimensions['B'].width = 12
        ws1.auto_filter.ref = f'A1:B{ws1.max_row}'

        # ── Sheet 2: By Client ID ──────────────────────────────────────────
        ws2 = wb.create_sheet('By Client ID')
        ws2.append(['Client ID', 'Client Name', 'PIC'])
        ws2.freeze_panes = 'A2'
        style_header(ws2, ref_cols=['A', 'B'], input_cols=['C'])

        client_pics = db.session.query(MasterClientPIC).order_by(MasterClientPIC.client_id).all()
        for m in client_pics:
            cid = clean(m.client_id) or ''
            cname = clean(m.client_name) or ''
            pic = clean(m.pic_name) or ''
            if cid or cname:
                ws2.append([cid, cname, pic])
        pad_to_min_rows(ws2, 20, 3)
        style_body(ws2, max_col=3, left_cols=('A', 'B'))
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 40
        ws2.column_dimensions['C'].width = 12
        ws2.auto_filter.ref = f'A1:C{ws2.max_row}'

        # ── Sheet 3: By Vendor ─────────────────────────────────────────────
        ws3 = wb.create_sheet('By Vendor')
        ws3.append(['Vendor ID', 'Vendor Name', 'PIC'])
        ws3.freeze_panes = 'A2'
        style_header(ws3, ref_cols=['A', 'B'], input_cols=['C'])

        vendor_pics = db.session.query(MasterVendorPIC).order_by(MasterVendorPIC.vendor_id).all()
        for m in vendor_pics:
            vid = normalize_vendor_id(m.vendor_id) or ''  # FIX: normalize strips .0 suffix
            vname = clean(m.vendor_name) or ''
            pic = clean(m.pic_name) or ''
            if vid or vname:
                ws3.append([vid, vname, pic])
        # Also include vendors already referenced in SO_data that don't yet
        # have a MasterVendorPIC entry — gives the user a complete list.
        existing_vendor_ids = {normalize_vendor_id(m.vendor_id) for m in vendor_pics if normalize_vendor_id(m.vendor_id)}  # FIX: normalize
        so_vendors = db.session.query(SOData.vendor_id, SOData.vendor_name).filter(
            SOData.vendor_id.isnot(None), SOData.vendor_id != ''
        ).distinct().all()
        for vid_raw, vname in so_vendors:
            vid = normalize_vendor_id(vid_raw) or ''  # FIX: normalize strips .0 suffix, prevents duplicates
            if vid and vid not in existing_vendor_ids:
                ws3.append([vid, clean(vname) or '', ''])
                existing_vendor_ids.add(vid)
        pad_to_min_rows(ws3, 20, 3)
        style_body(ws3, max_col=3, left_cols=('A', 'B'))
        ws3.column_dimensions['A'].width = 20
        ws3.column_dimensions['B'].width = 40
        ws3.column_dimensions['C'].width = 12
        ws3.auto_filter.ref = f'A1:C{ws3.max_row}'

        # ── Sheet 4: By Bid Type ───────────────────────────────────────────
        ws4 = wb.create_sheet('By Bid Type')
        ws4.append(['Bid Type', 'PIC'])
        ws4.freeze_panes = 'A2'
        style_header(ws4, ref_cols=['A', 'B'], input_cols=[])

        bid_type_pics = db.session.query(MasterBidTypePIC).order_by(MasterBidTypePIC.bid_type).all()
        for m in bid_type_pics:
            bt = clean(m.bid_type) or ''
            pic = clean(m.pic_name) or ''
            if bt:
                ws4.append([bt, pic])
        pad_to_min_rows(ws4, 20, 2)
        style_body(ws4, max_col=2, left_cols=('A',))
        ws4.column_dimensions['A'].width = 40
        ws4.column_dimensions['B'].width = 12
        ws4.auto_filter.ref = f'A1:B{ws4.max_row}'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Master_PIC_Update_Template.xlsx'
        )
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════
#  ADMIN: ONE-TIME CLEANUP
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/api/admin/fix-vendor-pic-ids', methods=['POST'])
def fix_vendor_pic_ids():
    """One-time cleanup: normalize vendor_id di MasterVendorPIC yang ada suffix .0
    Contoh: "443395.0" → "443395". Duplikat setelah normalisasi dihapus (yang lebih
    lama dihapus, yang lebih baru dipertahankan).
    Panggil sekali via: POST /api/admin/fix-vendor-pic-ids
    """
    try:
        fixed = 0
        dupes = 0
        all_rows = db.session.query(MasterVendorPIC).order_by(MasterVendorPIC.updated_at.desc()).all()
        seen = {}
        to_delete = []
        for row in all_rows:
            norm = normalize_vendor_id(row.vendor_id)
            if not norm:
                continue
            if norm in seen:
                # duplikat setelah normalisasi → tandai untuk dihapus (yang lebih lama)
                to_delete.append(row)
                dupes += 1
            else:
                if row.vendor_id != norm:
                    row.vendor_id = norm
                    row.updated_at = datetime.utcnow()
                    fixed += 1
                seen[norm] = row
        for row in to_delete:
            db.session.delete(row)
        db.session.commit()
        invalidate_master_pic_cache()
        clear_runtime_caches()
        return jsonify({
            'ok': True,
            'fixed': fixed,
            'duplicates_removed': dupes,
            'message': f'Normalized {fixed} vendor_id(s), removed {dupes} duplicate(s).'
        })
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════
#  DELIVERY MONITORING
# ═══════════════════════════════════════════════════════════════════════════

DLV_PROCESS_STAGES = [
    ('so_erp_create_date', 'SO(ERP) Create Date', 'SO ERP Created'),
    ('po_create_date',     'PO Create Date',      'PO Created'),
    ('po_rcvd_date',       'PO Rcvd. Date',       'PO Received'),
    ('ship_odr_date',      'Ship. Odr. Date',     'Shipping Order'),
    ('ship_compl_date',    'Ship. Compl. Date',   'Shipping Confirmed'),
    ('hub_rcv_date',       'HUB Rcv. Date',       'HUB Received'),
    ('hub_ship_date',      'HUB Ship. Date',      'HUB Shipped'),
    ('dlv_compl_date',     'Dlv. Compl. Date',    'Delivery Completed'),
]

DLV_LOCAL_PROCESS_STAGES = [
    ('so_erp_create_date', 'SO(ERP) Create Date', 'SO ERP Created'),
    ('po_create_date',     'PO Create Date',      'PO Created'),
    ('po_rcvd_date',       'PO Rcvd. Date',       'PO Received'),
    ('ship_odr_date',      'Ship. Odr. Date',     'Shipping Order'),
    ('ship_compl_date',    'Ship. Compl. Date',   'Shipping Confirmed'),
    ('dlv_compl_date',     'Dlv. Compl. Date',    'Delivery Completed'),
]

def _is_import_delivery(row):
    pur_curr = (getattr(row, 'pur_curr', None) or 'IDR').strip().upper()
    return bool(pur_curr and pur_curr != 'IDR')

def _delivery_stage_flow(row):
    return DLV_PROCESS_STAGES if _is_import_delivery(row) else DLV_LOCAL_PROCESS_STAGES

def _delivery_stage_pairs_for_row(row):
    stages = _delivery_stage_flow(row)
    return [
        (stages[i][0], stages[i + 1][0], stages[i][2], stages[i + 1][2])
        for i in range(len(stages) - 1)
    ]

def _delivery_stage_pairs_all():
    return [
        ('so_erp_create_date', 'po_create_date', 'SO ERP Created', 'PO Created'),
        ('po_create_date', 'po_rcvd_date', 'PO Created', 'PO Received'),
        ('po_rcvd_date', 'ship_odr_date', 'PO Received', 'Shipping Order'),
        ('ship_odr_date', 'ship_compl_date', 'Shipping Order', 'Shipping Confirmed'),
        ('ship_compl_date', 'dlv_compl_date', 'Shipping Confirmed', 'Delivery Completed'),
        ('ship_compl_date', 'hub_rcv_date', 'Shipping Confirmed', 'HUB Received'),
        ('hub_rcv_date', 'hub_ship_date', 'HUB Received', 'HUB Shipped'),
        ('hub_ship_date', 'dlv_compl_date', 'HUB Shipped', 'Delivery Completed'),
    ]

def _parse_dt(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (datetime, date)):
        return val
    try:
        import pandas as _pd
        ts = _pd.to_datetime(val, errors='coerce')
        if _pd.isna(ts):
            return None
        return ts.to_pydatetime()
    except Exception:
        return None

def _dt_to_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None

def _calc_stage_leadtimes(row):
    if row.po_status and 'cancel' in row.po_status.lower():
        return []

    results = []
    for field_from, field_to, label_from, label_to in _delivery_stage_pairs_for_row(row):
        dt_from = getattr(row, field_from, None)
        dt_to   = getattr(row, field_to,   None)
        d_from  = _dt_to_date(dt_from) if dt_from else None
        d_to    = _dt_to_date(dt_to)   if dt_to   else None

        if d_from is None:
            continue

        if d_to is None:
            today = date.today()
            wdays = count_workdays(d_from, today)
            results.append({
                'stage_from': field_from,
                'stage_to':   field_to,
                'label_from': label_from,
                'label_to':   label_to,
                'workdays':   wdays,
                'pending':    True,
            })
        else:
            wdays = count_workdays(d_from, d_to)
            results.append({
                'stage_from': field_from,
                'stage_to':   field_to,
                'label_from': label_from,
                'label_to':   label_to,
                'workdays':   wdays,
                'pending':    False,
            })
    return results


_COMPLETED_WARMUP_STARTED = False
_RFQ_WARMUP_STARTED = False
_DASHBOARD_STATS_WARMUP_STARTED = False


def warm_dashboard_stats_cache_async():
    global _DASHBOARD_STATS_WARMUP_STARTED
    if _DASHBOARD_STATS_WARMUP_STARTED or os.environ.get('PO_MONITOR_DISABLE_WARMUP') == '1':
        return
    _DASHBOARD_STATS_WARMUP_STARTED = True

    def _worker():
        try:
            time.sleep(20)
            with app.app_context():
                client = app.test_client()
                client.get('/api/dashboard/stats')
                client.get('/api/data/aging')
        except Exception as exc:
            print(f'Dashboard stats warmup skipped: {exc}')

    threading.Thread(target=_worker, daemon=True, name='dashboard-stats-warmup').start()


def warm_completed_summary_cache_async():
    global _COMPLETED_WARMUP_STARTED
    if _COMPLETED_WARMUP_STARTED or os.environ.get('PO_MONITOR_DISABLE_WARMUP') == '1':
        return
    _COMPLETED_WARMUP_STARTED = True

    def _worker():
        try:
            time.sleep(30)
            current_year = datetime.utcnow().year
            urls = [
                '/api/completed/summary?mode=dashboard',
                f'/api/completed/summary?date_year={current_year}&yoy_base_year={current_year}&mode=dashboard',
            ]
            with app.app_context():
                client = app.test_client()
                for url in urls:
                    client.get(url)
        except Exception as exc:
            print(f'Completed summary warmup skipped: {exc}')

    threading.Thread(target=_worker, daemon=True, name='completed-summary-warmup').start()

def warm_rfq_dashboard_cache_async():
    global _RFQ_WARMUP_STARTED
    if _RFQ_WARMUP_STARTED or os.environ.get('PO_MONITOR_DISABLE_WARMUP') == '1':
        return
    _RFQ_WARMUP_STARTED = True

    def _worker():
        try:
            time.sleep(8)
            with app.app_context():
                if RFQDashboardRow.query.count() == 0:
                    return
                rows, fetched_at = load_rfq_dashboard_rows()
                set_rfq_runtime_rows(rows, fetched_at)
                app.test_client().get('/api/rfq/data?page=1&per_page=10')
        except Exception as exc:
            print(f'RFQ dashboard warmup skipped: {exc}')

    threading.Thread(target=_worker, daemon=True, name='rfq-dashboard-warmup').start()

@app.route('/api/ping', methods=['GET'])
def ping():
    if request.args.get('db') != '1':
        return jsonify({'ok': True, 'db_checked': False, 'ts': datetime.utcnow().isoformat()})
    try:
        total_so = db.session.query(func.count(SOData.id)).scalar() or 0
        return jsonify({'ok': True, 'db_checked': True, 'total_so': total_so, 'ts': datetime.utcnow().isoformat()})
    except Exception as e:
        return jsonify({'ok': False, 'db_checked': True, 'error': str(e)}), 200


@app.route('/api/diagnostics/google-sheets', methods=['GET'])
def diagnostics_google_sheets():
    """Lightweight, secret-safe diagnostic for Google Sheets sync failures.
    Checks: service-account credential presence/parse, token refresh, proxy
    env vars, and one real metadata call against the RFQ sheet — without ever
    echoing the credential, token, or proxy URL contents back to the caller.
    """
    result = {'credential_configured': False, 'token_ok': False, 'proxy_env': {}, 'sheets_api_reachable': False}
    result['proxy_env'] = {
        'HTTPS_PROXY': bool(os.environ.get('HTTPS_PROXY')),
        'HTTP_PROXY': bool(os.environ.get('HTTP_PROXY')),
        'https_proxy': bool(os.environ.get('https_proxy')),
        'http_proxy': bool(os.environ.get('http_proxy')),
    }
    try:
        credentials_info = rfq_sheet_sync_credentials()
        result['credential_configured'] = bool(credentials_info)
        if not credentials_info:
            result['error'] = 'No GOOGLE_SERVICE_ACCOUNT_JSON / GOOGLE_SERVICE_ACCOUNT_FILE configured'
            return jsonify(result), 200
        result['service_account_email'] = credentials_info.get('client_email')
    except Exception as e:
        result['error'] = f'Failed to load/parse credential: {e}'
        return jsonify(result), 200
    try:
        google_sheets_access_token()
        result['token_ok'] = True
    except Exception as e:
        result['error'] = f'Failed to obtain access token: {e}'
        return jsonify(result), 200
    try:
        meta = google_sheets_metadata(RFQ_SHEET_ID)
        result['sheets_api_reachable'] = True
        result['rfq_sheet_title'] = (meta.get('properties') or {}).get('title')
    except Exception as e:
        result['error'] = f'Metadata call to RFQ sheet failed: {e}'
        return jsonify(result), 200
    return jsonify(result), 200


FRONTEND_DIST_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'frontend', 'dist'))

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_frontend(path):
    if path.startswith('api/'):
        return jsonify({'error': 'Not found'}), 404
    if os.path.isdir(FRONTEND_DIST_DIR):
        target = os.path.join(FRONTEND_DIST_DIR, path)
        if path and os.path.isfile(target):
            return send_from_directory(FRONTEND_DIST_DIR, path)
        index_path = os.path.join(FRONTEND_DIST_DIR, 'index.html')
        if os.path.isfile(index_path):
            return send_from_directory(FRONTEND_DIST_DIR, 'index.html')
    return jsonify({'status': 'ok', 'message': 'PO Monitoring API running'}), 200

warm_dashboard_stats_cache_async()
warm_completed_summary_cache_async()
warm_rfq_dashboard_cache_async()

_SCHEDULER_STARTED = False

def _auto_copy_sheet_job():
    try:
        with app.app_context():
            print(f'[scheduler] Auto copy-sheet started at {datetime.utcnow().isoformat()} UTC')
            result = sync_import_sheet_to_dashboard()
            print(f'[scheduler] Auto copy-sheet done: added={result.get("added")}, '
                  f'updated={result.get("updated")}, seen={result.get("seen")}')
    except Exception as exc:
        print(f'[scheduler] Auto copy-sheet error: {exc}')


def start_auto_copy_scheduler():
    global _SCHEDULER_STARTED
    if _SCHEDULER_STARTED:
        return
    if not _APSCHEDULER_AVAILABLE:
        print('[scheduler] APScheduler not available – skipping daily auto copy-sheet.')
        return
    if os.environ.get('PO_MONITOR_DISABLE_SCHEDULER') == '1':
        print('[scheduler] Daily scheduler disabled via PO_MONITOR_DISABLE_SCHEDULER=1')
        return
    try:
        scheduler = BackgroundScheduler(timezone='Asia/Jakarta')
        scheduler.add_job(
            _auto_copy_sheet_job,
            trigger=CronTrigger(hour=7, minute=0, timezone='Asia/Jakarta'),
            id='auto_copy_sheet',
            name='Daily Import Copy Sheet at 07:00 WIB',
            replace_existing=True,
        )
        scheduler.start()
        _SCHEDULER_STARTED = True
        print('[scheduler] Daily auto copy-sheet scheduled at 07:00 WIB (Asia/Jakarta).')
    except Exception as exc:
        print(f'[scheduler] Failed to start scheduler: {exc}')


start_auto_copy_scheduler()


@app.route('/api/import/scheduler-status', methods=['GET'])
def import_scheduler_status():
    return jsonify({
        'scheduler_available': _APSCHEDULER_AVAILABLE,
        'scheduler_started': _SCHEDULER_STARTED,
        'schedule': '07:00 WIB (Asia/Jakarta) daily',
        'disable_env': 'PO_MONITOR_DISABLE_SCHEDULER=1',
        'last_copy_at': import_meta_get('last_copy_at') or '',
    })


if __name__ == '__main__':
    load_similarity_cache()
    warm_dashboard_stats_cache_async()
    warm_completed_summary_cache_async()
    warm_rfq_dashboard_cache_async()
    print("Backend: http://127.0.0.1:5001")
    app.run(debug=True, host='0.0.0.0', port=5001)
